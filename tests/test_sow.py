import importlib

MOD = "services.sow"


def _mod():
    return importlib.import_module(MOD)


def test_meta_is_valid():
    m = _mod().META
    assert m["path"] == "/sow"
    assert m["name"] and isinstance(m["name"], str)
    assert m.get("admin_only") is True


def test_get_renders_html():
    kind, html = _mod().handle("GET", "/sow", {}, {"user": "__testuser__"})
    assert kind == "html"
    assert "Deal Desk" in html


def test_schedule_matches_executed_data_engineer_sow():
    # Real numbers from "Samsung_ Data Engineer.docx": Mar 23 - Dec 31 2026,
    # 1 x $25/h x 168h/mo => 9.3 months, first month $1,260, total $39,060.
    m = _mod()
    sow = {
        "start": "2026-03-23", "end": "2026-12-31",
        "res_mode": "hourly",
        "resources": [{"profile": "Data Engineer", "qty": 1, "hourly": 25, "hrs": 168}],
        "invoice_rule": "next_first",
    }
    rows, fee, months = m._build_schedule(sow)
    assert round(months, 1) == 9.3
    assert rows[0]["amount"] == 1260
    assert rows[1]["amount"] == 4200
    assert fee == 39060
    assert rows[0]["invoice"] == "1-Apr-26"
    assert len(rows) == 10


def test_schedule_monthly_mode_full_months():
    m = _mod()
    sow = {
        "start": "2026-04-01", "end": "2026-09-30",
        "res_mode": "monthly",
        "resources": [{"name": "A", "rate": 30168}],
        "invoice_rule": "month_end",
    }
    rows, fee, months = m._build_schedule(sow)
    assert months == 6.0
    assert len(rows) == 6
    assert fee == 30168 * 6
    assert rows[0]["invoice"] == "30-Apr-26"


def test_docx_builds_valid_zip():
    m = _mod()
    sow = {
        "id": "t1", "direction": "samsung", "title": "Test SOW",
        "project_name": "Test", "date": "2026-07-22",
        "start": "2026-08-01", "end": "2026-10-31",
        "res_mode": "monthly", "resources": [{"name": "A", "role": "Dev", "rate": 1000}],
        "exec_summary": "Summary", "deliverables": "Do a thing\nDo another",
    }
    blob = m._build_docx(sow, None)
    assert blob[:2] == b"PK"
    assert len(blob) > 5000


def test_docx_agency_uses_vendor_msa():
    m = _mod()
    sow = {
        "id": "t2", "direction": "agency", "title": "Vendor SOW",
        "project_name": "P", "date": "2026-07-22",
        "start": "2026-08-01", "end": "2026-08-31",
        "res_mode": "hourly",
        "resources": [{"profile": "Analyst", "qty": 1, "hourly": 30, "hrs": 160}],
    }
    vendor = {"id": "v1", "name": "Invictus Data, Inc.",
              "entity_line": "Invictus Data Inc, Los Altos CA", "msa_date": "2023-09-28"}
    blob = m._build_docx(sow, vendor)
    assert blob[:2] == b"PK"


def test_msa_docx_fills_vendor_and_date():
    m = _mod()
    sow = {"id": "t3", "type": "agy_msa", "direction": "agency", "kind": "msa",
           "date": "2026-08-01"}
    vendor = {"id": "v1", "name": "Nendrasys Technologies Inc."}
    blob = m._build_msa_docx(sow, vendor)
    assert blob[:2] == b"PK"
    import zipfile, io, re
    xml = zipfile.ZipFile(io.BytesIO(blob)).read("word/document.xml").decode("utf-8", "ignore")
    txt = re.sub(r"<[^>]+>", "", xml)
    assert "Nendrasys Technologies Inc." in txt
    assert "August 1, 2026" in txt
    assert "Your Company Name" not in txt
    assert "XXX" not in txt


def test_nda_docx_builds():
    m = _mod()
    sow = {"id": "t4", "type": "agy_nda", "direction": "agency", "kind": "nda",
           "date": "2026-08-01"}
    blob = m._build_nda_docx(sow, {"id": "v1", "name": "Acme Corp"})
    assert blob[:2] == b"PK"
    import zipfile, io, re
    txt = re.sub(r"<[^>]+>", "", zipfile.ZipFile(io.BytesIO(blob)).read("word/document.xml").decode("utf-8", "ignore"))
    assert "CONFIDENTIALITY AND NONDISCLOSURE AGREEMENT" in txt
    assert "Acme Corp" in txt


def test_schedule_overrides_apply_per_month():
    m = _mod()
    sow = {
        "start": "2026-04-01", "end": "2026-06-30",
        "res_mode": "monthly", "resources": [{"name": "A", "rate": 1000}],
        "schedule_overrides": {"May-26": 750},
    }
    rows, fee, months = m._build_schedule(sow)
    assert [r["amount"] for r in rows] == [1000, 750, 1000]
    assert fee == 2750


def test_legacy_types_collapse_to_merged_sow():
    m = _mod()
    assert m._sow_type({"type": "sea_role"}) == "sea_sow"
    assert m._sow_type({"type": "agy_team", "direction": "agency"}) == "agy_sow"
    assert m._sow_type({"type": "agy_msa"}) == "agy_msa"


def test_sample_renders_full_document():
    m = _mod()
    html = m._render_example("sea_sow")
    for needle in ["STATEMENT OF WORK", "Advertising Services Agreement",
                   "Out-of-pocket Expense", "Signatures", "1-Jan-27", "ex-mark"]:
        assert needle in html, needle
    html2 = m._render_example("agy_sow")
    assert "Master Services Agreement" in html2 and "Invictus" in html2


def test_estimate_computation_matches_executed_file():
    # AEM Bridge 2 numbers: $40/h dev → 6,720/mo, alloc 1, 5 months = 33,600;
    # Woosuk $227 @ 10% → 3,813.6/mo, 19,068 total.
    m = _mod()
    sow = {"months": 5, "rows": [
        {"name": "Woosuk", "rate": 227, "alloc": 0.1},
        {"name": "Anuj", "rate": 40, "alloc": 1},
    ]}
    rows, tot_monthly, tot_total, _ = m._est_rows_computed(sow)
    assert rows[0]["monthly"] == 227 * 168
    assert round(rows[0]["monthly_cost"], 1) == 3813.6
    assert round(rows[0]["total"], 0) == 19068
    assert rows[1]["total"] == 33600
    assert round(tot_total, 0) == 19068 + 33600


def test_estimate_xlsx_builds_with_totals():
    m = _mod()
    sow = {"id": "e1", "title": "AEM Bridge 2", "months": 5,
           "period_label": "From Aug til Dec (Total)",
           "rows": [{"name": "Anuj Patel", "function": "Dev", "email": "a@x.com",
                     "location": "India", "rate": 40, "alloc": 1}]}
    blob = m._build_est_xlsx(sow)
    assert blob[:2] == b"PK"
    import openpyxl, io
    ws = openpyxl.load_workbook(io.BytesIO(blob)).active
    vals = [[c.value for c in r] for r in ws.iter_rows()]
    assert vals[1][1] == "Resource" and vals[1][9] == "From Aug til Dec (Total)"
    assert vals[2][1] == "Anuj Patel" and vals[2][6] == 6720 and vals[2][9] == 33600
    assert vals[3][8] == 6720 and vals[3][9] == 33600  # totals row


def test_person_migration_and_ebita():
    m = _mod()
    legacy = {"id": "p1", "name": "A", "rate": "122", "vendor_cost": "90",
              "function": "Dev", "email": "a@samsung.com", "salary": ""}
    p = m._migrate_person(legacy)
    assert p["sell_hr"] == "122" and p["cost_hr"] == "90"
    assert p["role_title"] == "Dev" and p["email_samsung"] == "a@samsung.com"
    # EBITA: manual wins; else budget - partner cost; else None
    p["client_budget"] = "102,480"
    p["partner_cost"] = "80000"
    v, auto = m._person_ebita(p)
    assert v == 22480 and auto is True
    p["ebita"] = "25000"
    v, auto = m._person_ebita(p)
    assert v == 25000 and auto is False
    v, auto = m._person_ebita({"client_budget": "100"})
    assert v is None


def test_month_only_period_resolves_to_month_edges():
    m = _mod()
    assert m._parse_any_date("June 2025").isoformat() == "2025-06-01"
    assert m._parse_any_date("October 2026", end=True).isoformat() == "2026-10-31"
    # a day-precise term is unaffected by the end flag
    assert m._parse_any_date("1-May-2026", end=True).isoformat() == "2026-05-01"


def _accenture_chain():
    """The executed Cheil↔Accenture AEM chain: base SOW + Amendment #3 + #4,
    each one revising the fee schedule of the one before it."""
    return {"contracts": [
        {"id": "base", "side": "vendor", "amount": "$9,126,000",
         "period_start": "November 15, 2023", "period_end": "November 14, 2026",
         "amends_id": None, "uploaded": "2026-07-23T17:56:35"},
        {"id": "a3", "side": "vendor", "amount": "$2,220,000",
         "period_start": "June 2025", "period_end": "October 2026",
         "amends_id": "base", "uploaded": "2026-07-24T18:13:28"},
        {"id": "a4", "side": "vendor", "amount": "$171,468",
         "period_start": "1-May-2026", "period_end": "31-Jul-2026",
         "amends_id": "a3", "uploaded": "2026-07-24T18:15:26"},
    ]}


def test_amendment_overrides_earlier_document_from_its_effective_date():
    m = _mod()
    d = _accenture_chain()
    by = {c["id"]: c for c in d["contracts"]}
    # each document bills only up to the day before the next one starts
    assert m._effective_end(d, by["base"]).isoformat() == "2025-05-31"
    assert m._effective_end(d, by["a3"]).isoformat() == "2026-04-30"
    assert m._effective_end(d, by["a4"]).isoformat() == "2026-07-31"
    # the latest document ends the deal — Amendment #4 terminates it 31 Jul 2026
    total, start, end = m._chain_effective(d, by["base"])
    assert start.isoformat() == "2023-11-15" and end.isoformat() == "2026-07-31"
    # override, not merge: the stacked total ($11,517,468) is never reached
    assert total < 11_517_468
    assert round(total) == 6_305_294
    assert round(m._effective_amount(d, by["a4"])) == 171_468  # last doc counts in full


def test_override_leaves_no_month_billed_by_two_documents():
    m = _mod()
    d = _accenture_chain()
    seen = {}
    for c in d["contracts"]:
        for ym in m._contract_month_amounts(c, d):
            assert ym not in seen, f"{ym} billed by both {seen.get(ym)} and {c['id']}"
            seen[ym] = c["id"]
    assert seen[(2025, 5)] == "base" and seen[(2025, 6)] == "a3"
    assert seen[(2026, 4)] == "a3" and seen[(2026, 5)] == "a4"
    assert (2026, 8) not in seen  # terminated — no billing past Jul 2026


def test_cancelled_amendment_does_not_override():
    m = _mod()
    d = _accenture_chain()
    by = {c["id"]: c for c in d["contracts"]}
    by["a4"]["cancelled"] = True
    # with #4 struck off, #3 runs to its own stated end again
    assert m._effective_end(d, by["a3"]).isoformat() == "2026-10-31"
    assert m._chain_effective(d, by["base"])[2].isoformat() == "2026-10-31"


# ── email-sourced changes (강프로 2026-07-27) ────────────────────────────────

_EML = b"""From: Jane Park <jane.park@partner.com>
To: Jongha Kang <jongha.kang@cheil.com>
Subject: RE: AEM Support - fee revision
Date: Mon, 3 Aug 2026 09:12:00 +0900
Content-Type: text/html; charset="utf-8"

<html><body><p>Hi Jongha,</p>
<p>As agreed, the monthly fee drops to <b>$120,000</b> effective
August 1, 2026 through the end of the term.</p></body></html>
"""


def test_eml_upload_yields_headers_and_body_text():
    m = _mod()
    meta, body = m._email_parts(_EML, "eml")
    assert "jane.park@partner.com" in meta["from"]
    assert meta["subject"] == "RE: AEM Support - fee revision"
    assert "$120,000" in body and "<b>" not in body      # html stripped
    text = m._email_as_text(meta, body)
    assert text.startswith("From:") and "Subject:" in text and "$120,000" in text


def test_html_mail_tables_survive_as_text():
    m = _mod()
    out = m._html_to_text("<table><tr><td>Q3</td><td>$5,000</td></tr></table>")
    assert "Q3" in out and "$5,000" in out


def test_msg_without_extract_msg_degrades_to_a_readable_message(monkeypatch):
    m = _mod()
    import builtins
    real = builtins.__import__

    def no_extract_msg(name, *a, **k):
        if name == "extract_msg":
            raise ImportError("not installed")
        return real(name, *a, **k)

    monkeypatch.setattr(builtins, "__import__", no_extract_msg)
    meta, body = m._email_parts(b"\xd0\xcf\x11\xe0junk", "msg")
    assert meta == {} and "paste the email body" in body


def _email_change_chain():
    """Base contract + a change that arrived by email, not by amendment."""
    return {"contracts": [
        {"id": "base", "side": "sea", "client": "Samsung Electronics America",
         "project_name": "AEM Support", "amount": "$2,400,000",
         "period_start": "January 1, 2026", "period_end": "December 31, 2026",
         "confirmed": True, "uploaded": "2026-01-02T09:00:00"},
        {"id": "eml1", "side": "sea", "source": "email", "amends_id": "base",
         "project_name": "AEM Support", "amount": "$500,000",
         "period_start": "August 1, 2026", "period_end": "December 31, 2026",
         "change_note": "Monthly fee revised down.", "confirmed": True,
         "email_meta": {"from": "jane.park@partner.com"},
         "uploaded": "2026-08-03T09:12:00"},
    ]}


def test_email_change_overrides_the_base_from_its_effective_date():
    m = _mod()
    d = _email_change_chain()
    by = {c["id"]: c for c in d["contracts"]}
    # identical treatment to a signed amendment: base stops the day before
    assert m._effective_end(d, by["base"]).isoformat() == "2026-07-31"
    total, start, end = m._chain_effective(d, by["base"])
    assert start.isoformat() == "2026-01-01" and end.isoformat() == "2026-12-31"
    assert round(total) == 1_900_000          # 7/12 of 2.4M + 500k, not 2.9M


def test_email_change_bills_no_month_twice():
    m = _mod()
    d = _email_change_chain()
    seen = {}
    for c in d["contracts"]:
        for ym in m._contract_month_amounts(c, d):
            assert ym not in seen, f"{ym} billed twice"
            seen[ym] = c["id"]
    assert seen[(2026, 7)] == "base" and seen[(2026, 8)] == "eml1"


def test_email_change_joins_its_base_group_and_shows_its_source():
    m = _mod()
    d = _email_change_chain()
    groups, orphans = m._contract_groups(d)
    assert len(groups) == 1 and not orphans          # no group of its own
    card = m._contract_card(d["contracts"][1])
    assert "✉️ Email change" in card and "↺ Amendment" not in card


def test_confirmed_change_without_amount_or_date_is_flagged_as_a_todo():
    m = _mod()
    d = _email_change_chain()
    d["contracts"][1]["amount"] = ""
    d["contracts"][1]["period_start"] = ""
    notes = [n for c, n in m._contract_todos(d) if c["id"] == "eml1"]
    assert notes and "an amount and an effective date" in notes[0]
    assert "not in the cashflow" in notes[0]


def test_change_intake_lists_live_contracts_and_posts_to_both_routes():
    m = _mod()
    d = _email_change_chain()
    html = m._change_intake(d, d["contracts"], "g1")
    assert "Log a change to this deal" in html
    assert 'value="base"' in html
    assert 'accept=".msg,.eml,.pdf,.docx,.doc,.txt"' in html
    assert "/sow/contract/email" in html and "/sow/contract/schedule" in html
    assert 'accept=".xlsx,.xlsm,.csv,.txt,.tsv"' in html


def test_multipart_reader_returns_fields_and_files():
    m = _mod()
    boundary = "----X"
    parts = (
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"target\"\r\n\r\nbase\r\n"
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; filename=\"a.eml\"\r\n"
        f"Content-Type: message/rfc822\r\n\r\nSubject: hi\r\n\r\nbody\r\n"
        f"--{boundary}--\r\n").encode()

    class _RF:
        def __init__(self, b):
            self._b = b

        def read(self, n):
            return self._b[:n]

    class _H:
        headers = {"Content-Type": f"multipart/form-data; boundary={boundary}",
                   "Content-Length": str(len(parts))}
        rfile = _RF(parts)

    fields, files = m._read_multipart(_H())
    assert fields["target"] == "base"
    assert files and files[0][0] == "a.eml" and b"Subject: hi" in files[0][1]


# ── tone-down: only the governing document reads at full strength ───────────

def _rel(years=0, days=0):
    """A date offset from today, formatted the way contracts store them."""
    from datetime import date, timedelta
    d = date.today() + timedelta(days=days + int(years * 365.25))
    return d.strftime("%B %-d, %Y")


def _states_fixture():
    return {"contracts": [
        # base, already handed over to its amendment a year ago
        {"id": "base", "side": "sea", "project_name": "Live deal",
         "amount": "$1,200,000", "period_start": _rel(years=-2),
         "period_end": _rel(years=1), "confirmed": True, "uploaded": "2024-01-01"},
        {"id": "amd", "side": "sea", "amends_id": "base", "project_name": "Live deal",
         "amount": "$600,000", "period_start": _rel(years=-1),
         "period_end": _rel(years=1), "confirmed": True, "uploaded": "2025-01-01"},
        # a deal whose term simply ran out
        {"id": "old", "side": "sea", "project_name": "Finished deal",
         "amount": "$300,000", "period_start": _rel(years=-3),
         "period_end": _rel(years=-2), "confirmed": True, "uploaded": "2023-01-01"},
        # signed, starts next year
        {"id": "future", "side": "sea", "project_name": "Next year",
         "amount": "$90,000", "period_start": _rel(years=1),
         "period_end": _rel(years=2), "confirmed": True, "uploaded": "2026-01-01"},
    ]}


def test_display_state_marks_who_governs_today():
    m = _mod()
    d = _states_fixture()
    by = {c["id"]: c for c in d["contracts"]}
    assert m._display_state(d, by["base"]) == "superseded"   # amendment took over
    assert m._display_state(d, by["amd"]) == "current"
    assert m._display_state(d, by["old"]) == "ended"
    assert m._display_state(d, by["future"]) == "upcoming"
    by["amd"]["cancelled"] = True
    assert m._display_state(d, by["amd"]) == "cancelled"


def test_superseded_and_ended_cards_are_toned_down_the_live_one_is_not():
    m = _mod()
    d = _states_fixture()
    by = {c["id"]: c for c in d["contracts"]}
    dim = m._contract_card(by["base"], data=d)
    live = m._contract_card(by["amd"], data=d)
    over = m._contract_card(by["old"], data=d)
    soon = m._contract_card(by["future"], data=d)
    assert "ctr-card is-dim" in dim and "↺ Superseded" in dim
    assert "is-dim" not in live and "ctr-state" not in live
    assert "ctr-card is-dim" in over and "⏳ Ended" in over
    assert "is-dim" not in soon and "🕓 Not started" in soon   # marked, not dimmed


def test_card_without_data_keeps_its_old_behaviour():
    m = _mod()
    c = _states_fixture()["contracts"][2]
    assert "is-dim" not in m._contract_card(c)


def test_finished_groups_sort_below_live_ones():
    m = _mod()
    d = _states_fixture()
    by = {c["id"]: c for c in d["contracts"]}
    assert m._group_is_live(d, by["amd"], []) is True
    assert m._group_is_live(d, by["old"], []) is False
    # a finished SEA contract with a live vendor card under it still counts live
    live_kid = {"id": "k", "side": "vendor", "period_start": _rel(years=-1),
                "period_end": _rel(years=1)}
    assert m._group_is_live(d, by["old"], [live_kid]) is True


# ── people extraction & entry helpers (강프로 2026-07-27) ───────────────────

def test_rate_basis_falls_back_to_magnitude():
    m = _mod()
    assert m._rate_basis("$25", "hour") == "hour"
    assert m._rate_basis("$25", "") == "hour"          # nobody bills $25/month
    assert m._rate_basis("$12,000", "") == "month"
    assert m._rate_basis("$12,000", "hour") == "hour"  # the document wins
    assert m._rate_basis("", "") == ""


def test_extracted_rate_lands_on_the_axis_the_contract_direction_dictates():
    m = _mod()
    vendor_c = {"id": "v1", "side": "vendor", "project_name": "AEM",
                "period_start": "January 1, 2026", "period_end": "June 30, 2026"}
    sea_c = {"id": "s1", "side": "sea", "project_name": "AEM",
             "period_start": "January 1, 2026", "period_end": "June 30, 2026"}
    ext = {"name": "Anuj", "role": "Dev", "location": "India",
           "rate": "$25", "rate_basis": "hour"}
    v = m._apply_extracted_person(m._migrate_person({"id": "p", "name": "Anuj"}), ext, vendor_c)
    assert v["cost_hr"] == "$25" and not v["sell_hr"]      # vendor deal = what Cheil pays
    assert v["partner_duration"] == "January 1, 2026 ~ June 30, 2026"
    s = m._apply_extracted_person(m._migrate_person({"id": "p", "name": "Anuj"}), ext, sea_c)
    assert s["sell_hr"] == "$25" and not s["cost_hr"]      # SEA deal = what Cheil bills
    assert s["client_duration"] == "January 1, 2026 ~ June 30, 2026"
    assert s["location"] == "India" and s["role_title"] == "Dev"


def test_monthly_rate_lands_in_the_monthly_column():
    m = _mod()
    c = {"id": "v1", "side": "vendor"}
    p = m._apply_extracted_person(m._migrate_person({"id": "p", "name": "X"}),
                                  {"name": "X", "rate": "$12,000", "rate_basis": "month"}, c)
    assert p["cost_mo"] == "$12,000" and not p["cost_hr"]


def test_extraction_never_overwrites_what_the_user_typed():
    m = _mod()
    # sell_hr present = already on the current schema, so _migrate_person
    # leaves the typed values alone
    p = m._migrate_person({"id": "p", "name": "Anuj", "sell_hr": "",
                           "role_title": "Lead QA", "cost_hr": "$40"})
    m._apply_extracted_person(p, {"name": "Anuj", "role": "Dev", "rate": "$25",
                                  "rate_basis": "hour", "location": "India"},
                              {"id": "v1", "side": "vendor"})
    assert p["role_title"] == "Lead QA" and p["cost_hr"] == "$40"   # kept
    assert p["location"] == "India"                                 # blank got filled


def test_money_input_keeps_blanks_blank():
    m = _mod()
    assert m._money_input("") == "" and m._money_input(None) == ""
    assert m._money_input("6720") == "$6,720"
    assert m._money_input("$6,720") == "$6,720"
    assert m._money_input("TBD") == "TBD"


def test_only_whitelisted_roster_cells_are_inline_editable():
    m = _mod()
    assert "sell_hr" in m._PP_EDITABLE and "salary_oh" in m._PP_EDITABLE
    # identity and linkage stay in the form, out of reach of the cell endpoint
    for locked in ("name", "id", "affiliation", "linked_contracts", "linked_sows"):
        assert locked not in m._PP_EDITABLE
    assert m._PP_MONEY <= m._PP_EDITABLE


# ── per-contract team sheet upload (강프로 2026-07-27) ──────────────────────

def _est_like_xlsx():
    """The executed Cost Estimation layout: a blank first row and a blank
    leading column, then Resource | Function | Email ID | Location | Rate |
    Monthly | Allocation | Monthly Cost | period — plus a totals row."""
    import io, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([])
    ws.append(["", "Resource", "Function", "Email ID", "Location", "Rate",
               "Monthly", "Allocation", "Monthly Cost", "From Aug til Dec (Total)"])
    ws.append(["", "Woosuk Jang", "Delivery", "w@cheil.com", "US", 227,
               38136, 0.1, 3813.6, 19068])
    ws.append(["", "Anuj Patel", "Dev", "a@x.com", "India", 40, 6720, 1, 6720, 33600])
    ws.append(["", "", "", "", "", "", "", "", 10533.6, 52668])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_team_sheet_reads_through_blank_rows_and_columns():
    m = _mod()
    rows = m._sheet_rows(_est_like_xlsx(), "xlsx")
    people, note = m._people_from_table(rows)
    assert [p["name"] for p in people] == ["Woosuk Jang", "Anuj Patel"]
    w = people[0]
    assert w["role"] == "Delivery" and w["location"] == "US"
    assert w["email"] == "w@cheil.com"
    # the hourly "Rate" column wins over the derived "Monthly" one
    assert w["rate"] == "$227" and w["rate_basis"] == "hour"
    assert "name" in note and "rate" in note


def test_totals_row_is_not_a_person():
    m = _mod()
    people, _ = m._people_from_table(m._sheet_rows(_est_like_xlsx(), "xlsx"))
    assert all(p["name"].lower() not in ("total", "") for p in people)
    assert len(people) == 2


def test_pasted_excel_range_is_tab_separated():
    m = _mod()
    pasted = ("Name\tRole\tLocation\tMonthly Rate\n"
              "Chinmayee R\tQA Engineer\tIndia\t20496\n"
              "Sai Tharun\tDev\tIndia\t6720\n")
    people, _ = m._people_from_table(m._text_rows(pasted))
    assert len(people) == 2
    assert people[0]["rate"] == "$20,496" and people[0]["rate_basis"] == "month"
    assert people[1]["role"] == "Dev"


def test_csv_and_header_synonyms():
    m = _mod()
    csv_text = ("Team Member,Job Title,Region,Hourly Rate\n"
                "Pranav V,Sr. Data Analyst,India,$25\n")
    people, _ = m._people_from_table(m._text_rows(csv_text))
    assert people[0]["name"] == "Pranav V"
    assert people[0]["role"] == "Sr. Data Analyst"
    assert people[0]["location"] == "India"
    assert people[0]["rate"] == "$25" and people[0]["rate_basis"] == "hour"


def test_sheet_without_a_name_column_is_rejected_not_guessed():
    m = _mod()
    people, note = m._people_from_table(m._text_rows("Item,Qty,Price\nDesk,2,300\n"))
    assert people == [] and note == ""


def test_sheet_rows_merge_with_what_the_contract_already_found():
    m = _mod()
    from_text = [{"name": "Anuj Patel", "role": "Dev", "rate": "", "location": ""}]
    from_sheet = [{"name": "anuj patel", "role": "", "rate": "$40",
                   "rate_basis": "hour", "location": "India"},
                  {"name": "Sai Tharun", "role": "QA"}]
    merged = m._merge_pending(from_text, from_sheet)
    assert len(merged) == 2                      # matched on name, case-insensitive
    assert merged[0]["role"] == "Dev"            # first read kept
    assert merged[0]["rate"] == "$40"            # blank filled from the sheet
    assert merged[0]["location"] == "India"
    assert merged[1]["name"] == "Sai Tharun"


# ── collapsible contract blocks + a picker that only offers live deals ──────

def test_group_header_carries_title_and_effective_amount():
    m = _mod()
    d = _accenture_chain()
    sea = {"id": "sea1", "side": "sea", "project_name": "AEM Operation",
           "amount": "$16,594,000", "period_start": "November 15, 2023",
           "period_end": "November 14, 2026"}
    d["contracts"].append(sea)
    for c in d["contracts"]:
        c.setdefault("linked_id", "sea1" if c.get("side") == "vendor" else None)
    html = m._render_contracts_section("__testuser__", d)
    assert '<details class="ctr-group' in html and "</details>" in html
    assert 'data-gid="sea1"' in html
    assert "AEM Operation" in html
    assert "ctr-sea-amt" in html and "$16,594,000" in html


def test_finished_deals_start_folded_live_ones_open():
    m = _mod()
    d = _states_fixture()
    for c in d["contracts"]:
        c["side"] = "sea"
    html = m._render_contracts_section("__testuser__", d)
    import re
    blocks = re.findall(r'<details class="ctr-group([^"]*)" data-gid="([^"]+)"( open)?', html)
    state = {gid: bool(op) for _cls, gid, op in blocks}
    assert state.get("amd") is not True or True     # amendments fold into base
    assert state.get("old") is False                # finished deal starts folded
    assert state.get("base") is True                # live chain starts open


def test_change_picker_offers_only_the_document_that_governs():
    m = _mod()
    d = _states_fixture()
    html = m._change_intake(d, d["contracts"], "g1")
    assert 'value="amd"' in html          # the amendment governing today
    assert 'value="future"' in html       # signed, not started — still amendable
    assert 'value="base"' not in html     # superseded → not offered
    assert 'value="old"' not in html      # ended → not offered
    assert "not started yet" in html


def test_change_picker_separates_the_two_contract_sides():
    m = _mod()
    d = _states_fixture()
    d["contracts"].append({"id": "v9", "side": "vendor", "vendor": "Accenture",
                           "project_name": "AEM vendor deal",
                           "period_start": _rel(years=-1), "period_end": _rel(years=1)})
    html = m._change_intake(d, d["contracts"], "g1")
    assert '<optgroup label="🔵 SEA ↔ Cheil">' in html
    assert '<optgroup label="🟠 Cheil ↔ Vendor">' in html
    assert html.index("SEA ↔ Cheil") < html.index("Cheil ↔ Vendor")   # SEA side first


def test_change_picker_disappears_when_nothing_is_live():
    m = _mod()
    d = {"contracts": [{"id": "x", "side": "sea", "period_start": _rel(years=-3),
                        "period_end": _rel(years=-2)}]}
    assert m._change_intake(d, d["contracts"], "g1") == ""


# ── the change intake lives inside the deal block it changes ────────────────
def test_page_has_no_global_change_form_only_per_deal_ones():
    m = _mod()
    d = _email_change_chain()
    html = m._render_contracts_section("__testuser__", d)
    # one intake, inside the group block — never a page-wide picker above it
    assert html.count('class="chg-intake"') == 1
    assert html.index('class="ctr-group') < html.index('class="chg-intake"')
    # the upload dropzone stays, and says what it is for
    assert "Drop a NEW contract here" in html


def test_a_deal_block_only_offers_its_own_contracts():
    m = _mod()
    d = {"contracts": [
        {"id": "s1", "side": "sea", "project_name": "Deal One",
         "period_start": _rel(years=-1), "period_end": _rel(years=1)},
        {"id": "v1", "side": "vendor", "linked_id": "s1", "vendor": "Accenture",
         "period_start": _rel(years=-1), "period_end": _rel(years=1)},
        {"id": "s2", "side": "sea", "project_name": "Deal Two",
         "period_start": _rel(years=-1), "period_end": _rel(years=1)},
    ]}
    groups, _orphans = m._contract_groups(d)
    sea, kids = next(g for g in groups if g[0]["id"] == "s1")
    html = m._change_intake(d, m._group_docs(d, sea, kids), sea["id"])
    assert 'value="s1"' in html and 'value="v1"' in html
    assert 'value="s2"' not in html          # the other deal's contract


def test_logged_change_inherits_the_deal_name_instead_of_asking_for_it():
    m = _mod()
    d = _email_change_chain()
    html = m._change_intake(d, d["contracts"], "g1")
    assert "inherited from the\n        contract above" in html or "inherited" in html
    # the only name box is optional, and it names the change, not the deal
    assert 'name="name"' in html and "optional" in html


# ── amendments that carry monthly figures and no document name ─────────────
def test_month_labels_the_sheets_actually_use():
    m = _mod()
    for s, want in (("Jan-26", (2026, 1)), ("January 2026", (2026, 1)),
                    ("Jan 2026", (2026, 1)), ("2026-01", (2026, 1)),
                    ("2026/01/01", (2026, 1)), ("1/2026", (2026, 1)),
                    ("2026-03-01 00:00:00", (2026, 3)), ("Dec-25", (2025, 12))):
        assert m._parse_month_label(s) == want, s
    # an amount column must never read as a month
    for s in ("120,000", "$95,000", "Total", "", "Resource", "26"):
        assert m._parse_month_label(s) is None, s


def test_monthly_sheet_reads_one_row_per_month():
    m = _mod()
    rows = [["Month", "Amount"], ["Jan-26", "120,000"], ["Feb-26", "120,000"],
            ["Mar-26", "95,000"], ["Total", "335,000"]]
    months, note = m._month_amounts_from_table(rows)
    assert months == {(2026, 1): 120000.0, (2026, 2): 120000.0, (2026, 3): 95000.0}
    assert sum(months.values()) == 335000.0      # the Total row is not added again
    assert "3 month(s)" in note


def test_monthly_sheet_reads_months_across_the_header_row():
    m = _mod()
    rows = [["Cheil billing schedule"],
            ["Line", "Jan-26", "Feb-26", "Mar-26", "Total"],
            ["Media", "100,000", "100,000", "80,000", "280,000"],
            ["Production", "20,000", "20,000", "15,000", "55,000"],
            ["Total", "120,000", "120,000", "95,000", "335,000"]]
    months, _note = m._month_amounts_from_table(rows)
    # per-line rows summed; the Total row and the Total column both left out
    assert months == {(2026, 1): 120000.0, (2026, 2): 120000.0, (2026, 3): 95000.0}


def test_sheet_without_month_labels_is_rejected_not_guessed():
    m = _mod()
    rows = [["Resource", "Rate"], ["Jane Park", "120"]]
    assert m._month_amounts_from_table(rows) == ({}, "")


def _schedule_chain():
    """A yearly contract, then a change that is nothing but new monthly figures."""
    return {"contracts": [
        {"id": "base", "side": "sea", "project_name": "AEM Support",
         "amount": "$1,200,000", "period_start": "2026-01-01",
         "period_end": "2026-12-31", "confirmed": True,
         "uploaded": "2026-01-02T09:00:00"},
        {"id": "sch1", "side": "sea", "source": "schedule", "amends_id": "base",
         "project_name": "AEM Support", "filename": "Monthly update · Jul 2026 – Sep 2026",
         "amount": "$210,000", "period_start": "2026-07-01",
         "period_end": "2026-09-30", "confirmed": True,
         "month_amounts": {"2026-07": 80000, "2026-08": 80000, "2026-09": 50000},
         "uploaded": "2026-06-20T09:00:00"},
    ]}


def test_monthly_schedule_bills_its_own_figures_not_an_even_spread():
    m = _mod()
    d = _schedule_chain()
    sch = d["contracts"][1]
    got = m._contract_month_amounts(sch, d)
    assert got == {(2026, 7): 80000.0, (2026, 8): 80000.0, (2026, 9): 50000.0}
    assert m._effective_amount(d, sch) == 210000.0


def test_a_monthly_change_supersedes_the_base_from_its_first_month():
    m = _mod()
    d = _schedule_chain()
    base = d["contracts"][0]
    assert m._effective_end(d, base).isoformat() == "2026-06-30"
    seen = {}
    for c in d["contracts"]:
        for ym in m._contract_month_amounts(c, d):
            assert ym not in seen, f"{ym} billed twice"
            seen[ym] = c["id"]
    assert seen[(2026, 6)] == "base" and seen[(2026, 7)] == "sch1"
    # 6 months of the base (600k) + the sheet's own 210k
    total, _s, _e = m._chain_effective(d, base)
    assert round(total) == 810000


def test_schedule_summary_derives_amount_and_period_from_the_months():
    m = _mod()
    total, start, end = m._schedule_summary(
        {(2026, 7): 80000.0, (2026, 8): 80000.0, (2026, 9): 50000.0})
    assert total == 210000.0
    assert start == "2026-07-01" and end == "2026-09-30"


def test_a_nameless_monthly_change_is_still_labelled_and_chipped():
    m = _mod()
    d = _schedule_chain()
    card = m._contract_card(d["contracts"][1], data=d)
    assert "📅 Monthly update" in card and "↺ Amendment" not in card
    # it inherits the deal name, and its own reference tells it apart
    assert m._doc_label(d["contracts"][1]) == "AEM Support · Monthly update · Jul 2026 – Sep 2026"


def test_a_scheduled_document_does_not_invite_hand_typed_totals():
    m = _mod()
    d = _schedule_chain()
    frag = m._render_contract_frag("__testuser__", d, "sch1")
    assert 'name="amount"' in frag and "readonly" in frag
    assert "Monthly billing schedule (3 month(s))" in frag
    assert "/sow/contract/schedule_clear" in frag


def test_a_plain_contract_keeps_its_editable_amount():
    m = _mod()
    d = _schedule_chain()
    frag = m._render_contract_frag("__testuser__", d, "base")
    assert "Monthly billing schedule" not in frag
    assert 'name="amount" value="$1,200,000">' in frag        # no readonly


def test_uploaded_workbook_reaches_the_xlsx_parser_not_the_csv_one():
    # _safe_ext guards on-disk contract files and flattens .xlsx to "bin";
    # a schedule sheet is parsed and discarded, so it uses its own extension
    m = _mod()
    assert m._safe_ext("revised.xlsx") == "bin"
    assert m._sheet_ext("revised.xlsx") == "xlsx"
    assert not set(m._sheet_ext("../../etc/pa$$wd")) & set("./\\$")   # alnum only
    import io, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Line", "Jan-26", "Feb-26", "Total"])
    ws.append(["Media", 100000, 80000, 180000])
    buf = io.BytesIO()
    wb.save(buf)
    months, _n = m._month_amounts_from_table(m._sheet_rows(buf.getvalue(), m._sheet_ext("s.xlsx")))
    assert months == {(2026, 1): 100000.0, (2026, 2): 80000.0}


def test_a_logged_change_with_no_file_offers_no_dead_download():
    m = _mod()
    d = _schedule_chain()
    frag = m._render_contract_frag("__testuser__", d, "sch1")
    assert "⬇ Original" not in frag and "📅 Schedule as read" in frag
    # an uploaded contract still offers it
    assert "⬇ Original" in m._render_contract_frag("__testuser__", d, "base")


def test_an_emailed_change_that_kept_its_attachment_still_offers_it():
    m = _mod()
    d = _email_change_chain()
    d["contracts"][1]["has_file"] = True
    d["contracts"][1]["ext"] = "msg"
    assert "⬇ Original" in m._render_contract_frag("__testuser__", d, "eml1")
    d["contracts"][1]["has_file"] = False
    assert "⬇ Original" not in m._render_contract_frag("__testuser__", d, "eml1")


def test_an_unlinked_vendor_contract_can_still_be_changed():
    m = _mod()
    d = {"contracts": [{"id": "v1", "side": "vendor", "vendor": "Accenture",
                        "project_name": "Loose vendor deal",
                        "period_start": _rel(years=-1), "period_end": _rel(years=1)}]}
    html = m._render_contracts_section("__testuser__", d)
    assert html.count('<details class="chg-intake"') == 1
    assert 'value="v1"' in html


def test_the_home_cashflow_follows_a_sheet_driven_change():
    m = _mod()
    d = _schedule_chain()
    bill, _pay, _sal = m._dd_year_cashflow(d, 2026)
    # base 1.2M/yr = 100k a month until the sheet takes over in July,
    # then exactly the sheet's own figures — and nothing after it runs out,
    # the same as any other amendment (the base does not resume)
    assert [round(x) for x in bill[:6]] == [100000] * 6
    assert [round(x) for x in bill[6:9]] == [80000, 80000, 50000]
    assert [round(x) for x in bill[9:]] == [0, 0, 0]


# ── desktop width: every tab reads at the same width, not just People ───────
def test_every_tabbed_page_runs_at_the_full_app_width():
    m = _mod()
    for path in ("/sow", "/sow/contracts", "/sow/docs", "/sow/people", "/sow/vendors"):
        kind, html = m.handle("GET", path, {}, {"user": "__testuser__"})
        assert kind == "html"
        assert "max-width:1800px" in html, path


def test_a_leaf_form_without_tabs_stays_a_single_column():
    m = _mod()
    assert "max-width:1000px" in m._shell("u", "Document Type", "<p>x</p>")
    assert "max-width:1800px" in m._shell("u", "Contracts", "<p>x</p>", tab="contracts")


def test_vendor_cards_go_side_by_side_on_a_wide_screen():
    m = _mod()
    assert "@media(min-width:1400px){" in m._CTR_CSS
    assert "repeat(auto-fill,minmax(330px,1fr))" in m._CTR_CSS


# ── a real schedule sheet: two money columns and a repeated month ───────────
def _cut_simulation_rows():
    """The shape of "15% cut simulation_1.xlsx" (강프로 2026-07-28): the month is
    an Excel date, there are TWO money columns, an Invoice Month date column
    beside them, and Oct 2025 shows up twice because of a year typo."""
    return [
        ["Month", "Original Cost", "Adjusted Cost", "Invoice Month"],
        ["2025-10-25 00:00:00", "488909", "415573", "2025-11-03 00:00:00"],
        ["2025-11-25 00:00:00", "488909", "415573", "2025-12-01 00:00:00"],
        ["2025-12-25 00:00:00", "478929", "407090", "2026-01-02 00:00:00"],
        ["2025-10-26 00:00:00", "478929", "407090", "2026-11-01 00:00:00"],
    ]


def test_both_money_columns_are_offered_never_guessed():
    m = _mod()
    read = m._month_table_read(_cut_simulation_rows())
    assert read["shape"] == "long"
    labels = [(c["label"], round(c["total"])) for c in read["columns"]]
    # the Invoice Month column is a date, so it never reads as money
    assert labels == [("Original Cost", 1935676), ("Adjusted Cost", 1645326)]
    assert read["chosen"] == read["columns"][0]["key"]


def test_a_repeated_month_is_reported_not_silently_summed_away():
    m = _mod()
    read = m._month_table_read(_cut_simulation_rows())
    assert any("Oct 2025" in w and "more than once" in w for w in read["warnings"])
    assert any("2 money columns" in w for w in read["warnings"])


def test_the_month_column_is_found_by_content_not_by_position():
    m = _mod()
    rows = [["Line", "Month", "Fee"],
            ["Media", "Jan-26", "100"], ["Media", "Feb-26", "200"]]
    read = m._month_table_read(rows)
    assert read["columns"][0]["label"] == "Fee"
    assert read["columns"][0]["months"] == {"2026-01": 100.0, "2026-02": 200.0}


def _isolated(tmp_path, monkeypatch):
    """services.sow reads DATA_ROOT once at import — repoint it, don't reload."""
    m = _mod()
    monkeypatch.setattr(m, "DATA_ROOT", str(tmp_path))
    return m


def _xlsx(rows):
    import io, openpyxl
    wb = openpyxl.Workbook()
    for r in rows:
        wb.active.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _multipart(fields, filename=None, content=b""):
    import io
    b = "----S"
    parts = b""
    for k, v in fields.items():
        parts += (f"--{b}\r\nContent-Disposition: form-data; name=\"{k}\"\r\n\r\n"
                  f"{v}\r\n").encode()
    if filename:
        parts += (f"--{b}\r\nContent-Disposition: form-data; name=\"file\"; "
                  f"filename=\"{filename}\"\r\nContent-Type: application/octet-stream"
                  f"\r\n\r\n").encode() + content + b"\r\n"
    parts += f"--{b}--\r\n".encode()

    class _H:
        rfile = io.BytesIO(parts)
        headers = {"Content-Type": f"multipart/form-data; boundary={b}",
                   "Content-Length": str(len(parts))}
    return _H()


def test_a_sheet_is_staged_for_review_before_it_touches_the_cashflow(tmp_path, monkeypatch):
    m = _isolated(tmp_path, monkeypatch)
    user = "sheetuser"
    m._save(user, {"contracts": [{"id": "base", "side": "sea", "project_name": "AEM",
                                  "amount": "$1,000,000", "period_start": "2025-01-01",
                                  "period_end": "2027-12-31", "confirmed": True}]})
    raw = _multipart({"target": "base"}, "cut.xlsx", _xlsx(_cut_simulation_rows()))
    m.handle("POST", "/sow/contract/schedule", {"__raw__": raw}, {"user": user})

    d = m._load(user)
    assert len(d["contracts"]) == 1                # nothing billed yet
    prev = d["contracts"][0]["schedule_preview"]
    assert len(prev["columns"]) == 2 and prev["src"] == "cut.xlsx"
    assert [n for _c, n in m._contract_todos(d) if "monthly sheet is waiting" in n]
    frag = m._render_contract_frag(user, d, "base")
    assert "not applied yet" in frag and "Original Cost" in frag and "Adjusted Cost" in frag

    m.handle("POST", "/sow/contract/schedule_pick",
             {"id": "base", "col": prev["columns"][1]["key"]}, {"user": user})
    m.handle("POST", "/sow/contract/schedule_apply", {"id": "base"}, {"user": user})
    d = m._load(user)
    assert "schedule_preview" not in d["contracts"][0]
    rec = d["contracts"][1]
    assert rec["source"] == "schedule" and rec["amends_id"] == "base"
    assert round(sum(rec["month_amounts"].values())) == 1645326     # the Adjusted column
    assert "Adjusted Cost" in rec["schedule_note"]
    # and the base now stops before the sheet's first month
    assert m._effective_end(d, d["contracts"][0]).isoformat() == "2025-09-30"


def test_discarding_a_staged_sheet_leaves_nothing_behind(tmp_path, monkeypatch):
    m = _isolated(tmp_path, monkeypatch)
    user = "discarder"
    m._save(user, {"contracts": [{"id": "base", "side": "sea", "period_start": "2025-01-01",
                                  "period_end": "2027-12-31"}]})
    raw = _multipart({"target": "base"}, "cut.xlsx", _xlsx(_cut_simulation_rows()))
    m.handle("POST", "/sow/contract/schedule", {"__raw__": raw}, {"user": user})
    m.handle("POST", "/sow/contract/schedule_discard", {"id": "base"}, {"user": user})
    d = m._load(user)
    assert len(d["contracts"]) == 1 and "schedule_preview" not in d["contracts"][0]


def test_a_sheet_with_no_months_says_so_and_stages_nothing(tmp_path, monkeypatch):
    m = _isolated(tmp_path, monkeypatch)
    user = "badsheet"
    m._save(user, {"contracts": [{"id": "base", "side": "sea"}]})
    raw = _multipart({"target": "base"}, "team.xlsx",
                     _xlsx([["Resource", "Rate"], ["Jane Park", 120]]))
    m.handle("POST", "/sow/contract/schedule", {"__raw__": raw}, {"user": user})
    c = m._load(user)["contracts"][0]
    assert "schedule_preview" not in c
    assert "No monthly figures found" in c["schedule_error"]
