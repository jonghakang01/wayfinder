import csv, io, json, os, re, base64, uuid
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ── Web Push helpers ──────────────────────────────────────────────────────────
def _push_subs_file(username: str) -> Path:
    return DATA_DIR / f"push_subs_{username}.json"

def _load_push_subs(username: str) -> list:
    f = _push_subs_file(username)
    if not f.exists():
        return []
    try:
        return json.loads(f.read_text())
    except Exception:
        return []

def _save_push_subs(username: str, subs: list):
    _push_subs_file(username).write_text(json.dumps(subs, ensure_ascii=False))

def _send_push_notification(username: str, title: str, body: str, url: str = "/cardconv/ledger"):
    """Send Web Push to all stored subscriptions for the user. Silently skips on error."""
    subs = _load_push_subs(username)
    if not subs:
        return
    pem_path = DATA_DIR / "vapid_private.pem"
    if not pem_path.exists():
        return
    try:
        from pywebpush import webpush, WebPushException
        payload = json.dumps({"title": title, "body": body, "url": url})
        pem = pem_path.read_text()
        dead = []
        for sub in subs:
            try:
                webpush(
                    subscription_info=sub,
                    data=payload,
                    vapid_private_key=pem,
                    vapid_claims={"sub": "mailto:jongha.kang01@gmail.com"},
                    content_encoding="aes128gcm",
                )
            except WebPushException as e:
                if e.response and e.response.status_code in (404, 410):
                    dead.append(sub)
        if dead:
            _save_push_subs(username, [s for s in subs if s not in dead])
    except Exception:
        pass

DATA_DIR   = Path(os.path.expanduser("~/.appdata/cardconv"))
KW_FILE    = DATA_DIR / "keywords.json"
HIST_FILE  = DATA_DIR / "history.json"
OUT_DIR    = DATA_DIR / "outputs"
TOKENS_DIR = DATA_DIR / "tokens"
CREDS_FILE = DATA_DIR / "google_credentials.json"
_SVC_DIR   = Path(__file__).parent
TEMPLATE   = _SVC_DIR / "cardconv_template.xlsx"  # bundled with service
TEMPLATE_FALLBACK = Path(os.path.expanduser(
    "~/Desktop/US업무/법카 정산/Automation/for upload.xlsx"
))

SCOPES = [
    'https://www.googleapis.com/auth/drive.file',
    'https://www.googleapis.com/auth/drive.readonly',
]
ADMIN  = "jongha.kang"
TARGET_NAMES = {"JONG KANG", "JONGHA KANG"}

META = {
    "name": "Cheil USA AMEX Converter",
    "path": "/cardconv",
    "icon": "💳",
    "description": "Corporate card CSV → SAP upload xlsx",
    "admin_only": True,
}

DEFAULT_KW = [
    {"kw": "STARBUCKS",         "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "CAFE",              "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "COFFEE",            "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "BAKERY",            "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "SWEETGREEN",        "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "SRA CAFE",          "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "KOBRICK",           "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "DOORDASH",          "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "GRUBHUB",           "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "WHOLEFDS",          "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "TRADER JOE",        "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "RESTAURANT",        "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "GRILL",             "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "PIZZA",             "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "BURGER",            "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "SUSHI",             "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "PARIS BAGUETTE",    "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "AMAZON",            "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "HOME DEPOT",        "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "BEST BUY",          "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "TARGET",            "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "COSTCO",            "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "NESPRESSO",         "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "PG&E",              "gl": 53210177, "ser": "021", "purpose": "Office Utilities"},
    {"kw": "UBER",              "gl": 53270377, "ser": "306", "purpose": "Uber travel"},
    {"kw": "LYFT",              "gl": 53270377, "ser": "306", "purpose": "Lyft travel"},
    {"kw": "ENTERPRISE",        "gl": 53270377, "ser": "306", "purpose": "Car rental"},
    {"kw": "HERTZ",             "gl": 53270377, "ser": "306", "purpose": "Car rental"},
    {"kw": "AVIS",              "gl": 53270377, "ser": "306", "purpose": "Car rental"},
    {"kw": "CHEVRON",           "gl": 53270377, "ser": "306", "purpose": "Gas for car"},
    {"kw": "SHELL",             "gl": 53270377, "ser": "306", "purpose": "Gas for car"},
    {"kw": "ROBBIE",            "gl": 53270377, "ser": "306", "purpose": "Gas for car"},
    {"kw": "GARAGE",            "gl": 53270377, "ser": "306", "purpose": "Parking"},
    {"kw": "PARKING",           "gl": 53270377, "ser": "306", "purpose": "Parking"},
    {"kw": "FASTRAK",           "gl": 53270377, "ser": "306", "purpose": "Car Toll"},
    {"kw": "TOLL",              "gl": 53270377, "ser": "306", "purpose": "Rental Toll"},
    {"kw": "SPOTHERO",          "gl": 53270377, "ser": "306", "purpose": "Parking"},
    {"kw": "CAR WASH",          "gl": 53270377, "ser": "306", "purpose": "Car wash for Car"},
    {"kw": "XPRESS",            "gl": 53270377, "ser": "306", "purpose": "Car wash for Car"},
    {"kw": "HYATT",             "gl": 53270377, "ser": "306", "purpose": "Hotel accommodation"},
    {"kw": "MARRIOTT",          "gl": 53270377, "ser": "306", "purpose": "Hotel accommodation"},
    {"kw": "HILTON",            "gl": 53270377, "ser": "306", "purpose": "Hotel accommodation"},
    {"kw": "HOTEL",             "gl": 53270377, "ser": "306", "purpose": "Hotel accommodation"},
    {"kw": "UNITED AIRLINES",   "gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "KOREAN AIR",        "gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "SINGAPORE AIRLINES","gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "BRITISH AIRWAYS",   "gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "AMERICAN AIRLINES", "gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "TOTAL WINE",        "gl": 53410103, "ser": "159", "purpose": "Purchase for client"},
    {"kw": "WINE",              "gl": 53410103, "ser": "159", "purpose": "Purchase for client"},
    {"kw": "OPENAI",            "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "ANTHROPIC",         "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "CLAUDE",            "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "CHATGPT",           "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "PERPLEXITY",        "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "LOVABLE",           "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "MANUS AI",          "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "GOOGLE *GOOGLE",    "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "GOOGLE ONE",        "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "RHINO NETWORK",     "gl": 53290177, "ser": "052", "purpose": "Office Network setup"},
    {"kw": "RITZ",              "gl": 53470177, "ser": "289", "purpose": "Coffee, Snack and meal"},
    {"kw": "JO MALONE",         "gl": 53470177, "ser": "289", "purpose": "Coffee, Snack and meal"},
]

FIXED = {
    "receipt_type": "D",
    "employee_id":  20170321,
    "payee":        "A0016672",
    "domestic":     "D",
    "currency":     "USD",
    "tax_code":     "VV",
    "cost_center":  "AG010238",
}


# ── Data helpers ─────────────────────────────────────────────────────────────

def _ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    TOKENS_DIR.mkdir(parents=True, exist_ok=True)


def _load_kw():
    _ensure_dirs()
    if KW_FILE.exists():
        try:
            return json.loads(KW_FILE.read_text())
        except Exception:
            pass
    _save_kw(DEFAULT_KW)
    return DEFAULT_KW


def _save_kw(kws):
    _ensure_dirs()
    KW_FILE.write_text(json.dumps(kws, ensure_ascii=False, indent=2))


def _load_hist():
    if HIST_FILE.exists():
        try:
            return json.loads(HIST_FILE.read_text())
        except Exception:
            pass
    return []


def _save_hist(hist):
    HIST_FILE.write_text(json.dumps(hist, ensure_ascii=False, indent=2))


def _add_hist(entry: dict):
    if "id" not in entry:
        entry["id"] = uuid.uuid4().hex[:10]
    hist = _load_hist()
    hist.insert(0, entry)
    _save_hist(hist[:100])


def _receipts_file(username: str) -> Path:
    return DATA_DIR / f"receipts_{username}.json"


def _migrate_entry(e: dict) -> dict:
    """Ensure a receipt entry has v2 fields (id, ocr_status, match_status)."""
    e = dict(e)
    if not e.get("id"):
        e["id"] = "rcpt_" + (e.get("file_id") or uuid.uuid4().hex)[:8]
    # Multi-receipt: legacy entries are the only receipt on their image → index 0.
    e.setdefault("sub_index", 0)
    if "ocr_status" not in e:
        e["ocr_status"] = "done" if e.get("ocr_amount") is not None else "pending"
    # v2.1: handwritten-priority OCR. Legacy entries keep ocr_amount as printed total.
    e.setdefault("ocr_printed_amount", e.get("ocr_amount"))
    e.setdefault("ocr_handwritten_amount", None)
    # v2.2: multi-receipt bounding box overlay. Legacy entries have no bbox.
    e.setdefault("ocr_bbox", None)
    # v2.3: card brand (amex/visa/other, OCR-detected, user-editable), usage tag
    # (default "Regular"), and completion state (completed entries are hidden from
    # the default Ledger view, Sync and Mapping, and their Drive originals are
    # moved to a "Completed" folder).
    e.setdefault("card_brand", None)
    e.setdefault("usage", "Regular")
    e.setdefault("completed", False)
    e.setdefault("completed_at", None)
    e.setdefault("archived_drive", False)
    if "match_status" not in e:
        if e.get("matched"):
            e["match_status"] = "matched"
        else:
            # Initial state is always pending_match. unmatched is reserved for
            # receipts that were tried against a CSV but failed to match.
            e["match_status"] = "pending_match"
    # Backfill OCR date from the matched CSV transaction for legacy entries that
    # were matched while OCR failed to read a date.
    if e.get("match_status") == "matched" and e.get("ocr_date") in (None, "", "unknown"):
        mt_date = (e.get("matched_transaction") or {}).get("date")
        if mt_date:
            e["ocr_date_original"] = e.get("ocr_date")
            e["ocr_date"] = mt_date
    return e


def _load_ledger(username: str) -> dict:
    """Load ledger, auto-migrating v1 (list) to v2 (dict) format."""
    f = _receipts_file(username)
    raw = None
    if f.exists():
        try:
            raw = json.loads(f.read_text())
        except Exception:
            raw = None
    if raw is None:
        return {"version": 2, "last_batch_at": None, "entries": []}
    if isinstance(raw, list):  # v1 → v2
        return {"version": 2, "last_batch_at": None,
                "entries": [_migrate_entry(e) for e in raw]}
    raw["entries"] = [_migrate_entry(e) for e in raw.get("entries", [])]
    raw.setdefault("version", 2)
    raw.setdefault("last_batch_at", None)
    return raw


def _save_ledger(username: str, ledger: dict):
    _ensure_dirs()
    _receipts_file(username).write_text(json.dumps(ledger, ensure_ascii=False, indent=2))


def _ledger_entries(username: str) -> list:
    return _load_ledger(username)["entries"]


def _ledger_stats(entries: list) -> dict:
    matched = sum(1 for e in entries if e.get("match_status") == "matched")
    unmatched = sum(1 for e in entries if e.get("match_status") == "unmatched")
    pending = sum(1 for e in entries if e.get("match_status") == "pending_match")
    return {"total": len(entries), "matched": matched,
            "unmatched": unmatched, "pending_match": pending}


def _load_receipts(username: str) -> list:
    """Backward-compatible entries accessor (returns ledger entries list)."""
    return _ledger_entries(username)


def _save_receipts(username: str, receipts: list):
    """Persist entries list into v2 ledger, preserving wrapper metadata."""
    ledger = _load_ledger(username)
    ledger["entries"] = receipts
    _save_ledger(username, ledger)


def _drive_meta_file(username: str) -> Path:
    return DATA_DIR / f"drive_meta_{username}.json"


def _load_drive_meta(username: str) -> dict:
    f = _drive_meta_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return {}


def _save_drive_meta(username: str, meta: dict):
    _ensure_dirs()
    _drive_meta_file(username).write_text(json.dumps(meta, ensure_ascii=False, indent=2))


def _review_file(username: str) -> Path:
    return DATA_DIR / f"review_{username}.json"


def _load_review(username: str) -> dict:
    f = _review_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return {}


def _save_review(username: str, review: dict):
    _ensure_dirs()
    _review_file(username).write_text(json.dumps(review, ensure_ascii=False, indent=2))


# ── OCR Staging (pre-ledger visual review) ────────────────────────────────────

def _ocr_staging_file(username: str) -> Path:
    return DATA_DIR / f"ocr_staging_{username}.json"

def _load_ocr_staging(username: str) -> dict:
    f = _ocr_staging_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return {"entries": []}

def _save_ocr_staging(username: str, data: dict):
    _ensure_dirs()
    _ocr_staging_file(username).write_text(json.dumps(data, ensure_ascii=False, indent=2))

def _clear_ocr_staging(username: str):
    f = _ocr_staging_file(username)
    if f.exists():
        f.unlink()

def _handle_ocr_staging_confirm(username: str, body: dict):
    """POST /cardconv/receipts/review/confirm (JSON) — apply manual edits and add to ledger.

    Expects JSON body: {"confirmed": [{"id":..., "ocr_date":..., "ocr_merchant":...,
                                       "ocr_amount":..., "ocr_handwritten_amount":...}, ...]}
    Also accepts legacy FormData with just confirmed[] IDs (no edits).
    """
    staging = _load_ocr_staging(username)
    all_entries = staging.get("entries", [])

    # Build a correction map from the JSON payload.
    confirmed_list = body.get("confirmed", [])
    if isinstance(confirmed_list, list) and confirmed_list and isinstance(confirmed_list[0], dict):
        # JSON path: each item has id + corrected fields
        corrections = {item["id"]: item for item in confirmed_list if "id" in item}
        confirmed_ids = set(corrections.keys())
    else:
        # Legacy FormData path: just a list of IDs, no corrections
        if isinstance(confirmed_list, str):
            confirmed_list = [confirmed_list]
        confirmed_ids = set(confirmed_list)
        corrections = {}

    confirmed = []
    for e in all_entries:
        if e.get("id") not in confirmed_ids:
            continue
        entry = dict(e)
        fix = corrections.get(e["id"], {})
        if fix.get("ocr_date") is not None:
            entry["ocr_date"] = fix["ocr_date"] or None
        if fix.get("ocr_merchant") is not None:
            entry["ocr_merchant"] = fix["ocr_merchant"] or None
        for amt_key in ("ocr_amount", "ocr_handwritten_amount"):
            raw = fix.get(amt_key)
            if raw is not None:
                try:
                    entry[amt_key] = float(raw) if str(raw).strip() else None
                except (ValueError, TypeError):
                    pass
        # Recompute final amount after manual correction.
        hw = entry.get("ocr_handwritten_amount")
        pr = entry.get("ocr_amount")
        entry["ocr_final_amount"] = hw if hw is not None else pr
        confirmed.append(entry)

    if confirmed:
        receipts = _load_receipts(username)
        receipts.extend(confirmed)
        _save_receipts(username, receipts)

    _clear_ocr_staging(username)
    return ("json", {"ok": True, "added": len(confirmed)})


# ── User settings: card member names ──────────────────────────────────────────

DEFAULT_CARD_NAMES = ["JONG KANG", "JONGHA KANG"]


def _user_settings_file(username: str) -> Path:
    return DATA_DIR / f"user_settings_{username}.json"


def _load_user_settings(username: str) -> dict:
    f = _user_settings_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return {}


def _save_user_settings(username: str, data: dict):
    _ensure_dirs()
    _user_settings_file(username).write_text(json.dumps(data, ensure_ascii=False, indent=2))


def _get_card_member_names(username: str) -> list:
    """User's card member names (uppercased). Defaults to JONG/JONGHA KANG."""
    if not username:
        return list(DEFAULT_CARD_NAMES)
    names = _load_user_settings(username).get("card_member_names")
    if not names:
        return list(DEFAULT_CARD_NAMES)
    return [n.strip().upper() for n in names if str(n).strip()]


# ── Uploaded CSV store (Convert page reuse) ────────────────────────────────────

def _uploads_dir(username: str) -> Path:
    return DATA_DIR / f"uploads_{username}"


def _uploads_index_file(username: str) -> Path:
    return DATA_DIR / f"uploads_{username}.json"


def _load_uploads(username: str) -> list:
    f = _uploads_index_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return []


def _save_uploads(username: str, items: list):
    _ensure_dirs()
    _uploads_index_file(username).write_text(json.dumps(items, ensure_ascii=False, indent=2))


def _save_uploaded_csv(username: str, csv_bytes: bytes, orig_name: str,
                       rows: int, out_filename: str) -> dict:
    """Persist the raw CSV under uploads_{user}/ and index it for reuse."""
    d = _uploads_dir(username)
    d.mkdir(parents=True, exist_ok=True)
    uid = "csv_" + uuid.uuid4().hex[:8]
    (d / f"{uid}.csv").write_bytes(csv_bytes)
    entry = {
        "id":           uid,
        "filename":     orig_name,
        "stored_name":  f"{uid}.csv",
        "uploaded_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
        "rows":         rows,
        "out_filename": out_filename,
    }
    items = _load_uploads(username)
    items.insert(0, entry)
    _save_uploads(username, items)
    return entry


# ── Drive OAuth helpers ───────────────────────────────────────────────────────

def _get_creds(username: str):
    """Load and auto-refresh OAuth credentials for user. Returns None if not connected."""
    _ensure_dirs()
    token_file = TOKENS_DIR / f"{username}.json"
    if not CREDS_FILE.exists() or not token_file.exists():
        return None
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        creds = Credentials.from_authorized_user_file(str(token_file), SCOPES)
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            token_file.write_text(creds.to_json())
        if creds and creds.valid:
            return creds
    except Exception:
        pass
    return None


def _is_drive_connected(username: str) -> bool:
    return _get_creds(username) is not None


def _get_drive_service(username: str):
    from googleapiclient.discovery import build
    creds = _get_creds(username)
    if not creds:
        return None
    return build('drive', 'v3', credentials=creds)


def _get_or_create_folder(service, name: str, parent_id: str = None) -> str:
    """Return Drive folder ID, creating it if it doesn't exist."""
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        q += f" and '{parent_id}' in parents"
    res = service.files().list(q=q, fields='files(id)').execute()
    files = res.get('files', [])
    if files:
        return files[0]['id']
    meta = {'name': name, 'mimeType': 'application/vnd.google-apps.folder'}
    if parent_id:
        meta['parents'] = [parent_id]
    return service.files().create(body=meta, fields='id').execute()['id']


def _get_receipts_folder_ids(service, username: str) -> tuple:
    """Return (receipts_folder_id, matched_folder_id), creating folders if needed."""
    wayfinder_id = _get_or_create_folder(service, 'Wayfinder')
    receipts_id  = _get_or_create_folder(service, 'Receipts', wayfinder_id)
    matched_id   = _get_or_create_folder(service, 'Matched', receipts_id)
    # Cache receipts folder ID for UI link
    meta = _load_drive_meta(username)
    if meta.get('receipts_folder_id') != receipts_id:
        meta['receipts_folder_id'] = receipts_id
        _save_drive_meta(username, meta)
    return receipts_id, matched_id


def _upload_file_to_drive(username: str, file_bytes: bytes, filename: str,
                           mime_type: str) -> tuple:
    """Upload to Drive under Wayfinder/Receipts/. Returns (file_id, drive_url)."""
    from googleapiclient.http import MediaIoBaseUpload
    service = _get_drive_service(username)
    if not service:
        return None, None
    receipts_id, _ = _get_receipts_folder_ids(service, username)
    media  = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type)
    result = service.files().create(
        body={'name': filename, 'parents': [receipts_id]},
        media_body=media,
        fields='id,webViewLink'
    ).execute()
    fid = result.get('id')
    url = result.get('webViewLink') or f'https://drive.google.com/file/d/{fid}/view'
    return fid, url


def _move_to_matched_folder(username: str, file_id: str) -> bool:
    """DEPRECATED — no longer called. Drive is a plain store and the ledger is the
    single source of truth, so receipts are never auto-moved to Matched/ anymore.
    Kept for backward compatibility; Sync scans both Receipts/ and Matched/.
    Move receipt file from Wayfinder/Receipts/ to Wayfinder/Receipts/Matched/."""
    try:
        service = _get_drive_service(username)
        if not service:
            return False
        receipts_id, matched_id = _get_receipts_folder_ids(service, username)
        service.files().update(
            fileId=file_id,
            addParents=matched_id,
            removeParents=receipts_id,
            fields='id,parents'
        ).execute()
        return True
    except Exception:
        return False


# ── OCR ───────────────────────────────────────────────────────────────────────

_OCR_PROMPT = (
    'This image may contain ONE OR MULTIPLE receipts (e.g. several small '
    'receipts scanned together on a single page). Identify EACH distinct receipt. '
    'For EACH receipt look CAREFULLY for handwritten numbers (e.g. tip amount, '
    'final total written by hand on top of the printed receipt). '
    'Inspect tip line, total line, and any margin notes for handwriting. '
    'For each receipt extract: '
    '1) date (YYYY-MM-DD), '
    '2) merchant name, '
    '3) printed_amount: the PRINTED/typed total (number only), '
    '4) handwritten_amount: any HAND-WRITTEN final amount including tip (number only, '
    'null if none visible). '
    'Handwritten amount, when present, is the REAL final amount and overrides printed. '
    '5) card_brand: the payment card brand used, normalized to one of "amex", '
    '"visa", "other", or null only if there is NO card information at all (e.g. '
    'cash payment). Determine it from BOTH of these signals: '
    '(a) brand text/logo on the receipt — "AMERICAN EXPRESS"/"AMEX"/"AMX" -> "amex", '
    '"VISA" -> "visa", any other brand (Mastercard, Discover, etc.) -> "other"; AND '
    '(b) the card account number, even when masked (e.g. "XXXX-XXXXXX-X1234", '
    '"************1234", "ending in 1234", "AETC 3759"): use the FIRST visible digit '
    '- a number starting with 3 (15-digit, 4-6-5 grouping) -> "amex"; starting with 4 '
    '-> "visa"; starting with 2 or 5 -> "other". If (a) and (b) disagree, prefer the '
    'explicit brand text. AMEX receipts often show "AMEX", "AETC", or a 15-digit / '
    '3-prefixed card number — treat any of these as "amex". '
    'For each receipt, ALSO return its bounding box in the image as bbox: '
    '[ymin, xmin, ymax, xmax] using a 0-1000 normalized coordinate system '
    '(0=top/left, 1000=bottom/right). '
    'Return a JSON ARRAY ONLY, one object per receipt: '
    '[{"date":"YYYY-MM-DD","merchant":"name","printed_amount":0.00,"handwritten_amount":null,'
    '"card_brand":"amex","bbox":[100,50,800,950]}]. '
    'If only one receipt is visible, return an array with a single element.'
)


def _normalize_ocr(result: dict) -> dict:
    """Derive the final `amount` (handwritten priority) and coerce numeric fields.

    Keeps backward compat with the old single-`amount` shape: if a model still returns
    only `amount`, it is treated as the printed total.
    """
    if not result:
        return result

    def _num(v):
        if v is None:
            return None
        try:
            return round(float(v), 2)
        except (ValueError, TypeError):
            return None

    printed = _num(result.get("printed_amount"))
    handwritten = _num(result.get("handwritten_amount"))
    legacy = _num(result.get("amount"))
    if printed is None and legacy is not None:
        printed = legacy
    result["printed_amount"] = printed
    result["handwritten_amount"] = handwritten
    # Handwritten (tip-inclusive) total wins; fall back to printed.
    result["amount"] = handwritten if handwritten is not None else printed
    # bbox: [ymin, xmin, ymax, xmax] in 0-1000 normalized coords (None if absent/invalid).
    result["bbox"] = _coerce_bbox(result.get("bbox"))
    result["card_brand"] = _coerce_card_brand(result.get("card_brand"))
    return result


def _coerce_card_brand(v):
    """Normalize a model's card-brand string to 'amex'/'visa'/'other' or None."""
    if not v:
        return None
    s = str(v).strip().lower()
    if not s or s in ("null", "none", "unknown"):
        return None
    if "amex" in s or "american express" in s or s == "amx":
        return "amex"
    if "visa" in s:
        return "visa"
    return "other"


def _coerce_bbox(v):
    """Validate a model bbox into [ymin, xmin, ymax, xmax] ints, or None."""
    if not isinstance(v, (list, tuple)) or len(v) != 4:
        return None
    try:
        box = [int(round(float(x))) for x in v]
    except (ValueError, TypeError):
        return None
    # Clamp to the 0-1000 normalized range and require a positive area.
    box = [max(0, min(1000, c)) for c in box]
    ymin, xmin, ymax, xmax = box
    if ymax <= ymin or xmax <= xmin:
        return None
    return box


def _extract_ocr_list(text: str) -> list:
    """Parse a model OCR response into a list of normalized receipt dicts.

    Accepts either a JSON ARRAY (multi-receipt, current prompt) or a single
    JSON OBJECT (back-compat). Returns [] when nothing parseable is found.
    """
    if not text:
        return []
    text = text.strip()
    parsed = None
    # Prefer a top-level array (one image may hold multiple receipts).
    m = re.search(r'\[.*\]', text, re.DOTALL)
    if m:
        try:
            parsed = json.loads(m.group(0))
        except Exception:
            parsed = None
    if parsed is None:  # fall back to a single object
        m = re.search(r'\{.*\}', text, re.DOTALL)
        if m:
            try:
                parsed = json.loads(m.group(0))
            except Exception:
                parsed = None
    if parsed is None:
        return []
    if isinstance(parsed, dict):
        parsed = [parsed]
    return [_normalize_ocr(item) for item in parsed if isinstance(item, dict)]


def _ocr_receipt(file_bytes: bytes, mime_type: str) -> list:
    """OCR receipt(s) using Claude Vision. Returns a list of receipt dicts (may be empty)."""
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return []
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        b64 = base64.standard_b64encode(file_bytes).decode()
        if mime_type == 'application/pdf':
            block = {
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": b64}
            }
        else:
            block = {
                "type": "image",
                "source": {"type": "base64", "media_type": mime_type, "data": b64}
            }
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,  # room for multiple receipts in one image
            messages=[{"role": "user", "content": [
                block,
                {"type": "text", "text": _OCR_PROMPT}
            ]}]
        )
        results = _extract_ocr_list(resp.content[0].text)
        for r in results:
            r["_model"] = "Claude"
        return results
    except Exception:
        pass
    return []


# Configurable via GEMINI_OCR_MODEL (.env). gemini-2.5-flash: fast/cheap, strong on
# KR+EN receipts. Bump to gemini-2.5-pro if accuracy issues arise (~10x cost).
_DEFAULT_GEMINI_OCR_MODEL = "gemini-2.5-flash"


def _ocr_receipt_gemini(file_bytes: bytes, mime_type: str) -> list:
    """OCR receipt(s) using Gemini Vision. Returns a list of receipt dicts (may be empty)."""
    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        return []
    try:
        # New unified SDK (google-genai). Old google.generativeai is deprecated.
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=api_key)
        model_name = os.environ.get('GEMINI_OCR_MODEL', _DEFAULT_GEMINI_OCR_MODEL)
        # Inline bytes handle both images and PDFs (<20MB)
        resp = client.models.generate_content(
            model=model_name,
            contents=[
                types.Part.from_bytes(data=file_bytes, mime_type=mime_type),
                _OCR_PROMPT,
            ],
        )
        results = _extract_ocr_list(resp.text or "")
        for r in results:
            r["_model"] = "Gemini"
        return results
    except Exception:
        pass
    return []


def _ocr_receipt_auto(file_bytes: bytes, mime_type: str) -> list:
    """Primary OCR: Gemini first, fall back to Claude. Returns a list of receipt dicts.

    One image may contain multiple receipts, so callers must iterate the result.
    """
    results = _ocr_receipt_gemini(file_bytes, mime_type)
    if results and any(r.get("amount") is not None for r in results):
        return results
    claude = _ocr_receipt(file_bytes, mime_type)
    return claude or results


def _sub_entry_id(file_id: str, sub_index: int) -> str:
    """Stable ledger entry id for the Nth receipt found in one image."""
    base = (file_id or uuid.uuid4().hex)[:8]
    return f"rcpt_{base}_{sub_index}"


def _ocr_entry_fields(ocr: dict) -> dict:
    """Map a normalized OCR dict to the ledger entry's ocr_* fields."""
    ocr_date = ocr.get("date")
    if ocr_date == "YYYY-MM-DD":  # treat placeholder as missing
        ocr_date = None
    has_ocr = ocr.get("amount") is not None
    return {
        "ocr_status":             "done" if has_ocr else "failed",
        "ocr_date":               ocr_date,
        "ocr_amount":             ocr.get("amount"),
        "ocr_printed_amount":     ocr.get("printed_amount"),
        "ocr_handwritten_amount": ocr.get("handwritten_amount"),
        "ocr_merchant":           ocr.get("merchant"),
        "ocr_model":              ocr.get("_model"),
        "ocr_bbox":               ocr.get("bbox"),
        "card_brand":             ocr.get("card_brand"),
    }


# ── Multipart file parser ─────────────────────────────────────────────────────

_MIME_MAP = {
    'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
    'png': 'image/png',  'pdf': 'application/pdf',
    'gif': 'image/gif',  'webp': 'image/webp',
}


def _parse_multipart_files(raw_handler) -> list:
    """Parse multipart body; returns list of (filename, bytes, mime_type)."""
    ct = raw_handler.headers.get("Content-Type", "")
    m  = re.search(r'boundary=([^\s;]+)', ct)
    if not m:
        return []
    boundary = ("--" + m.group(1)).encode()
    length   = int(raw_handler.headers.get("Content-Length", 0))
    data     = raw_handler.rfile.read(length)
    files    = []
    for part in data.split(boundary):
        if b'filename="' not in part:
            continue
        fn_m = re.search(rb'filename="([^"]+)"', part)
        if not fn_m:
            continue
        filename = fn_m.group(1).decode('utf-8', errors='replace')
        hdr_end  = part.find(b"\r\n\r\n")
        if hdr_end == -1:
            continue
        content = part[hdr_end + 4:]
        if content.endswith(b"\r\n"):
            content = content[:-2]
        ext  = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        mime = _MIME_MAP.get(ext, 'application/octet-stream')
        if content:
            files.append((filename, content, mime))
    return files


# ── Conversion logic ──────────────────────────────────────────────────────────

def _classify(merchant: str, rules: list):
    m = merchant.upper()
    for r in rules:
        if r["kw"].upper() in m:
            return r["gl"], r["ser"], r["purpose"]
    return None, None, None


def _fmt_card(acct: str):
    acct = re.sub(r'\D', '', acct)
    if len(acct) >= 8:
        masked = acct[:4] + "********" + acct[-4:]
        supp   = f"{acct[:4]}-{acct[4:10]}-{acct[10:]}" if len(acct) == 15 else acct
    else:
        masked = acct; supp = acct
    return masked, supp


def _name_parts(full: str):
    p = full.strip().split()
    return (p[-1], " ".join(p[:-1])) if len(p) >= 2 else (full, "")


def _parse_date(s: str):
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            pass
    return None


def convert(csv_bytes: bytes, filename: str, username: str = None) -> tuple:
    """Returns (xlsx_bytes, out_filename, total_rows, unmatched_count)."""
    rules  = _load_kw()
    today  = date.today()
    m      = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    tag    = m.group(1) if m else today.strftime("%Y-%m-%d")
    out_fn = f"for_upload_{tag}.xlsx"

    if TEMPLATE.exists():
        template_path = TEMPLATE
    elif TEMPLATE_FALLBACK.exists():
        template_path = TEMPLATE_FALLBACK
    else:
        raise FileNotFoundError("Template file not found. Please upload 'for upload.xlsx' to ~/.appdata/cardconv/template.xlsx")

    target_names = set(_get_card_member_names(username))
    reader = csv.DictReader(io.TextIOWrapper(io.BytesIO(csv_bytes), encoding='utf-8-sig', newline=''))
    rows   = [r for r in reader if r.get("Card Member Name","").strip().upper() in target_names]

    wb = openpyxl.load_workbook(template_path)
    ws = wb["sheetMst"]

    start = 2
    while ws.cell(start, 1).value is not None:
        start += 1

    posting_dt = datetime(today.year, today.month, today.day)
    unmatched  = 0
    review_rows     = []   # per-transaction rows for the Review page
    receipt_matched = 0    # transactions with a matched receipt

    # Build receipt lookup: (date_str, rounded_amount) → receipt record
    receipts      = []
    receipts_map  = {}
    receipts_dirty = False
    if username:
        receipts = _load_receipts(username)
        for r in receipts:
            if r.get("completed"):
                continue  # completed receipts are archived — excluded from mapping
            rdate = r.get("ocr_date")
            if rdate == "YYYY-MM-DD":
                rdate = None
            # Handwritten amount takes priority as the final match target.
            hw = r.get("ocr_handwritten_amount")
            pr = r.get("ocr_printed_amount") or r.get("ocr_amount")
            ramount = hw if hw is not None else pr
            if ramount is not None:
                try:
                    key = (rdate, round(float(ramount), 2))
                    receipts_map[key] = r
                except (ValueError, TypeError):
                    pass

    for row in rows:
        merchant = row.get("Merchant Name", "").strip()
        dba      = row.get("Merchant Doing Business As", "").strip()
        vendor   = dba if (dba and dba != merchant) else merchant
        try:
            amount = float(row.get("Amount", 0))
        except ValueError:
            amount = 0.0
        inv_dt       = _parse_date(row.get("Date", ""))
        masked, supp = _fmt_card(row.get("Account Number", ""))
        last, first  = _name_parts(row.get("Card Member Name", ""))

        gl, ser, purpose = _classify(vendor, rules)
        if gl is None:
            gl, ser, purpose = _classify(merchant, rules)
        if gl is None:
            unmatched += 1
            gl, ser, purpose = 53410177, "160", "Coffee, Snack and meal"

        ws.cell(start,  1).value = FIXED["receipt_type"]
        ws.cell(start,  2).value = FIXED["employee_id"]
        ws.cell(start,  3).value = FIXED["payee"]
        ws.cell(start,  4).value = None
        ws.cell(start,  5).value = inv_dt
        ws.cell(start,  6).value = FIXED["domestic"]
        ws.cell(start,  7).value = vendor
        ws.cell(start,  8).value = posting_dt
        ws.cell(start,  9).value = gl
        ws.cell(start, 10).value = ser
        ws.cell(start, 11).value = FIXED["currency"]
        ws.cell(start, 12).value = FIXED["tax_code"]
        ws.cell(start, 13).value = None
        ws.cell(start, 14).value = amount
        ws.cell(start, 15).value = FIXED["cost_center"]
        ws.cell(start, 16).value = None
        ws.cell(start, 17).value = None
        ws.cell(start, 18).value = purpose
        ws.cell(start, 19).value = None
        ws.cell(start, 20).value = masked
        ws.cell(start, 21).value = last
        ws.cell(start, 22).value = first
        ws.cell(start, 23).value = supp
        ws.cell(start, 24).value = None
        ws.cell(start, 25).value = None
        ws.cell(start, 26).value = amount

        # Receipt match in column 27; move matched receipts to Matched folder
        inv_date_str = inv_dt.strftime("%Y-%m-%d") if inv_dt else None
        amt_rounded  = round(amount, 2)
        receipt_match = None
        if receipts_map:
            receipt_match = receipts_map.get((inv_date_str, amt_rounded))
            if not receipt_match:
                for (rdate, ramt), r in receipts_map.items():
                    amt_match = abs(ramt - amt_rounded) <= 0.01
                    if not amt_match:
                        continue
                    if rdate is None:
                        date_match = True
                    elif rdate == inv_date_str:
                        date_match = True
                    else:
                        # Allow ±1 day for card posting date vs transaction date skew
                        try:
                            from datetime import timedelta
                            rd = date.fromisoformat(rdate)
                            id_ = date.fromisoformat(inv_date_str) if inv_date_str else None
                            date_match = id_ is not None and abs((rd - id_).days) <= 1
                        except Exception:
                            date_match = False
                    if date_match:
                        receipt_match = r
                        break
            if receipt_match:
                fn  = receipt_match.get('filename', '')
                url = receipt_match.get('drive_url', '')
                ws.cell(start, 27).value = f"✅ {fn} ({url})" if url else f"✅ {fn}"
                # Update match metadata always (even if Drive folder move fails).
                if not receipt_match.get('matched'):
                    receipt_match['matched'] = True
                    receipt_match['match_status'] = 'matched'
                    receipt_match['matched_at'] = datetime.now().isoformat()
                    receipt_match['matched_transaction'] = {
                        'date':   inv_date_str,
                        'amount': amt_rounded,
                        'vendor': vendor,
                    }
                    # Backfill OCR date from the matched CSV transaction when OCR
                    # failed to read it. Keep the original in ocr_date_original.
                    if receipt_match.get('ocr_date') in (None, '', 'unknown') and inv_date_str:
                        receipt_match['ocr_date_original'] = receipt_match.get('ocr_date')
                        receipt_match['ocr_date'] = inv_date_str
                    receipts_dirty = True
                    # NOTE: Drive folder auto-move is deprecated. Drive is a plain
                    # store; the Wayfinder ledger is the single source of truth.
            else:
                ws.cell(start, 27).value = "❌ Missing"

        # Collect a Review row for this transaction
        rcpt_info = None
        if receipt_match:
            receipt_matched += 1
            mfid = receipt_match.get("file_id")
            # Sibling receipts sharing the same source image (multi-receipt page).
            # Used to draw every bbox in the lightbox and highlight the matched one.
            siblings = [
                {"id": s.get("id"), "ocr_bbox": s.get("ocr_bbox")}
                for s in receipts
                if mfid and s.get("file_id") == mfid and s.get("ocr_bbox")
            ]
            rcpt_info = {
                "file_id":      mfid,
                "id":           receipt_match.get("id"),
                "filename":     receipt_match.get("filename"),
                "drive_url":    receipt_match.get("drive_url"),
                "ocr_amount":   receipt_match.get("ocr_amount"),
                "ocr_date":     receipt_match.get("ocr_date"),
                "ocr_merchant": receipt_match.get("ocr_merchant"),
                "ocr_bbox":     receipt_match.get("ocr_bbox"),
                "siblings":     siblings if len(siblings) > 1 else [],
            }
        review_rows.append({
            "id":          "rv_" + uuid.uuid4().hex[:8],
            "date":        inv_date_str,
            "merchant":    vendor,
            "amount":      amt_rounded,
            "gl":          gl,
            "ser":         ser,
            "purpose":     purpose,
            "matched":     bool(receipt_match),
            "receipt":     rcpt_info,
            "loss_reason": "",
        })

        start += 1

    if receipts_dirty and username:
        _save_receipts(username, receipts)

    # Persist Review snapshot for the Review page
    if username:
        _save_review(username, {
            "generated_at": datetime.now().isoformat(),
            "source":       filename,
            "out_filename": out_fn,
            "total":        len(rows),
            "matched":      receipt_matched,
            "unmatched":    len(rows) - receipt_matched,
            "kw_unmatched": unmatched,
            "rows":         review_rows,
        })

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    (OUT_DIR / out_fn).write_bytes(xlsx_bytes)
    return xlsx_bytes, out_fn, len(rows), unmatched


# ── Ledger ──────────────────────────────────────────────────────────────────

_SUPPORTED_MIME = {'image/jpeg', 'image/png', 'application/pdf', 'image/gif', 'image/webp'}


def _run_batch_ocr(username: str) -> dict:
    """Scan Drive, OCR new/pending files, update ledger. Returns stats dict."""
    service = _get_drive_service(username)
    if not service:
        return {"error": "drive not connected"}
    try:
        # Scan both Receipts/ and legacy Matched/ (back-compat + safety): the
        # auto-move is deprecated, but old setups may still have files in Matched/.
        receipts_id, matched_id = _get_receipts_folder_ids(service, username)
        results = service.files().list(
            q=(f"('{receipts_id}' in parents or '{matched_id}' in parents) "
               f"and trashed=false and mimeType!='application/vnd.google-apps.folder'"),
            fields="files(id,name,mimeType,webViewLink)"
        ).execute()
        drive_files = results.get('files', [])

        ledger  = _load_ledger(username)
        entries = ledger["entries"]

        # Skip files already confirmed in the ledger.
        done_fids = {e.get("file_id") for e in entries
                     if (e.get("multi_ocr") and e.get("ocr_amount") is not None)
                     or e.get("matched") or e.get("completed")}
        # Also skip files already waiting in the staging queue (accumulated from
        # previous runs the user hasn't reviewed yet).
        staging = _load_ocr_staging(username)
        staged_fids = {e.get("file_id") for e in staging.get("entries", [])}
        skip_fids = done_fids | staged_fids

        new_staged = []
        now_disp = datetime.now().strftime("%Y-%m-%d %H:%M")
        now_iso  = datetime.now().isoformat()

        for f in drive_files:
            fid  = f.get('id')
            mime = f.get('mimeType', '')
            if mime not in _SUPPORTED_MIME:
                continue
            if fid in skip_fids:
                continue
            content  = service.files().get_media(fileId=fid).execute()
            ocr_list = _ocr_receipt_auto(content, mime) or [{}]
            url      = f.get('webViewLink') or f'https://drive.google.com/file/d/{fid}/view'
            for sub_index, ocr in enumerate(ocr_list):
                entry = {
                    "id":           _sub_entry_id(fid, sub_index),
                    "file_id":      fid,
                    "sub_index":    sub_index,
                    "filename":     f.get('name', ''),
                    "drive_url":    url,
                    "mime_type":    mime,
                    "match_status": "pending_match",
                    "multi_ocr":    True,
                    "uploaded_at":  now_disp,
                    "synced_at":    now_iso,
                }
                entry.update(_ocr_entry_fields(ocr))
                new_staged.append(entry)

        # Accumulate into staging queue — never replace, always append.
        if new_staged:
            staging.setdefault("entries", []).extend(new_staged)
            staging["staged_at"] = now_iso
            _save_ocr_staging(username, staging)

        ledger["last_batch_at"] = now_iso
        _save_ledger(username, ledger)

        if new_staged:
            total_pending = len(staging.get("entries", []))
            _send_push_notification(
                username,
                title="🧾 New receipts ready to review",
                body=f"{len(new_staged)} new receipt{'s' if len(new_staged) != 1 else ''} synced from Drive — {total_pending} total pending review.",
                url="/cardconv/ledger",
            )
        return {"staged": len(new_staged), "total_pending": len(staging.get("entries", []))}
    except Exception as e:
        return {"error": str(e)}


def _handle_batch_run(username: str, ctx):
    """POST /cardconv/batch/run — X-Batch-Secret authenticated batch OCR trigger."""
    secret = os.environ.get("CARDCONV_BATCH_SECRET", "")
    headers = (ctx or {}).get("headers")
    provided = headers.get("X-Batch-Secret", "") if headers else ""
    if not secret or provided != secret:
        return ("json", {"error": "unauthorized"}, 401)
    return ("json", _run_batch_ocr(username))


def _apply_ledger_filters(entries: list, status: str, dfrom: str, dto: str,
                          card_brand: str = "", usage: str = "",
                          completed: str = "hide") -> list:
    """Filter + sort ledger entries by status, OCR-date range, card brand, usage
    and completion state.

    Shared by the JSON API, the PDF export and the xlsx export so all honor
    identical filters. Date filters keep entries without an OCR date always
    visible. `completed` is one of: "hide" (default — exclude completed),
    "only" (completed only), "all" (include both).
    """
    filtered = entries
    if status and status != "all":
        filtered = [e for e in filtered if e.get("match_status") == status]
    if dfrom:
        filtered = [e for e in filtered if not e.get("ocr_date") or e["ocr_date"] >= dfrom]
    if dto:
        filtered = [e for e in filtered if not e.get("ocr_date") or e["ocr_date"] <= dto]
    if card_brand and card_brand != "all":
        if card_brand == "unknown":
            filtered = [e for e in filtered if not e.get("card_brand")]
        else:
            filtered = [e for e in filtered if e.get("card_brand") == card_brand]
    if usage and usage != "all":
        filtered = [e for e in filtered if (e.get("usage") or "Regular") == usage]
    if completed == "only":
        filtered = [e for e in filtered if e.get("completed")]
    elif completed != "all":  # "hide" (default)
        filtered = [e for e in filtered if not e.get("completed")]
    return sorted(
        filtered,
        key=lambda e: e.get("ocr_date") or e.get("uploaded_at") or "",
        reverse=True,
    )


def _mark_duplicates(entries: list):
    """Flag likely-duplicate receipts in-place.

    Two receipts are duplicates when their final amount and merchant match AND
    their dates are compatible — equal, or at least one side missing (OCR often
    fails to read a date on one copy of the same receipt). Date is deliberately
    NOT part of the bucket key so a date-less entry still groups with its dated
    twin; multi-receipt sub-entries (different file_id, same OCR values) group
    too since file_id is ignored. Within a group the keeper / head is chosen by
    match_status priority (matched > unmatched > pending_match), then a present
    date; the rest are flagged for easy bulk-deletion.
    Sets e['dup'] (bool), e['dup_keep'] (bool) and e['dup_group_id'] (str|None)
    on each entry. dup_group_id lets the UI collapse a group into one row.
    """
    for e in entries:
        e["dup"] = False
        e["dup_keep"] = False
        e["dup_group_id"] = None

    # Bucket by (amount, merchant); date handled per-pair below.
    buckets = {}
    for e in entries:
        amt = e.get("ocr_amount")
        if amt is None:
            continue
        merch = (e.get("ocr_merchant") or "").strip().lower()
        buckets.setdefault((round(amt, 2), merch), []).append(e)

    gi = 0
    for bucket in buckets.values():
        if len(bucket) < 2:
            continue
        # Partition into date-compatible groups (equal date, or a missing date
        # absorbed into the first dated group it meets).
        used = [False] * len(bucket)
        for i in range(len(bucket)):
            if used[i]:
                continue
            group, used[i] = [bucket[i]], True
            anchor = bucket[i].get("ocr_date")
            for j in range(i + 1, len(bucket)):
                if used[j]:
                    continue
                dj = bucket[j].get("ocr_date")
                if anchor is None or dj is None or anchor == dj:
                    group.append(bucket[j])
                    used[j] = True
                    if anchor is None:
                        anchor = dj  # lock onto the first known date
            if len(group) < 2:
                continue
            gid = f"dg_{gi}"
            gi += 1
            # Keeper / group head: status priority first
            # (matched > unmatched > pending_match), then a present date,
            # then the earliest uploaded_at / id for a stable tiebreak.
            status_rank = {"matched": 0, "unmatched": 1, "pending_match": 2}
            group.sort(key=lambda e: (
                status_rank.get(e.get("match_status"), 3),
                e.get("ocr_date") is None,
                str(e.get("uploaded_at") or ""),
                str(e.get("id") or ""),
            ))
            for k, e in enumerate(group):
                e["dup"] = True
                e["dup_keep"] = (k == 0)
                e["dup_group_id"] = gid


def _handle_ledger_delete(username: str, body: dict):
    """POST /cardconv/ledger/delete — remove entries by id from the ledger.

    Body: {"ids": [...], "also_drive": bool}. When also_drive is true, the
    corresponding Drive originals are moved to the trash (best-effort).
    """
    raw = body.get("ids", [])
    ids = {str(i) for i in (raw if isinstance(raw, list) else [raw]) if i}
    if not ids:
        return ("json", {"error": "no ids"}, 400)
    also_drive = bool(body.get("also_drive"))

    ledger = _load_ledger(username)
    removed_entries = [e for e in ledger["entries"] if e.get("id") in ids]
    ledger["entries"] = [e for e in ledger["entries"] if e.get("id") not in ids]
    _save_ledger(username, ledger)

    trashed = 0
    if also_drive and removed_entries:
        service = _get_drive_service(username)
        if service:
            for e in removed_entries:
                fid = e.get("file_id")
                if not fid:
                    continue
                try:
                    service.files().update(fileId=fid, body={"trashed": True}).execute()
                    trashed += 1
                except Exception:
                    pass  # best-effort; ledger removal already succeeded
    return ("json", {"ok": True, "removed": len(removed_entries), "trashed": trashed})


def _get_completed_folder_id(service, username: str) -> str:
    """Drive folder ID for Wayfinder/Receipts/Completed/, created on demand."""
    receipts_id, _ = _get_receipts_folder_ids(service, username)
    return _get_or_create_folder(service, 'Completed', receipts_id)


def _archive_to_completed(username: str, file_ids: set) -> int:
    """Best-effort move of Drive originals into the Completed folder.

    Returns the count moved. The ledger `completed` flag is the source of truth,
    so any Drive failure is swallowed and reported only via the returned count.
    """
    if not file_ids:
        return 0
    service = _get_drive_service(username)
    if not service:
        return 0
    try:
        completed_id = _get_completed_folder_id(service, username)
        receipts_id, matched_id = _get_receipts_folder_ids(service, username)
    except Exception:
        return 0
    moved = 0
    for fid in file_ids:
        if not fid:
            continue
        try:
            # Drop both possible source parents; addParents is idempotent.
            cur = service.files().get(fileId=fid, fields='parents').execute()
            parents = set(cur.get('parents', []))
            remove = ",".join(parents & {receipts_id, matched_id}) or None
            service.files().update(
                fileId=fid, addParents=completed_id,
                removeParents=remove, fields='id,parents'
            ).execute()
            moved += 1
        except Exception:
            pass  # best-effort; ledger flag already reflects completion
    return moved


def _restore_from_completed(username: str, file_ids: set) -> int:
    """Best-effort move of Drive originals back from Completed to Receipts."""
    if not file_ids:
        return 0
    service = _get_drive_service(username)
    if not service:
        return 0
    try:
        completed_id = _get_completed_folder_id(service, username)
        receipts_id, _ = _get_receipts_folder_ids(service, username)
    except Exception:
        return 0
    moved = 0
    for fid in file_ids:
        if not fid:
            continue
        try:
            service.files().update(
                fileId=fid, addParents=receipts_id,
                removeParents=completed_id, fields='id,parents'
            ).execute()
            moved += 1
        except Exception:
            pass
    return moved


def _handle_ledger_complete(username: str, body: dict):
    """POST /cardconv/ledger/complete — mark entries complete (or undo).

    Body: {"ids": [...], "undo": bool}. Completing flags entries and moves their
    Drive originals to the Completed folder; the ledger flag is authoritative so
    Drive moves are best-effort. Multiple ledger entries can share one Drive file
    (multi-receipt page) — that file is only moved when ALL its entries agree.
    """
    raw = body.get("ids", [])
    ids = {str(i) for i in (raw if isinstance(raw, list) else [raw]) if i}
    if not ids:
        return ("json", {"error": "no ids"}, 400)
    undo = bool(body.get("undo"))

    ledger = _load_ledger(username)
    now = datetime.now().isoformat()
    touched = []
    for e in ledger["entries"]:
        if e.get("id") in ids:
            e["completed"] = not undo
            e["completed_at"] = None if undo else now
            touched.append(e)

    # A Drive file is moved only when none of its sibling entries remain in the
    # opposite state, so a multi-receipt page isn't archived prematurely.
    by_file = {}
    for e in ledger["entries"]:
        fid = e.get("file_id")
        if fid:
            by_file.setdefault(fid, []).append(e)
    move_fids = set()
    for e in touched:
        fid = e.get("file_id")
        sibs = by_file.get(fid, [e])
        all_completed = all(s.get("completed") for s in sibs)
        if not undo and all_completed:
            move_fids.add(fid)
        elif undo:
            move_fids.add(fid)  # any sibling un-completed → bring file back

    if undo:
        moved = _restore_from_completed(username, move_fids)
        for e in touched:
            e["archived_drive"] = False
    else:
        moved = _archive_to_completed(username, move_fids)
        for e in touched:
            if e.get("file_id") in move_fids:
                e["archived_drive"] = True
    _save_ledger(username, ledger)
    return ("json", {"ok": True, "count": len(touched), "moved": moved})


def _parse_filter_params(query: dict) -> dict:
    """Extract the shared Ledger filter params from a query dict.

    Used by the JSON API, PDF export and xlsx export so all interpret the
    status/date/card-brand/usage/completed filters identically.
    """
    def _q(key, default):
        return (query.get(key, [default]) or [default])[0]
    return {
        "status":     _q("status", "all"),
        "dfrom":      _q("from", ""),
        "dto":        _q("to", ""),
        "card_brand": _q("card_brand", "all"),
        "usage":      _q("usage", "all"),
        "completed":  _q("completed", "hide"),
    }


def _handle_ledger_api(username: str, query: dict):
    """GET /cardconv/ledger/api — filtered JSON data."""
    ledger  = _load_ledger(username)
    entries = ledger["entries"]
    f = _parse_filter_params(query)
    try:
        page = max(1, int((query.get("page", ["1"]) or ["1"])[0]))
    except ValueError:
        page = 1
    try:
        limit = max(1, int((query.get("limit", ["50"]) or ["50"])[0]))
    except ValueError:
        limit = 50

    filtered = _apply_ledger_filters(entries, f["status"], f["dfrom"], f["dto"],
                                     f["card_brand"], f["usage"], f["completed"])
    _mark_duplicates(filtered)

    stats   = _ledger_stats(filtered)
    total_f = len(filtered)
    pages   = max(1, (total_f + limit - 1) // limit)
    start   = (page - 1) * limit
    # Distinct usage tags across the whole ledger (for the filter dropdown), and
    # a completed count so the UI can surface how many are archived.
    usages = sorted({(e.get("usage") or "Regular") for e in entries})
    completed_n = sum(1 for e in entries if e.get("completed"))
    return ("json", {
        "total":       stats["total"],
        "matched":     stats["matched"],
        "unmatched":   stats["unmatched"],
        "pending_match": stats["pending_match"],
        "completed":   completed_n,
        "usages":      usages,
        "page":        page,
        "pages":       pages,
        "last_synced": ledger.get("last_batch_at"),
        "entries":     filtered[start:start + limit],
    })


def _handle_status_change(username: str, entry_id: str, body: dict):
    """POST /cardconv/ledger/<id>/status — manual status override."""
    raw = body.get("status", "")
    status = (raw[0] if isinstance(raw, list) else str(raw)).strip()
    if status not in ("matched", "unmatched", "pending_match"):
        return ("json", {"error": "invalid status"}, 400)
    ledger = _load_ledger(username)
    for e in ledger["entries"]:
        if e.get("id") == entry_id:
            e["match_status"] = status
            e["matched"] = (status == "matched")
            if status == "matched":
                e["matched_at"] = datetime.now().isoformat()
            _save_ledger(username, ledger)
            return ("json", {"ok": True})
    return ("json", {"error": "not found"}, 404)


def _handle_rematch(username: str, entry_id: str):
    """POST /cardconv/ledger/<id>/rematch — re-try CSV matching for a pending/unmatched entry."""
    ledger = _load_ledger(username)
    entries = ledger["entries"]
    entry = next((e for e in entries if e.get("id") == entry_id), None)
    if not entry:
        return ("json", {"error": "not found"}, 404)

    rdate = entry.get("ocr_date")
    if rdate == "YYYY-MM-DD":
        rdate = None
    # Handwritten amount takes priority over printed as the final match target.
    hw = entry.get("ocr_handwritten_amount")
    pr = entry.get("ocr_printed_amount") or entry.get("ocr_amount")
    raw_amount = hw if hw is not None else pr
    if raw_amount is None:
        return ("json", {"error": "no OCR amount to match against"}, 400)
    try:
        ramount = round(float(raw_amount), 2)
    except (ValueError, TypeError):
        return ("json", {"error": "invalid amount"}, 400)

    uploads = _load_uploads(username)
    uploads_dir_path = _uploads_dir(username)
    target_names = set(_get_card_member_names(username))
    matched_tx = None

    for upload in uploads:
        csv_path = uploads_dir_path / upload.get("stored_name", "")
        if not csv_path.exists():
            continue
        try:
            csv_bytes = csv_path.read_bytes()
            reader = csv.DictReader(
                io.TextIOWrapper(io.BytesIO(csv_bytes), encoding='utf-8-sig', newline=''))
            rows = [r for r in reader
                    if r.get("Card Member Name", "").strip().upper() in target_names]
        except Exception:
            continue

        for row in rows:
            try:
                amount = round(float(row.get("Amount", 0)), 2)
            except ValueError:
                continue
            if abs(amount - ramount) > 0.01:
                continue

            inv_dt = _parse_date(row.get("Date", ""))
            inv_date_str = inv_dt.strftime("%Y-%m-%d") if inv_dt else None

            if rdate is None or inv_date_str is None:
                date_match = True
            elif rdate == inv_date_str:
                date_match = True
            else:
                try:
                    from datetime import timedelta as _td
                    rd  = date.fromisoformat(rdate)
                    id_ = date.fromisoformat(inv_date_str)
                    date_match = abs((rd - id_).days) <= 1
                except Exception:
                    date_match = False

            if date_match:
                merchant = row.get("Merchant Name", "").strip()
                dba      = row.get("Merchant Doing Business As", "").strip()
                matched_tx = {
                    "date":   inv_date_str,
                    "amount": amount,
                    "vendor": dba if (dba and dba != merchant) else merchant,
                }
                break
        if matched_tx:
            break

    now_iso = datetime.now().isoformat()
    if matched_tx:
        entry["matched"]             = True
        entry["match_status"]        = "matched"
        entry["matched_at"]          = now_iso
        entry["matched_transaction"] = matched_tx
        if entry.get("ocr_date") in (None, "", "unknown") and matched_tx.get("date"):
            entry["ocr_date_original"] = entry.get("ocr_date")
            entry["ocr_date"]          = matched_tx["date"]
    else:
        entry["match_status"] = "unmatched"
        entry["matched"]      = False

    _save_ledger(username, ledger)
    return ("json", {"ok": True, "matched": bool(matched_tx), "entry": entry})


def _handle_reocr(username: str, entry_id: str):
    """POST /cardconv/ledger/<id>/reocr — re-run OCR for all entries sharing the same file_id."""
    service = _get_drive_service(username)
    if not service:
        return ("json", {"error": "Drive not connected"}, 401)
    ledger = _load_ledger(username)
    entries = ledger["entries"]

    # Find the target entry to get its file_id
    target = next((e for e in entries if e.get("id") == entry_id), None)
    if not target:
        return ("json", {"error": "not found"}, 404)
    file_id = target.get("file_id")
    if not file_id:
        return ("json", {"error": "no file_id on entry"}, 400)

    try:
        meta    = service.files().get(fileId=file_id, fields="mimeType,name,webViewLink").execute()
        mime    = meta.get("mimeType", "image/jpeg")
        content = service.files().get_media(fileId=file_id).execute()
    except Exception as e:
        return ("json", {"error": f"Drive fetch failed: {e}"}, 500)

    ocr_list = _ocr_receipt_auto(content, mime) or [{}]

    # Preserve non-OCR fields from the existing sibling entries (matched at same sub_index)
    siblings = {e.get("sub_index", 0): e for e in entries if e.get("file_id") == file_id}

    # Remove all existing entries for this file_id
    entries[:] = [e for e in entries if e.get("file_id") != file_id]

    url     = meta.get("webViewLink") or f"https://drive.google.com/file/d/{file_id}/view"
    now_iso = datetime.now().isoformat()
    updated = []
    for sub_index, ocr in enumerate(ocr_list):
        fields  = _ocr_entry_fields(ocr)
        sibling = siblings.get(sub_index, {})
        entry   = {
            "id":           _sub_entry_id(file_id, sub_index),
            "file_id":      file_id,
            "sub_index":    sub_index,
            "filename":     sibling.get("filename") or meta.get("name", ""),
            "drive_url":    url,
            "mime_type":    mime,
            "match_status": sibling.get("match_status", "pending_match"),
            "matched":      sibling.get("matched", False),
            "multi_ocr":    True,
            "uploaded_at":  sibling.get("uploaded_at", now_iso),
            "synced_at":    now_iso,
            "reocr_at":     now_iso,
        }
        if sibling.get("matched_at"):
            entry["matched_at"] = sibling["matched_at"]
        if sibling.get("matched_transaction"):
            entry["matched_transaction"] = sibling["matched_transaction"]
        entry.update(fields)
        entries.append(entry)
        updated.append(entry)

    _save_ledger(username, ledger)
    return ("json", {"ok": True, "updated": updated})


def _handle_image_proxy(username: str, file_id: str, bbox: list = None):
    """GET /cardconv/receipts/image/<file_id>[?bbox=ymin,xmin,ymax,xmax] — Drive proxy with optional crop.
    Always auto-rotates via EXIF so receipts display upright.
    """
    service = _get_drive_service(username)
    if not service:
        return ("html", "<p>Drive not connected</p>", 401)
    try:
        from PIL import Image as _Img, ImageOps as _IOP
        content = service.files().get_media(fileId=file_id).execute()
        img = _Img.open(io.BytesIO(content))
        img = _IOP.exif_transpose(img)   # auto-rotate
        img = img.convert("RGB")

        if bbox and len(bbox) == 4:
            w, h = img.size
            ymin, xmin, ymax, xmax = [max(0, min(1000, v)) for v in bbox]
            left   = int(xmin / 1000 * w)
            upper  = int(ymin / 1000 * h)
            right  = int(xmax / 1000 * w)
            lower  = int(ymax / 1000 * h)
            pad = max(6, min(w, h) // 40)
            left, upper = max(0, left - pad), max(0, upper - pad)
            right, lower = min(w, right + pad), min(h, lower + pad)
            img = img.crop((left, upper, right, lower))

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=88, optimize=True)
        return ("binary", buf.getvalue(), "image/jpeg", None)
    except Exception as e:
        return ("html", f"<p>Image load error: {e}</p>", 404)


# ── PDF export ────────────────────────────────────────────────────────────────

_DEJAVU_DIR = "/usr/share/fonts/truetype/dejavu"
# Font candidates, preferred first. A CJK font (e.g. apt install fonts-nanum on
# the server) renders Korean merchant names; DejaVu covers Latin/symbols without
# crashing on non-ASCII. Each entry is (family, regular_path, bold_path).
_PDF_FONT_CANDIDATES = [
    ("Nanum", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
              "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
    ("NotoKR", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
               "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"),
    ("DejaVu", f"{_DEJAVU_DIR}/DejaVuSans.ttf",
               f"{_DEJAVU_DIR}/DejaVuSans-Bold.ttf"),
]
_PDF_STATUS_LABEL = {"matched": "Matched", "unmatched": "Unmatched",
                     "pending_match": "Pending Match"}
_PDF_STATUS_COLOR = {"matched": (22, 163, 74), "unmatched": (220, 38, 38),
                     "pending_match": (217, 119, 6)}


def _compress_receipt_image(raw: bytes, max_w: int = 700, quality: int = 72):
    """Resize + recompress a receipt image into a small JPEG.

    Returns (jpeg_bytes, (w, h)) or None on failure. Auto-rotates via EXIF,
    caps width at max_w px, and re-encodes as JPEG to keep the PDF small.
    """
    from PIL import Image, ImageOps
    try:
        img = Image.open(io.BytesIO(raw))
        img = ImageOps.exif_transpose(img)   # honour EXIF orientation tag
        img = img.convert("RGB")             # flatten alpha/CMYK/palette for JPEG
        if img.width > max_w:
            h = max(1, round(img.height * max_w / img.width))
            img = img.resize((max_w, h), Image.LANCZOS)
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=quality, optimize=True)
        return out.getvalue(), img.size
    except Exception:
        return None


def _fetch_drive_image(service, file_id: str):
    """Fetch raw image bytes from Drive, or None if unavailable."""
    if not service or not file_id:
        return None
    try:
        return service.files().get_media(fileId=file_id).execute()
    except Exception:
        return None


def _handle_ledger_pdf(username: str, query: dict):
    """GET /cardconv/ledger/download.pdf — landscape A4, 5 cols × 2 rows = 10 per page.
    Images only (no text), EXIF auto-rotated, maximally sized with minimal gaps.
    Multi-receipt source images span extra columns.
    """
    from fpdf import FPDF

    f = _parse_filter_params(query)
    status, dfrom, dto = f["status"], f["dfrom"], f["dto"]
    entries = _apply_ledger_filters(_ledger_entries(username), status, dfrom, dto,
                                    f["card_brand"], f["usage"], f["completed"])
    stats   = _ledger_stats(entries)
    service = _get_drive_service(username)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False)
    MRG  = 5          # tight margins to maximise image area
    GAP  = 1.0        # gap between cells
    PAD  = 0.5        # inner padding within each cell
    pdf.set_margins(MRG, MRG, MRG)

    FONT, unicode_ok = "Helvetica", False
    for fam, reg, bold in _PDF_FONT_CANDIDATES:
        if os.path.exists(reg) and os.path.exists(bold):
            pdf.add_font(fam, "", reg)
            pdf.add_font(fam, "B", bold)
            FONT, unicode_ok = fam, True
            break

    def S(t):
        t = "" if t is None else str(t)
        return t if unicode_ok else t.encode("latin-1", "replace").decode("latin-1")

    # ── Group entries by file_id ──
    seen: dict  = {}
    groups: list = []
    for e in entries:
        fid = e.get("file_id") or ""
        if fid and fid in seen:
            groups[seen[fid]][1].append(e)
        else:
            seen[fid] = len(groups)
            groups.append((fid, [e]))

    # ── Grid: 5 cols × 2 rows, images only ──
    N_COLS = 5
    N_ROWS = 2
    HDR_H  = 9.0      # one-line header (first page only)
    PAGE_W = 297 - 2 * MRG   # 287mm
    PAGE_H = 210 - 2 * MRG   # 200mm
    CELL_W = (PAGE_W - GAP * (N_COLS - 1)) / N_COLS    # ≈56.6mm
    ROW_H  = (PAGE_H - HDR_H - GAP * (N_ROWS - 1)) / N_ROWS  # ≈94mm (first page)
    ROW_H_CONT = (PAGE_H - GAP * (N_ROWS - 1)) / N_ROWS      # ≈99.5mm (no header)

    def cell_inner_w(span: int) -> float:
        return CELL_W * span + GAP * (span - 1)

    def _best_orientation(jpeg: bytes, dim: tuple, cell_w: float, cell_h: float):
        """Return (jpeg, dim) rotated 90° if that gives significantly better cell coverage."""
        from PIL import Image as _PILImg
        def coverage(iw, ih, cw, ch):
            dw = cw
            dh = dw * ih / iw
            if dh > ch:
                dh, dw = ch, ch * iw / ih
            return dw * dh
        normal  = coverage(dim[0], dim[1], cell_w, cell_h)
        rotated = coverage(dim[1], dim[0], cell_w, cell_h)
        if rotated > normal * 1.15:          # rotate if ≥15% better fill
            img = _PILImg.open(io.BytesIO(jpeg)).rotate(90, expand=True)
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=82, optimize=True)
            return buf.getvalue(), img.size
        return jpeg, dim

    # ── Pre-fetch + compress images ──
    img_cache: dict = {}
    for fid, elist in groups:
        if not fid:
            img_cache[fid] = None
            continue
        span   = min(len(elist), N_COLS)
        cw     = cell_inner_w(span) - 2 * PAD
        ch     = ROW_H_CONT - 2 * PAD
        max_px = max(500, int(cell_inner_w(span) * 5))
        raw    = _fetch_drive_image(service, fid) or b""
        comp   = _compress_receipt_image(raw, max_w=max_px, quality=82)
        if comp:
            jpeg, dim = comp
            jpeg, dim = _best_orientation(jpeg, dim, cw, ch)
            img_cache[fid] = (jpeg, dim)
        else:
            img_cache[fid] = None

    # ── Pack groups into grid rows ──
    grid_rows: list = []
    cur_row:   list = []
    col = 0
    for fid, elist in groups:
        span = min(len(elist), N_COLS)
        if col + span > N_COLS:
            grid_rows.append(cur_row)
            cur_row, col = [], 0
        cur_row.append((fid, elist, span))
        col += span
    if cur_row:
        grid_rows.append(cur_row)

    if not grid_rows:
        pdf.add_page()
        pdf.set_font(FONT, "", 10)
        pdf.set_xy(MRG, MRG)
        pdf.cell(0, 10, S("No receipts for the selected filter."))
        out = pdf.output()
        return ("binary", bytes(out), "application/pdf",
                f"receipts_{username}_{date.today().isoformat()}.pdf")

    # ── Draw ──
    def draw_header(first: bool):
        if not first:
            return
        pdf.set_font(FONT, "", 7)
        pdf.set_text_color(130, 130, 130)
        pdf.set_xy(MRG, MRG)
        flabel = {"all": "All"}.get(status) or _PDF_STATUS_LABEL.get(status, status)
        rng = f"{dfrom or '…'} ~ {dto or '…'}"
        pdf.cell(0, HDR_H - 1, S(
            f"Receipt Ledger  ·  {username}  ·  {flabel}  ·  {rng}  ·  "
            f"Total {stats['total']}  Matched {stats['matched']}  "
            f"Unmatched {stats['unmatched']}  Pending {stats['pending_match']}"
        ))
        pdf.set_text_color(0, 0, 0)

    first_page  = True
    row_in_page = 0   # which row on the current page (0 or 1)
    y           = 0.0
    n_rows      = len(grid_rows)

    for row_idx, row in enumerate(grid_rows):
        # Start a new page when both rows are filled
        if row_in_page == 0:
            pdf.add_page()
            draw_header(first_page)
            y = MRG + (HDR_H if first_page else 0)
            first_page = False

        # Determine row height: if this is the only row on this page (last item or
        # next item would start a new page), expand to full usable height.
        is_only_row_on_page = (row_in_page == 0 and (row_idx == n_rows - 1))
        if is_only_row_on_page:
            cur_row_h = PAGE_H - (HDR_H if pdf.page == 1 else 0) - 2 * MRG
        elif row_in_page == 0 and pdf.page == 1:
            cur_row_h = ROW_H
        else:
            cur_row_h = ROW_H_CONT

        x = MRG
        for fid, elist, span in row:
            iw = cell_inner_w(span)

            # Thin border
            pdf.set_draw_color(200, 200, 200)
            pdf.set_line_width(0.15)
            pdf.rect(x, y, iw, cur_row_h)

            # Image — fill the cell as fully as possible
            comp = img_cache.get(fid)
            if comp:
                jpeg, dim = comp
                avail_w = iw - 2 * PAD
                avail_h = cur_row_h - 2 * PAD
                dw = avail_w
                dh = dw * dim[1] / dim[0]
                if dh > avail_h:
                    dh, dw = avail_h, avail_h * dim[0] / dim[1]
                try:
                    pdf.image(io.BytesIO(jpeg),
                              x=x + (iw - dw) / 2,
                              y=y + (cur_row_h - dh) / 2,
                              w=dw, h=dh)
                except Exception:
                    pass
            else:
                pdf.set_xy(x, y)
                pdf.set_font(FONT, "", 6)
                pdf.set_text_color(180, 180, 180)
                pdf.cell(iw, cur_row_h, S("no image"), align="C")
                pdf.set_text_color(0, 0, 0)

            x += iw + GAP

        y += cur_row_h + GAP
        row_in_page = (row_in_page + 1) % N_ROWS

    out  = pdf.output()
    fname = f"receipts_{username}_{date.today().isoformat()}.pdf"
    _add_hist({
        "type":     "pdf_download",
        "date":     datetime.now().strftime("%Y-%m-%d %H:%M"),
        "filename": fname,
        "filter":   flabel if 'flabel' in dir() else status,
        "count":    stats.get("total", 0),
        "user":     username,
    })
    return ("binary", bytes(out), "application/pdf", fname)


# ── HTTP handler ──────────────────────────────────────────────────────────────

def handle(method, path, body, ctx=None):
    user = (ctx or {}).get("user")
    if user != ADMIN:
        return ("html", "<h2 style='padding:40px;color:#f87171'>Access denied</h2>")

    # Tab pages: Ledger (main) | Convert | Review | Keywords
    if method == "GET" and path == "/cardconv":
        return ("redirect", "/cardconv/ledger")
    if method == "GET" and path == "/cardconv/convert":
        return ("html", _render_convert(user))
    if method == "GET" and path == "/cardconv/review":
        return ("html", _render_review(user))
    if method == "GET" and path == "/cardconv/history":
        return ("html", _render_history(user))
    if method == "POST" and path == "/cardconv/history/delete":
        ids = body.get("ids", [])
        if isinstance(ids, str): ids = [ids]
        hist = [h for h in _load_hist() if h.get("id") not in ids]
        _save_hist(hist)
        return ("json", {"ok": True})
    if method == "POST" and path == "/cardconv/history/clear":
        _save_hist([])
        return ("json", {"ok": True})
    if method == "GET" and path == "/cardconv/keywords":
        return ("html", _render_keywords(user))
    if method == "POST" and path == "/cardconv/review/reason":
        return _handle_review_reason(user, body)
    if method == "POST" and path == "/cardconv/review/match":
        return _handle_review_manual_match(user, body)
    if method == "GET" and path == "/cardconv/review/download":
        return _handle_review_download(user, body)  # GET passes query dict as body

    # Ledger
    if method == "GET" and path == "/cardconv/ledger":
        return ("html", _render_ledger(user))
    if method == "GET" and path == "/cardconv/ledger/api":
        return _handle_ledger_api(user, body)  # GET passes query dict as body
    if method == "GET" and path == "/cardconv/ledger/download.pdf":
        return _handle_ledger_pdf(user, body)  # GET passes query dict as body
    if method == "POST" and path == "/cardconv/ledger/delete":
        return _handle_ledger_delete(user, body)
    if method == "POST" and path == "/cardconv/ledger/complete":
        return _handle_ledger_complete(user, body)
    if method == "GET" and path == "/cardconv/ledger/download.xlsx":
        return _handle_ledger_xlsx(user, body)  # GET passes query dict as body
    if method == "POST" and path.startswith("/cardconv/ledger/") and path.endswith("/status"):
        entry_id = path[len("/cardconv/ledger/"):-len("/status")]
        return _handle_status_change(user, entry_id, body)
    if method == "POST" and path.startswith("/cardconv/ledger/") and path.endswith("/update"):
        entry_id = path[len("/cardconv/ledger/"):-len("/update")]
        return _handle_ledger_update(user, entry_id, body)
    if method == "POST" and path.startswith("/cardconv/ledger/") and path.endswith("/reocr"):
        entry_id = path[len("/cardconv/ledger/"):-len("/reocr")]
        return _handle_reocr(user, entry_id)
    if method == "POST" and path.startswith("/cardconv/ledger/") and path.endswith("/rematch"):
        entry_id = path[len("/cardconv/ledger/"):-len("/rematch")]
        return _handle_rematch(user, entry_id)
    if method == "GET" and path.startswith("/cardconv/receipts/image/"):
        file_id = path[len("/cardconv/receipts/image/"):]
        bbox = None
        raw_bbox = (ctx.get("query") or {}).get("bbox")
        if raw_bbox:
            try:
                bbox = [float(v) for v in raw_bbox.split(",")]
            except Exception:
                bbox = None
        return _handle_image_proxy(user, file_id, bbox)
    if method == "POST" and path == "/cardconv/batch/run":
        return _handle_batch_run(user, ctx)

    # Drive OAuth
    if method == "GET" and path == "/cardconv/drive/connect":
        return _handle_drive_connect(user)
    if method == "POST" and path == "/cardconv/drive/auth":
        return _handle_drive_auth(user, body)

    # Drive sync (background)
    if method == "POST" and path == "/cardconv/drive/sync":
        return _handle_drive_sync(user)
    if method == "GET" and path == "/cardconv/drive/sync/status":
        return _handle_drive_sync_status(user, body)  # GET passes query as body

    # Receipt upload
    if method == "POST" and path == "/cardconv/receipts/upload":
        return _handle_receipt_upload(user, body)

    # Manual receipt addition (for OCR-missed sub-receipts on an existing image)
    if method == "POST" and path == "/cardconv/receipts/manual-add":
        return _handle_manual_receipt_add(user, body)

    # Web Push subscription management
    if method == "POST" and path == "/cardconv/push/subscribe":
        sub = body if isinstance(body, dict) else {}
        if not sub.get("endpoint"):
            return ("json", {"error": "invalid subscription"}, 400)
        subs = _load_push_subs(user)
        endpoints = {s.get("endpoint") for s in subs}
        if sub["endpoint"] not in endpoints:
            subs.append(sub)
            _save_push_subs(user, subs)
        return ("json", {"ok": True})
    if method == "POST" and path == "/cardconv/push/unsubscribe":
        endpoint = (body or {}).get("endpoint", "")
        subs = [s for s in _load_push_subs(user) if s.get("endpoint") != endpoint]
        _save_push_subs(user, subs)
        return ("json", {"ok": True})

    # OCR staging review
    if method == "GET" and path == "/cardconv/receipts/review":
        return ("html", _render_ocr_staging_review(user))
    if method == "GET" and path == "/cardconv/receipts/review/api":
        staging = _load_ocr_staging(user)
        return ("json", {"entries": staging.get("entries", [])})
    if method == "POST" and path == "/cardconv/receipts/review/confirm":
        return _handle_ocr_staging_confirm(user, body)
    if method == "POST" and path == "/cardconv/receipts/review/discard":
        _clear_ocr_staging(user)
        return ("redirect", "/cardconv/ledger")

    # File download
    if method == "GET" and path.startswith("/cardconv/download/"):
        fname = path[len("/cardconv/download/"):]
        fpath = OUT_DIR / fname
        if fpath.exists() and fpath.suffix == ".xlsx":
            return ("file", str(fpath),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fname)
        return ("html", "<h2>File not found</h2>", 404)

    # CSV upload
    if method == "POST" and path == "/cardconv/upload":
        return _handle_upload(body, user)

    # Uploaded CSV reuse (re-run / delete)
    if method == "POST" and path == "/cardconv/upload/rerun":
        uid = (body.get("id", [""])[0]).strip()
        return _handle_upload_rerun(user, uid)
    if method == "POST" and path == "/cardconv/upload/delete":
        uid = (body.get("id", [""])[0]).strip()
        return _handle_upload_delete(user, uid)

    # Card member names (My Card Names)
    if method == "POST" and path == "/cardconv/cardnames/add":
        # Section moved from Keywords to Convert page; redirect there.
        name = (body.get("name", [""])[0]).strip().upper()
        if name:
            s = _load_user_settings(user)
            names = s.get("card_member_names") or list(DEFAULT_CARD_NAMES)
            if name not in [n.strip().upper() for n in names]:
                names.append(name)
            s["card_member_names"] = names
            _save_user_settings(user, s)
        return ("redirect", "/cardconv/convert")
    if method == "POST" and path == "/cardconv/cardnames/delete":
        name = (body.get("name", [""])[0]).strip().upper()
        s = _load_user_settings(user)
        names = s.get("card_member_names") or list(DEFAULT_CARD_NAMES)
        s["card_member_names"] = [n for n in names if n.strip().upper() != name]
        _save_user_settings(user, s)
        return ("redirect", "/cardconv/convert")

    # Keyword add
    if method == "POST" and path == "/cardconv/keyword/add":
        kw      = (body.get("kw",      [""])[0]).strip().upper()
        gl      = (body.get("gl",      [""])[0]).strip()
        ser     = (body.get("ser",     [""])[0]).strip()
        purpose = (body.get("purpose", [""])[0]).strip()
        if kw and gl and purpose:
            kws = _load_kw()
            if not any(k["kw"] == kw for k in kws):
                kws.insert(0, {"kw": kw, "gl": int(gl), "ser": ser, "purpose": purpose})
                _save_kw(kws)
        return ("redirect", "/cardconv/keywords")

    # Keyword delete
    if method == "POST" and path == "/cardconv/keyword/delete":
        kw  = (body.get("kw", [""])[0]).strip().upper()
        kws = [k for k in _load_kw() if k["kw"].upper() != kw]
        _save_kw(kws)
        return ("redirect", "/cardconv/keywords")

    return ("redirect", "/cardconv/ledger")


def _get_client_info():
    import json as _json
    data = _json.loads(CREDS_FILE.read_text())
    return data.get("installed", data.get("web", {}))


def _handle_drive_connect(username: str):
    if not CREDS_FILE.exists():
        return ("html", f"<p style='padding:20px;color:var(--danger)'>Credentials file not found: {CREDS_FILE}</p>")
    try:
        import urllib.parse as _up
        c = _get_client_info()
        params = {
            "client_id":     c["client_id"],
            "redirect_uri":  "urn:ietf:wg:oauth:2.0:oob",
            "response_type": "code",
            "scope":         " ".join(SCOPES),
            "access_type":   "offline",
            "prompt":        "consent",
        }
        auth_url = "https://accounts.google.com/o/oauth2/v2/auth?" + _up.urlencode(params)
        return ("html", _render_drive_connect(username, auth_url))
    except Exception as e:
        return ("html", f"<p style='padding:20px;color:var(--danger)'>Drive connect error: {e}</p>")


def _handle_drive_auth(username: str, body):
    raw = body.get("code", "")
    code = (raw[0] if isinstance(raw, list) else str(raw)).strip()
    if not code:
        return ("redirect", "/cardconv/ledger")
    try:
        import json as _json, urllib.request as _req, urllib.parse as _up
        c = _get_client_info()
        token_data = _up.urlencode({
            "code":          code,
            "client_id":     c["client_id"],
            "client_secret": c["client_secret"],
            "redirect_uri":  "urn:ietf:wg:oauth:2.0:oob",
            "grant_type":    "authorization_code",
        }).encode()
        req = _req.Request("https://oauth2.googleapis.com/token", data=token_data,
                           headers={"Content-Type": "application/x-www-form-urlencoded"})
        resp = _req.urlopen(req)
        token_json = _json.loads(resp.read())
        if "error" in token_json:
            raise Exception(f"{token_json['error']}: {token_json.get('error_description','')}")
        c = _get_client_info()
        save_data = {
            "token":         token_json.get("access_token"),
            "refresh_token": token_json.get("refresh_token"),
            "token_uri":     "https://oauth2.googleapis.com/token",
            "client_id":     c["client_id"],
            "client_secret": c["client_secret"],
            "scopes":        SCOPES,
        }
        _ensure_dirs()
        (TOKENS_DIR / f"{username}.json").write_text(_json.dumps(save_data))
    except Exception as e:
        return ("html", f"<p style='padding:20px;color:var(--danger)'>Auth error: {e} "
                        f"<a href='/cardconv/ledger' style='color:var(--accent)'>Back</a></p>")
    return ("redirect", "/cardconv/ledger")


import threading as _threading, uuid as _uuid

_sync_jobs: dict = {}   # job_id → {"status": "running"|"done"|"error", "staged": int, "error": str}

def _do_drive_sync_work(username: str, job_id: str):
    """Background worker: scans Drive, OCRs new files, stages for review."""
    try:
        service = _get_drive_service(username)
        if not service:
            _sync_jobs[job_id] = {"status": "error", "staged": 0, "error": "Drive not connected"}
            return

        receipts_id, matched_id = _get_receipts_folder_ids(service, username)
        results = service.files().list(
            q=(f"('{receipts_id}' in parents or '{matched_id}' in parents) "
               f"and trashed=false and mimeType!='application/vnd.google-apps.folder'"),
            fields="files(id,name,mimeType,webViewLink)"
        ).execute()
        drive_files = results.get('files', [])

        existing = _load_receipts(username)
        done_fids = {r.get('file_id') for r in existing
                     if (r.get('multi_ocr') and r.get('ocr_amount') is not None)
                     or r.get('matched')}
        # Skip files already in the staging queue.
        staging = _load_ocr_staging(username)
        staged_fids = {e.get("file_id") for e in staging.get("entries", [])}
        skip_fids = done_fids | staged_fids
        supported = {'image/jpeg', 'image/png', 'application/pdf', 'image/gif', 'image/webp'}

        new_staged = []
        now_disp = datetime.now().strftime("%Y-%m-%d %H:%M")
        now_iso  = datetime.now().isoformat()

        for f in drive_files:
            fid = f.get('id')
            mime = f.get('mimeType', '')
            if mime not in supported:
                continue
            if fid in skip_fids:
                continue
            content  = service.files().get_media(fileId=fid).execute()
            ocr_list = _ocr_receipt_auto(content, mime) or [{}]
            url      = f.get('webViewLink') or f'https://drive.google.com/file/d/{fid}/view'
            for sub_index, ocr in enumerate(ocr_list):
                entry = {
                    "id":           _sub_entry_id(fid, sub_index),
                    "file_id":      fid,
                    "sub_index":    sub_index,
                    "filename":     f.get('name', ''),
                    "drive_url":    url,
                    "mime_type":    mime,
                    "match_status": "pending_match",
                    "multi_ocr":    True,
                    "uploaded_at":  now_disp,
                    "synced_at":    now_iso,
                }
                entry.update(_ocr_entry_fields(ocr))
                new_staged.append(entry)

        # Accumulate — append to existing staging queue, never replace.
        if new_staged:
            staging.setdefault("entries", []).extend(new_staged)
            staging["staged_at"] = now_iso
            _save_ocr_staging(username, staging)

        ledger = _load_ledger(username)
        ledger["last_batch_at"] = now_iso
        _save_ledger(username, ledger)

        _sync_jobs[job_id] = {"status": "done", "staged": len(new_staged)}

    except Exception as e:
        _sync_jobs[job_id] = {"status": "error", "staged": 0, "error": str(e)}


def _handle_drive_sync(username: str):
    """POST /cardconv/drive/sync — start background sync, return job_id immediately."""
    if not _get_drive_service(username):
        return ("json", {"error": "Drive not connected"}, 400)
    job_id = _uuid.uuid4().hex
    _sync_jobs[job_id] = {"status": "running", "staged": 0}
    t = _threading.Thread(target=_do_drive_sync_work, args=(username, job_id), daemon=True)
    t.start()
    return ("json", {"job_id": job_id})


def _handle_drive_sync_status(username: str, query: dict):
    """GET /cardconv/drive/sync/status?job=<id> — poll sync progress."""
    job_id = (query.get("job", [""])[0]).strip()
    job = _sync_jobs.get(job_id)
    if not job:
        return ("json", {"status": "not_found"})
    return ("json", job)


def _handle_receipt_upload(username: str, body):
    raw = body.get("__raw_handler__")
    if raw is None:
        return ("redirect", "/cardconv/ledger")
    if not _is_drive_connected(username):
        return ("html", "<p style='padding:20px;color:var(--danger)'>Connect Google Drive first. "
                        "<a href='/cardconv/ledger' style='color:var(--accent)'>Back</a></p>")

    files = _parse_multipart_files(raw)
    if not files:
        return ("redirect", "/cardconv/ledger")

    receipts  = _load_receipts(username)
    supported = {'image/jpeg', 'image/png', 'application/pdf', 'image/gif', 'image/webp'}

    staged_entries = []
    now_disp = datetime.now().strftime("%Y-%m-%d %H:%M")
    now_iso  = datetime.now().isoformat()

    for filename, content, mime_type in files:
        if mime_type not in supported:
            continue
        file_id, drive_url = _upload_file_to_drive(username, content, filename, mime_type)
        ocr_list = _ocr_receipt_auto(content, mime_type) or [{}]
        for sub_index, ocr in enumerate(ocr_list):
            entry = {
                "id":           _sub_entry_id(file_id, sub_index),
                "file_id":      file_id,
                "sub_index":    sub_index,
                "filename":     filename,
                "drive_url":    drive_url,
                "mime_type":    mime_type,
                "uploaded_at":  now_disp,
                "synced_at":    now_iso,
                "match_status": "pending_match",
                "multi_ocr":    True,
            }
            entry.update(_ocr_entry_fields(ocr))
            staged_entries.append(entry)

    # Stage for visual review; only confirmed entries go to the ledger.
    existing = _load_ocr_staging(username)
    existing.setdefault("entries", []).extend(staged_entries)
    existing["staged_at"] = now_iso
    _save_ocr_staging(username, existing)
    return ("redirect", "/cardconv/receipts/review")


def _handle_manual_receipt_add(username: str, body: dict):
    """Add a manually-entered receipt sub-entry to an existing Drive image in the ledger."""
    file_id   = (body.get("file_id") or "").strip()
    filename  = (body.get("filename") or "").strip()
    drive_url = (body.get("drive_url") or "").strip()
    mime_type = (body.get("mime_type") or "image/jpeg").strip()

    def _num(v):
        try: return float(v) if v not in (None, "") else None
        except: return None

    ocr_date   = (body.get("ocr_date") or "").strip() or None
    merchant   = (body.get("ocr_merchant") or "").strip() or None
    printed    = _num(body.get("ocr_printed_amount"))
    handw      = _num(body.get("ocr_handwritten_amount"))
    final_amt  = handw if handw is not None else printed

    if not file_id:
        return ("json", {"error": "file_id required"}, 400)

    entries = _load_receipts(username)
    # Assign the next sub_index for this file_id.
    existing_subs = [e.get("sub_index", 0) for e in entries if e.get("file_id") == file_id]
    sub_index = (max(existing_subs) + 1) if existing_subs else 0

    now_disp = datetime.now().strftime("%Y-%m-%d %H:%M")
    now_iso  = datetime.now().isoformat()
    entry = {
        "id":                    _sub_entry_id(file_id, sub_index),
        "file_id":               file_id,
        "sub_index":             sub_index,
        "filename":              filename,
        "drive_url":             drive_url,
        "mime_type":             mime_type,
        "uploaded_at":           now_disp,
        "synced_at":             now_iso,
        "match_status":          "pending_match",
        "multi_ocr":             True,
        "ocr_status":            "manual",
        "ocr_model":             "Manual",
        "ocr_date":              ocr_date,
        "ocr_merchant":          merchant,
        "ocr_printed_amount":    printed,
        "ocr_handwritten_amount": handw,
        "ocr_amount":            final_amt,
        "ocr_bbox":              None,
    }
    entries.append(entry)
    _save_receipts(username, entries)
    return ("json", {"ok": True, "id": entry["id"]})


def _handle_upload(body, user=None):
    raw = body.get("__raw_handler__")
    if raw is None:
        return ("html", "<p>Upload error: no raw handler</p>")
    ct = raw.headers.get("Content-Type", "")
    m  = re.search(r'boundary=([^\s;]+)', ct)
    if not m:
        return ("html", "<p>Upload error: no boundary</p>")
    boundary  = ("--" + m.group(1)).encode()
    length    = int(raw.headers.get("Content-Length", 0))
    data      = raw.rfile.read(length)

    csv_bytes = None
    csv_name  = "upload.csv"
    for part in data.split(boundary):
        if b'filename="' not in part:
            continue
        fn_m = re.search(rb'filename="([^"]+)"', part)
        if fn_m:
            csv_name = fn_m.group(1).decode()
        hdr_end = part.find(b"\r\n\r\n")
        if hdr_end == -1:
            continue
        content = part[hdr_end + 4:]
        if content.endswith(b"\r\n"):
            content = content[:-2]
        csv_bytes = content
        break

    if not csv_bytes:
        return ("redirect", "/cardconv/ledger")

    try:
        xlsx_bytes, out_fn, total, unmatched = convert(csv_bytes, csv_name, username=user)
        _add_hist({
            "type":      "conversion",
            "filename":  out_fn,
            "source":    csv_name,
            "rows":      total,
            "unmatched": unmatched,
            "date":      datetime.now().strftime("%Y-%m-%d %H:%M"),
            "user":      user or "",
        })
        if user:
            _save_uploaded_csv(user, csv_bytes, csv_name, total, out_fn)
        # Conversion result is staged in review_{user}.json; review before download.
        return ("redirect", "/cardconv/review")
    except Exception as e:
        return ("html", f"<p style='color:red;padding:20px'>Error: {e}</p>")


def _handle_upload_rerun(username: str, uid: str):
    """Re-run conversion (re-match receipts) from a previously uploaded CSV."""
    items = _load_uploads(username)
    entry = next((i for i in items if i.get("id") == uid), None)
    if not entry:
        return ("redirect", "/cardconv/convert")
    stored = _uploads_dir(username) / entry.get("stored_name", "")
    if not stored.exists():
        return ("redirect", "/cardconv/convert")
    try:
        csv_bytes = stored.read_bytes()
        fn = entry.get("filename", "upload.csv")
        xlsx_bytes, out_fn, total, unmatched = convert(csv_bytes, fn, username=username)
        _add_hist({
            "filename":  out_fn,
            "source":    fn,
            "rows":      total,
            "unmatched": unmatched,
            "date":      datetime.now().strftime("%Y-%m-%d %H:%M"),
        })
        entry["rows"] = total
        entry["out_filename"] = out_fn
        _save_uploads(username, items)
        return ("redirect", "/cardconv/review")
    except Exception as e:
        return ("html", f"<p style='color:red;padding:20px'>Error: {e}</p>")


def _handle_upload_delete(username: str, uid: str):
    """Delete a stored CSV and its index entry."""
    items = _load_uploads(username)
    entry = next((i for i in items if i.get("id") == uid), None)
    if entry:
        stored = _uploads_dir(username) / entry.get("stored_name", "")
        try:
            if stored.exists():
                stored.unlink()
        except Exception:
            pass
        _save_uploads(username, [i for i in items if i.get("id") != uid])
    return ("redirect", "/cardconv/convert")


# ── Render helpers ─────────────────────────────────────────────────────────────

def _render_drive_connect(username: str, auth_url: str) -> str:
    from server import CSS_VER
    return f'''<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>🔗 Connect Google Drive · Wayfinder</title>
<link rel="stylesheet" href="/static/style.css?v={CSS_VER}">
</head><body>
<nav>
  <span class="nav-brand">🔗 Connect Google Drive</span>
  <span class="nav-user"><a href="/cardconv/ledger" class="nav-back">← Back to Ledger</a></span>
</nav>
<div class="container" style="max-width:640px">
  <div class="notepad-card">
    <div class="notepad-header">
      <span style="font-size:var(--text-xs);font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--accent)">Authorize Google Drive Access</span>
    </div>
    <div class="notepad-body" style="padding:28px;display:flex;flex-direction:column;gap:20px">
      <ol style="color:var(--text-muted);font-size:.88rem;line-height:2.2;padding-left:22px">
        <li>Click the button below to open Google authorization page</li>
        <li>Sign in and grant <b style="color:var(--text)">Drive file access</b></li>
        <li>Copy the authorization code shown on that page</li>
        <li>Paste it in the field below and click Confirm</li>
      </ol>
      <a href="{auth_url}" target="_blank" class="btn btn-primary" style="width:fit-content">
        🔗 Open Google Authorization Page
      </a>
      <form method="POST" action="/cardconv/drive/auth" style="display:flex;flex-direction:column;gap:12px">
        <div style="display:flex;flex-direction:column;gap:6px">
          <label style="font-size:.72rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:.06em">Authorization Code</label>
          <input type="text" name="code" placeholder="Paste the code from Google here..." required
            style="padding:10px 14px;border:1px solid var(--border);border-radius:8px;background:var(--surface-2);color:var(--text);font-size:.88rem;outline:none">
        </div>
        <button type="submit" class="btn btn-accent" style="width:fit-content">✅ Confirm &amp; Connect</button>
      </form>
    </div>
  </div>
</div>
</body></html>'''


# ── Shared tab bar ───────────────────────────────────────────────────────────

_CC_TAB_CSS = (
    ".cc-tabs{display:flex;gap:0;border-bottom:1px solid var(--border);margin-bottom:20px;flex-wrap:wrap}"
    ".cc-tab{padding:10px 20px;font-size:.82rem;font-weight:600;color:var(--text-muted);"
    "border-bottom:2px solid transparent;text-decoration:none;transition:color .15s,border-color .15s}"
    ".cc-tab:hover{color:var(--text)}"
    ".cc-tab.active{color:var(--accent);border-bottom-color:var(--accent)}"
    ".tab-badge{display:inline-flex;align-items:center;justify-content:center;min-width:16px;height:16px;"
    "background:#ef4444;border-radius:8px;font-size:.62rem;font-weight:700;color:#fff;padding:0 4px;"
    "margin-left:5px;vertical-align:middle}"
    # Workflow step bar
    ".cc-wf-bar{display:flex;align-items:center;background:var(--surface-2);"
    "border:1px solid var(--border);border-radius:var(--radius-md);"
    "padding:0 10px 0 8px;margin-bottom:16px;height:36px}"
    ".cc-wf-close{margin-left:auto;flex-shrink:0;background:none;border:none;"
    "color:var(--text-muted);cursor:pointer;font-size:.85rem;"
    "padding:2px 6px;border-radius:4px;line-height:1}"
    ".cc-wf-close:hover{color:var(--text);background:var(--surface-3)}"
    ".cc-wf-steps{display:flex;align-items:center;flex:1;min-width:0;overflow:hidden}"
    ".cc-wf-step{display:flex;align-items:center;gap:5px;padding:0 6px;"
    "font-size:.74rem;color:var(--text-muted);white-space:nowrap;flex-shrink:0}"
    ".cc-wf-step .sn{display:inline-flex;align-items:center;justify-content:center;"
    "width:18px;height:18px;border-radius:50%;border:1.5px solid currentColor;"
    "font-size:.6rem;font-weight:700;flex-shrink:0}"
    ".cc-wf-step.wf-active{color:var(--text);font-weight:600}"
    ".cc-wf-step.wf-active .sn{background:var(--accent);border-color:var(--accent);color:#000}"
    ".cc-wf-step.wf-done .sn{background:#22c55e;border-color:#22c55e;color:#fff}"
    ".cc-wf-step.wf-done{color:var(--text-muted)}"
    ".cc-wf-sep{color:var(--border);font-size:.65rem;flex-shrink:0;padding:0 2px}"
    "@media(max-width:600px){.cc-wf-label{display:none}}"
    # Inline info tooltip
    ".cc-info-wrap{position:relative;display:inline-flex;align-items:center;vertical-align:middle}"
    ".cc-info{display:inline-flex;align-items:center;justify-content:center;"
    "width:14px;height:14px;border-radius:50%;font-size:.6rem;font-weight:700;"
    "color:var(--text-muted);cursor:pointer;border:1px solid var(--border);"
    "margin-left:6px;flex-shrink:0;line-height:1;user-select:none}"
    ".cc-info:hover{color:var(--accent);border-color:var(--accent)}"
    ".cc-tip{display:none;position:absolute;z-index:200;"
    "background:var(--surface-3);border:1px solid var(--border);"
    "border-radius:var(--radius-md);padding:10px 12px;font-size:.78rem;"
    "color:var(--text);max-width:280px;width:max-content;"
    "box-shadow:0 4px 16px rgba(0,0,0,.18);line-height:1.5;"
    "white-space:normal;text-align:left;top:calc(100% + 6px);left:0}"
    ".cc-tip.tip-right{left:auto;right:0}"
)


def _tab_bar(active: str, user: str) -> str:
    """Shared Card Converter tab bar. active ∈ ledger|convert|review|history|keywords."""
    unmatched_n = _ledger_stats(_ledger_entries(user))["unmatched"]
    ledger_badge = f'<span class="tab-badge">{unmatched_n}</span>' if unmatched_n else ''
    staged_n = len(_load_ocr_staging(user).get("entries", []))
    ocr_badge = f'<span class="tab-badge" style="background:#f59e0b;cursor:pointer" onclick="openOcrModal();return false;">{staged_n}</span>' if staged_n else ''
    tabs = [
        ("ledger",   "/cardconv/ledger",   "Receipt Ledger" + ledger_badge + ocr_badge),
        ("convert",  "/cardconv/convert",  "Convert"),
        ("review",   "/cardconv/review",   "Review"),
        ("history",  "/cardconv/history",  "History"),
        ("keywords", "/cardconv/keywords", "Keywords"),
    ]
    out = ['<div class="cc-tabs">']
    for key, href, label in tabs:
        cls = "cc-tab active" if key == active else "cc-tab"
        out.append(f'<a href="{href}" class="{cls}">{label}</a>')
    out.append('</div>')
    return "".join(out) + _workflow_bar(active, user)


def _info_icon(tip: str, right: bool = False) -> str:
    """Inline ℹ icon with click-toggled tooltip."""
    tip_cls = "cc-tip tip-right" if right else "cc-tip"
    return (f'<span class="cc-info-wrap">'
            f'<span class="cc-info" onclick="ccTipToggle(this)">ℹ</span>'
            f'<span class="{tip_cls}">{tip}</span>'
            f'</span>')


def _workflow_bar(active: str, user: str) -> str:
    """5-step onboarding workflow bar shown below the tab bar.
    Hidden for history/keywords tabs and after user dismisses with ×."""
    if active in ("history", "keywords", "ocr_review"):
        return ""

    drive_done = _is_drive_connected(user)

    # (step_num, label, tabs_where_active)
    steps = [
        (1, "Connect Drive",    ["ledger"]),
        (2, "Add Receipts",     ["ledger"]),
        (3, "Review Ledger",    ["ledger"]),
        (4, "Convert CSV",      ["convert"]),
        (5, "Review & Download",["review"]),
    ]

    parts = ['<div class="cc-wf-steps">']
    for i, (num, label, active_tabs) in enumerate(steps):
        if num == 1 and drive_done:
            cls = "cc-wf-step wf-done"
            badge = "✓"
        elif active in active_tabs:
            cls = "cc-wf-step wf-active"
            badge = str(num)
        else:
            cls = "cc-wf-step"
            badge = str(num)
        parts.append(
            f'<span class="{cls}">'
            f'<span class="sn">{badge}</span>'
            f'<span class="cc-wf-label">{label}</span>'
            f'</span>'
        )
        if i < len(steps) - 1:
            parts.append('<span class="cc-wf-sep">›</span>')
    parts.append('</div>')

    js = """<script>
(function(){
  if(localStorage.getItem('cc_guide_hidden')){
    var b=document.getElementById('ccWfBar');
    if(b) b.style.display='none';
  }
  window.ccHideGuide=function(){
    localStorage.setItem('cc_guide_hidden','1');
    document.getElementById('ccWfBar').style.display='none';
  };
  window.ccTipToggle=function(el){
    var tip=el.nextElementSibling;
    var open=tip.style.display==='block';
    document.querySelectorAll('.cc-tip').forEach(function(t){t.style.display='none';});
    if(!open){
      tip.style.display='block';
      var r=tip.getBoundingClientRect();
      if(r.right>window.innerWidth-20) tip.classList.add('tip-right');
      else tip.classList.remove('tip-right');
    }
  };
  document.addEventListener('click',function(e){
    if(!e.target.classList.contains('cc-info')){
      document.querySelectorAll('.cc-tip').forEach(function(t){t.style.display='none';});
    }
  });
})();
</script>"""

    return (
        f'<div class="cc-wf-bar" id="ccWfBar">'
        + "".join(parts)
        + '<button class="cc-wf-close" onclick="ccHideGuide()" title="Hide guide">×</button>'
        + '</div>'
        + js
    )


# Shared upload-zone CSS (used by Convert and Ledger register section)
_UPLOAD_CSS = (
    ".upload-zone{border:2px dashed var(--border);border-radius:var(--radius-lg);padding:40px 20px;"
    "text-align:center;cursor:pointer;transition:.2s;background:var(--surface)}"
    ".upload-zone:hover,.upload-zone.drag-over{border-color:var(--accent);background:var(--surface-2)}"
    ""  # file input now uses opacity:0 overlay instead of display:none
)


# ── Ledger register section (Drive + receipt upload) ───────────────────────────

def _register_section(user: str) -> str:
    """Drive status + receipt upload — moved onto the Ledger page."""
    connected = _is_drive_connected(user)
    meta = _load_drive_meta(user)
    receipts_folder_id = meta.get('receipts_folder_id')
    last_synced = _load_ledger(user).get("last_batch_at") or ""
    if connected:
        folder_link = ""
        if receipts_folder_id:
            folder_link = (f'<a href="https://drive.google.com/drive/folders/{receipts_folder_id}" '
                           f'target="_blank" class="btn btn-ghost btn-sm">📂 Open in Drive →</a>')
        sync_tip = _info_icon(
            'Fetches new receipts from Drive and uses AI (Gemini/Claude) to automatically extract date, amount, and merchant.')
        drive_status_html = f'''
      <span style="font-size:.88rem;font-weight:600;color:var(--success)">✅ Connected</span>
      {folder_link}
      <button class="btn btn-ghost btn-sm" onclick="startDriveSync(this)" style="margin-left:4px">🔄 Sync from Drive</button>{sync_tip}
      <span id="lastSynced" data-ts="{last_synced}" style="font-size:.78rem;color:var(--text-muted);margin-left:4px"></span>'''
        receipt_upload_html = '''
      <form id="rcptForm" method="POST" action="/cardconv/receipts/upload" enctype="multipart/form-data">
        <div class="upload-zone" id="rcptZone" onclick="document.getElementById('rcptFiles').click()">
          <div style="font-size:2rem;margin-bottom:8px">🧾</div>
          <div style="font-weight:700;color:var(--text);margin-bottom:4px">Drop receipts here</div>
          <div style="font-size:.8rem;color:var(--text-muted)">JPG · PNG · PDF &nbsp;·&nbsp; Multiple files supported &nbsp;·&nbsp; OCR runs automatically</div>
          <input type="file" id="rcptFiles" name="files" multiple accept=".jpg,.jpeg,.png,.pdf" onchange="handleRcptFiles(this)">
        </div>
        <div id="rcptInfo" style="display:none;margin-top:12px;padding:12px 16px;background:var(--surface-2);border-radius:var(--radius-md)">
          <div id="rcptFileList" style="font-size:.82rem;color:var(--text);margin-bottom:10px;display:flex;flex-wrap:wrap;gap:6px"></div>
          <button type="submit" class="btn btn-primary">📤 Upload &amp; OCR</button>
        </div>
      </form>'''
    else:
        drive_status_html = (
            '<span style="font-size:.88rem;font-weight:600;color:var(--danger)">❌ Not connected</span>'
            '<a href="/cardconv/drive/connect" class="btn btn-primary btn-sm">Connect Google Drive</a>'
        )
        receipt_upload_html = ('<p style="color:var(--text-muted);font-size:.85rem">'
                               'Connect Google Drive above to enable receipt upload.</p>')
    drive_tip = _info_icon(
        'Links to your Wayfinder/Receipts/ folder in Google Drive. '
        'After connecting, click Sync to automatically OCR all receipts in that folder.')
    return f'''
  <div class="notepad-card" style="margin-bottom:20px">
    <div class="notepad-header">
      <span style="font-size:var(--text-xs);font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--slate-400)">Google Drive</span>{drive_tip}
    </div>
    <div class="notepad-body" style="padding:14px 20px;display:flex;align-items:center;gap:16px;flex-wrap:wrap">
      {drive_status_html}
    </div>
  </div>
  <div class="notepad-card" style="margin-bottom:20px">
    <div class="notepad-header">
      <span style="font-size:var(--text-xs);font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--accent)">Register Receipts</span>
    </div>
    <div class="notepad-body" style="padding:20px">
      {receipt_upload_html}
    </div>
  </div>'''


# Receipt-upload drop-zone JS (injected into the Ledger page)
_RCPT_JS = r'''
const rcptZone = document.getElementById('rcptZone');
const rcptInfo = document.getElementById('rcptInfo');
const rcptList = document.getElementById('rcptFileList');
function handleRcptFiles(input){
  if(input.files.length>0){
    rcptList.innerHTML = Array.from(input.files).map(f =>
      '<span style="background:var(--surface-3);padding:3px 8px;border-radius:4px;font-size:.78rem">'+f.name+'</span>').join('');
    rcptInfo.style.display='block';
    rcptZone.style.borderColor='var(--accent)';
  }
}
if(rcptZone){
  rcptZone.addEventListener('dragover', e => { e.preventDefault(); rcptZone.classList.add('drag-over'); });
  rcptZone.addEventListener('dragleave', () => rcptZone.classList.remove('drag-over'));
  rcptZone.addEventListener('drop', e => {
    e.preventDefault(); rcptZone.classList.remove('drag-over');
    const input = document.getElementById('rcptFiles');
    if(e.dataTransfer.files.length>0){ input.files = e.dataTransfer.files; handleRcptFiles(input); }
  });
}
'''


# ── Convert page ───────────────────────────────────────────────────────────────

def _render_convert(user: str) -> str:
    from server import CSS_VER
    uploads = _load_uploads(user)
    up_rows = ""
    for u in uploads:
        up_rows += (
            '<div style="display:flex;align-items:center;gap:12px;padding:10px 0;border-bottom:1px solid var(--border)">'
            f'<span style="font-size:.8rem;color:var(--text-muted);min-width:130px">{_esc(u.get("uploaded_at"))}</span>'
            f'<span style="flex:1;font-size:.85rem;color:var(--text);font-weight:600">{_esc(u.get("filename"))}</span>'
            f'<span style="font-size:.78rem;color:var(--success)">{u.get("rows", 0)} rows</span>'
            f'<form method="POST" action="/cardconv/upload/rerun" style="display:inline">'
            f'<input type="hidden" name="id" value="{_esc(u.get("id"))}">'
            f'<button class="btn btn-secondary btn-sm">🔄 Re-run</button></form>'
            f'<form method="POST" action="/cardconv/upload/delete" style="display:inline" '
            f'onsubmit="return confirm(\'Delete this CSV?\')">'
            f'<input type="hidden" name="id" value="{_esc(u.get("id"))}">'
            f'<button class="btn btn-danger btn-sm">✕</button></form>'
            '</div>')
    if not up_rows:
        up_rows = '<div style="color:var(--text-muted);font-size:.85rem;padding:16px 0">No uploaded CSVs yet</div>'

    # My Card Names section (moved from Keywords page)
    names = _get_card_member_names(user)
    name_chips = ""
    for n in names:
        name_chips += (
            '<form method="POST" action="/cardconv/cardnames/delete" '
            'style="display:inline-flex;align-items:center;gap:6px;background:var(--surface-2);'
            'border:1px solid var(--border);border-radius:999px;padding:4px 6px 4px 12px;margin:0">'
            f'<span style="font-size:.82rem;font-weight:600;color:var(--accent)">{_esc(n)}</span>'
            f'<input type="hidden" name="name" value="{_esc(n)}">'
            '<button class="btn btn-danger btn-sm" style="padding:0 7px;line-height:1.5">✕</button>'
            '</form>')

    return f'''<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>💳 Convert · Wayfinder</title>
<link rel="stylesheet" href="/static/style.css?v={CSS_VER}">
<style>{_CC_TAB_CSS}{_UPLOAD_CSS}</style>
</head><body>
<nav>
  <span class="nav-brand">💳 Card Converter</span>
  <span class="nav-user">👤 {user} &nbsp;·&nbsp; <a href="/logout">Logout</a></span>
</nav>
<div class="container" style="max-width:860px">
  {_tab_bar("convert", user)}

  <div class="notepad-card" style="margin-bottom:20px">
    <div class="notepad-header">
      <span style="font-size:var(--text-xs);font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--accent)">My Card Names</span>{_info_icon('Only transactions whose Card Member Name matches one of these names will be converted. Enter your name exactly as it appears in the AMEX CSV.')}
    </div>
    <div class="notepad-body" style="padding:12px 16px">
      <p style="font-size:.78rem;color:var(--text-muted);margin-bottom:12px">CSV의 'Card Member Name'이 아래 이름과 일치하는 거래만 변환됩니다.</p>
      <form method="POST" action="/cardconv/cardnames/add" style="display:flex;gap:8px;margin-bottom:14px">
        <input name="name" placeholder="e.g. JOHN DOE" required style="flex:1;padding:7px 10px;border:1px solid var(--border);border-radius:6px;background:var(--surface-2);color:var(--text);font-size:.82rem">
        <button type="submit" class="btn btn-primary btn-sm">+ Add</button>
      </form>
      <div id="cardNamesWrap" style="display:flex;flex-wrap:wrap;gap:8px">{name_chips}</div>
    </div>
  </div>

  <div class="notepad-card" style="margin-bottom:20px">
    <div class="notepad-header">
      <span style="font-size:var(--text-xs);font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--accent)">Upload CSV</span>{_info_icon('Upload the Posted_*.csv downloaded from your AMEX statement. It will be automatically matched with receipts and exported as an SAP-ready xlsx.', right=True)}
    </div>
    <div class="notepad-body" style="padding:20px">
      <form id="upForm" method="POST" action="/cardconv/upload" enctype="multipart/form-data">
        <div class="upload-zone" id="dropZone" style="position:relative">
          <input type="file" id="csvFile" name="file" accept=".csv"
            onchange="handleCsvFile(this)"
            style="position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%;z-index:2">
          <div style="font-size:2rem;margin-bottom:8px">📎</div>
          <div style="font-weight:700;color:var(--text);margin-bottom:4px">Drop Posted_*.csv here</div>
          <div style="font-size:.8rem;color:var(--text-muted)">or click to browse</div>
        </div>
        <div id="fileInfo" style="display:none;margin-top:12px;padding:12px 16px;background:var(--surface-2);border-radius:var(--radius-md);align-items:center;gap:12px">
          <span style="font-size:1.2rem">📄</span>
          <span id="fileName" style="flex:1;font-size:.85rem;font-weight:600;color:var(--text)"></span>
          <button type="submit" class="btn btn-primary">Convert → Review</button>
        </div>
      </form>
      <div id="nameSuggest" style="display:none;margin-top:14px;padding:10px 14px;background:var(--surface-2);border:1px solid var(--border);border-radius:var(--radius-md)">
        <div style="font-size:.76rem;color:var(--text-muted);margin-bottom:8px">👤 Found in CSV — click to add to My Card Names:</div>
        <div id="nameChips" style="display:flex;flex-wrap:wrap;gap:6px"></div>
      </div>
      <p style="font-size:.78rem;color:var(--text-muted);margin-top:14px">
        Conversion matches receipts from the Ledger and opens the <b>Review</b> page before download.
      </p>
    </div>
  </div>

  <div class="notepad-card" style="margin-bottom:20px">
    <div class="notepad-header">
      <span style="font-size:var(--text-xs);font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--accent)">Uploaded CSVs</span>
    </div>
    <div class="notepad-body" style="padding:8px 16px 12px">
      {up_rows}
    </div>
  </div>
</div>
<script>
const csvZone = document.getElementById('dropZone');
const csvInfo = document.getElementById('fileInfo');
const csvName = document.getElementById('fileName');
const existingNames = new Set({json.dumps(names)});

function parseCsvSuggest(text) {{
  const lines = text.split(/\\r?\\n/);
  if (!lines.length) return;
  const hdr = lines[0].split(',').map(h => h.trim().replace(/^"|"$/g,''));
  const col = hdr.findIndex(h => /cardmember/i.test(h) || /card.?member/i.test(h));
  if (col < 0) return;
  const counts = {{}};
  lines.slice(1).forEach(line => {{
    if (!line.trim()) return;
    const cells = [];
    let cur = '', inQ = false;
    for (const ch of line + ',') {{
      if (ch === '"') {{ inQ = !inQ; }}
      else if (ch === ',' && !inQ) {{ cells.push(cur.trim()); cur = ''; }}
      else cur += ch;
    }}
    const name = (cells[col] || '').trim().toUpperCase();
    if (name && !existingNames.has(name)) counts[name] = (counts[name] || 0) + 1;
  }});
  const PRIORITY = ['EUISUN', 'DAE KIM', 'CHRIS CHO'];
  const sorted = Object.entries(counts).sort((a,b) => b[1]-a[1]).map(e => e[0]);
  const priority = sorted.filter(n => PRIORITY.some(p => n.includes(p)));
  const rest = sorted.filter(n => !PRIORITY.some(p => n.includes(p)));
  const top = [...priority, ...rest].slice(0, 10);
  if (!top.length) return;
  const chips = document.getElementById('nameChips');
  chips.innerHTML = top.map(n =>
    `<button type="button" class="preset-btn" style="font-size:.78rem" onclick="addSuggestedName(this,'${{n.replace(/'/g,"\\'")}}')">+ ${{n}}</button>`
  ).join('');
  document.getElementById('nameSuggest').style.display = 'block';
}}

function addSuggestedName(btn, name) {{
  fetch('/cardconv/cardnames/add', {{
    method:'POST', headers:{{'Content-Type':'application/x-www-form-urlencoded'}},
    body:'name='+encodeURIComponent(name)
  }}).then(r => {{
    if(r.ok) {{
      btn.disabled=true; btn.style.opacity='.4'; existingNames.add(name);
      // Immediately add to the My Card Names section above
      var wrap = document.getElementById('cardNamesWrap');
      if(wrap) {{
        var chip = document.createElement('form');
        chip.method = 'POST'; chip.action = '/cardconv/cardnames/delete';
        chip.style.cssText = 'display:inline-flex;align-items:center;gap:6px;background:var(--surface-2);border:1px solid var(--border);border-radius:999px;padding:4px 6px 4px 12px;margin:0';
        chip.innerHTML = '<span style="font-size:.82rem;font-weight:600;color:var(--accent)">'+name+'</span>'
          + '<input type="hidden" name="name" value="'+name.replace(/"/g,'&quot;')+'">'
          + '<button class="btn btn-danger btn-sm" style="padding:0 7px;line-height:1.5">✕</button>';
        wrap.appendChild(chip);
      }}
    }}
  }});
}}

function handleCsvFile(input) {{
  if (!input.files[0]) return;
  csvName.textContent = input.files[0].name;
  csvInfo.style.display = 'flex';
  csvZone.style.display = 'none';
  const reader = new FileReader();
  reader.onload = e => parseCsvSuggest(e.target.result);
  reader.readAsText(input.files[0]);
}}
csvZone.addEventListener('dragover', e => {{ e.preventDefault(); csvZone.classList.add('drag-over'); }});
csvZone.addEventListener('dragleave', () => csvZone.classList.remove('drag-over'));
csvZone.addEventListener('drop', e => {{
  e.preventDefault(); csvZone.classList.remove('drag-over');
  const f = e.dataTransfer.files[0];
  if (f) {{
    document.getElementById('csvFile').files = e.dataTransfer.files;
    csvName.textContent = f.name;
    csvInfo.style.display = 'flex';
    csvZone.style.display = 'none';
    const reader = new FileReader();
    reader.onload = e2 => parseCsvSuggest(e2.target.result);
    reader.readAsText(f);
  }}
}});
</script>
</body></html>'''


# ── History page (Recent Conversions, moved off Convert) ─────────────────────────

def _render_history(user: str) -> str:
    from server import CSS_VER
    hist = _load_hist()

    rows_html = ""
    for h in hist:
        hid   = _esc(h.get("id", ""))
        htype = h.get("type", "conversion")
        hdate = _esc(h.get("date", ""))
        icon  = "📤" if htype == "conversion" else "📥"
        type_label = "Conversion" if htype == "conversion" else "PDF Download"
        type_color = "color:#38bdf8" if htype == "conversion" else "color:#a78bfa"

        if htype == "conversion":
            src  = _esc(h.get("source", ""))
            fn   = _esc(h.get("filename", ""))
            rows = h.get("rows", 0)
            unm  = h.get("unmatched", 0)
            dl   = f'<a href="/cardconv/download/{fn}" class="btn btn-ghost btn-sm" style="font-size:.74rem;padding:3px 10px">⬇ xlsx</a>'
            detail = (f'<span style="font-size:.8rem;color:var(--text);font-weight:600">{fn}</span>'
                      f'<span style="font-size:.74rem;color:var(--text-muted)">from {src}</span>'
                      f'<span style="font-size:.74rem;color:var(--success)">{rows} rows</span>')
            if unm:
                detail += f'<span style="font-size:.72rem;color:var(--warn)">{unm} unmatched</span>'
        else:
            fn     = _esc(h.get("filename", ""))
            count  = h.get("count", 0)
            filt   = _esc(h.get("filter", "All"))
            dl     = ""
            detail = (f'<span style="font-size:.8rem;color:var(--text);font-weight:600">{fn}</span>'
                      f'<span style="font-size:.74rem;color:var(--text-muted)">Filter: {filt} · {count} receipts</span>')

        rows_html += (
            f'<div class="hist-row" data-id="{hid}" style="display:flex;align-items:center;gap:10px;padding:9px 0;border-bottom:1px solid var(--border)">'
            f'<input type="checkbox" class="hist-cb" data-id="{hid}" style="width:14px;height:14px;accent-color:var(--accent);cursor:pointer;flex-shrink:0">'
            f'<span style="font-size:1rem;flex-shrink:0">{icon}</span>'
            f'<span style="font-size:.72rem;font-weight:600;padding:1px 7px;border-radius:8px;background:var(--surface-3);{type_color};flex-shrink:0">{type_label}</span>'
            f'<span style="font-size:.78rem;color:var(--text-muted);min-width:120px;flex-shrink:0">{hdate}</span>'
            f'<div style="display:flex;flex-wrap:wrap;gap:6px;align-items:center;flex:1;min-width:0">{detail}</div>'
            f'{dl}'
            f'</div>')

    if not rows_html:
        rows_html = '<div style="color:var(--text-muted);font-size:.85rem;padding:20px 0;text-align:center">No history yet</div>'

    return f'''<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>🕘 History · Wayfinder</title>
<link rel="stylesheet" href="/static/style.css?v={CSS_VER}">
<style>{_CC_TAB_CSS}</style>
</head><body>
<nav>
  <span class="nav-brand">💳 Card Converter</span>
  <span class="nav-user">👤 {user} &nbsp;·&nbsp; <a href="/logout">Logout</a></span>
</nav>
<div class="container" style="max-width:860px">
  {_tab_bar("history", user)}

  <div class="notepad-card" style="margin-bottom:20px">
    <div class="notepad-header" style="display:flex;align-items:center;justify-content:space-between">
      <span style="font-size:var(--text-xs);font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--slate-400)">Upload &amp; Download History</span>
      <div style="display:flex;gap:8px">
        <button onclick="delSelected()" class="btn btn-ghost btn-sm" style="font-size:.74rem" id="delSelBtn" disabled>🗑 Delete Selected</button>
        <button onclick="clearAll()" class="btn btn-danger btn-sm" style="font-size:.74rem">✕ Clear All</button>
      </div>
    </div>
    <div class="notepad-body" style="padding:4px 16px 12px">
      <label style="display:flex;align-items:center;gap:6px;padding:6px 0;font-size:.78rem;color:var(--text-muted);cursor:pointer;border-bottom:1px solid var(--border);margin-bottom:2px">
        <input type="checkbox" id="checkAll" style="width:14px;height:14px;accent-color:var(--accent);cursor:pointer"> Select all
      </label>
      <div id="histList">{rows_html}</div>
    </div>
  </div>
</div>
<script>
const checkAll = document.getElementById('checkAll');
const delSelBtn = document.getElementById('delSelBtn');

function updateBtn(){{
  var n = document.querySelectorAll('.hist-cb:checked').length;
  delSelBtn.disabled = n === 0;
  delSelBtn.textContent = n ? '🗑 Delete Selected (' + n + ')' : '🗑 Delete Selected';
}}

checkAll.addEventListener('change', function(){{
  document.querySelectorAll('.hist-cb').forEach(cb => cb.checked = checkAll.checked);
  updateBtn();
}});

document.getElementById('histList').addEventListener('change', function(e){{
  if(e.target.classList.contains('hist-cb')) updateBtn();
}});

function delSelected(){{
  var ids = Array.from(document.querySelectorAll('.hist-cb:checked')).map(cb => cb.dataset.id);
  if(!ids.length) return;
  fetch('/cardconv/history/delete', {{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{ids}}) }})
    .then(() => location.reload());
}}

function clearAll(){{
  if(!confirm('Clear all history?')) return;
  fetch('/cardconv/history/clear', {{method:'POST'}}).then(() => location.reload());
}}
</script>
</body></html>'''


# ── Keywords page ────────────────────────────────────────────────────────────

def _render_keywords(user: str) -> str:
    from server import CSS_VER
    kws = _load_kw()
    kw_rows = ""
    for k in kws:
        kw_rows += f'''<tr>
      <td style="font-weight:600;color:var(--accent)">{k["kw"]}</td>
      <td style="color:var(--text-muted)">{k["gl"]}</td>
      <td style="color:var(--text-muted)">{k["ser"]}</td>
      <td style="flex:1;color:var(--text)">{k["purpose"]}</td>
      <td><form method="POST" action="/cardconv/keyword/delete" style="display:inline">
        <input type="hidden" name="kw" value="{k["kw"]}">
        <button class="btn btn-danger btn-sm">✕</button>
      </form></td>
    </tr>'''

    return f'''<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>🔑 Keywords · Wayfinder</title>
<link rel="stylesheet" href="/static/style.css?v={CSS_VER}">
<style>{_CC_TAB_CSS}
.kw-table{{width:100%;border-collapse:collapse;font-size:.82rem}}
.kw-table td{{padding:8px 10px;border-bottom:1px solid var(--border)}}
.kw-table tr:last-child td{{border-bottom:none}}
</style>
</head><body>
<nav>
  <span class="nav-brand">💳 Card Converter</span>
  <span class="nav-user">👤 {user} &nbsp;·&nbsp; <a href="/logout">Logout</a></span>
</nav>
<div class="container" style="max-width:860px">
  {_tab_bar("keywords", user)}

  <div class="notepad-card" id="keywords">
    <div class="notepad-header" style="display:flex;align-items:center;justify-content:space-between">
      <span style="font-size:var(--text-xs);font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--slate-400)">Keywords ({len(kws)})</span>
    </div>
    <div class="notepad-body" style="padding:12px 16px">
      <form method="POST" action="/cardconv/keyword/add" style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:16px;align-items:flex-end">
        <div style="display:flex;flex-direction:column;gap:4px">
          <label style="font-size:.7rem;color:var(--text-muted);font-weight:600">KEYWORD</label>
          <input name="kw" placeholder="e.g. STARBUCKS" required style="padding:7px 10px;border:1px solid var(--border);border-radius:6px;background:var(--surface-2);color:var(--text);font-size:.82rem;width:160px">
        </div>
        <div style="display:flex;flex-direction:column;gap:4px">
          <label style="font-size:.7rem;color:var(--text-muted);font-weight:600">G/L ACCOUNT</label>
          <input name="gl" placeholder="53410177" required style="padding:7px 10px;border:1px solid var(--border);border-radius:6px;background:var(--surface-2);color:var(--text);font-size:.82rem;width:110px">
        </div>
        <div style="display:flex;flex-direction:column;gap:4px">
          <label style="font-size:.7rem;color:var(--text-muted);font-weight:600">SER.</label>
          <input name="ser" placeholder="160" style="padding:7px 10px;border:1px solid var(--border);border-radius:6px;background:var(--surface-2);color:var(--text);font-size:.82rem;width:70px">
        </div>
        <div style="display:flex;flex-direction:column;gap:4px;flex:1;min-width:180px">
          <label style="font-size:.7rem;color:var(--text-muted);font-weight:600">PURPOSE</label>
          <input name="purpose" placeholder="Coffee, Snack and meal" required style="padding:7px 10px;border:1px solid var(--border);border-radius:6px;background:var(--surface-2);color:var(--text);font-size:.82rem;width:100%">
        </div>
        <button type="submit" class="btn btn-primary btn-sm" style="align-self:flex-end">+ Add</button>
      </form>
      <div style="max-height:480px;overflow-y:auto">
        <table class="kw-table">
          <thead style="position:sticky;top:0;background:var(--surface)">
            <tr style="border-bottom:1px solid var(--border)">
              <th style="padding:6px 10px;text-align:left;font-size:.7rem;color:var(--text-muted);font-weight:700;text-transform:uppercase">Keyword</th>
              <th style="padding:6px 10px;text-align:left;font-size:.7rem;color:var(--text-muted);font-weight:700;text-transform:uppercase">G/L</th>
              <th style="padding:6px 10px;text-align:left;font-size:.7rem;color:var(--text-muted);font-weight:700;text-transform:uppercase">Ser.</th>
              <th style="padding:6px 10px;text-align:left;font-size:.7rem;color:var(--text-muted);font-weight:700;text-transform:uppercase">Purpose</th>
              <th></th>
            </tr>
          </thead>
          <tbody>{kw_rows}</tbody>
        </table>
      </div>
    </div>
  </div>
</div>
</body></html>'''


# ── Review page ──────────────────────────────────────────────────────────────

def _handle_ledger_update(username: str, entry_id: str, body: dict):
    """POST /cardconv/ledger/<id>/update — manually edit OCR fields of an entry."""
    ledger = _load_ledger(username)
    updated = False
    for e in ledger["entries"]:
        if e.get("id") != entry_id:
            continue
        for field in ("ocr_date", "ocr_merchant"):
            raw = body.get(field)
            if raw is not None:
                val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
                e[field] = val or None
        # Card brand: normalize to amex/visa/other, or clear when blank.
        raw = body.get("card_brand")
        if raw is not None:
            val = (raw[0] if isinstance(raw, list) else str(raw)).strip().lower()
            e["card_brand"] = val if val in ("amex", "visa", "other") else None
        # Usage: free-text tag, defaults back to "Regular" when cleared.
        raw = body.get("usage")
        if raw is not None:
            val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
            e["usage"] = val or "Regular"
        for field in ("ocr_printed_amount", "ocr_handwritten_amount"):
            raw = body.get(field)
            if raw is not None:
                val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
                try:
                    e[field] = float(val) if val else None
                except ValueError:
                    pass
        # ocr_amount always reflects handwritten-priority final amount.
        hw = e.get("ocr_handwritten_amount")
        pr = e.get("ocr_printed_amount")
        e["ocr_amount"] = hw if hw is not None else pr
        updated = True
        break
    if not updated:
        return ("json", {"error": "not found"}, 404)
    _save_ledger(username, ledger)
    return ("json", {"ok": True})


def _handle_review_manual_match(username: str, body: dict):
    """POST /cardconv/review/match — manually link an unmatched review row to a receipt."""
    def val(k):
        v = body.get(k, "")
        return (v[0] if isinstance(v, list) else str(v)).strip()
    row_id   = val("row_id")
    rcpt_id  = val("receipt_id")
    if not row_id or not rcpt_id:
        return ("json", {"error": "missing params"}, 400)

    # Find the receipt in the ledger
    receipts = _load_receipts(username)
    receipt  = next((r for r in receipts if r.get("id") == rcpt_id), None)
    if not receipt:
        return ("json", {"error": "receipt not found"}, 404)

    # Update review row
    review = _load_review(username)
    matched_row = None
    for r in review.get("rows", []):
        if r.get("id") == row_id:
            r["matched"] = True
            r["receipt"] = {
                "id":           receipt.get("id"),
                "file_id":      receipt.get("file_id"),
                "drive_url":    receipt.get("drive_url"),
                "ocr_date":     receipt.get("ocr_date"),
                "ocr_merchant": receipt.get("ocr_merchant"),
                "ocr_amount":   receipt.get("ocr_amount"),
                "filename":     receipt.get("filename"),
            }
            matched_row = r
            break
    if not matched_row:
        return ("json", {"error": "row not found"}, 404)

    # Mark receipt as matched in ledger
    receipt["matched"] = True
    receipt["match_status"] = "matched"
    receipt["matched_at"] = datetime.now().isoformat()
    receipt["matched_transaction"] = {
        "date":   matched_row.get("date"),
        "amount": matched_row.get("amount"),
        "vendor": matched_row.get("merchant"),
    }
    _save_receipts(username, receipts)
    _save_review(username, review)
    return ("json", {"ok": True, "receipt": matched_row["receipt"]})


def _handle_review_reason(username: str, body: dict):
    """POST /cardconv/review/reason — save a loss reason for an unmatched row."""
    def _val(k):
        v = body.get(k, "")
        return (v[0] if isinstance(v, list) else str(v))
    rid = _val("id").strip()
    reason = _val("reason")
    if not rid:
        return ("json", {"error": "missing id"}, 400)
    review = _load_review(username)
    for r in review.get("rows", []):
        if r.get("id") == rid:
            r["loss_reason"] = reason
            _save_review(username, review)
            return ("json", {"ok": True})
    return ("json", {"error": "not found"}, 404)


def _handle_review_download(username: str, query: dict):
    """GET /cardconv/review/download — staged xlsx filtered to a date range.

    from/to are 'YYYY-MM-DD' query params. Rows are filtered by their invoice
    date (column 5). Rows without an invoice date are always kept, matching the
    Ledger/Review filter convention. No range → full file.
    """
    review = _load_review(username)
    out_fn = review.get("out_filename") or ""
    src    = OUT_DIR / out_fn
    if not out_fn or not src.exists():
        return ("html", "<h2 style='padding:40px'>No staged conversion to download.</h2>", 404)

    dfrom = (query.get("from", [""]) or [""])[0]
    dto   = (query.get("to", [""]) or [""])[0]
    if not dfrom and not dto:
        # Unfiltered — serve the original file untouched.
        return ("file", str(src),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", out_fn)

    wb = openpyxl.load_workbook(src)
    ws = wb["sheetMst"]

    def _row_date_str(v):
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        if isinstance(v, date):
            return v.strftime("%Y-%m-%d")
        return str(v)[:10] if v else ""

    drop = []
    r = 2
    while ws.cell(r, 1).value is not None:
        d = _row_date_str(ws.cell(r, 5).value)
        keep = (not d) or ((not dfrom or d >= dfrom) and (not dto or d <= dto))
        if not keep:
            drop.append(r)
        r += 1
    for idx in reversed(drop):
        ws.delete_rows(idx, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return ("file_inline", buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", out_fn)


_CARD_BRAND_LABEL = {"amex": "AMEX", "visa": "Visa", "other": "Other"}


def _handle_ledger_xlsx(username: str, query: dict):
    """GET /cardconv/ledger/download.xlsx — filtered ledger rows as a settlement xlsx.

    Maps the currently-filtered Ledger entries onto the same template `convert`
    uses, sourcing date/vendor/amount from each receipt (matched-transaction
    values preferred when present). Card Type and Usage are appended as columns
    28/29 for visibility. Honors all Ledger filters via the shared parser.
    """
    f = _parse_filter_params(query)
    entries = _apply_ledger_filters(_ledger_entries(username),
                                    f["status"], f["dfrom"], f["dto"],
                                    f["card_brand"], f["usage"], f["completed"])
    if TEMPLATE.exists():
        template_path = TEMPLATE
    elif TEMPLATE_FALLBACK.exists():
        template_path = TEMPLATE_FALLBACK
    else:
        return ("html", "<h2 style='padding:40px'>Template file not found.</h2>", 404)

    rules = _load_kw()
    today = date.today()
    posting_dt = datetime(today.year, today.month, today.day)
    wb = openpyxl.load_workbook(template_path)
    ws = wb["sheetMst"]
    # Extra columns for receipt-centric metadata.
    ws.cell(1, 28).value = "Card Type"
    ws.cell(1, 29).value = "Usage"

    start = 2
    while ws.cell(start, 1).value is not None:
        start += 1

    for e in entries:
        mt       = e.get("matched_transaction") or {}
        vendor   = mt.get("vendor") or e.get("ocr_merchant") or ""
        amount   = mt.get("amount")
        if amount is None:
            amount = e.get("ocr_amount") or 0.0
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            amount = 0.0
        date_str = mt.get("date") or e.get("ocr_date")
        inv_dt   = _parse_date(date_str) if date_str else None

        gl, ser, purpose = _classify(vendor, rules)
        if gl is None:
            gl, ser, purpose = 53410177, "160", "Coffee, Snack and meal"

        ws.cell(start,  1).value = FIXED["receipt_type"]
        ws.cell(start,  2).value = FIXED["employee_id"]
        ws.cell(start,  3).value = FIXED["payee"]
        ws.cell(start,  5).value = inv_dt
        ws.cell(start,  6).value = FIXED["domestic"]
        ws.cell(start,  7).value = vendor
        ws.cell(start,  8).value = posting_dt
        ws.cell(start,  9).value = gl
        ws.cell(start, 10).value = ser
        ws.cell(start, 11).value = FIXED["currency"]
        ws.cell(start, 12).value = FIXED["tax_code"]
        ws.cell(start, 14).value = amount
        ws.cell(start, 15).value = FIXED["cost_center"]
        ws.cell(start, 18).value = purpose
        ws.cell(start, 26).value = amount
        ws.cell(start, 28).value = _CARD_BRAND_LABEL.get(e.get("card_brand"), "")
        ws.cell(start, 29).value = e.get("usage") or "Regular"
        start += 1

    buf = io.BytesIO()
    wb.save(buf)
    out_fn = f"ledger_export_{today.strftime('%Y-%m-%d')}.xlsx"
    return ("file_inline", buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", out_fn)


def _esc(s) -> str:
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;")) if s is not None else ""


def _render_review(user: str) -> str:
    from server import CSS_VER
    review    = _load_review(user)
    rows      = review.get("rows", [])
    total     = review.get("total", len(rows))
    matched   = review.get("matched", 0)
    unmatched = review.get("unmatched", 0)
    out_fn    = review.get("out_filename", "")
    source    = review.get("source", "")
    gen_at    = (review.get("generated_at", "") or "")[:19].replace("T", " ")

    def _money(a):
        return f'${a:,.2f}' if isinstance(a, (int, float)) else (_esc(a) or '–')

    if not rows:
        body_html = ('<div style="text-align:center;color:var(--text-muted);padding:40px">'
                     'No conversion yet — upload a CSV on the '
                     '<a href="/cardconv/convert" style="color:var(--accent)">Convert</a> tab.</div>')
    else:
        items = []
        for r in rows:
            is_matched = r.get("matched")
            rc = r.get("receipt") or {}
            # Transaction (CSV line item) header
            txn = (
                '<div class="rv-txn">'
                  '<div class="rv-txn-main">'
                    f'<span class="rv-date">{_esc(r.get("date")) or "–"}</span>'
                    f'<span class="rv-merchant">{_esc(r.get("merchant"))}</span>'
                  '</div>'
                  '<div class="rv-txn-meta">'
                    f'<span class="rv-amt">{_money(r.get("amount"))}</span>'
                    f'<span class="rv-gl">G/L {_esc(r.get("gl"))}</span>'
                  '</div>'
                '</div>')
            # Inline matched-receipt mini card, or unmatched + loss-reason input
            if is_matched and rc.get("file_id"):
                fid   = rc["file_id"]
                tn    = f'https://drive.google.com/thumbnail?id={fid}&sz=w240'
                proxy = f'/cardconv/receipts/image/{fid}'
                link  = (f'<a href="{_esc(rc.get("drive_url"))}" target="_blank" '
                         f'class="rv-drive-link">🔗 Drive</a>' if rc.get("drive_url") else '')
                # Payload for the lightbox: full image proxy + bbox of this entry and
                # its siblings (multi-receipt page) so the clicked one can be highlighted.
                rv_data = _esc(json.dumps({
                    "fid":      fid,
                    "id":       rc.get("id"),
                    "merchant": rc.get("ocr_merchant"),
                    "bbox":     rc.get("ocr_bbox"),
                    "siblings": rc.get("siblings") or [],
                    "drive":    rc.get("drive_url"),
                }, ensure_ascii=False))
                receipt_block = (
                    '<div class="rv-receipt matched">'
                      f'<img class="rv-thumb" src="{tn}" loading="lazy" '
                      f'data-rv="{rv_data}" title="Click to enlarge" '
                      f'onerror="this.onerror=null;this.src=\'{proxy}\'">'
                      '<div class="rv-card-info">'
                        f'<div class="rv-card-line">🗓 {_esc(rc.get("ocr_date")) or "–"}</div>'
                        f'<div class="rv-card-line rv-card-merchant">{_esc(rc.get("ocr_merchant")) or "–"}</div>'
                        f'<div class="rv-card-line rv-card-amt">{_money(rc.get("ocr_amount"))}</div>'
                        f'{link}'
                      '</div>'
                    '</div>')
            else:
                row_id_esc = _esc(r.get("id", ""))
                txn_json   = _esc(json.dumps({
                    "id": r.get("id",""), "date": r.get("date",""),
                    "merchant": r.get("merchant",""), "amount": r.get("amount"),
                }, ensure_ascii=False))
                receipt_block = (
                    '<div class="rv-receipt unmatched">'
                      '<div class="rv-nomatch">❌ No receipt matched</div>'
                      f'<button type="button" '
                      f'onclick="rvOpenMatchPanel(this)" '
                      f'data-txn="{txn_json}" '
                      f'class="btn btn-ghost btn-sm" '
                      f'style="margin-top:6px;font-size:.74rem;color:var(--accent);width:100%">'
                      f'🔗 Match manually</button>'
                    '</div>')
            item_cls = 'rv-item' + ('' if is_matched else ' unmatched')
            row_date = _esc(r.get("date")) or ""
            items.append(
                f'<div class="{item_cls}" data-date="{row_date}" '
                f'data-matched="{"1" if is_matched else "0"}">{txn}{receipt_block}</div>')
        body_html = "".join(items)

    download_btn = ('<button id="rvDownload" class="btn btn-primary">⬇ Download xlsx</button>'
                    if out_fn else '')
    meta_line = f'{_esc(source)} &nbsp;·&nbsp; {gen_at}' if source else 'No conversion staged'

    return f'''<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>🔍 Review · Wayfinder</title>
<link rel="stylesheet" href="/static/style.css?v={CSS_VER}">
<style>{_CC_TAB_CSS}
.stat-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:16px}}
.stat-card{{background:var(--surface-2);border:1px solid var(--border);border-radius:var(--radius-md);padding:16px 20px;text-align:center}}
.stat-value{{font-size:1.6rem;font-weight:700;color:var(--text);line-height:1.2}}
.stat-label{{font-size:.73rem;color:var(--text-muted);margin-top:4px;text-transform:uppercase;letter-spacing:.06em}}
.filter-bar{{display:flex;align-items:center;gap:10px;padding:10px 16px;background:var(--surface-2);
  border:1px solid var(--border);border-radius:var(--radius-md);margin-bottom:14px;flex-wrap:wrap}}
.filter-bar input[type=date]{{background:var(--surface);border:1px solid var(--border);
  border-radius:6px;color:var(--text);font-size:.82rem;padding:5px 8px;outline:none}}
.filter-bar input[type=date]:focus{{border-color:var(--accent)}}
.preset-btn{{background:var(--surface);border:1px solid var(--border);border-radius:6px;color:var(--text);
  font-size:.76rem;padding:4px 9px;cursor:pointer}}
.preset-btn:hover{{border-color:var(--accent)}}
.preset-btn.active{{background:rgba(250,204,21,.18);border-color:#facc15;color:#b45309;font-weight:700}}
.rv-list{{display:flex;flex-direction:column;gap:10px}}
.rv-item{{display:flex;gap:14px;align-items:stretch;background:var(--surface-2);border:1px solid var(--border);border-radius:var(--radius-md);padding:12px 14px}}
.rv-item.unmatched{{border-color:rgba(239,68,68,.35);background:rgba(239,68,68,.06)}}
.rv-txn{{flex:1;min-width:0;display:flex;flex-direction:column;justify-content:center;gap:6px}}
.rv-txn-main{{display:flex;flex-direction:column;gap:2px}}
.rv-date{{font-size:.74rem;color:var(--text-muted)}}
.rv-merchant{{font-size:.95rem;font-weight:700;color:var(--text)}}
.rv-txn-meta{{display:flex;gap:12px;align-items:baseline;flex-wrap:wrap}}
.rv-amt{{font-size:1.05rem;font-weight:700;color:var(--text)}}
.rv-gl{{font-size:.74rem;color:var(--text-muted)}}
.rv-receipt{{flex:0 0 230px;border-left:1px solid var(--border);padding-left:14px}}
.rv-receipt.matched{{display:flex;gap:12px;align-items:flex-start}}
.rv-thumb{{width:120px;height:120px;border-radius:8px;object-fit:cover;border:1px solid var(--border);background:var(--surface-3);cursor:zoom-in;transition:border-color .12s}}
.rv-thumb:hover{{border-color:var(--accent)}}
.rv-lb{{position:fixed;inset:0;background:rgba(2,6,23,.82);display:none;align-items:center;justify-content:center;z-index:1000;padding:24px}}
.rv-lb.open{{display:flex}}
.rv-lb-box{{position:relative;max-width:92vw;max-height:92vh;display:flex;flex-direction:column;gap:10px}}
.rv-lb-img-wrap{{position:relative;display:inline-block;max-width:92vw;max-height:80vh}}
.rv-lb-img{{max-width:92vw;max-height:80vh;border-radius:8px;display:block}}
.rv-lb-svg{{position:absolute;top:0;left:0;pointer-events:none;display:none}}
.rv-lb-bar{{display:flex;align-items:center;justify-content:space-between;gap:14px;color:#e2e8f0;font-size:.85rem}}
.rv-lb-close{{position:absolute;top:-12px;right:-12px;width:34px;height:34px;border-radius:50%;border:none;
  background:var(--surface);color:var(--text);font-size:1.1rem;cursor:pointer;box-shadow:0 2px 8px rgba(0,0,0,.4)}}
.rv-lb-open{{color:var(--accent);text-decoration:none;font-size:.82rem}}
.rv-lb-open:hover{{text-decoration:underline}}
.rv-card-info{{display:flex;flex-direction:column;gap:4px;min-width:0}}
.rv-card-line{{font-size:.8rem;color:var(--text-muted)}}
.rv-card-merchant{{font-weight:600;color:var(--text);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.rv-card-amt{{font-size:1rem;font-weight:700;color:#22c55e}}
.rv-drive-link{{font-size:.76rem;color:var(--accent);text-decoration:none;margin-top:2px}}
.rv-drive-link:hover{{text-decoration:underline}}
.rv-nomatch{{color:var(--danger);font-size:.84rem;font-weight:700;margin-bottom:6px}}

@media(max-width:600px){{.rv-item{{flex-direction:column;gap:10px}}.rv-receipt{{flex:none;border-left:none;border-top:1px solid var(--border);padding-left:0;padding-top:10px}}}}
.rv-foot{{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:16px 4px;flex-wrap:wrap}}
@media(max-width:600px){{.stat-grid{{grid-template-columns:1fr 1fr 1fr}}}}
</style>
</head><body>
<nav>
  <span class="nav-brand">💳 Card Converter</span>
  <span class="nav-user">👤 {user} &nbsp;·&nbsp; <a href="/logout">Logout</a></span>
</nav>
<div class="container" style="max-width:920px">
  {_tab_bar("review", user)}

  <div style="font-size:.8rem;color:var(--text-muted);margin-bottom:12px;display:flex;align-items:center;gap:6px">
    <span>{meta_line}</span>{_info_icon('Shows converted transactions with receipt matching results. Unmatched rows (red) can be linked via 🔗 Match manually, or carried over to the next billing cycle.', right=True)}
  </div>

  <div class="stat-grid">
    <div class="stat-card"><div class="stat-value" id="rvTotal">{total}</div><div class="stat-label">Total</div></div>
    <div class="stat-card"><div class="stat-value" id="rvMatched" style="color:#22c55e">{matched}</div><div class="stat-label">Matched</div></div>
    <div class="stat-card"><div class="stat-value" id="rvUnmatched" style="color:#ef4444">{unmatched}</div><div class="stat-label">Unmatched</div></div>
  </div>

  <div class="filter-bar">
    📅 <input type="date" id="rvFrom"> ~ <input type="date" id="rvTo">
    <button class="btn btn-ghost btn-sm" id="rvReset">Reset</button>
    <span style="flex:1"></span>
    {download_btn}
  </div>

  <div class="filter-bar" style="gap:8px">
    <span style="font-size:.76rem;color:var(--text-muted)">Quick range:</span>
    <button class="preset-btn" data-preset="month">This month</button>
    <button class="preset-btn" data-preset="30d">Last 30 days</button>
    <button class="preset-btn" data-preset="3m">Last 3 months</button>
    <button class="preset-btn" data-preset="ytd">YTD</button>
    <button class="preset-btn" data-preset="all">All time</button>
  </div>

  <div class="notepad-card">
    <div class="notepad-body" style="padding:12px 14px">
      <div class="rv-list">{body_html}</div>
    </div>
  </div>

  <div class="rv-foot">
    <span style="font-size:.78rem;color:var(--text-muted)">각 거래 옆에 매칭된 영수증이 표시됩니다. 미매칭 거래는 빨간색으로 표시됩니다.</span>
  </div>
</div>

<div class="rv-lb" id="rvLb">
  <div class="rv-lb-box">
    <button class="rv-lb-close" id="rvLbClose" title="Close">×</button>
    <div class="rv-lb-img-wrap">
      <img class="rv-lb-img" id="rvLbImg" alt="receipt">
      <svg class="rv-lb-svg" id="rvLbSvg"></svg>
    </div>
    <div class="rv-lb-bar">
      <span id="rvLbCaption"></span>
      <a id="rvLbDrive" class="rv-lb-open" target="_blank" rel="noopener">🔗 Open in Drive</a>
    </div>
  </div>
</div>
<script>
// ── Date filter + presets (mirrors the Ledger page) ──────────────────────────
const $ = id => document.getElementById(id);
function iso(d){{ return d.toISOString().slice(0,10); }}

// Date filters keep rows without an invoice date always visible, matching Ledger.
function applyFilter(){{
  const from = $('rvFrom').value, to = $('rvTo').value;
  let total=0, matched=0, unmatched=0;
  document.querySelectorAll('.rv-item').forEach(it => {{
    const d = it.dataset.date || '';
    const show = (!from || !d || d >= from) && (!to || !d || d <= to);
    it.style.display = show ? '' : 'none';
    if(show){{
      total++;
      if(it.dataset.matched === '1') matched++; else unmatched++;
    }}
  }});
  $('rvTotal').textContent = total;
  $('rvMatched').textContent = matched;
  $('rvUnmatched').textContent = unmatched;
}}

function applyPreset(p){{
  const now = new Date();
  let from = '', to = iso(now);
  if(p==='month')    from = iso(new Date(now.getFullYear(), now.getMonth(), 1));
  else if(p==='30d') from = iso(new Date(now.getTime() - 29*86400000));
  else if(p==='3m')  from = iso(new Date(now.getFullYear(), now.getMonth()-3, now.getDate()));
  else if(p==='ytd') from = iso(new Date(now.getFullYear(), 0, 1));
  else if(p==='all'){{ from = ''; to = ''; }}
  $('rvFrom').value = from;
  $('rvTo').value = to;
  document.querySelectorAll('.preset-btn').forEach(b =>
    b.classList.toggle('active', b.dataset.preset === p));
  applyFilter();
}}

$('rvFrom').addEventListener('change', () => {{ clearPresetActive(); applyFilter(); }});
$('rvTo').addEventListener('change', () => {{ clearPresetActive(); applyFilter(); }});
$('rvReset').addEventListener('click', () => {{
  $('rvFrom').value = ''; $('rvTo').value = ''; clearPresetActive(); applyFilter();
}});
function clearPresetActive(){{ document.querySelectorAll('.preset-btn').forEach(b => b.classList.remove('active')); }}
document.querySelectorAll('.preset-btn').forEach(b =>
  b.addEventListener('click', () => applyPreset(b.dataset.preset)));

// Filtered xlsx download — only the currently filtered transactions are included.
const rvDl = $('rvDownload');
if(rvDl){{
  rvDl.addEventListener('click', () => {{
    const p = new URLSearchParams();
    if($('rvFrom').value) p.set('from', $('rvFrom').value);
    if($('rvTo').value)   p.set('to', $('rvTo').value);
    window.location = '/cardconv/review/download?' + p.toString();
  }});
}}

// ── Receipt lightbox ─────────────────────────────────────────────────────────
// Click a thumbnail → full image with bbox overlay (multi-receipt highlights the
// matched entry). ocr_bbox is [ymin,xmin,ymax,xmax] in a 0-1000 coord system.
const rvLb = $('rvLb'), rvLbImg = $('rvLbImg'), rvLbSvg = $('rvLbSvg');

function rvEscSvg(s){{
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}

function rvPaintBoxes(boxes, curId){{
  const img = rvLbImg, svg = rvLbSvg;
  if(!boxes.length || !img.naturalWidth){{ svg.innerHTML=''; svg.style.display='none'; return; }}
  const cW=img.clientWidth, cH=img.clientHeight, nW=img.naturalWidth, nH=img.naturalHeight;
  const scale=Math.min(cW/nW, cH/nH), dW=nW*scale, dH=nH*scale;
  svg.style.left=((cW-dW)/2)+'px'; svg.style.top=((cH-dH)/2)+'px';
  svg.style.width=dW+'px'; svg.style.height=dH+'px';
  svg.setAttribute('viewBox','0 0 '+dW+' '+dH);
  svg.style.display='block';
  svg.innerHTML = boxes.map(function(x,i){{
    const b=x.ocr_bbox;
    const x0=b[1]/1000*dW, y0=b[0]/1000*dH, x1=b[3]/1000*dW, y1=b[2]/1000*dH;
    const isCur=(x.id===curId);
    const col=isCur?'#38bdf8':'#64748b', sw=isCur?2.5:1.5;
    const ty=Math.max(y0+13,13);
    return '<rect x="'+x0.toFixed(1)+'" y="'+y0.toFixed(1)+'" width="'+(x1-x0).toFixed(1)+
      '" height="'+(y1-y0).toFixed(1)+'" rx="4" fill="'+col+'" fill-opacity="'+(isCur?0.12:0.05)+
      '" stroke="'+col+'" stroke-width="'+sw+'"/>'+
      '<text x="'+(x0+4).toFixed(1)+'" y="'+ty.toFixed(1)+'" fill="'+col+'" font-size="11" '+
      'font-weight="700" style="paint-order:stroke;stroke:rgba(2,6,23,.75);stroke-width:3px">'+
      rvEscSvg(''+(i+1))+'</text>';
  }}).join('');
}}

function rvOpenLightbox(data){{
  // Build the list of boxes to draw: siblings if present, else this entry alone.
  let boxes = (data.siblings && data.siblings.length)
    ? data.siblings.filter(s => Array.isArray(s.ocr_bbox))
    : (Array.isArray(data.bbox) ? [{{id: data.id, ocr_bbox: data.bbox}}] : []);
  rvLbSvg.innerHTML=''; rvLbSvg.style.display='none';
  rvLbImg.src = '/cardconv/receipts/image/' + data.fid;
  $('rvLbCaption').textContent = data.merchant || '';
  const drive = $('rvLbDrive');
  if(data.drive){{ drive.href = data.drive; drive.style.display=''; }}
  else drive.style.display='none';
  const render = () => rvPaintBoxes(boxes, data.id);
  rvLbImg.onload = render;
  if(rvLbImg.complete && rvLbImg.naturalWidth) render();
  rvLb.classList.add('open');
}}

function rvCloseLightbox(){{ rvLb.classList.remove('open'); rvLbImg.src=''; }}

document.querySelectorAll('.rv-thumb[data-rv]').forEach(img => {{
  img.addEventListener('click', () => {{
    try {{ rvOpenLightbox(JSON.parse(img.dataset.rv)); }} catch(e) {{}}
  }});
}});
$('rvLbClose').addEventListener('click', rvCloseLightbox);
rvLb.addEventListener('click', e => {{ if(e.target === rvLb) rvCloseLightbox(); }});
document.addEventListener('keydown', e => {{ if(e.key === 'Escape') rvCloseLightbox(); }});
window.addEventListener('resize', () => {{ if(rvLb.classList.contains('open')) rvLbImg.onload && rvLbImg.onload(); }});
applyPreset('all');

// ── Manual match — right-side panel ──────────────────────────────────────────
var _mmTxn = null;

function rvOpenMatchPanel(btn) {{
  var txnRaw = btn ? btn.getAttribute('data-txn') : '{{}}';
  try {{ _mmTxn = JSON.parse(txnRaw); }} catch(x) {{ _mmTxn = {{}}; }}

  var pop = document.getElementById('rvMatchPop');
  var list = document.getElementById('rvMatchList');

  // Position popover near the button
  var rect = btn.getBoundingClientRect();
  var popW = 320;
  var left = rect.right + 8;
  if (left + popW > window.innerWidth - 8) left = rect.left - popW - 8;
  if (left < 8) left = 8;
  var top = rect.top;
  var maxH = window.innerHeight - top - 16;
  if (maxH < 200) {{ top = Math.max(8, rect.bottom - 300); maxH = Math.min(300, window.innerHeight - top - 8); }}

  pop.style.left  = left + 'px';
  pop.style.top   = top + 'px';
  pop.style.maxHeight = Math.max(200, maxH) + 'px';
  pop.style.display = 'flex';

  document.getElementById('rvPopTitle').textContent =
    (_mmTxn.merchant || '') + (_mmTxn.date ? '  ' + _mmTxn.date : '') +
    (_mmTxn.amount != null ? '  $' + Number(_mmTxn.amount).toFixed(2) : '');

  list.innerHTML = '<div style="padding:16px;text-align:center;color:var(--text-muted);font-size:.8rem">Loading…</div>';

  fetch('/cardconv/ledger/api?status=all')
    .then(function(r) {{ return r.json(); }})
    .then(function(d) {{ rvRenderMatchList(d.entries || []); }})
    .catch(function() {{ list.innerHTML = '<div style="padding:12px;color:red;font-size:.8rem">Load failed</div>'; }});
}}

function rvCloseMatchPanel() {{
  document.getElementById('rvMatchPop').style.display = 'none';
}}

function rvRenderMatchList(entries) {{
  var list = document.getElementById('rvMatchList');
  var pending = entries.filter(function(e) {{ return e.match_status !== 'matched'; }});
  if (!pending.length) {{
    list.innerHTML = '<div style="padding:16px;text-align:center;color:var(--text-muted);font-size:.8rem">No unmatched receipts</div>';
    return;
  }}
  list.innerHTML = pending.map(function(e) {{
    var fid = e.file_id || '';
    var tn  = fid ? 'https://drive.google.com/thumbnail?id=' + fid + '&sz=w80' : '';
    var img = tn ? '<img src="' + tn + '" width="44" height="44" style="object-fit:cover;border-radius:4px;flex-shrink:0">' : '';
    var badge = e.match_status === 'pending_match' ? 'Pending' : 'Unmatched';
    return '<div style="display:flex;gap:8px;align-items:center;padding:8px 12px;border-bottom:1px solid var(--border);cursor:pointer;font-size:.8rem" '
      + 'onclick="rvDoMatch(this)" data-rcpt="' + e.id.replace(/"/g, '') + '">'
      + img
      + '<div style="flex:1;min-width:0">'
      +   '<div style="font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">' + (e.ocr_merchant || '–') + '</div>'
      +   '<div style="color:var(--text-muted);font-size:.72rem">' + (e.ocr_date || '–') + '  ' + badge + '</div>'
      + '</div>'
      + '<div style="font-weight:700;color:var(--accent);flex-shrink:0">' + (e.ocr_amount != null ? '$' + Number(e.ocr_amount).toFixed(2) : '–') + '</div>'
      + '</div>';
  }}).join('');
}}

function rvDoMatch(el) {{
  if (!_mmTxn) return;
  var rcptId = el.dataset.rcpt;
  fetch('/cardconv/review/match', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
    body: 'row_id=' + encodeURIComponent(_mmTxn.id || '') + '&receipt_id=' + encodeURIComponent(rcptId)
  }}).then(function(r) {{ return r.json(); }}).then(function(d) {{
    if (d.ok) {{ rvCloseMatchPanel(); location.reload(); }}
    else {{ alert('Match failed: ' + (d.error || 'unknown')); }}
  }});
}}

document.addEventListener('keydown', function(e) {{
  if (e.key === 'Escape') rvCloseMatchPanel();
}});
document.addEventListener('click', function(e) {{
  var pop = document.getElementById('rvMatchPop');
  if (pop && pop.style.display !== 'none' && !pop.contains(e.target) && !e.target.closest('[onclick*="rvOpenMatchPanel"]'))
    rvCloseMatchPanel();
}});
</script>

<!-- Manual match popover -->
<div id="rvMatchPop" style="display:none;position:fixed;width:320px;background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-md);z-index:9999;flex-direction:column;box-shadow:0 8px 32px rgba(0,0,0,.35);overflow:hidden">
  <div style="display:flex;align-items:center;justify-content:space-between;padding:10px 12px;border-bottom:1px solid var(--border);flex-shrink:0">
    <div id="rvPopTitle" style="font-size:.78rem;font-weight:600;color:var(--text);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:250px"></div>
    <button onclick="rvCloseMatchPanel()" style="background:none;border:none;color:var(--text-muted);font-size:1.2rem;cursor:pointer;line-height:1;padding:0 2px;flex-shrink:0">&times;</button>
  </div>
  <div id="rvMatchList" style="overflow-y:auto;max-height:inherit"></div>
</div>
</body></html>'''


def _render_ocr_staging_review(user: str) -> str:
    from server import CSS_VER
    staging  = _load_ocr_staging(user)
    entries  = staging.get("entries", [])

    def money(a):
        return f'${a:,.2f}' if isinstance(a, (int, float)) else (_esc(str(a)) if a else '–')

    cards = []
    for e in entries:
        eid     = _esc(e.get("id", ""))
        fid     = e.get("file_id", "")
        proxy   = f'/cardconv/receipts/image/{_esc(fid)}' if fid else ''
        fn      = _esc(e.get("filename", ""))
        date_v  = _esc(e.get("ocr_date") or '–')
        merch_v = _esc(e.get("ocr_merchant") or '–')
        amt_v   = money(e.get("ocr_amount"))
        hw_v    = money(e.get("ocr_handwritten_amount"))
        status  = _esc(e.get("ocr_status", ""))

        img_html = (f'<img src="{proxy}" class="stg-thumb" loading="lazy" '
                    f'onerror="this.style.display=\'none\'">'
                    if proxy else '<div class="stg-nophoto">No image</div>')

        ocr_ok = e.get("ocr_status") == "done" and e.get("ocr_merchant")
        badge  = ('<span class="stg-badge ok">OCR OK</span>' if ocr_ok
                  else '<span class="stg-badge warn">OCR partial</span>')

        cards.append(f'''
<div class="stg-card" id="card-{eid}">
  <label class="stg-check-wrap">
    <input type="checkbox" name="confirmed" value="{eid}" {"checked" if ocr_ok else ""}>
    <span class="stg-check-lbl">Include</span>
  </label>
  <div class="stg-img-wrap">{img_html}</div>
  <div class="stg-info">
    <div class="stg-filename">{fn} {badge}</div>
    <div class="stg-row"><span class="stg-lbl">Date</span><span class="stg-val">{date_v}</span></div>
    <div class="stg-row"><span class="stg-lbl">Merchant</span><span class="stg-val">{merch_v}</span></div>
    <div class="stg-row"><span class="stg-lbl">Printed</span><span class="stg-val">{amt_v}</span></div>
    <div class="stg-row"><span class="stg-lbl">Handwritten</span><span class="stg-val">{hw_v}</span></div>
  </div>
</div>''')

    body_html = ''.join(cards) if cards else (
        '<div class="stg-empty">No pending OCR entries. '
        '<a href="/cardconv/ledger">Go to Ledger</a></div>')

    count = len(entries)
    return f'''<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>OCR Review · Wayfinder</title>
<link rel="stylesheet" href="/static/style.css?v={CSS_VER}">
<style>
{_CC_TAB_CSS}
.stg-header{{display:flex;align-items:center;justify-content:space-between;margin-bottom:18px;flex-wrap:wrap;gap:10px}}
.stg-title{{font-size:1.2rem;font-weight:700}}
.stg-count{{font-size:.85rem;color:var(--text-muted)}}
.stg-actions{{display:flex;gap:10px;align-items:center;flex-wrap:wrap}}
.stg-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:16px}}
.stg-card{{background:var(--surface-2);border:1px solid var(--border);border-radius:var(--radius-md);
  overflow:hidden;display:flex;flex-direction:column}}
.stg-card:has(input:not(:checked)){{opacity:.55;border-style:dashed}}
.stg-check-wrap{{display:flex;align-items:center;gap:8px;padding:10px 12px;border-bottom:1px solid var(--border);
  cursor:pointer;background:var(--surface)}}
.stg-check-wrap input{{width:16px;height:16px;accent-color:var(--accent);cursor:pointer}}
.stg-check-lbl{{font-size:.82rem;font-weight:600;color:var(--text)}}
.stg-img-wrap{{background:#000;display:flex;align-items:center;justify-content:center;min-height:160px;max-height:260px;overflow:hidden}}
.stg-thumb{{max-width:100%;max-height:260px;object-fit:contain;display:block}}
.stg-nophoto{{color:var(--text-muted);font-size:.8rem;padding:20px}}
.stg-info{{padding:12px}}
.stg-filename{{font-size:.78rem;color:var(--text-muted);margin-bottom:8px;display:flex;align-items:center;gap:6px;flex-wrap:wrap}}
.stg-row{{display:flex;justify-content:space-between;font-size:.83rem;padding:3px 0;border-bottom:1px solid var(--border-faint,var(--border))}}
.stg-row:last-child{{border-bottom:none}}
.stg-lbl{{color:var(--text-muted);font-size:.76rem}}
.stg-val{{font-weight:600}}
.stg-badge{{font-size:.65rem;font-weight:700;padding:2px 7px;border-radius:10px}}
.stg-badge.ok{{background:rgba(34,197,94,.15);color:#22c55e}}
.stg-badge.warn{{background:rgba(245,158,11,.15);color:#f59e0b}}
.stg-empty{{text-align:center;color:var(--text-muted);padding:60px 20px}}
</style>
</head><body>
{_tab_bar("ocr_review", user)}
<div style="max-width:1100px;margin:0 auto;padding:20px 16px">
  <div class="stg-header">
    <div>
      <div class="stg-title">OCR Review</div>
      <div class="stg-count">{count} receipt(s) pending — check what to add to the ledger</div>
    </div>
    <div class="stg-actions">
      <button type="button" onclick="toggleAll(true)" class="btn btn-secondary" style="font-size:.82rem;padding:6px 14px">Check All</button>
      <button type="button" onclick="toggleAll(false)" class="btn btn-secondary" style="font-size:.82rem;padding:6px 14px">Uncheck All</button>
    </div>
  </div>
  <form method="POST" action="/cardconv/receipts/review/confirm" id="stgForm">
    <div class="stg-grid">
      {body_html}
    </div>
    <div style="display:flex;gap:12px;margin-top:24px;justify-content:flex-end;flex-wrap:wrap">
      <form method="POST" action="/cardconv/receipts/review/discard" style="margin:0">
        <button type="submit" class="btn btn-danger" onclick="return confirm('Discard all staged entries?')"
          style="font-size:.85rem">Discard All</button>
      </form>
      <button type="submit" form="stgForm" class="btn btn-primary" style="font-size:.85rem">
        ✓ Confirm Selected → Add to Ledger
      </button>
    </div>
  </form>
</div>
<script>
function toggleAll(on) {{
  document.querySelectorAll('#stgForm input[type=checkbox]').forEach(cb => cb.checked = on);
}}
</script>
</body></html>'''


def _render_ledger(user: str) -> str:
    from server import CSS_VER
    vapid_pub = os.environ.get("VAPID_PUBLIC_KEY", "")
    return (_LEDGER_HTML
            .replace("__CSSVER__", str(CSS_VER))
            .replace("__USER__", user)
            .replace("__TABS__", _tab_bar("ledger", user))
            .replace("__REGISTER__", _register_section(user))
            .replace("__TABCSS__", _CC_TAB_CSS + _UPLOAD_CSS)
            .replace("__RCPTJS__", _RCPT_JS)
            .replace("__VAPID_PUB__", vapid_pub))


# Raw (non-f) template so CSS/JS braces need no escaping; only __TOKENS__ are filled.
_LEDGER_HTML = r'''<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>🧾 Receipt Ledger · Wayfinder</title>
<link rel="stylesheet" href="/static/style.css?v=__CSSVER__">
<style>
__TABCSS__
.stat-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px}
.stat-card{background:var(--surface-2);border:1px solid var(--border);border-radius:var(--radius-md);
  padding:16px 20px;text-align:center}
.stat-value{font-size:1.6rem;font-weight:700;color:var(--text);line-height:1.2}
.stat-label{font-size:.73rem;color:var(--text-muted);margin-top:4px;text-transform:uppercase;letter-spacing:.06em}
.filter-bar{display:flex;align-items:center;gap:10px;padding:10px 16px;background:var(--surface-2);
  border:1px solid var(--border);border-radius:var(--radius-md);margin-bottom:14px;flex-wrap:wrap}
.filter-bar input[type=date],.filter-bar select{background:var(--surface);border:1px solid var(--border);
  border-radius:6px;color:var(--text);font-size:.82rem;padding:5px 8px;outline:none}
.filter-bar input[type=date]:focus,.filter-bar select:focus{border-color:var(--accent)}
.ledger-table{width:100%;border-collapse:collapse;font-size:.83rem}
.ledger-table th{padding:8px 12px;text-align:left;font-size:.72rem;font-weight:700;text-transform:uppercase;
  letter-spacing:.07em;color:var(--text-muted);border-bottom:1px solid var(--border)}
.ledger-table td{padding:10px 12px;border-bottom:1px solid var(--border);vertical-align:middle}
.ledger-table tbody tr:hover td{background:var(--surface-2);cursor:pointer}
.ledger-table tr:last-child td{border-bottom:none}
.ledger-table tr.dup-row td{background:rgba(250,204,21,.10)}
.ledger-table tr.dup-row:hover td{background:rgba(250,204,21,.18)}
.dup-tag{display:inline-block;margin-left:6px;padding:1px 6px;border-radius:10px;font-size:.6rem;
  font-weight:700;background:rgba(250,204,21,.22);color:#b45309;white-space:nowrap}
.keep-tag{display:inline-block;margin-left:6px;padding:1px 6px;border-radius:10px;font-size:.6rem;
  font-weight:700;background:rgba(34,197,94,.16);color:#16a34a;white-space:nowrap}
.row-check{width:15px;height:15px;cursor:pointer;accent-color:var(--accent)}
.preset-btn{background:var(--surface);border:1px solid var(--border);border-radius:6px;color:var(--text);
  font-size:.76rem;padding:4px 9px;cursor:pointer}
.preset-btn:hover{border-color:var(--accent)}
.preset-btn.active{background:rgba(250,204,21,.18);border-color:#facc15;color:#b45309;font-weight:700}
.grp-toggle{display:inline-block;margin-left:6px;padding:1px 7px;border-radius:10px;font-size:.62rem;
  font-weight:700;background:rgba(99,102,241,.18);color:#6366f1;cursor:pointer;white-space:nowrap}
.grp-toggle:hover{background:rgba(99,102,241,.3)}
.ledger-table tr.dup-child{display:none}
.ledger-table tr.dup-child.show{display:table-row}
.del-modal{position:fixed;top:50%;left:50%;transform:translate(-50%,-50%) scale(.96);z-index:120;
  width:380px;max-width:92vw;background:var(--surface);border:1px solid var(--border);
  border-radius:var(--radius-md);padding:22px 22px 18px;opacity:0;pointer-events:none;
  transition:opacity .18s,transform .18s;box-shadow:0 12px 40px rgba(0,0,0,.4)}
.del-modal.open{opacity:1;pointer-events:all;transform:translate(-50%,-50%) scale(1)}
.del-title{font-size:1rem;font-weight:700;margin-bottom:10px}
.del-body{font-size:.86rem;color:var(--text);margin-bottom:14px}
.del-check{display:flex;align-items:center;gap:8px;font-size:.82rem;color:var(--text-muted);
  cursor:pointer;margin-bottom:18px}
.del-check input{width:15px;height:15px;cursor:pointer;accent-color:var(--danger)}
.del-actions{display:flex;justify-content:flex-end;gap:10px}
.receipt-thumb{width:40px;height:40px;border-radius:6px;object-fit:cover;border:1px solid var(--border);
  background:var(--surface-3);cursor:zoom-in}
.receipt-thumb-placeholder{width:40px;height:40px;border-radius:6px;border:1px dashed var(--border);
  display:flex;align-items:center;justify-content:center;font-size:.75rem;color:var(--text-muted)}
.status-badge{display:inline-flex;align-items:center;gap:4px;padding:3px 8px;border-radius:999px;
  font-size:.72rem;font-weight:700;white-space:nowrap}
.status-matched{background:rgba(34,197,94,.15);color:#22c55e}
.status-unmatched{background:rgba(239,68,68,.15);color:#ef4444}
.status-pending_match{background:rgba(245,158,11,.15);color:#f59e0b}
.ai-badge{font-size:.62rem;font-weight:700;padding:1px 6px;border-radius:10px;white-space:nowrap}
.ai-badge.gemini{color:#1a73e8;background:rgba(26,115,232,.1)}
.ai-badge.claude{color:#7c3aed;background:rgba(124,58,237,.1)}
.card-badge{font-size:.66rem;font-weight:700;padding:2px 8px;border-radius:10px;white-space:nowrap}
.card-amex{color:#1e40af;background:rgba(37,99,235,.14)}
.card-visa{color:#6d28d9;background:rgba(124,58,237,.12)}
.card-other{color:#64748b;background:rgba(100,116,139,.14)}
.comp-tag{display:inline-block;margin-left:6px;padding:1px 6px;border-radius:10px;font-size:.6rem;
  font-weight:700;background:rgba(129,140,248,.18);color:#818cf8;white-space:nowrap}
.ledger-table tr.completed-row td{opacity:.62}
.ledger-table tr.completed-row:hover td{opacity:.85}
.overlay-bg{position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:100;opacity:0;pointer-events:none;transition:opacity .2s}
.overlay-bg.open{opacity:1;pointer-events:all}
.detail-panel{position:fixed;top:0;right:0;width:420px;max-width:100vw;height:100vh;background:var(--surface);
  border-left:1px solid var(--border);z-index:101;transform:translateX(100%);
  transition:transform .25s cubic-bezier(.4,0,.2,1);overflow-y:auto;display:flex;flex-direction:column}
.detail-panel.open{transform:translateX(0)}
.detail-panel-header{display:flex;align-items:center;justify-content:space-between;padding:16px 20px;
  border-bottom:1px solid var(--border)}
.detail-section{padding:16px 20px;border-bottom:1px solid var(--border)}
.detail-section-title{font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;
  color:var(--text-muted);margin-bottom:10px}
.detail-row{display:flex;justify-content:space-between;font-size:.84rem;padding:4px 0}
.detail-row .key{color:var(--text-muted)}
.detail-row .val{font-weight:600;color:var(--text)}
.receipt-image-full{width:100%;border-radius:var(--radius-md);border:1px solid var(--border);
  object-fit:contain;max-height:320px;background:var(--surface-2)}
.receipt-image-wrap{position:relative;line-height:0}
/* SVG is sized/positioned in JS to overlap the rendered (object-fit:contain) image rect. */
.receipt-bbox-overlay{position:absolute;top:0;left:0;pointer-events:none;display:none}
.detail-actions{padding:16px 20px;display:flex;flex-direction:column;gap:8px;margin-top:auto}
.pagination{display:flex;align-items:center;justify-content:center;gap:12px;padding:14px;font-size:.82rem;color:var(--text-muted)}
.pagination button{background:var(--surface-2);border:1px solid var(--border);border-radius:6px;color:var(--text);
  padding:4px 12px;font-size:.8rem;cursor:pointer}
.pagination button:disabled{opacity:.4;cursor:default}
@media(max-width:600px){.detail-panel{width:100vw}.stat-grid{grid-template-columns:1fr 1fr}}
</style>
</head><body>
<nav>
  <span class="nav-brand">💳 Card Converter</span>
  <span class="nav-user">👤 __USER__ &nbsp;·&nbsp; <a href="/logout">Logout</a></span>
</nav>
<div class="container" style="max-width:1100px">

  __TABS__

  __REGISTER__

  <div class="stat-grid">
    <div class="stat-card"><div class="stat-value" id="statTotal">–</div><div class="stat-label">Total</div></div>
    <div class="stat-card"><div class="stat-value" id="statMatched" style="color:#22c55e">–</div><div class="stat-label">Matched</div></div>
    <div class="stat-card"><div class="stat-value" id="statUnmatched" style="color:#ef4444">–</div><div class="stat-label">Unmatched</div></div>
    <div class="stat-card"><div class="stat-value" id="statCompleted" style="color:#818cf8">–</div><div class="stat-label">Completed</div></div>
  </div>

  <div class="filter-bar">
    📅 <input type="date" id="fFrom"> ~ <input type="date" id="fTo">
    <span style="margin-left:8px">Status:</span>
    <select id="fStatus">
      <option value="all">All</option>
      <option value="matched">Matched</option>
      <option value="unmatched">Unmatched</option>
      <option value="pending_match">Pending Match</option>
    </select>
    <span>Card:</span>
    <select id="fCard">
      <option value="all">All</option>
      <option value="amex">AMEX</option>
      <option value="visa">Visa</option>
      <option value="other">Other</option>
      <option value="unknown">Unknown</option>
    </select>
    <span>Usage:</span>
    <select id="fUsage"><option value="all">All</option></select>
    <span>Show:</span>
    <select id="fCompleted">
      <option value="hide">Active (hide completed)</option>
      <option value="only">Completed only</option>
      <option value="all">All</option>
    </select>
    <button class="btn btn-ghost btn-sm" id="fReset">Reset</button>
    <span style="margin-left:auto;display:flex;gap:6px">
      <button class="btn btn-secondary btn-sm" id="fDownloadXlsx">⬇ Download xlsx</button>
      <button class="btn btn-ghost btn-sm" id="fDownload">📄 PDF</button>
    </span>
  </div>

  <div class="filter-bar" style="gap:8px">
    <span style="font-size:.76rem;color:var(--text-muted)">Quick range:</span>
    <button class="preset-btn" data-preset="month">This month</button>
    <button class="preset-btn" data-preset="30d">Last 30 days</button>
    <button class="preset-btn" data-preset="3m">Last 3 months</button>
    <button class="preset-btn" data-preset="ytd">YTD</button>
    <button class="preset-btn" data-preset="all">All time</button>
    <button class="preset-btn" id="viewToggle" title="Collapse duplicate receipts into one row">🔁 Group Duplicates</button>
    <span class="cc-info-wrap"><span class="cc-info" onclick="ccTipToggle(this)">ℹ</span><span class="cc-tip">같은 영수증이 여러 장 인식된 경우 그룹으로 묶어 표시합니다. 불필요한 중복은 삭제하세요.</span></span>
    <button class="btn btn-sm" id="fComplete" disabled
      style="margin-left:auto;background:rgba(129,140,248,.15);color:#818cf8;border:1px solid rgba(129,140,248,.3)">
      ✓ Complete Selected (0)</button>
    <button class="btn btn-sm" id="fUncomplete" disabled
      style="background:rgba(148,163,184,.15);color:#94a3b8;border:1px solid rgba(148,163,184,.3)">
      ↩ Un-complete (0)</button>
    <button class="btn btn-sm" id="fDelete" disabled
      style="background:rgba(239,68,68,.15);color:#ef4444;border:1px solid rgba(239,68,68,.3)">
      🗑 Delete Selected (0)</button>
  </div>

  <div class="notepad-card">
    <div class="notepad-body" style="padding:8px 16px 4px">
      <table class="ledger-table">
        <thead><tr>
          <th style="width:24px"><input type="checkbox" class="row-check" id="checkAll" title="Select all"></th>
          <th>Date</th><th>Printed</th><th>Handwritten</th><th>Final</th><th>Merchant</th><th>Card</th><th>Usage</th><th>Receipt</th><th>Status</th><th>AI</th><th>Action</th>
        </tr></thead>
        <tbody id="ledgerBody"></tbody>
      </table>
      <div class="pagination">
        <button id="pPrev">Prev</button>
        <span id="pInfo">1 / 1</span>
        <button id="pNext">Next</button>
      </div>
    </div>
  </div>

</div>

<div class="overlay-bg" id="overlay"></div>
<div class="detail-panel" id="panel">
  <div class="detail-panel-header">
    <span style="font-weight:700;font-size:.95rem">Receipt Detail</span>
    <button class="btn btn-ghost btn-sm" id="panelClose">× Close</button>
  </div>
  <div class="detail-section">
    <div class="detail-section-title">Receipt Image <span style="font-size:.7rem;color:var(--text-muted);font-weight:400">(click to enlarge)</span></div>
    <div class="receipt-image-wrap" id="dImageWrap" style="cursor:zoom-in" onclick="openImgLb()">
      <img class="receipt-image-full" id="dImage" alt="receipt">
      <svg class="receipt-bbox-overlay" id="dBboxOverlay"></svg>
    </div>
  </div>
  <div class="detail-section">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:6px">
      <div class="detail-section-title" style="margin-bottom:0">OCR Result</div>
      <button id="dEditBtn" class="btn btn-ghost btn-sm" onclick="togglePanelEdit()" style="font-size:.74rem;padding:3px 10px">✏️ Edit</button>
    </div>
    <div class="detail-row"><span class="key">Date</span>
      <span class="val" id="dDate">–</span>
      <input id="eDate" type="date" style="display:none;width:130px;background:var(--surface);border:1px solid var(--border);border-radius:4px;color:var(--text);font-size:.82rem;padding:2px 6px">
    </div>
    <div class="detail-row"><span class="key">Merchant</span>
      <span class="val" id="dMerchant">–</span>
      <input id="eMerchant" type="text" style="display:none;width:100%;background:var(--surface);border:1px solid var(--border);border-radius:4px;color:var(--text);font-size:.82rem;padding:2px 6px">
    </div>
    <div class="detail-row"><span class="key">Printed $</span>
      <span class="val" id="dPrinted">–</span>
      <input id="ePrinted" type="number" step="0.01" style="display:none;width:100px;background:var(--surface);border:1px solid var(--border);border-radius:4px;color:var(--text);font-size:.82rem;padding:2px 6px">
    </div>
    <div class="detail-row"><span class="key">Handwritten $</span>
      <span class="val" id="dHand">–</span>
      <input id="eHand" type="number" step="0.01" style="display:none;width:100px;background:var(--surface);border:1px solid var(--border);border-radius:4px;color:var(--text);font-size:.82rem;padding:2px 6px">
    </div>
    <div class="detail-row"><span class="key">Amount (final)</span><span class="val" id="dAmount">–</span></div>
    <div class="detail-row"><span class="key">Card Type</span>
      <span class="val" id="dCard">–</span>
      <select id="eCard" style="display:none;width:120px;background:var(--surface);border:1px solid var(--border);border-radius:4px;color:var(--text);font-size:.82rem;padding:2px 6px">
        <option value="">–</option><option value="amex">AMEX</option><option value="visa">Visa</option><option value="other">Other</option>
      </select>
    </div>
    <div class="detail-row"><span class="key">Usage</span>
      <span class="val" id="dUsage">Regular</span>
      <input id="eUsage" type="text" placeholder="Regular" style="display:none;width:130px;background:var(--surface);border:1px solid var(--border);border-radius:4px;color:var(--text);font-size:.82rem;padding:2px 6px">
    </div>
    <div class="detail-row"><span class="key">AI Model</span><span class="val" id="dModel">–</span></div>
    <div id="dSaveRow" style="display:none;margin-top:8px;display:none">
      <button class="btn btn-primary btn-sm" style="width:100%;font-size:.82rem" onclick="savePanelEdit()">💾 Save Changes</button>
    </div>
  </div>
  <div class="detail-section" id="dMatchSection">
    <div class="detail-section-title">Matched CSV Transaction</div>
    <div class="detail-row"><span class="key">Date</span><span class="val" id="dmDate">–</span></div>
    <div class="detail-row"><span class="key">Amount</span><span class="val" id="dmAmount">–</span></div>
    <div class="detail-row"><span class="key">Vendor</span><span class="val" id="dmVendor">–</span></div>
  </div>
  <div class="detail-section">
    <div class="detail-section-title">Status</div>
    <div id="dStatus"></div>
  </div>
  <div class="detail-actions">
    <button class="btn btn-ghost btn-sm" data-set="matched" style="color:#22c55e">✅ Mark Matched</button>
    <button class="btn btn-ghost btn-sm" data-set="unmatched" style="color:#ef4444">❌ Mark Unmatched</button>
    <button class="btn btn-ghost btn-sm" data-set="pending_match" style="color:#f59e0b">⏳ Mark Pending</button>
    <button class="btn btn-ghost btn-sm" id="dCompleteBtn" onclick="togglePanelComplete()" style="color:#818cf8;margin-top:6px;width:100%">✓ Mark Complete</button>
    <button class="btn btn-ghost btn-sm" id="reOcrBtn" onclick="reOCR()" style="color:#818cf8;margin-top:6px;width:100%">🔄 Re-OCR</button>
    <button class="btn btn-ghost btn-sm" id="manualAddBtn" onclick="openManualAdd()" style="color:#34d399;margin-top:2px;width:100%">➕ 이 이미지에 영수증 수동 추가</button>
  </div>
</div>

<!-- Manual Receipt Add Modal -->
<div class="overlay-bg" id="manualOverlay"></div>
<div class="del-modal" id="manualModal" style="width:420px">
  <div class="del-title">➕ 영수증 수동 추가</div>
  <div style="font-size:.82rem;color:var(--text-muted);margin-bottom:14px">OCR에서 누락된 영수증을 이 이미지에 직접 추가합니다.</div>
  <div style="display:flex;flex-direction:column;gap:10px;margin-bottom:16px">
    <div>
      <label style="font-size:.76rem;color:var(--text-muted);display:block;margin-bottom:4px">날짜</label>
      <input id="mDate" type="date" style="width:100%;background:var(--surface);border:1px solid var(--border);border-radius:6px;color:var(--text);font-size:.85rem;padding:6px 10px;outline:none;box-sizing:border-box">
    </div>
    <div>
      <label style="font-size:.76rem;color:var(--text-muted);display:block;margin-bottom:4px">가맹점</label>
      <input id="mMerchant" type="text" placeholder="가맹점명" style="width:100%;background:var(--surface);border:1px solid var(--border);border-radius:6px;color:var(--text);font-size:.85rem;padding:6px 10px;outline:none;box-sizing:border-box">
    </div>
    <div style="display:flex;gap:10px">
      <div style="flex:1">
        <label style="font-size:.76rem;color:var(--text-muted);display:block;margin-bottom:4px">인쇄 금액 ($)</label>
        <input id="mPrinted" type="number" step="0.01" min="0" placeholder="0.00" style="width:100%;background:var(--surface);border:1px solid var(--border);border-radius:6px;color:var(--text);font-size:.85rem;padding:6px 10px;outline:none;box-sizing:border-box">
      </div>
      <div style="flex:1">
        <label style="font-size:.76rem;color:var(--text-muted);display:block;margin-bottom:4px">수기 금액 ✍️</label>
        <input id="mHandw" type="number" step="0.01" min="0" placeholder="선택" style="width:100%;background:var(--surface);border:1px solid var(--border);border-radius:6px;color:var(--text);font-size:.85rem;padding:6px 10px;outline:none;box-sizing:border-box">
      </div>
    </div>
  </div>
  <div class="del-actions">
    <button class="btn btn-ghost btn-sm" onclick="closeManualAdd()">취소</button>
    <button class="btn btn-primary btn-sm" id="manualAddConfirm" onclick="submitManualAdd()">추가</button>
  </div>
</div>

<div class="overlay-bg" id="delOverlay"></div>
<div class="del-modal" id="delModal">
  <div class="del-title">🗑 영수증 삭제</div>
  <div class="del-body" id="delBody">체크된 영수증을 Ledger에서 삭제할까요?</div>
  <label class="del-check"><input type="checkbox" id="delDrive"> Drive 원본도 함께 휴지통으로 이동</label>
  <div class="del-actions">
    <button class="btn btn-ghost btn-sm" id="delCancel">취소</button>
    <button class="btn btn-sm" id="delConfirm"
      style="background:rgba(239,68,68,.15);color:#ef4444;border:1px solid rgba(239,68,68,.3)">삭제</button>
  </div>
</div>

<script>
let CUR_PAGE = 1, CUR_ID = null, CUR_FILE_ID = null, ENTRIES = [], VIEW_MODE = 'all';
const $ = id => document.getElementById(id);
const STATUS_LABEL = {matched:'✅ Matched', unmatched:'❌ Unmatched', pending_match:'⏳ Pending Match'};

function fmtAmt(a){ return (a===null||a===undefined) ? '–' : '$' + Number(a).toFixed(2); }

function thumb(e){
  if(!e.file_id) return '<div class="receipt-thumb-placeholder">🧾</div>';
  const proxy = '/cardconv/receipts/image/' + e.file_id;
  const tn = 'https://drive.google.com/thumbnail?id=' + e.file_id + '&sz=w80';
  return '<img class="receipt-thumb" src="' + tn + '" loading="lazy" ' +
         'onerror="this.onerror=null;this.src=\'' + proxy + '\'">';
}

function aiBadge(m){
  if(m==='Gemini') return '<span class="ai-badge gemini">Gemini</span>';
  if(m==='Claude') return '<span class="ai-badge claude">Claude</span>';
  if(m==='Manual') return '<span class="ai-badge" style="color:#34d399;background:rgba(52,211,153,.12)">Manual</span>';
  return '<span style="color:var(--text-muted);font-size:.72rem">–</span>';
}

const CARD_LABEL = {amex:'AMEX', visa:'Visa', other:'Other'};
function cardBadge(b){
  if(!b || !CARD_LABEL[b]) return '<span style="color:var(--text-muted);font-size:.72rem">–</span>';
  return '<span class="card-badge card-' + b + '">' + CARD_LABEL[b] + '</span>';
}

function matchInfo(e){
  const mt = e.matched_transaction;
  if(!mt) return '';
  const parts = [];
  if(mt.vendor) parts.push(mt.vendor);
  if(mt.amount!==null && mt.amount!==undefined) parts.push(fmtAmt(mt.amount));
  if(mt.date) parts.push(mt.date);
  if(!parts.length) return '';
  return '<div style="font-size:.7rem;color:var(--text-muted);margin-top:3px">↳ ' + parts.join(' · ') + '</div>';
}

function fmtAgo(iso){
  if(!iso) return '';
  const t = new Date(iso); if(isNaN(t)) return '';
  const sec = Math.floor((Date.now() - t) / 1000);
  let rel;
  if(sec < 60) rel = 'just now';
  else if(sec < 3600) rel = Math.floor(sec/60) + ' min ago';
  else if(sec < 86400) rel = Math.floor(sec/3600) + ' hr ago';
  else rel = Math.floor(sec/86400) + ' days ago';
  const pad = n => String(n).padStart(2,'0');
  const stamp = t.getFullYear()+'-'+pad(t.getMonth()+1)+'-'+pad(t.getDate())+' '+pad(t.getHours())+':'+pad(t.getMinutes());
  return 'Last synced: ' + stamp + ' (' + rel + ')';
}

function renderLastSynced(iso){
  const el = $('lastSynced');
  if(!el) return;
  const ts = iso || el.dataset.ts;
  if(ts){ el.dataset.ts = ts; }
  el.textContent = fmtAgo(el.dataset.ts);
}

// Build one ledger <tr>. opts.groupHead adds a clickable '+N' badge that toggles
// the group's child rows; opts.groupChild marks a collapsed duplicate row.
function rowHtml(e, i, opts){
  opts = opts || {};
  const h = e.ocr_handwritten_amount;
  const handCell = (h===null||h===undefined)
    ? '<td style="color:var(--text-muted)">–</td>'
    : '<td style="color:#f59e0b;font-weight:600">' + fmtAmt(h) + ' ✍️</td>';
  const actionCell = (e.match_status==='matched')
    ? '<td><button class="btn btn-ghost btn-sm act-undo" data-id="' + e.id +
      '" style="color:#f59e0b;padding:2px 8px" title="Undo match — reset to pending">↩ Undo</button></td>'
    : '<td><button class="btn btn-ghost btn-sm act-rematch" data-id="' + e.id +
      '" style="color:#818cf8;padding:2px 8px" title="Re-try CSV matching for this receipt">🔗 Rematch</button></td>';
  // Duplicate group: non-keeper rows are pre-checked for quick cleanup.
  const preCheck = (e.dup && !e.dup_keep) ? ' checked' : '';
  const checkCell = '<td><input type="checkbox" class="row-check sel" data-id="' +
    e.id + '"' + preCheck + '></td>';
  let dupTag = e.dup
    ? (e.dup_keep ? '<span class="keep-tag">KEEP</span>'
                  : '<span class="dup-tag">🔁 Duplicate</span>')
    : '';
  if(opts.groupHead){
    dupTag = '<span class="grp-toggle" data-gid="' + opts.groupHead + '">+' +
      opts.extra + ' duplicate' + (opts.extra>1?'s':'') + '</span>';
  }
  let cls = e.dup ? 'dup-row' : '';
  if(opts.groupChild) cls += ' dup-child gc-' + opts.groupChild;
  if(e.completed) cls += ' completed-row';
  const compTag = e.completed ? '<span class="comp-tag">✓ Done</span>' : '';
  return '<tr data-i="' + i + '"' + (cls?(' class="'+cls.trim()+'"'):'') + '>' +
    checkCell +
    '<td>' + (e.ocr_date||'–') + dupTag + compTag + '</td>' +
    '<td style="color:var(--text-muted)">' + fmtAmt(e.ocr_printed_amount) + '</td>' +
    handCell +
    '<td style="font-weight:700">' + fmtAmt(e.ocr_amount) + '</td>' +
    '<td>' + (e.ocr_merchant||'–') + '</td>' +
    '<td>' + cardBadge(e.card_brand) + '</td>' +
    '<td style="color:var(--text-muted)">' + (e.usage||'Regular') + '</td>' +
    '<td>' + thumb(e) + '</td>' +
    '<td><span class="status-badge status-' + (e.match_status||'unmatched') + '">' +
      (STATUS_LABEL[e.match_status]||e.match_status||'–') + '</span>' + matchInfo(e) + '</td>' +
    '<td>' + aiBadge(e.ocr_model) + '</td>' +
    actionCell +
  '</tr>';
}

function renderBody(entries){
  if(VIEW_MODE === 'all') return entries.map((e,i) => rowHtml(e,i)).join('');
  // Group mode: collapse each dup_group_id into a head row + hidden child rows.
  const groups = {}, order = [];
  entries.forEach((e,i) => {
    const gid = e.dup_group_id;
    if(gid){
      if(!groups[gid]){ groups[gid] = []; order.push({type:'group', gid}); }
      groups[gid].push(i);
    } else { order.push({type:'single', i}); }
  });
  // Head index of a group = dup_keep (matched-preferred by backend) entry.
  const headIdx = gid => groups[gid].slice().sort((a, b) =>
    (entries[a].dup_keep ? 0 : 1) - (entries[b].dup_keep ? 0 : 1))[0];
  // Sort all rows (group + single) by head's ocr_date, newest first.
  // None/'unknown' dates sink to the bottom.
  const dateVal = d => {
    if(!d || d === 'unknown') return null;
    const t = Date.parse(d);
    return isNaN(t) ? null : t;
  };
  const headDate = o => dateVal(o.type === 'single'
    ? entries[o.i].ocr_date
    : entries[headIdx(o.gid)].ocr_date);
  order.sort((a, b) => {
    const da = headDate(a), db = headDate(b);
    if(da === null && db === null) return 0;
    if(da === null) return 1;
    if(db === null) return -1;
    return db - da;
  });
  let html = '';
  order.forEach(o => {
    if(o.type==='single'){ html += rowHtml(entries[o.i], o.i); return; }
    // Reorder: dup_keep (matched-preferred by backend) entry becomes head.
    const idxs = groups[o.gid].slice().sort((a, b) => {
      const ka = entries[a].dup_keep ? 0 : 1;
      const kb = entries[b].dup_keep ? 0 : 1;
      return ka - kb;
    });
    const head = idxs[0];
    html += rowHtml(entries[head], head, {groupHead:o.gid, extra:idxs.length-1});
    idxs.slice(1).forEach(ci => { html += rowHtml(entries[ci], ci, {groupChild:o.gid}); });
  });
  return html;
}

function rerender(){
  const body = $('ledgerBody');
  if(!ENTRIES.length){
    body.innerHTML = '<tr><td colspan="12" style="text-align:center;color:var(--text-muted);padding:30px">No receipts</td></tr>';
  } else {
    body.innerHTML = renderBody(ENTRIES);
    body.querySelectorAll('tr[data-i]').forEach(tr =>
      tr.addEventListener('click', () => openPanel(ENTRIES[+tr.dataset.i])));
    body.querySelectorAll('.act-undo').forEach(b =>
      b.addEventListener('click', ev => { ev.stopPropagation(); unmatchRow(b.dataset.id); }));
    body.querySelectorAll('.act-rematch').forEach(b =>
      b.addEventListener('click', ev => { ev.stopPropagation(); quickRematch(b.dataset.id, b); }));
    body.querySelectorAll('.sel').forEach(c =>
      c.addEventListener('click', ev => ev.stopPropagation()));
    body.querySelectorAll('.sel').forEach(c =>
      c.addEventListener('change', updateDeleteBtn));
    body.querySelectorAll('.grp-toggle').forEach(g =>
      g.addEventListener('click', ev => {
        ev.stopPropagation();
        body.querySelectorAll('.gc-' + g.dataset.gid).forEach(r => r.classList.toggle('show'));
      }));
  }
  $('checkAll').checked = false;
  updateDeleteBtn();
}

// Build the shared filter query string from the current control values.
function filterParams(){
  const p = new URLSearchParams();
  if($('fFrom').value) p.set('from', $('fFrom').value);
  if($('fTo').value)   p.set('to', $('fTo').value);
  p.set('status', $('fStatus').value);
  p.set('card_brand', $('fCard').value);
  p.set('usage', $('fUsage').value);
  p.set('completed', $('fCompleted').value);
  return p;
}

// Rebuild the Usage dropdown from the distinct tags the API reports, keeping
// the current selection if it still exists.
function syncUsageOptions(usages){
  const sel = $('fUsage'), cur = sel.value;
  const opts = ['<option value="all">All</option>']
    .concat((usages||[]).map(u => '<option value="' + u.replace(/"/g,'&quot;') + '">' + u + '</option>'));
  sel.innerHTML = opts.join('');
  sel.value = (usages||[]).includes(cur) || cur==='all' ? cur : 'all';
}

async function load(){
  const p = filterParams();
  p.set('page', CUR_PAGE);
  const r = await fetch('/cardconv/ledger/api?' + p.toString());
  const d = await r.json();
  $('statTotal').textContent = d.total;
  $('statMatched').textContent = d.matched;
  $('statUnmatched').textContent = d.unmatched;
  $('statCompleted').textContent = (d.completed!=null ? d.completed : '–');
  syncUsageOptions(d.usages);
  ENTRIES = d.entries;
  rerender();
  renderLastSynced(d.last_synced);
  $('pInfo').textContent = d.page + ' / ' + d.pages;
  $('pPrev').disabled = d.page <= 1;
  $('pNext').disabled = d.page >= d.pages;
  CUR_PAGE = d.page;
  window._pages = d.pages;
}

function escSvg(s){
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

// Draw bbox rectangles for every receipt sharing this entry's source image.
// ocr_bbox is [ymin,xmin,ymax,xmax] in a 0-1000 normalized coord system.
function drawBoxes(cur){
  const svg = $('dBboxOverlay'), img = $('dImage');
  if(!svg || !img) return;
  const sibs = (typeof ENTRIES!=='undefined'?ENTRIES:[]).filter(
    x => x.file_id && cur.file_id && x.file_id===cur.file_id && Array.isArray(x.ocr_bbox));
  const render = () => paintBoxes(svg, img, sibs, cur);
  if(img.complete && img.naturalWidth) render();
  // First paint may run before the image has dimensions; redraw on load.
  img.onload = render;
}

function paintBoxes(svg, img, sibs, cur){
  if(!sibs.length || !img.naturalWidth){ svg.innerHTML=''; svg.style.display='none'; return; }
  // object-fit:contain → find the rendered image rect inside the <img> box.
  const cW=img.clientWidth, cH=img.clientHeight, nW=img.naturalWidth, nH=img.naturalHeight;
  const scale=Math.min(cW/nW, cH/nH);
  const dW=nW*scale, dH=nH*scale;
  svg.style.left=(img.offsetLeft+(cW-dW)/2)+'px';
  svg.style.top =(img.offsetTop +(cH-dH)/2)+'px';
  svg.style.width=dW+'px'; svg.style.height=dH+'px';
  svg.setAttribute('viewBox','0 0 '+dW+' '+dH);
  svg.style.display='block';
  svg.innerHTML = sibs.map(function(x,i){
    const b=x.ocr_bbox;                       // [ymin,xmin,ymax,xmax] 0-1000
    const x0=b[1]/1000*dW, y0=b[0]/1000*dH, x1=b[3]/1000*dW, y1=b[2]/1000*dH;
    const isCur=(x.id===cur.id);
    const col=isCur?'#38bdf8':'#64748b';      // selected = bright, others = muted
    const sw=isCur?2.5:1.5;
    const n=(x.sub_index!=null?x.sub_index:i)+1;
    const label=escSvg((x.ocr_merchant ? (n+'. '+x.ocr_merchant) : (''+n)).slice(0,22));
    const ty=Math.max(y0+13, 13);
    return '<rect x="'+x0.toFixed(1)+'" y="'+y0.toFixed(1)+'" width="'+(x1-x0).toFixed(1)+
      '" height="'+(y1-y0).toFixed(1)+'" rx="4" fill="'+col+'" fill-opacity="'+(isCur?0.12:0.05)+
      '" stroke="'+col+'" stroke-width="'+sw+'"/>'+
      '<text x="'+(x0+4).toFixed(1)+'" y="'+ty.toFixed(1)+'" fill="'+col+'" font-size="11" '+
      'font-weight="700" style="paint-order:stroke;stroke:rgba(2,6,23,.75);stroke-width:3px">'+
      label+'</text>';
  }).join('');
}
function openPanel(e){
  CUR_ID = e.id;
  CUR_FILE_ID = e.file_id || null;
  $('dDate').textContent = e.ocr_date || '–';
  $('dAmount').textContent = fmtAmt(e.ocr_amount);
  $('dPrinted').textContent = fmtAmt(e.ocr_printed_amount);
  const hand = (e.ocr_handwritten_amount===null||e.ocr_handwritten_amount===undefined);
  $('dHand').textContent = hand ? '–' : (fmtAmt(e.ocr_handwritten_amount) + ' ✍️');
  $('dHand').style.color = hand ? '' : '#f59e0b';
  $('dMerchant').textContent = e.ocr_merchant || '–';
  $('dModel').textContent = e.ocr_model || '–';
  $('dCard').innerHTML = cardBadge(e.card_brand);
  $('dUsage').textContent = e.usage || 'Regular';
  // Pre-fill edit inputs
  $('eDate').value = e.ocr_date || '';
  $('eMerchant').value = e.ocr_merchant || '';
  $('ePrinted').value = (e.ocr_printed_amount != null) ? e.ocr_printed_amount : (e.ocr_amount || '');
  $('eHand').value = (e.ocr_handwritten_amount != null) ? e.ocr_handwritten_amount : '';
  $('eCard').value = e.card_brand || '';
  $('eUsage').value = (e.usage && e.usage !== 'Regular') ? e.usage : '';
  // Complete toggle button reflects current state.
  var cb = $('dCompleteBtn');
  if(cb){ cb.textContent = e.completed ? '↩ Un-complete' : '✓ Mark Complete'; }
  // Reset edit mode
  exitPanelEdit();

  const img = $('dImage');
  if(e.file_id){
    let url = '/cardconv/receipts/image/' + e.file_id;
    if(Array.isArray(e.ocr_bbox) && e.ocr_bbox.length===4)
      url += '?bbox=' + e.ocr_bbox.join(',');
    img.src = url;
    img.style.display = 'block';
  } else { img.style.display='none'; }
  // SVG overlay not needed — multi-receipt images are cropped per entry
  const svg = $('dBboxOverlay');
  if(svg){ svg.innerHTML=''; svg.style.display='none'; }
  const mt = e.matched_transaction;
  $('dMatchSection').style.display = mt ? 'block' : 'none';
  if(mt){
    $('dmDate').textContent = mt.date || '–';
    $('dmAmount').textContent = fmtAmt(mt.amount);
    $('dmVendor').textContent = mt.vendor || '–';
  }
  $('dStatus').innerHTML = '<span class="status-badge status-' + (e.match_status||'unmatched') + '">' +
    (STATUS_LABEL[e.match_status]||e.match_status||'–') + '</span>';
  $('overlay').classList.add('open');
  $('panel').classList.add('open');
}

function closePanel(){
  $('overlay').classList.remove('open');
  $('panel').classList.remove('open');
  CUR_ID = null;
  CUR_FILE_ID = null;
  exitPanelEdit();
}

const PANEL_EDIT_FIELDS = ['dDate','dMerchant','dPrinted','dHand','dCard','dUsage'];
function exitPanelEdit(){
  PANEL_EDIT_FIELDS.forEach(function(id){
    var s = $(id); var e = $(id.replace('d','e')); if(!s||!e) return;
    s.style.display = ''; e.style.display = 'none';
  });
  var sr = $('dSaveRow'); if(sr) sr.style.display = 'none';
  var eb = $('dEditBtn'); if(eb) eb.textContent = '✏️ Edit';
}

function togglePanelEdit(){
  var editing = $('eDate').style.display !== 'none';
  if(editing){ exitPanelEdit(); return; }
  PANEL_EDIT_FIELDS.forEach(function(id){
    var s = $(id); var e = $(id.replace('d','e')); if(!s||!e) return;
    s.style.display = 'none'; e.style.display = '';
  });
  var sr = $('dSaveRow'); if(sr) sr.style.display = 'block';
  var eb = $('dEditBtn'); if(eb) eb.textContent = '✕ Cancel';
}

async function savePanelEdit(){
  if(!CUR_ID) return;
  var body = new URLSearchParams({
    ocr_date:                 $('eDate').value,
    ocr_merchant:             $('eMerchant').value,
    ocr_printed_amount:       $('ePrinted').value,
    ocr_handwritten_amount:   $('eHand').value,
    card_brand:               $('eCard').value,
    usage:                    $('eUsage').value,
  });
  var r = await fetch('/cardconv/ledger/' + CUR_ID + '/update', {method:'POST', body});
  var d = await r.json();
  if(!d.ok){ alert('Save failed: '+(d.error||r.status)); return; }
  exitPanelEdit();
  load();
  closePanel();
}

async function togglePanelComplete(){
  if(!CUR_ID) return;
  var e = ENTRIES.find(x => x.id === CUR_ID) || {};
  var undo = !!e.completed;
  if(!undo && !confirm('이 영수증을 완료 처리할까요?\n(기본 목록·Sync·Mapping에서 제외되고 Drive Completed 폴더로 이동됩니다)')) return;
  var r = await fetch('/cardconv/ledger/complete', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({ids:[CUR_ID], undo: undo})
  });
  var d = await r.json().catch(() => ({}));
  if(!d.ok){ alert('실패: '+(d && d.error||r.status)); return; }
  closePanel();
  load();
}

async function setStatus(status){
  if(!CUR_ID) return;
  await fetch('/cardconv/ledger/' + CUR_ID + '/status', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({status: status})
  });
  closePanel();
  load();
}

async function reOCR(){
  if(!CUR_ID) return;
  const btn = $('reOcrBtn');
  btn.disabled = true;
  btn.textContent = 'Processing...';
  try {
    const r = await fetch('/cardconv/ledger/' + CUR_ID + '/reocr', {method:'POST'});
    const d = await r.json();
    if(!r.ok || d.error){
      alert('Re-OCR failed: ' + (d.error || r.status));
      return;
    }
    // Update ENTRIES with the returned updated entries
    if(Array.isArray(d.updated)){
      d.updated.forEach(function(u){
        const idx = ENTRIES.findIndex(function(e){ return e.id === u.id; });
        if(idx >= 0) ENTRIES[idx] = u; else ENTRIES.push(u);
      });
      // Remove entries that were replaced (different sub_index count)
      const updIds = new Set(d.updated.map(function(u){ return u.id; }));
      const updFid = d.updated[0] && d.updated[0].file_id;
      if(updFid){
        ENTRIES = ENTRIES.filter(function(e){ return e.file_id !== updFid || updIds.has(e.id); });
      }
    }
    closePanel();
    load();
  } finally {
    btn.disabled = false;
    btn.textContent = '🔄 Re-OCR';
  }
}

// Quick Re-OCR from table row button (no panel needed).
async function quickReOCR(id, btn){
  if(!id) return;
  btn.disabled = true; btn.textContent = '⏳';
  try {
    const r = await fetch('/cardconv/ledger/' + id + '/reocr', {method:'POST'});
    const d = await r.json();
    if(!r.ok || d.error){ btn.textContent = '❌'; setTimeout(()=>{ btn.disabled=false; btn.textContent='🔄 Re-run'; }, 2000); return; }
    if(Array.isArray(d.updated)){
      d.updated.forEach(u => { const idx = ENTRIES.findIndex(e => e.id===u.id); if(idx>=0) ENTRIES[idx]=u; else ENTRIES.push(u); });
    }
    load();
  } catch(e) { btn.disabled=false; btn.textContent='🔄 Re-run'; }
}

// Re-try CSV matching for a pending/unmatched row.
async function quickRematch(id, btn){
  if(!id) return;
  btn.disabled = true; btn.textContent = '⏳';
  try {
    const r = await fetch('/cardconv/ledger/' + id + '/rematch', {method:'POST'});
    const d = await r.json();
    if(!r.ok || d.error){ btn.textContent = '❌'; setTimeout(()=>{ btn.disabled=false; btn.textContent='🔗 Rematch'; }, 2000); return; }
    if(d.matched){
      btn.textContent = '✅';
      setTimeout(()=>{ btn.disabled=false; btn.textContent='🔗 Rematch'; }, 1500);
    } else {
      btn.textContent = '❌ No match';
      setTimeout(()=>{ btn.disabled=false; btn.textContent='🔗 Rematch'; }, 2000);
    }
    load();
  } catch(e) { btn.disabled=false; btn.textContent='🔗 Rematch'; }
}

// Undo Match from the ledger table — reset row to pending_match.
async function unmatchRow(id){
  if(!id) return;
  await fetch('/cardconv/ledger/' + id + '/status', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({status: 'pending_match'})
  });
  load();
}

function selectedIds(){
  return [...document.querySelectorAll('.sel:checked')].map(c => c.dataset.id);
}

function updateDeleteBtn(){
  const ids = selectedIds();
  const n = ids.length;
  $('fDelete').textContent = '🗑 Delete Selected (' + n + ')';
  $('fDelete').disabled = n === 0;
  // Split selection into active vs already-completed to drive the two buttons.
  const sel = new Set(ids);
  let active = 0, done = 0;
  ENTRIES.forEach(e => { if(sel.has(e.id)){ e.completed ? done++ : active++; } });
  $('fComplete').textContent = '✓ Complete Selected (' + active + ')';
  $('fComplete').disabled = active === 0;
  $('fUncomplete').textContent = '↩ Un-complete (' + done + ')';
  $('fUncomplete').disabled = done === 0;
}

async function completeSelected(undo){
  const sel = new Set(selectedIds());
  // Only act on entries in the relevant state (active→complete, done→un-complete).
  const ids = ENTRIES.filter(e => sel.has(e.id) && (undo ? e.completed : !e.completed))
                     .map(e => e.id);
  if(!ids.length) return;
  const verb = undo ? '완료 해제' : '완료 처리';
  if(!confirm(ids.length + '건을 ' + verb + '할까요?' +
      (undo ? '' : '\n(완료 항목은 기본 목록·Sync·Mapping에서 제외되고 Drive Completed 폴더로 이동됩니다)'))) return;
  const r = await fetch('/cardconv/ledger/complete', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({ids: ids, undo: !!undo})
  });
  const d = await r.json().catch(() => ({}));
  if(d && d.moved!=null && !undo && d.moved < ids.length){
    // Some Drive moves may not apply (e.g. multi-receipt page partially completed).
  }
  load();
}

function deleteSelected(){
  const ids = selectedIds();
  if(!ids.length) return;
  $('delBody').textContent = '체크된 영수증 ' + ids.length + '건을 Ledger에서 삭제할까요?';
  $('delDrive').checked = false;
  $('delOverlay').classList.add('open');
  $('delModal').classList.add('open');
}

function closeDelModal(){
  $('delOverlay').classList.remove('open');
  $('delModal').classList.remove('open');
}

async function confirmDelete(){
  const ids = selectedIds();
  const alsoDrive = $('delDrive').checked;
  closeDelModal();
  if(!ids.length) return;
  await fetch('/cardconv/ledger/delete', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({ids: ids, also_drive: alsoDrive})
  });
  load();
}

function iso(d){ return d.toISOString().slice(0,10); }

function applyPreset(p){
  const now = new Date();
  let from = '', to = iso(now);
  if(p==='month')    from = iso(new Date(now.getFullYear(), now.getMonth(), 1));
  else if(p==='30d') from = iso(new Date(now.getTime() - 29*86400000));
  else if(p==='3m')  from = iso(new Date(now.getFullYear(), now.getMonth()-3, now.getDate()));
  else if(p==='ytd') from = iso(new Date(now.getFullYear(), 0, 1));
  else if(p==='all'){ from = ''; to = ''; }
  $('fFrom').value = from;
  $('fTo').value = to;
  CUR_PAGE = 1;
  load();
}

function setDefaultDates(){
  // Default: All Time (no date filter)
  $('fFrom').value = '';
  $('fTo').value = '';
}

document.querySelectorAll('.detail-actions button[data-set]').forEach(b =>
  b.addEventListener('click', () => setStatus(b.dataset.set)));
$('panelClose').addEventListener('click', closePanel);
$('overlay').addEventListener('click', closePanel);
document.addEventListener('keydown', e => { if(e.key==='Escape') closePanel(); });
$('fFrom').addEventListener('change', () => { CUR_PAGE=1; load(); });
$('fTo').addEventListener('change', () => { CUR_PAGE=1; load(); });
$('fStatus').addEventListener('change', () => { CUR_PAGE=1; load(); });
$('fCard').addEventListener('change', () => { CUR_PAGE=1; load(); });
$('fUsage').addEventListener('change', () => { CUR_PAGE=1; load(); });
$('fCompleted').addEventListener('change', () => { CUR_PAGE=1; load(); });
$('fReset').addEventListener('click', () => {
  $('fStatus').value='all'; $('fCard').value='all'; $('fUsage').value='all';
  $('fCompleted').value='hide'; setDefaultDates(); CUR_PAGE=1; load();
});
// Both downloads respect the currently applied filters.
$('fDownload').addEventListener('click', () => {
  window.location = '/cardconv/ledger/download.pdf?' + filterParams().toString();
});
$('fDownloadXlsx').addEventListener('click', () => {
  window.location = '/cardconv/ledger/download.xlsx?' + filterParams().toString();
});
$('pPrev').addEventListener('click', () => { if(CUR_PAGE>1){CUR_PAGE--; load();} });
$('pNext').addEventListener('click', () => { if(CUR_PAGE<window._pages){CUR_PAGE++; load();} });
$('fDelete').addEventListener('click', deleteSelected);
$('fComplete').addEventListener('click', () => completeSelected(false));
$('fUncomplete').addEventListener('click', () => completeSelected(true));
$('checkAll').addEventListener('change', () => {
  document.querySelectorAll('.sel').forEach(c => { c.checked = $('checkAll').checked; });
  updateDeleteBtn();
});
document.querySelectorAll('.preset-btn:not(#viewToggle)').forEach(b =>
  b.addEventListener('click', () => applyPreset(b.dataset.preset)));

// Group Duplicates toggle — re-renders the current page without refetching.
$('viewToggle').addEventListener('click', () => {
  VIEW_MODE = (VIEW_MODE === 'all') ? 'group' : 'all';
  $('viewToggle').textContent = (VIEW_MODE === 'all') ? '🔁 Group Duplicates' : '☰ Show All';
  $('viewToggle').classList.toggle('active', VIEW_MODE === 'group');
  rerender();
});

// Manual receipt add modal.
let _MANUAL_FILE_ID = null, _MANUAL_META = {};
function openManualAdd(){
  if(!CUR_FILE_ID) return;
  _MANUAL_FILE_ID = CUR_FILE_ID;
  const cur = ENTRIES.find(e => e.id === CUR_ID) || {};
  _MANUAL_META = { filename: cur.filename||'', drive_url: cur.drive_url||'', mime_type: cur.mime_type||'image/jpeg' };
  $('mDate').value = ''; $('mMerchant').value = ''; $('mPrinted').value = ''; $('mHandw').value = '';
  $('manualOverlay').classList.add('open');
  $('manualModal').classList.add('open');
}
function closeManualAdd(){
  $('manualOverlay').classList.remove('open');
  $('manualModal').classList.remove('open');
}
async function submitManualAdd(){
  if(!_MANUAL_FILE_ID) return;
  const btn = $('manualAddConfirm');
  btn.disabled = true; btn.textContent = '추가 중...';
  const body = new URLSearchParams({
    file_id:               _MANUAL_FILE_ID,
    filename:              _MANUAL_META.filename,
    drive_url:             _MANUAL_META.drive_url,
    mime_type:             _MANUAL_META.mime_type,
    ocr_date:              $('mDate').value,
    ocr_merchant:          $('mMerchant').value,
    ocr_printed_amount:    $('mPrinted').value,
    ocr_handwritten_amount: $('mHandw').value,
  });
  try {
    const r = await fetch('/cardconv/receipts/manual-add', {method:'POST', body});
    const d = await r.json();
    if(!d.ok){ alert('추가 실패: ' + (d.error||r.status)); return; }
    closeManualAdd();
    closePanel();
    load();
  } catch(e){ alert('오류: '+e); }
  finally { btn.disabled=false; btn.textContent='추가'; }
}
$('manualOverlay').addEventListener('click', closeManualAdd);

// Delete confirmation modal (with optional Drive trashing).
$('delCancel').addEventListener('click', closeDelModal);
$('delOverlay').addEventListener('click', closeDelModal);
$('delConfirm').addEventListener('click', confirmDelete);

setDefaultDates();
load();
</script>
<script>__RCPTJS__</script>
<script>
(function(){
  const VAPID_PUB = '__VAPID_PUB__';
  if(!VAPID_PUB || !('serviceWorker' in navigator) || !('PushManager' in window)) return;
  function urlB64ToUint8Array(b64){
    const pad = '='.repeat((4 - b64.length%4)%4);
    const raw = atob((b64+pad).replace(/-/g,'+').replace(/_/g,'/'));
    return Uint8Array.from(raw, c => c.charCodeAt(0));
  }
  navigator.serviceWorker.ready.then(function(reg){
    reg.pushManager.getSubscription().then(function(existing){
      if(existing) return;  // already subscribed
      if(Notification.permission === 'denied') return;
      return reg.pushManager.subscribe({
        userVisibleOnly: true,
        applicationServerKey: urlB64ToUint8Array(VAPID_PUB)
      }).then(function(sub){
        return fetch('/cardconv/push/subscribe', {
          method:'POST',
          headers:{'Content-Type':'application/json'},
          body: JSON.stringify(sub.toJSON())
        });
      });
    });
  });
})();
</script>

<!-- Sync Loading Overlay -->
<div id="syncOverlay" style="display:none;position:fixed;inset:0;background:rgba(2,6,23,.82);z-index:500;flex-direction:column;align-items:center;justify-content:center;gap:20px">
  <style>
  @keyframes spin{to{transform:rotate(360deg)}}
  @keyframes pulse{0%,100%{opacity:.6}50%{opacity:1}}
  .sync-spinner{width:52px;height:52px;border:4px solid rgba(56,189,248,.2);border-top-color:#38bdf8;border-radius:50%;animation:spin .9s linear infinite}
  .sync-dots span{display:inline-block;width:6px;height:6px;border-radius:50%;background:#38bdf8;margin:0 3px;animation:pulse 1.2s ease-in-out infinite}
  .sync-dots span:nth-child(2){animation-delay:.2s}
  .sync-dots span:nth-child(3){animation-delay:.4s}
  </style>
  <div class="sync-spinner"></div>
  <div style="color:#e2e8f0;font-size:1rem;font-weight:600;letter-spacing:.02em">Syncing from Drive…</div>
  <div style="color:#94a3b8;font-size:.82rem">Powered by Google Gemini — OCR is running, this may take a moment ✨</div>
  <div class="sync-dots"><span></span><span></span><span></span></div>
</div>
<script>
function startDriveSync(btn) {
  btn.disabled = true;
  document.getElementById('syncOverlay').style.display = 'flex';
  fetch('/cardconv/drive/sync', {method: 'POST'})
    .then(function(r) { return r.json(); })
    .then(function(d) {
      if (d.error) { syncFail(d.error); return; }
      pollSync(d.job_id);
    })
    .catch(function(e) { syncFail(String(e)); });
}

function pollSync(jobId) {
  setTimeout(function() {
    fetch('/cardconv/drive/sync/status?job=' + jobId)
      .then(function(r) { return r.json(); })
      .then(function(d) {
        if (d.status === 'running') { pollSync(jobId); return; }
        document.getElementById('syncOverlay').style.display = 'none';
        if (d.status === 'error') { alert('Sync error: ' + (d.error || 'unknown')); load(); return; }
        // done
        if (d.staged > 0) {
          openOcrModal();
        } else {
          load();
        }
      })
      .catch(function() { pollSync(jobId); });  // retry on network hiccup
  }, 2500);
}

function syncFail(msg) {
  document.getElementById('syncOverlay').style.display = 'none';
  alert('Sync failed: ' + msg);
}
</script>

<!-- Image Lightbox -->
<div id="imgLb" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.92);z-index:300;align-items:center;justify-content:center;cursor:zoom-out" onclick="closeImgLb()">
  <img id="imgLbImg" style="max-width:95vw;max-height:95vh;object-fit:contain;border-radius:4px;box-shadow:0 0 40px rgba(0,0,0,.8)" alt="receipt">
  <button onclick="closeImgLb()" style="position:fixed;top:16px;right:20px;background:rgba(255,255,255,.15);border:none;color:#fff;font-size:1.6rem;line-height:1;width:36px;height:36px;border-radius:50%;cursor:pointer">&times;</button>
</div>
<script>
function openImgLb(fid){
  var fileId = fid || CUR_FILE_ID;
  if(!fileId) return;
  var url = '/cardconv/receipts/image/' + fileId;
  var lb = document.getElementById('imgLb');
  var img = document.getElementById('imgLbImg');
  img.src = url;
  lb.style.display = 'flex';
  document.addEventListener('keydown', _imgLbKey);
}
function closeImgLb(){
  var lb = document.getElementById('imgLb');
  lb.style.display = 'none';
  document.getElementById('imgLbImg').src = '';
  document.removeEventListener('keydown', _imgLbKey);
}
function _imgLbKey(e){ if(e.key==='Escape') closeImgLb(); }
</script>

<!-- OCR Review Modal -->
<div id="ocrReviewOverlay" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:200;align-items:flex-start;justify-content:center;overflow-y:auto;padding:30px 16px">
  <div style="background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-md);width:100%;max-width:860px;margin:auto;display:flex;flex-direction:column">
    <div style="display:flex;align-items:center;justify-content:space-between;padding:18px 20px;border-bottom:1px solid var(--border)">
      <div>
        <div style="font-size:1.1rem;font-weight:700">OCR Review</div>
        <div id="ocrReviewCount" style="font-size:.82rem;color:var(--text-muted);margin-top:2px"></div>
      </div>
      <div style="display:flex;gap:8px;align-items:center">
        <button onclick="ocrToggleAll(true)" class="btn btn-secondary" style="font-size:.78rem;padding:5px 12px">Check All</button>
        <button onclick="ocrToggleAll(false)" class="btn btn-secondary" style="font-size:.78rem;padding:5px 12px">Uncheck All</button>
        <button onclick="closeOcrModal()" style="background:none;border:none;color:var(--text-muted);font-size:1.4rem;cursor:pointer;line-height:1">&#x2715;</button>
      </div>
    </div>
    <div id="ocrReviewBody" style="padding:16px;display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:12px;max-height:70vh;overflow-y:auto">
      <div style="grid-column:1/-1;text-align:center;color:var(--text-muted);padding:40px">Loading...</div>
    </div>
    <div style="display:flex;justify-content:flex-end;gap:10px;padding:14px 20px;border-top:1px solid var(--border);flex-wrap:wrap">
      <button onclick="ocrDiscardAll()" class="btn btn-danger" style="font-size:.84rem">Discard All</button>
      <button onclick="ocrConfirmSelected()" class="btn btn-primary" style="font-size:.84rem">&#10003; Confirm Selected &rarr; Add to Ledger</button>
    </div>
  </div>
</div>

<script>
(function() {
  var overlay = document.getElementById('ocrReviewOverlay');
  var body    = document.getElementById('ocrReviewBody');
  var countEl = document.getElementById('ocrReviewCount');
  var _entries = [];

  function money(v) {
    return (v == null) ? '–' : ('$' + parseFloat(v).toFixed(2));
  }

  var INPUT_STYLE = 'width:100%;box-sizing:border-box;background:var(--surface);border:1px solid var(--border);border-radius:4px;color:var(--text);font-size:.8rem;padding:3px 6px;outline:none';
  var LABEL_STYLE = 'font-size:.72rem;color:var(--text-muted);display:block;margin-bottom:2px';

  function renderEntries(entries) {
    _entries = entries;
    countEl.textContent = entries.length + ' receipt(s) — verify and edit before adding to ledger';
    if (!entries.length) {
      body.innerHTML = '<div style="grid-column:1/-1;text-align:center;color:var(--text-muted);padding:40px">No pending entries.</div>';
      return;
    }
    body.innerHTML = entries.map(function(e) {
      var proxy = e.file_id ? '/cardconv/receipts/image/' + encodeURIComponent(e.file_id) : '';
      var fid = e.file_id || '';
      var imgHtml = proxy
        ? '<img src="' + proxy + '" style="width:100%;max-height:200px;object-fit:contain;display:block;background:#000;cursor:zoom-in" loading="lazy" onclick="event.stopPropagation();openImgLb(\'' + fid + '\')" title="Click to enlarge">'
        : '<div style="height:120px;display:flex;align-items:center;justify-content:center;color:var(--text-muted);font-size:.8rem">No image</div>';
      var ocrOk = e.ocr_status === 'done' && e.ocr_merchant;
      var badge = ocrOk
        ? '<span style="font-size:.62rem;font-weight:700;padding:2px 6px;border-radius:8px;background:rgba(34,197,94,.15);color:#22c55e">OCR OK</span>'
        : '<span style="font-size:.62rem;font-weight:700;padding:2px 6px;border-radius:8px;background:rgba(245,158,11,.15);color:#f59e0b">Partial</span>';
      var eid = e.id;
      var amtVal = (e.ocr_amount != null) ? e.ocr_amount : '';
      var hwVal  = (e.ocr_handwritten_amount != null) ? e.ocr_handwritten_amount : '';
      return '<div class="ocr-card" data-id="' + eid + '" style="border:1px solid var(--border);border-radius:var(--radius-md);overflow:hidden;background:var(--surface-2)">'
        + '<label style="display:flex;align-items:center;gap:8px;padding:8px 10px;border-bottom:1px solid var(--border);cursor:pointer;background:var(--surface)">'
        + '<input type="checkbox" class="ocr-cb" data-id="' + eid + '" ' + (ocrOk ? 'checked' : '') + ' style="width:15px;height:15px;accent-color:var(--accent);cursor:pointer">'
        + '<span style="font-size:.8rem;font-weight:600">Include</span>' + badge
        + '</label>'
        + '<div style="background:#000;display:flex;align-items:center;justify-content:center;min-height:120px">' + imgHtml + '</div>'
        + '<div style="padding:10px;display:flex;flex-direction:column;gap:7px">'
        + '<div style="color:var(--text-muted);font-size:.7rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">' + (e.filename || '') + '</div>'
        + '<div><label style="' + LABEL_STYLE + '">Date</label>'
        +   '<input class="ocr-field" data-field="ocr_date" data-id="' + eid + '" type="date" value="' + (e.ocr_date || '') + '" style="' + INPUT_STYLE + '"></div>'
        + '<div><label style="' + LABEL_STYLE + '">Merchant</label>'
        +   '<input class="ocr-field" data-field="ocr_merchant" data-id="' + eid + '" type="text" value="' + ((e.ocr_merchant || '')).replace(/"/g,'&quot;') + '" placeholder="–" style="' + INPUT_STYLE + '"></div>'
        + '<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px">'
        +   '<div><label style="' + LABEL_STYLE + '">Printed $</label>'
        +     '<input class="ocr-field" data-field="ocr_amount" data-id="' + eid + '" type="number" step="0.01" value="' + amtVal + '" placeholder="–" style="' + INPUT_STYLE + '"></div>'
        +   '<div><label style="' + LABEL_STYLE + '">Handwritten $</label>'
        +     '<input class="ocr-field" data-field="ocr_handwritten_amount" data-id="' + eid + '" type="number" step="0.01" value="' + hwVal + '" placeholder="–" style="' + INPUT_STYLE + '"></div>'
        + '</div>'
        + '</div></div>';
    }).join('');
  }

  window.openOcrModal = function() {
    overlay.style.display = 'flex';
    fetch('/cardconv/receipts/review/api')
      .then(function(r) { return r.json(); })
      .then(function(d) { renderEntries(d.entries || []); })
      .catch(function() { body.innerHTML = '<div style="grid-column:1/-1;text-align:center;color:var(--danger);padding:40px">Failed to load.</div>'; });
  };
  var openOcrModal = window.openOcrModal;

  window.closeOcrModal = function() { overlay.style.display = 'none'; };

  window.ocrToggleAll = function(on) {
    document.querySelectorAll('.ocr-cb').forEach(function(cb) { cb.checked = on; });
  };

  function clearOcrBadge() {
    var badge = document.querySelector('a[href="/cardconv/receipts/review"] .tab-badge');
    if (badge) badge.remove();
  }

  window.ocrDiscardAll = function() {
    if (!confirm('Discard all staged entries?')) return;
    fetch('/cardconv/receipts/review/discard', {method:'POST'})
      .then(function() { closeOcrModal(); clearOcrBadge(); });
  };

  window.ocrConfirmSelected = function() {
    var checkedIds = new Set(
      Array.from(document.querySelectorAll('.ocr-cb:checked')).map(function(cb) { return cb.dataset.id; })
    );
    // Collect per-card field values for checked cards only.
    var confirmed = Array.from(document.querySelectorAll('.ocr-card')).reduce(function(acc, card) {
      var id = card.dataset.id;
      if (!checkedIds.has(id)) return acc;
      var item = {id: id};
      card.querySelectorAll('.ocr-field').forEach(function(inp) {
        item[inp.dataset.field] = inp.value;
      });
      acc.push(item);
      return acc;
    }, []);
    fetch('/cardconv/receipts/review/confirm', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({confirmed: confirmed})
    }).then(function() { closeOcrModal(); clearOcrBadge(); load(); });
  };

  overlay.addEventListener('click', function(e) { if (e.target === overlay) closeOcrModal(); });

  // Auto-open if redirected with ?ocr_review=1
  if (new URLSearchParams(location.search).get('ocr_review') === '1') {
    history.replaceState({}, '', location.pathname);
    openOcrModal();
  }

})();
</script>
</body></html>'''
