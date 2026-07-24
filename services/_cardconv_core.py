import csv, io, json, os, re, base64, uuid, zipfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ── Web Push helpers ──────────────────────────────────────────────────────────
def _push_subs_file(username: str) -> Path:
    return DATA_DIR / f"push_subs_{username}.json"

def _load_push_subs(username: str) -> list:
    f = _push_subs_file(username)
    if not f.exists():
        return []
    try:
        return json.loads(f.read_text())
    except Exception:
        return []

def _save_push_subs(username: str, subs: list):
    _push_subs_file(username).write_text(json.dumps(subs, ensure_ascii=False))

def _send_push_notification(username: str, title: str, body: str, url: str = "/cardconv/ledger"):
    """Send Web Push to all stored subscriptions for the user. Silently skips on error."""
    subs = _load_push_subs(username)
    if not subs:
        return
    pem_path = DATA_DIR / "vapid_private.pem"
    if not pem_path.exists():
        return
    try:
        from pywebpush import webpush, WebPushException
        payload = json.dumps({"title": title, "body": body, "url": url})
        pem = pem_path.read_text()
        dead = []
        for sub in subs:
            try:
                webpush(
                    subscription_info=sub,
                    data=payload,
                    vapid_private_key=pem,
                    vapid_claims={"sub": "mailto:jongha.kang01@gmail.com"},
                    content_encoding="aes128gcm",
                )
            except WebPushException as e:
                if e.response and e.response.status_code in (404, 410):
                    dead.append(sub)
        if dead:
            _save_push_subs(username, [s for s in subs if s not in dead])
    except Exception:
        pass

from services._paths import DATA_ROOT
DATA_DIR   = Path(DATA_ROOT) / "cardconv"
KW_FILE    = DATA_DIR / "keywords.json"
HIST_FILE  = DATA_DIR / "history.json"
OUT_DIR    = DATA_DIR / "outputs"
TOKENS_DIR = DATA_DIR / "tokens"
CREDS_FILE = DATA_DIR / "google_credentials.json"
_SVC_DIR   = Path(__file__).parent
TEMPLATE   = _SVC_DIR / "cardconv_template.xlsx"  # bundled with service
TEMPLATE_FALLBACK = Path(os.path.expanduser(
    "~/Desktop/US업무/법카 정산/Automation/for upload.xlsx"
))

SCOPES = [
    'https://www.googleapis.com/auth/drive.file',
    'https://www.googleapis.com/auth/drive.readonly',
]
ADMIN  = "jongha.kang"

META = {
    "name": "Cheil AMEX Expense Assistant",
    "path": "/cardconv",
    "icon": "💳",
    "description": "Corporate card CSV → SAP upload xlsx",
}

DEFAULT_KW = [
    {"kw": "STARBUCKS",         "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "CAFE",              "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "COFFEE",            "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "BAKERY",            "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "SWEETGREEN",        "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "SRA CAFE",          "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "KOBRICK",           "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "DOORDASH",          "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "GRUBHUB",           "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "WHOLEFDS",          "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "TRADER JOE",        "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "RESTAURANT",        "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "GRILL",             "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "PIZZA",             "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "BURGER",            "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "SUSHI",             "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "PARIS BAGUETTE",    "gl": 53410177, "ser": "160", "purpose": "Coffee, Snack and meal"},
    {"kw": "AMAZON",            "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "HOME DEPOT",        "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "BEST BUY",          "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "TARGET",            "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "COSTCO",            "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "NESPRESSO",         "gl": 53210177, "ser": "021", "purpose": "Office supply purchases"},
    {"kw": "PG&E",              "gl": 53210177, "ser": "021", "purpose": "Office Utilities"},
    {"kw": "UBER",              "gl": 53270377, "ser": "306", "purpose": "Uber travel"},
    {"kw": "LYFT",              "gl": 53270377, "ser": "306", "purpose": "Lyft travel"},
    {"kw": "ENTERPRISE",        "gl": 53270377, "ser": "306", "purpose": "Car rental"},
    {"kw": "HERTZ",             "gl": 53270377, "ser": "306", "purpose": "Car rental"},
    {"kw": "AVIS",              "gl": 53270377, "ser": "306", "purpose": "Car rental"},
    {"kw": "CHEVRON",           "gl": 53270377, "ser": "306", "purpose": "Gas for car"},
    {"kw": "SHELL",             "gl": 53270377, "ser": "306", "purpose": "Gas for car"},
    {"kw": "ROBBIE",            "gl": 53270377, "ser": "306", "purpose": "Gas for car"},
    {"kw": "GARAGE",            "gl": 53270377, "ser": "306", "purpose": "Parking"},
    {"kw": "PARKING",           "gl": 53270377, "ser": "306", "purpose": "Parking"},
    {"kw": "FASTRAK",           "gl": 53270377, "ser": "306", "purpose": "Car Toll"},
    {"kw": "TOLL",              "gl": 53270377, "ser": "306", "purpose": "Rental Toll"},
    {"kw": "SPOTHERO",          "gl": 53270377, "ser": "306", "purpose": "Parking"},
    {"kw": "CAR WASH",          "gl": 53270377, "ser": "306", "purpose": "Car wash for Car"},
    {"kw": "XPRESS",            "gl": 53270377, "ser": "306", "purpose": "Car wash for Car"},
    {"kw": "HYATT",             "gl": 53270377, "ser": "306", "purpose": "Hotel accommodation"},
    {"kw": "MARRIOTT",          "gl": 53270377, "ser": "306", "purpose": "Hotel accommodation"},
    {"kw": "HILTON",            "gl": 53270377, "ser": "306", "purpose": "Hotel accommodation"},
    {"kw": "HOTEL",             "gl": 53270377, "ser": "306", "purpose": "Hotel accommodation"},
    {"kw": "UNITED AIRLINES",   "gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "KOREAN AIR",        "gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "SINGAPORE AIRLINES","gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "BRITISH AIRWAYS",   "gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "AMERICAN AIRLINES", "gl": 53270377, "ser": "306", "purpose": "Flight travel"},
    {"kw": "TOTAL WINE",        "gl": 53410103, "ser": "159", "purpose": "Purchase for client"},
    {"kw": "WINE",              "gl": 53410103, "ser": "159", "purpose": "Purchase for client"},
    {"kw": "OPENAI",            "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "ANTHROPIC",         "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "CLAUDE",            "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "CHATGPT",           "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "PERPLEXITY",        "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "LOVABLE",           "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "MANUS AI",          "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "GOOGLE *GOOGLE",    "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "GOOGLE ONE",        "gl": 53311577, "ser": "085", "purpose": "AI subscription"},
    {"kw": "RHINO NETWORK",     "gl": 53290177, "ser": "052", "purpose": "Office Network setup"},
    {"kw": "RITZ",              "gl": 53470177, "ser": "289", "purpose": "Coffee, Snack and meal"},
    {"kw": "JO MALONE",         "gl": 53470177, "ser": "289", "purpose": "Coffee, Snack and meal"},
]

FIXED = {
    "receipt_type": "D",
    "employee_id":  20170321,
    "payee":        "A0016672",
    "domestic":     "D",
    "currency":     "USD",
    "tax_code":     "VV",
    "cost_center":  "AG010238",
}


# ── Data helpers ─────────────────────────────────────────────────────────────

def _ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    TOKENS_DIR.mkdir(parents=True, exist_ok=True)


def _load_kw():
    _ensure_dirs()
    if KW_FILE.exists():
        try:
            return json.loads(KW_FILE.read_text())
        except Exception:
            pass
    _save_kw(DEFAULT_KW)
    return DEFAULT_KW


def _save_kw(kws):
    _ensure_dirs()
    KW_FILE.write_text(json.dumps(kws, ensure_ascii=False, indent=2))


def _load_hist():
    if HIST_FILE.exists():
        try:
            return json.loads(HIST_FILE.read_text())
        except Exception:
            pass
    return []


def _save_hist(hist):
    HIST_FILE.write_text(json.dumps(hist, ensure_ascii=False, indent=2))


def _add_hist(entry: dict):
    if "id" not in entry:
        entry["id"] = uuid.uuid4().hex[:10]
    # tag with the active card profile so multi-card users can tell rows apart
    u = entry.get("user")
    if u and "profile" not in entry:
        try:
            profs, pid = _load_profiles(u)
            if pid:
                entry["profile"] = next(
                    (p["name"] for p in profs if p["id"] == pid), pid)
        except Exception:
            pass
    hist = _load_hist()
    hist.insert(0, entry)
    _save_hist(hist[:100])


# ── Card profiles (강프로 2026-07-24) ────────────────────────────────────────
# One login can manage several corporate cards. A profile scopes the whole
# pipeline — statement pool, ledger, OCR staging, card names, Drive folder,
# review — by suffixing the per-user data filenames with `@<pid>`. The default
# profile (pid='') keeps the legacy filenames, so existing data needs no
# migration. Profiles are self-service: users create them from the tab bar.
import threading as _prof_threading
_PROFILE_CTX = _prof_threading.local()


def _profiles_file(username: str) -> Path:
    return DATA_DIR / f"profiles_{username}.json"


def _load_profiles(username: str):
    """Returns (profiles, active_pid); the default profile (pid='') is always
    first. Extra profiles are [{'id','name'}...]."""
    try:
        d = json.loads(_profiles_file(username).read_text())
    except Exception:
        d = {}
    profs = [p for p in d.get("profiles", [])
             if isinstance(p, dict) and p.get("id")]
    profs.insert(0, {"id": "", "name": d.get("default_name") or "My card"})
    active = d.get("active", "")
    if active not in {p["id"] for p in profs}:
        active = ""
    return profs, active


def _save_profiles(username: str, extra_profiles: list, active: str,
                   default_name: str = None):
    _ensure_dirs()
    if default_name is None:  # preserve a previously renamed default profile
        try:
            default_name = json.loads(
                _profiles_file(username).read_text()).get("default_name", "")
        except Exception:
            default_name = ""
    data = {"profiles": extra_profiles, "active": active}
    if default_name:
        data["default_name"] = default_name
    _profiles_file(username).write_text(
        json.dumps(data, ensure_ascii=False, indent=2))


def _delete_profile_data(username: str, pid: str):
    """Remove every app-side file of one extra profile (Drive files stay —
    read-only policy). No-op for the default profile."""
    import shutil
    if not pid:
        return
    for f in DATA_DIR.glob(f"*_{username}@{pid}*"):
        try:
            shutil.rmtree(f, ignore_errors=True) if f.is_dir() else f.unlink()
        except Exception:
            pass


def _active_pid(username: str) -> str:
    forced = getattr(_PROFILE_CTX, "force_pid", None)
    if forced is not None:
        return forced
    return _load_profiles(username)[1]


def _pkey(username: str) -> str:
    """Storage key for the active card profile — legacy `<user>` for the
    default profile, `<user>@<pid>` for additional cards."""
    pid = _active_pid(username)
    return f"{username}@{pid}" if pid else username


def _export_tag(username: str) -> str:
    """user(+profile) tag for export filenames — e.g. 'jongha.kang' or
    'jongha.kang_CEO_card' (강프로 2026-07-24)."""
    tag = username
    profs, pid = _load_profiles(username)
    if pid:
        pname = next((p["name"] for p in profs if p["id"] == pid), pid)
        tag += "_" + pname
    return re.sub(r"[^\w.@-]+", "_", tag).strip("_")


def _receipts_file(username: str) -> Path:
    return DATA_DIR / f"receipts_{_pkey(username)}.json"


def _migrate_entry(e: dict) -> dict:
    """Ensure a receipt entry has v2 fields (id, ocr_status, match_status)."""
    e = dict(e)
    if not e.get("id"):
        e["id"] = "rcpt_" + (e.get("file_id") or uuid.uuid4().hex)[:8]
    # Multi-receipt: legacy entries are the only receipt on their image → index 0.
    e.setdefault("sub_index", 0)
    if "ocr_status" not in e:
        e["ocr_status"] = "done" if e.get("ocr_amount") is not None else "pending"
    # v2.1: handwritten-priority OCR. Legacy entries keep ocr_amount as printed total.
    e.setdefault("ocr_printed_amount", e.get("ocr_amount"))
    e.setdefault("ocr_handwritten_amount", None)
    # v2.2: multi-receipt bounding box overlay. Legacy entries have no bbox.
    e.setdefault("ocr_bbox", None)
    # v2.3: card brand (amex/visa/other, OCR-detected, user-editable), usage tag
    # (default "Regular"), and completion state (completed entries are hidden from
    # the default Ledger view, Sync and Mapping, and their Drive originals are
    # moved to a "Completed" folder).
    e.setdefault("card_brand", None)
    e.setdefault("usage", "Regular")
    # v2.5: handwritten "w/ NAME" companion note — flows into the SAP purpose.
    e.setdefault("ocr_companions", None)
    e.setdefault("ocr_handwriting", None)
    # v2.7: actual settled USD from the matched statement line (vs FX estimate).
    e.setdefault("usd_settled", None)
    # v2.8: 'Reason for Cash' on cash receipts — SAP xlsx column S.
    e.setdefault("cash_reason", None)
    # v2.6: printed transaction time (duplicate disambiguation) + user override
    # that pins an entry out of duplicate grouping ("this is a separate purchase").
    e.setdefault("ocr_time", None)
    e.setdefault("dup_exempt", False)
    # v2.4: foreign-currency receipts (KRW/INR business trips) — OCR currency +
    # USD estimate for band-matching against the USD AMEX statement.
    e.setdefault("ocr_currency", None)
    e.setdefault("usd_estimate", None)
    e.setdefault("fx_rate", None)
    e.setdefault("completed", False)
    e.setdefault("completed_at", None)
    e.setdefault("archived_drive", False)
    if "match_status" not in e:
        if e.get("matched"):
            e["match_status"] = "matched"
        else:
            # Initial state is always pending_match. unmatched is reserved for
            # receipts that were tried against a CSV but failed to match.
            e["match_status"] = "pending_match"
    # Backfill OCR date from the matched CSV transaction for legacy entries that
    # were matched while OCR failed to read a date.
    if e.get("match_status") == "matched" and e.get("ocr_date") in (None, "", "unknown"):
        mt_date = (e.get("matched_transaction") or {}).get("date")
        if mt_date:
            e["ocr_date_original"] = e.get("ocr_date")
            e["ocr_date"] = mt_date
    # A matched receipt is an AMEX transaction (the CSV is an AMEX statement);
    # backfill the brand for already-matched entries that predate card detection.
    if e.get("match_status") == "matched" and not e.get("card_brand"):
        e["card_brand"] = "amex"
    return e


def _load_ledger(username: str) -> dict:
    """Load ledger, auto-migrating v1 (list) to v2 (dict) format."""
    f = _receipts_file(username)
    raw = None
    if f.exists():
        try:
            raw = json.loads(f.read_text())
        except Exception:
            raw = None
    if raw is None:
        return {"version": 2, "last_batch_at": None, "entries": []}
    if isinstance(raw, list):  # v1 → v2
        return {"version": 2, "last_batch_at": None,
                "entries": [_migrate_entry(e) for e in raw]}
    raw["entries"] = [_migrate_entry(e) for e in raw.get("entries", [])]
    raw.setdefault("version", 2)
    raw.setdefault("last_batch_at", None)
    return raw


def _save_ledger(username: str, ledger: dict):
    _ensure_dirs()
    _receipts_file(username).write_text(json.dumps(ledger, ensure_ascii=False, indent=2))


def _ledger_entries(username: str) -> list:
    return _load_ledger(username)["entries"]


def _ledger_stats(entries: list) -> dict:
    matched = sum(1 for e in entries if e.get("match_status") == "matched")
    unmatched = sum(1 for e in entries if e.get("match_status") == "unmatched")
    pending = sum(1 for e in entries if e.get("match_status") == "pending_match")
    return {"total": len(entries), "matched": matched,
            "unmatched": unmatched, "pending_match": pending}


def _load_receipts(username: str) -> list:
    """Backward-compatible entries accessor (returns ledger entries list)."""
    return _ledger_entries(username)


def _save_receipts(username: str, receipts: list):
    """Persist entries list into v2 ledger, preserving wrapper metadata."""
    ledger = _load_ledger(username)
    ledger["entries"] = receipts
    _save_ledger(username, ledger)


def _drive_meta_file(username: str) -> Path:
    return DATA_DIR / f"drive_meta_{_pkey(username)}.json"


def _load_drive_meta(username: str) -> dict:
    f = _drive_meta_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return {}


def _save_drive_meta(username: str, meta: dict):
    _ensure_dirs()
    _drive_meta_file(username).write_text(json.dumps(meta, ensure_ascii=False, indent=2))


def _review_file(username: str) -> Path:
    return DATA_DIR / f"review_{_pkey(username)}.json"


def _load_review(username: str) -> dict:
    f = _review_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return {}


def _save_review(username: str, review: dict):
    _ensure_dirs()
    _review_file(username).write_text(json.dumps(review, ensure_ascii=False, indent=2))


# ── OCR Staging (pre-ledger visual review) ────────────────────────────────────

def _ocr_staging_file(username: str) -> Path:
    return DATA_DIR / f"ocr_staging_{_pkey(username)}.json"

def _load_ocr_staging(username: str) -> dict:
    f = _ocr_staging_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return {"entries": []}

def _save_ocr_staging(username: str, data: dict):
    _ensure_dirs()
    _ocr_staging_file(username).write_text(json.dumps(data, ensure_ascii=False, indent=2))

def _clear_ocr_staging(username: str):
    f = _ocr_staging_file(username)
    if f.exists():
        f.unlink()


# Discard tombstones: Drive trash can silently fail (drive.file scope can't
# modify files the user dropped into the folder themselves), and an untrashed
# file would be re-OCRed and re-staged by every sync — the "discarded receipt
# keeps coming back" loop (2026-07-22). Tombstoned file_ids are skipped by
# _list_new_drive_files regardless of the trash outcome.
_DISCARDED_CAP = 500

def _ocr_discarded_file(username: str) -> Path:
    return DATA_DIR / f"ocr_discarded_{_pkey(username)}.json"

def _load_discarded_fids(username: str) -> dict:
    f = _ocr_discarded_file(username)
    if f.exists():
        try:
            d = json.loads(f.read_text())
            if isinstance(d, dict):
                return d
        except Exception:
            pass
    return {}

def _mark_discarded_fid(username: str, fid: str, filename: str = None):
    if not fid:
        return
    _ensure_dirs()
    d = _load_discarded_fids(username)
    d[fid] = {"at": datetime.now().isoformat(), "filename": filename or ""}
    if len(d) > _DISCARDED_CAP:
        def _ts(v):
            return v.get("at", "") if isinstance(v, dict) else str(v)
        for k in sorted(d, key=lambda k: _ts(d[k]))[:len(d) - _DISCARDED_CAP]:
            d.pop(k, None)
    _ocr_discarded_file(username).write_text(json.dumps(d, ensure_ascii=False, indent=2))


def _unmark_discarded_fid(username: str, fid: str) -> bool:
    d = _load_discarded_fids(username)
    if fid not in d:
        return False
    d.pop(fid)
    _ensure_dirs()
    _ocr_discarded_file(username).write_text(json.dumps(d, ensure_ascii=False, indent=2))
    return True


def _discarded_items(username: str) -> list:
    """Tombstones as UI rows, newest first (legacy values were bare timestamps)."""
    out = []
    for fid, v in _load_discarded_fids(username).items():
        if isinstance(v, dict):
            out.append({"file_id": fid, "at": v.get("at", ""), "filename": v.get("filename", "")})
        else:
            out.append({"file_id": fid, "at": str(v), "filename": ""})
    return sorted(out, key=lambda x: x["at"], reverse=True)


def _handle_discarded_restore(username: str, body: dict):
    """POST /cardconv/receipts/review/restore — drop a tombstone so the file
    re-enters the OCR queue on the next Drive sync."""
    fid = body.get("file_id")
    if isinstance(fid, list):
        fid = fid[0] if fid else ""
    fid = (fid or "").strip()
    if not fid:
        return ("json", {"error": "file_id required"}, 400)
    if not _unmark_discarded_fid(username, fid):
        return ("json", {"error": "not found"}, 404)
    return ("json", {"ok": True})

def _handle_ocr_staging_discard_file(username: str, body: dict):
    """POST /cardconv/receipts/review/discard-file — drop every staged sub-entry
    of one photo and tombstone the file so the next sync doesn't re-stage it.
    Drive is never touched (uniform policy 2026-07-22 — drive.file couldn't
    modify user-dropped files anyway; restore via the Ledger Discarded list).
    Used by the 3+-receipts-per-photo warning flow (user re-uploads better shots)."""
    fid = body.get("file_id")
    if isinstance(fid, list):
        fid = fid[0] if fid else ""
    fid = (fid or "").strip()
    if not fid:
        return ("json", {"error": "file_id required"}, 400)
    staging = _load_ocr_staging(username)
    entries = staging.get("entries", [])
    keep    = [e for e in entries if e.get("file_id") != fid]
    removed = len(entries) - len(keep)
    fname = next((e.get("filename") for e in entries if e.get("file_id") == fid), "")
    staging["entries"] = keep
    _save_ocr_staging(username, staging)
    _mark_discarded_fid(username, fid, fname)
    return ("json", {"ok": True, "removed": removed})


def _handle_ocr_staging_discard_entry(username: str, body: dict):
    """POST /cardconv/receipts/review/discard-entry — reject one staged receipt
    before it reaches the Ledger. Drive is never touched; the file is
    tombstoned (skipped by future syncs) once nothing else references it —
    other sub-entries of a multi-receipt photo, or a ledger entry from an
    earlier confirm, keep the shared file un-tombstoned. Restorable from the
    Ledger's Discarded list."""
    eid = body.get("id")
    if isinstance(eid, list):
        eid = eid[0] if eid else ""
    eid = (eid or "").strip()
    if not eid:
        return ("json", {"error": "id required"}, 400)
    staging = _load_ocr_staging(username)
    entries = staging.get("entries", [])
    target = next((e for e in entries if e.get("id") == eid), None)
    if not target:
        return ("json", {"error": "entry not found"}, 404)
    staging["entries"] = [e for e in entries if e.get("id") != eid]
    _save_ocr_staging(username, staging)
    fid = target.get("file_id")
    if fid:
        still_staged = any(e.get("file_id") == fid for e in staging["entries"])
        in_ledger = any(le.get("file_id") == fid
                        for le in _load_ledger(username).get("entries", []))
        if not still_staged and not in_ledger:
            _mark_discarded_fid(username, fid, target.get("filename"))
    return ("json", {"ok": True, "remaining": len(staging["entries"])})


def _flag_staged_dups(username: str, entries: list) -> list:
    """Annotate staged entries in place with dup_hint:
    'ledger' — an active ledger receipt already carries the same amount +
    merchant on a compatible date/time (this upload is likely a re-scan);
    'staged' — an earlier entry in this same queue does (double upload).
    Mirrors the Ledger dup-group rule: (amount, merchant) bucket, dates equal
    or one side missing, printed times within tolerance (_times_close)."""
    def key(e):
        amt = e.get("ocr_amount")
        if amt is None:
            return None
        try:
            return (round(float(amt), 2), (e.get("ocr_merchant") or "").strip().lower())
        except (TypeError, ValueError):
            return None

    def compatible(a, b):
        da, db = a.get("ocr_date"), b.get("ocr_date")
        return ((da is None or db is None or da == db)
                and _times_close(a.get("ocr_time"), b.get("ocr_time")))

    ledger_by_key = {}
    for le in _load_ledger(username).get("entries", []):
        k = key(le)
        if k and not le.get("completed"):
            ledger_by_key.setdefault(k, []).append(le)

    seen = {}
    for e in entries:
        e["dup_hint"] = None
        k = key(e)
        if not k:
            continue
        if any(compatible(e, le) for le in ledger_by_key.get(k, [])):
            e["dup_hint"] = "ledger"
        elif any(compatible(e, se) for se in seen.get(k, [])):
            e["dup_hint"] = "staged"
        seen.setdefault(k, []).append(e)
    return entries


def _handle_ocr_staging_confirm(username: str, body: dict):
    """POST /cardconv/receipts/review/confirm (JSON) — apply manual edits and add to ledger.

    Expects JSON body: {"confirmed": [{"id":..., "ocr_date":..., "ocr_merchant":...,
                                       "ocr_amount":..., "ocr_handwritten_amount":...}, ...]}
    Also accepts legacy FormData with just confirmed[] IDs (no edits).
    """
    staging = _load_ocr_staging(username)
    all_entries = staging.get("entries", [])

    # Build a correction map from the JSON payload.
    confirmed_list = body.get("confirmed", [])
    is_ajax = (isinstance(confirmed_list, list) and confirmed_list
               and isinstance(confirmed_list[0], dict))
    if is_ajax:
        # JSON path (modal fetch): each item has id + corrected fields
        corrections = {item["id"]: item for item in confirmed_list if "id" in item}
        confirmed_ids = set(corrections.keys())
    else:
        # Legacy FormData path: a list of IDs, plus optional per-entry
        # 'cash_reason_<id>' inputs from the staging page's cash rows.
        if isinstance(confirmed_list, str):
            confirmed_list = [confirmed_list]
        confirmed_ids = set(confirmed_list)
        corrections = {}
        for k, v in body.items():
            if k.startswith("cash_reason_"):
                val = (v[0] if isinstance(v, list) else str(v)).strip()
                if val:
                    corrections[k[len("cash_reason_"):]] = {"cash_reason": val}

    confirmed = []
    for e in all_entries:
        if e.get("id") not in confirmed_ids:
            continue
        entry = dict(e)
        fix = corrections.get(e["id"], {})
        if fix.get("ocr_date") is not None:
            entry["ocr_date"] = fix["ocr_date"] or None
        if fix.get("ocr_merchant") is not None:
            entry["ocr_merchant"] = fix["ocr_merchant"] or None
        for amt_key in ("ocr_amount", "ocr_handwritten_amount"):
            raw = fix.get(amt_key)
            if raw is not None:
                try:
                    entry[amt_key] = float(raw) if str(raw).strip() else None
                except (ValueError, TypeError):
                    pass
        raw = fix.get("card_brand")
        if raw is not None:
            v = str(raw).strip().lower()
            entry["card_brand"] = v if v in ("amex", "visa", "other") else None
        raw = fix.get("ocr_companions")
        if raw is not None:
            entry["ocr_companions"] = _coerce_companions(str(raw))
        raw = fix.get("cash_reason")
        if raw is not None:
            entry["cash_reason"] = str(raw).strip()[:120] or None
        # Recompute final amount after manual correction.
        hw = entry.get("ocr_handwritten_amount")
        pr = entry.get("ocr_amount")
        entry["ocr_final_amount"] = hw if hw is not None else pr
        confirmed.append(entry)

    # Server-side guard (강프로 2026-07-24): a cash receipt cannot enter the
    # ledger without its Reason for Cash — it feeds SAP column S at export.
    no_reason = [e for e in confirmed
                 if e.get("card_brand") == "other"
                 and not (e.get("cash_reason") or "").strip()]
    if no_reason:
        if is_ajax:
            return ("json", {"ok": False, "error":
                             f"Reason for Cash is required for cash receipts "
                             f"({len(no_reason)} missing)."})
        return ("redirect", "/cardconv/receipts/review?need_reason=1")

    if confirmed:
        receipts = _load_receipts(username)
        receipts.extend(confirmed)
        _save_receipts(username, receipts)
        # Match the fresh receipts against open pool transactions right away —
        # otherwise they sit pending until some other flow happens to re-match.
        _rematch_pool(username)

    _clear_ocr_staging(username)
    # Native form (staging page) submissions navigate the browser, so send them
    # back to the Ledger; the modal's fetch path keeps getting JSON.
    if is_ajax:
        return ("json", {"ok": True, "added": len(confirmed)})
    return ("redirect", "/cardconv/ledger")


# ── User settings: card member names ──────────────────────────────────────────

# Shown as a placeholder example only — names are per-account with no built-in
# default, so each user must register their own before converting.
EXAMPLE_CARD_NAME = "JONGHA KANG"


def _user_settings_file(username: str) -> Path:
    return DATA_DIR / f"user_settings_{_pkey(username)}.json"


def _load_user_settings(username: str) -> dict:
    f = _user_settings_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return {}


def _save_user_settings(username: str, data: dict):
    _ensure_dirs()
    _user_settings_file(username).write_text(json.dumps(data, ensure_ascii=False, indent=2))


def _get_card_member_names(username: str) -> list:
    """User's saved card member names (uppercased). Empty until the user adds
    their own — there is no account-independent default."""
    if not username:
        return []
    names = _load_user_settings(username).get("card_member_names") or []
    return [n.strip().upper() for n in names if str(n).strip()]


# ── Uploaded CSV store (Convert page reuse) ────────────────────────────────────

def _uploads_dir(username: str) -> Path:
    return DATA_DIR / f"uploads_{_pkey(username)}"


def _uploads_index_file(username: str) -> Path:
    return DATA_DIR / f"uploads_{_pkey(username)}.json"


def _load_uploads(username: str) -> list:
    f = _uploads_index_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            pass
    return []


def _save_uploads(username: str, items: list):
    _ensure_dirs()
    _uploads_index_file(username).write_text(json.dumps(items, ensure_ascii=False, indent=2))


def _save_uploaded_csv(username: str, csv_bytes: bytes, orig_name: str,
                       rows: int, out_filename: str) -> dict:
    """Persist the raw CSV under uploads_{user}/ and index it for reuse."""
    d = _uploads_dir(username)
    d.mkdir(parents=True, exist_ok=True)
    uid = "csv_" + uuid.uuid4().hex[:8]
    (d / f"{uid}.csv").write_bytes(csv_bytes)
    entry = {
        "id":           uid,
        "filename":     orig_name,
        "stored_name":  f"{uid}.csv",
        "uploaded_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
        "rows":         rows,
        "out_filename": out_filename,
    }
    items = _load_uploads(username)
    items.insert(0, entry)
    _save_uploads(username, items)
    return entry


# ── Cross-upload duplicate detection ──────────────────────────────────────────
# Statement CSVs with overlapping periods repeat transactions. History keeps
# every uploaded CSV, so it is the source of truth: deleting an upload there
# automatically releases its transactions for re-conversion.

def _tx_key(row: dict):
    """Duplicate-detection key for one CSV transaction row."""
    try:
        amt = f"{float(row.get('Amount', 0)):.2f}"
    except (ValueError, TypeError):
        amt = str(row.get('Amount', '')).strip()
    return (
        (row.get('Date') or '').strip(),
        amt,
        (row.get('Account Number') or '').strip(),
        (row.get('Merchant Name') or '').strip().upper(),
    )


def _key_norm(k) -> tuple:
    """Cross-source comparison form of a tx key: (date, amount, merchant).

    The account number is dropped — reissued cards give the same transaction
    a different account across exports — and merchant whitespace is collapsed
    (the Master xlsx pads fields with space runs that the Posted CSV doesn't).
    """
    k = tuple(k or ())
    if len(k) < 2:
        # keyless pool rows (e.g. synthetic cash mirrors) can never dedupe a
        # real statement row — return a sentinel no CSV row will produce
        return ("", "", "")
    date, amt = k[0], k[1]
    merchant = re.sub(r"\s+", " ", k[3] if len(k) > 3 else "").strip().upper()
    return (date, amt, merchant)


def _dedup_rows(rows: list, prior: Counter) -> tuple:
    """(kept_rows, skipped_count). Multiset semantics: N prior occurrences
    absorb at most N new ones, so legitimate same-day/same-amount repeats
    within one statement survive. Keys are compared in normalized form so
    the same transaction dedupes across CSV and Master-xlsx sources."""
    if not prior:
        return rows, 0
    kept, skipped = [], 0
    remaining = Counter(_key_norm(k) for k in prior.elements())
    for r in rows:
        k = _key_norm(_tx_key(r))
        if remaining.get(k, 0) > 0:
            remaining[k] -= 1
            skipped += 1
        else:
            kept.append(r)
    return kept, skipped


# ── Transaction pool ──────────────────────────────────────────────────────────
# Uploaded CSVs are merged into one persistent pool (duplicates ingested once).
# Review shows every "open" transaction across uploads until the user marks it
# completed; the xlsx download is built from the open set on demand.

def _tx_pool_file(username: str) -> Path:
    return DATA_DIR / f"transactions_{_pkey(username)}.json"


def _save_tx_pool(username: str, pool: dict):
    _ensure_dirs()
    _tx_pool_file(username).write_text(json.dumps(pool, ensure_ascii=False, indent=2))


def _parse_member_rows(csv_bytes: bytes, username: str) -> list:
    """CSV rows filtered to the user's card member names."""
    target = set(_get_card_member_names(username))
    reader = csv.DictReader(io.TextIOWrapper(io.BytesIO(csv_bytes), encoding='utf-8-sig', newline=''))
    return [r for r in reader if r.get("Card Member Name", "").strip().upper() in target]


def _row_to_entry(row: dict, rules: dict, source: str) -> dict:
    merchant = row.get("Merchant Name", "").strip()
    dba      = row.get("Merchant Doing Business As", "").strip()
    vendor   = dba if (dba and dba != merchant) else merchant
    try:
        amount = float(row.get("Amount", 0))
    except ValueError:
        amount = 0.0
    inv_dt = _parse_date(row.get("Date", ""))
    gl, ser, purpose = _classify(vendor, rules)
    if gl is None:
        gl, ser, purpose = _classify(merchant, rules)
    kw_unmatched = gl is None
    if gl is None:
        gl, ser, purpose = 53410177, "160", "Coffee, Snack and meal"
    return {
        "id":           "tx_" + uuid.uuid4().hex[:8],
        "status":       "open",
        "key":          list(_tx_key(row)),
        "date":         inv_dt.strftime("%Y-%m-%d") if inv_dt else None,
        "merchant":     vendor,
        "amount":       round(amount, 2),
        "gl":           gl,
        "ser":          ser,
        "purpose":      purpose,
        "kw_unmatched": kw_unmatched,
        "account":      row.get("Account Number", ""),
        "member":       row.get("Card Member Name", ""),
        "matched":      False,
        "receipt":      None,
        "loss_reason":  "",
        "source":       source,
        "added_at":     datetime.now().isoformat(),
        "completed_at": None,
    }


def _load_tx_pool(username: str) -> dict:
    """Load the pool; on first use migrate from the CSVs stored in History.

    Migration: every transaction (deduped across uploads) starts open —
    completion is an explicit user action, never assumed. The open batch
    inherits matched/receipt info from the legacy review snapshot."""
    f = _tx_pool_file(username)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            return {"entries": []}

    pool = {"entries": [], "migrated_at": datetime.now().isoformat()}
    uploads = _load_uploads(username)  # newest first
    if uploads:
        rules = _load_kw()
        seen = Counter()
        d = _uploads_dir(username)
        for up in reversed(uploads):  # oldest → newest
            p = d / up.get("stored_name", "")
            if not p.exists():
                continue
            try:
                rows = _parse_member_rows(p.read_bytes(), username)
            except Exception:
                continue
            fresh, _ = _dedup_rows(rows, seen)
            for row in fresh:
                seen[_tx_key(row)] += 1
                pool["entries"].append(_row_to_entry(row, rules, up.get("filename", "")))

        # Enrich the open batch from the legacy review snapshot (match info).
        legacy = {}
        for r in _load_review(username).get("rows", []):
            legacy.setdefault((r.get("date"), r.get("amount"), r.get("merchant")), []).append(r)
        for e in pool["entries"]:
            if e["status"] != "open":
                continue
            cands = legacy.get((e["date"], e["amount"], e["merchant"]))
            if cands:
                r = cands.pop(0)
                e["matched"] = bool(r.get("matched"))
                e["receipt"] = r.get("receipt")
                e["loss_reason"] = r.get("loss_reason", "")
    _save_tx_pool(username, pool)
    # A stale legacy snapshot can leave the open batch unmatched — run live
    # receipt matching so migration never depends on snapshot freshness.
    if pool["entries"]:
        _rematch_pool(username)
        pool = json.loads(f.read_text())
    return pool


def _rematch_pool(username: str, only_receipt_ids=None) -> dict:
    """Match every open, unmatched pool transaction against the receipt ledger.

    Same matching as _ingest_csv applies to fresh rows. Returns {"matched": n}.
    only_receipt_ids restricts the candidate receipts (bulk re-match of a selection).
    """
    pool = _load_tx_pool(username)
    receipts = _load_receipts(username)
    # Heal orphaned links first, in both directions. Receipt side: matched
    # flags whose transaction link no longer exists (dropped by migration) —
    # left as-is the matcher skips the receipt forever. Transaction side: a
    # link to a receipt that was deleted (e.g. removing the matched copy of a
    # duplicate, 2026-07-22 STARBUCKS case) — left as-is the matcher skips the
    # transaction forever and the surviving duplicate can never pair with it.
    linked = {(e.get("receipt") or {}).get("id")
              for e in pool.get("entries", []) if e.get("matched")}
    dirty_heal = False
    for r in receipts:
        if r.get("matched") and r.get("id") not in linked:
            r["matched"] = False
            r["match_status"] = "pending_match"
            r["usd_settled"] = None
            r["matched_transaction"] = None  # stale ↳ snapshot confuses the row
            dirty_heal = True
    receipt_ids = {r.get("id") for r in receipts}
    dirty_pool_heal = False
    for e in pool.get("entries", []):
        rid = (e.get("receipt") or {}).get("id")
        if e.get("matched") and rid and rid not in receipt_ids:
            e["matched"] = False
            e["receipt"] = None
            dirty_pool_heal = True
    todo = [e for e in pool.get("entries", [])
            if e.get("status") == "open" and not e.get("matched")]
    if not todo:
        if dirty_heal:
            _save_receipts(username, receipts)
        if dirty_pool_heal:
            _save_tx_pool(username, pool)
        return {"matched": 0}
    receipts_map, fx_receipts, dirty = _build_receipt_index(receipts, username)
    if only_receipt_ids:
        sel = set(only_receipt_ids)
        receipts_map = {k: r for k, r in receipts_map.items() if r.get("id") in sel}
        fx_receipts = [t for t in fx_receipts if t[2].get("id") in sel]
    dirty = dirty or dirty_heal
    matched = 0
    for e in todo:
        r = _find_receipt_match(e["date"], e["amount"], receipts_map, fx_receipts)
        if r is not None:
            _apply_receipt_match(e, r, receipts)
            matched += 1
            dirty = True
    if matched or dirty_pool_heal:
        _save_tx_pool(username, pool)
    if dirty:
        _save_receipts(username, receipts)
    return {"matched": matched}


def _heal_orphan_matches(username: str):
    """Lazy self-heal for one-sided match links: a pool rebuild can drop the
    transaction-side link while the receipt keeps its matched flags, so the
    Ledger says "Matched" while Review shows the transaction open — and the
    matcher skips matched receipts forever (2026-07-21 Sumiya case). Cheap
    set check on every load; the full heal+rematch runs only on an actual
    orphan, so this is a no-op when consistent."""
    receipts = _load_receipts(username)
    pool = _load_tx_pool(username)
    linked = {(e.get("receipt") or {}).get("id")
              for e in pool.get("entries", []) if e.get("matched")}
    receipt_ids = {r.get("id") for r in receipts}
    receipt_orphan = any(r.get("matched") and r.get("id") not in linked
                         for r in receipts)
    ghost_tx = any(rid and rid not in receipt_ids for rid in linked)
    if receipt_orphan or ghost_tx:
        _rematch_pool(username)


def _build_receipt_index(receipts: list, username: str):
    """(receipts_map, fx_receipts, dirty) — same matching inputs convert used."""
    receipts_map, fx_receipts, dirty = {}, [], False
    for r in receipts:
        if r.get("completed"):
            continue
        # Cash never appears on the AMEX statement — a cash receipt matching a
        # statement line by date+amount coincidence silently flipped it to AMEX
        # (2026-07-21 mirroring incident). Rule: cash can't match; matched ⇒ AMEX.
        if (r.get("card_brand") or "") == "other":
            continue
        rdate = r.get("ocr_date")
        if rdate == "YYYY-MM-DD":
            rdate = None
        hw = r.get("ocr_handwritten_amount")
        pr = r.get("ocr_printed_amount") or r.get("ocr_amount")
        ramount = hw if hw is not None else pr
        cur = r.get("ocr_currency")
        if cur and cur != "USD":
            usd_est = r.get("usd_estimate")
            if usd_est is None:
                usd_est, fx = _fx_usd_estimate(ramount, cur, rdate)
                if usd_est is not None:
                    r["usd_estimate"], r["fx_rate"] = usd_est, fx
                    dirty = True
            if usd_est is not None:
                fx_receipts.append((rdate, usd_est, r))
        elif ramount is not None:
            try:
                receipts_map[(rdate, round(float(ramount), 2))] = r
            except (ValueError, TypeError):
                pass
    return receipts_map, fx_receipts, dirty


def _find_receipt_match(inv_date_str, amt_rounded, receipts_map, fx_receipts):
    match = receipts_map.get((inv_date_str, amt_rounded))
    if match:
        return match
    for (rdate, ramt), r in receipts_map.items():
        if abs(ramt - amt_rounded) > 0.01:
            continue
        if rdate is None or rdate == inv_date_str:
            return r
        try:  # ±1 day posting-date skew
            rd  = date.fromisoformat(rdate)
            id_ = date.fromisoformat(inv_date_str) if inv_date_str else None
            if id_ is not None and abs((rd - id_).days) <= 1:
                return r
        except Exception:
            pass
    for rdate, usd_est, r in fx_receipts:
        # Foreign receipts: ±FX_TOLERANCE USD band, ±3d (late intl posting).
        if r.get('matched') or not usd_est:
            continue
        if abs(amt_rounded - usd_est) > usd_est * FX_TOLERANCE:
            continue
        if rdate is None or rdate == inv_date_str:
            return r
        try:
            rd  = date.fromisoformat(rdate)
            id_ = date.fromisoformat(inv_date_str) if inv_date_str else None
            if id_ is not None and abs((rd - id_).days) <= 3:
                return r
        except Exception:
            pass
    return None


def _sync_cash_pool(username: str) -> None:
    """Mirror cash receipts (card_brand='other') into the transaction pool.

    Cash spend never appears on the AMEX statement CSV, so Review would
    otherwise hide it. Each cash receipt gets a synthetic matched pool row
    (id 'cash_<receipt id>', no G/L): visible, taggable and settle-tracked
    like any transaction, included in every export — the SAP xlsx fills the
    'Reason for Cash' column from cash_reason (blank = fill in by hand).
    Un-marking the receipt as cash removes the row again and frees the
    receipt for normal statement matching."""
    receipts = _load_receipts(username)
    pool = _load_tx_pool(username)
    entries = pool.get("entries", [])
    have = {e.get("id") for e in entries}
    # Receipts already linked to a REAL statement transaction are on the
    # statement — never cash, whatever the OCR thought.
    real_linked = {(e.get("receipt") or {}).get("id"): e for e in entries
                   if e.get("matched") and not e.get("cash")}
    dirty = False
    # Heal OCR 'other' misreads on statement-matched receipts (+ backfill the
    # settled USD amount for pre-existing matches).
    for r in receipts:
        e = real_linked.get(r.get("id"))
        if e is None:
            continue
        if (r.get("card_brand") or "") == "other":
            r["card_brand"] = "amex"
            dirty = True
        if r.get("usd_settled") is None and e.get("amount") is not None:
            r["usd_settled"] = e.get("amount")
            dirty = True
    cash = [r for r in receipts
            if (r.get("card_brand") or "") == "other"
            and r.get("id") not in real_linked]
    cash_rids = {r.get("id") for r in cash}
    for r in cash:
        cid = f"cash_{r.get('id')}"
        if cid in have:
            continue
        entry = {
            "id": cid, "cash": True,
            "date": r.get("ocr_date") or "",
            "merchant": r.get("ocr_merchant") or r.get("filename") or "Cash receipt",
            "amount": r.get("usd_estimate") or r.get("ocr_amount") or 0,
            "gl": "",
            "status": "completed" if r.get("completed") else "open",
            "added_at": datetime.now().isoformat(),
        }
        _apply_receipt_match(entry, r, receipts)
        entries.append(entry)
        dirty = True
    # Cash rows classify like statement rows: same keyword rules fill
    # G/L / Ser. / Purpose (backfills pre-2026-07-21 rows too, when the SAP
    # export didn't carry cash lines).
    rules = _load_kw()
    rcpt_by_id = {r.get("id"): r for r in receipts}
    for e in entries:
        if not e.get("cash"):
            continue
        if not e.get("gl"):
            gl, ser, purpose = _classify(e.get("merchant") or "", rules)
            if gl is None:
                gl, ser, purpose = 53410177, "160", "Coffee, Snack and meal"
            e.update(gl=gl, ser=ser, purpose=purpose)
            dirty = True
        # Receipt-side 'Reason for Cash' (OCR confirm / Ledger edit) is the
        # source of truth — mirror onto the pool row for Review + SAP col S.
        src = rcpt_by_id.get((e.get("receipt") or {}).get("id"))
        if src is not None and src.get("cash_reason") != e.get("cash_reason"):
            e["cash_reason"] = src.get("cash_reason")
            dirty = True
    kept = []
    for e in entries:
        if e.get("cash") and (e.get("receipt") or {}).get("id") not in cash_rids:
            rid = (e.get("receipt") or {}).get("id")
            if rid in real_linked:
                # Duplicate: the receipt is genuinely on the statement (FX
                # match) — the real transaction row wins, drop the mirror and
                # leave the receipt matched to it.
                dirty = True
                continue
            # Receipt deleted or no longer cash — drop the synthetic row and
            # release a still-existing receipt back to pending_match.
            for r in receipts:
                if r.get("id") == rid and r.get("matched"):
                    r["matched"] = False
                    r["match_status"] = "pending_match"
                    r["usd_settled"] = None
            dirty = True
            continue
        kept.append(e)
    if dirty:
        pool["entries"] = kept
        _save_tx_pool(username, pool)
        _save_receipts(username, receipts)


def _apply_receipt_match(entry: dict, receipt: dict, receipts: list):
    """Flag the ledger receipt + attach match info to the pool entry."""
    if not receipt.get('matched'):
        receipt['matched'] = True
        receipt['match_status'] = 'matched'
        receipt['matched_at'] = datetime.now().isoformat()
        receipt['matched_transaction'] = {
            'date': entry["date"], 'amount': entry["amount"], 'vendor': entry["merchant"],
        }
        if not entry.get('cash'):
            # A statement line IS ground truth: the purchase was charged to
            # AMEX (overrides an OCR 'other' misread of Korean card slips),
            # and its USD amount is the real settled figure — better than the
            # FX estimate the OCR produced.
            receipt['card_brand'] = 'amex'
            receipt['usd_settled'] = entry.get('amount')
        if receipt.get('ocr_date') in (None, '', 'unknown') and entry["date"]:
            receipt['ocr_date_original'] = receipt.get('ocr_date')
            receipt['ocr_date'] = entry["date"]
    # A usage tag set on the transaction before matching (Review allows
    # receipt-less tagging) carries over unless the receipt already has one.
    tx_usage = (entry.get("usage") or "").strip()
    if tx_usage and tx_usage != "Regular" and (receipt.get("usage") or "Regular") == "Regular":
        receipt["usage"] = tx_usage
    # Same carry-over for a w/ note typed on the transaction before matching.
    tx_comp = (entry.get("companions") or "").strip()
    if tx_comp and not receipt.get("ocr_companions"):
        receipt["ocr_companions"] = tx_comp
    mfid = receipt.get("file_id")
    siblings = [
        {"id": s.get("id"), "ocr_bbox": s.get("ocr_bbox")}
        for s in receipts
        if mfid and s.get("file_id") == mfid and s.get("ocr_bbox")
    ]
    entry["matched"] = True
    entry["no_receipt"] = False  # a receipt turned up after all
    entry["receipt"] = {
        "file_id":      mfid,
        "id":           receipt.get("id"),
        "filename":     receipt.get("filename"),
        "drive_url":    receipt.get("drive_url"),
        "ocr_amount":   receipt.get("ocr_amount"),
        "ocr_currency": receipt.get("ocr_currency"),
        "usd_estimate": receipt.get("usd_estimate"),
        "usd_settled":  receipt.get("usd_settled"),
        "ocr_date":     receipt.get("ocr_date"),
        "ocr_merchant": receipt.get("ocr_merchant"),
        "companions":   receipt.get("ocr_companions"),
        "ocr_bbox":     receipt.get("ocr_bbox"),
        "siblings":     siblings if len(siblings) > 1 else [],
    }


def _master_xlsx_to_csv_bytes(xlsx_bytes: bytes) -> bytes:
    """Convert an 'AMEX Master' xlsx OR a 'Billing Support File' .xls into
    Posted_*.csv-shaped bytes.

    The Master sheet ('I. Jongha' style) is a corporate recon export; the
    Billing Support File (Cardmember Monthly Account Detail, sent per control
    account as legacy .xls) shares the same column vocabulary with a metadata
    block above the header. Both map onto the CSV the rest of the pipeline
    expects. Dates are emitted ISO and account numbers digits-only so _tx_key
    stays comparable with rows ingested from real Posted CSVs.
    """
    # Sheet rows as plain tuples from either engine: openpyxl for xlsx, xlrd
    # for legacy BIFF .xls (the AMEX Billing Support File comes as .xls).
    if xlsx_bytes[:4] == b"\xd0\xcf\x11\xe0":
        import xlrd
        book = xlrd.open_workbook(file_contents=xlsx_bytes)
        all_sheets = []
        for sh in book.sheets():
            rows = []
            for r in range(sh.nrows):
                vals = []
                for c in range(sh.ncols):
                    cell_o = sh.cell(r, c)
                    v = cell_o.value
                    if cell_o.ctype == 3:  # XL_CELL_DATE
                        v = datetime(*xlrd.xldate_as_tuple(v, book.datemode))
                    vals.append(v)
                rows.append(tuple(vals))
            all_sheets.append(rows)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        all_sheets = [list(ws.iter_rows(values_only=True)) for ws in wb.worksheets]

    def norm(h):
        return re.sub(r"\s+", " ", str(h or "")).strip().lower()

    # The header row is row 1 in the Master export but sits below a metadata
    # block in the Billing Support File — scan the first rows of each sheet.
    sheet_rows, cols, hdr_i = None, {}, -1
    for rows in all_sheets:
        for i, header in enumerate(rows[:30]):
            idx = {norm(h): j for j, h in enumerate(header) if h}
            if (("desc" in idx or "transaction description 1" in idx)
                    and ("t.date" in idx or "transaction date" in idx)):
                sheet_rows, cols, hdr_i = rows, idx, i
                break
        if sheet_rows is not None:
            break
    if sheet_rows is None:
        raise ValueError(
            "No statement table found (needs a 'Desc'/'Transaction Description 1' "
            "+ 'T.Date'/'Transaction Date' header row)")

    def cell(row, *names):
        for n in names:
            i = cols.get(n)
            if i is not None and i < len(row) and row[i] not in (None, ""):
                return row[i]
        return ""

    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=[
        "Date", "Card Member Name", "Account Number", "Amount",
        "Merchant Name", "Merchant Doing Business As", "_recon"])
    w.writeheader()
    for row in sheet_rows[hdr_i + 1:]:
        desc = re.sub(r"\s+", " ", str(cell(row, "desc", "transaction description 1"))).strip()
        raw_date = cell(row, "transaction date", "t.date")
        raw_amt = cell(row, "transaction amount usd", "amt")
        if not desc or raw_date in ("", None) or raw_amt in ("", None):
            continue
        if isinstance(raw_date, datetime):
            iso = raw_date.strftime("%Y-%m-%d")
        else:
            dt = _parse_date(str(raw_date))
            if not dt:
                continue
            iso = dt.strftime("%Y-%m-%d")
        try:
            amt = f"{float(raw_amt):.2f}"
        except (ValueError, TypeError):
            continue
        first = str(cell(row, "supplemental cardmember first name",
                         "basic cardmember first name")).strip()
        last = str(cell(row, "supplemental cardmember last name",
                        "basic cardmember last name")).strip()
        acct = re.sub(r"\D", "", str(cell(row, "supplemental account number",
                                          "basic card account no.")))
        w.writerow({
            "Date": iso,
            "Card Member Name": f"{first} {last}".strip(),
            "Account Number": acct,
            "Amount": amt,
            "Merchant Name": desc,
            "Merchant Doing Business As": "",
            "_recon": str(cell(row, "recon")).strip().upper(),
        })
    return out.getvalue().encode("utf-8")


def _ingest_csv(username: str, csv_bytes: bytes, filename: str) -> dict:
    """Merge a CSV into the pool. Returns {"added", "dup_skipped", "matched"}."""
    pool = _load_tx_pool(username)
    rows = _parse_member_rows(csv_bytes, username)
    existing = Counter(k for k in (tuple(e.get("key") or ()) for e in pool["entries"])
                       if len(k) >= 2)
    fresh, dup_skipped = _dedup_rows(rows, existing)

    rules = _load_kw()
    receipts = _load_receipts(username)
    receipts_map, fx_receipts, dirty = _build_receipt_index(receipts, username)

    added, matched = 0, 0
    for row in fresh:
        e = _row_to_entry(row, rules, filename)
        if str(row.get("_recon", "")).strip().upper() == "Y":
            # Master xlsx marks reconciled transactions — enter as history,
            # not as open Review work.
            e["status"] = "completed"
            e["completed_at"] = e["added_at"]
        r = _find_receipt_match(e["date"], e["amount"], receipts_map, fx_receipts)
        if r is not None:
            _apply_receipt_match(e, r, receipts)
            matched += 1
            dirty = True
        pool["entries"].append(e)
        added += 1

    pool["last_ingest"] = {
        "filename": filename, "added": added, "dup_skipped": dup_skipped,
        "at": datetime.now().isoformat(),
    }
    _save_tx_pool(username, pool)
    if dirty:
        _save_receipts(username, receipts)
    return {"added": added, "dup_skipped": dup_skipped, "matched": matched}


def _inline_to_shared_strings(xlsx_bytes: bytes) -> bytes:
    """Rewrite an openpyxl-saved xlsx to use a sharedStrings table.

    openpyxl (3.1+) hardcodes strings as t="inlineStr", which strict parsers
    (SAP's Excel upload among them) don't understand — cells get misread and
    numeric columns render as dates. Excel itself always writes t="s" +
    xl/sharedStrings.xml, so we convert to that form.
    """
    src = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    parts = {n: src.read(n) for n in src.namelist()}

    strings: list = []
    index: dict = {}
    total = 0

    cell_re = re.compile(
        rb'(<c [^>]*?)t="inlineStr"([^>]*)><is><t(?: xml:space="preserve")?>'
        rb'(.*?)</t></is></c>', re.S)

    def convert(m):
        nonlocal total
        text = m.group(3)
        if text not in index:
            index[text] = len(strings)
            strings.append(text)
        total += 1
        idx = index[text]
        return m.group(1) + b't="s"' + m.group(2) + b'><v>' + str(idx).encode() + b'</v></c>'

    for name in list(parts):
        if name.startswith("xl/worksheets/") and name.endswith(".xml"):
            parts[name] = cell_re.sub(convert, parts[name])

    if strings:
        sst = [b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
               b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
               b'count="%d" uniqueCount="%d">' % (total, len(strings))]
        for s in strings:
            pre = b' xml:space="preserve"' if s.strip() != s else b''
            sst.append(b'<si><t%s>%s</t></si>' % (pre, s))
        sst.append(b'</sst>')
        parts["xl/sharedStrings.xml"] = b''.join(sst)

        ct = parts["[Content_Types].xml"]
        if b'sharedStrings' not in ct:
            parts["[Content_Types].xml"] = ct.replace(
                b'</Types>',
                b'<Override PartName="/xl/sharedStrings.xml" ContentType='
                b'"application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/></Types>')

        rels = parts["xl/_rels/workbook.xml.rels"]
        if b'sharedStrings' not in rels:
            rids = [int(x) for x in re.findall(rb'Id="rId(\d+)"', rels)]
            new_rid = max(rids or [0]) + 1
            parts["xl/_rels/workbook.xml.rels"] = rels.replace(
                b'</Relationships>',
                b'<Relationship Id="rId%d" Type="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships/sharedStrings" '
                b'Target="sharedStrings.xml"/></Relationships>' % new_rid)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as out:
        for name, data in parts.items():
            out.writestr(name, data)
    return buf.getvalue()


SAP_TEMPLATE = _SVC_DIR / "cardconv_sap_template.xlsx"  # SAP-accepted file, styles preserved

# Per-column style ids captured from the accepted file's data rows. SAP reads
# the workbook against ITS OWN style table, so the output must reuse these
# exact style indices - rebuilding styles (as openpyxl does) garbles parsing.
_SAP_COL_STYLE = {
    "A": "6", "B": "7", "C": "6", "D": "19", "E": "15", "F": "6", "G": "16",
    "H": "11", "I": "20", "J": "21", "K": "6", "L": "6", "M": "19", "N": "20",
    "O": "6", "P": "19", "Q": "19", "R": "22", "S": "19", "T": "6",
    "U": "23", "V": "23", "W": "23", "X": "23", "Y": "8", "Z": "20",
}
_SAP_SER_TEXT_STYLE = "29"  # J when Ser. keeps a leading zero (text, @)
_XLSX_EPOCH = date(1899, 12, 30)


def _xesc(s: str) -> str:
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))


def _build_xlsx_from_entries(entries: list, username: str = "") -> tuple:
    """(xlsx_bytes, out_filename) - surgical row replacement in the SAP-accepted
    workbook. Everything except sheet rows and sharedStrings stays byte-identical
    (styles.xml above all), because SAP's parser only understands files shaped
    exactly like its own template."""
    if not SAP_TEMPLATE.exists():
        raise FileNotFoundError("SAP template missing: services/cardconv_sap_template.xlsx")

    today  = date.today()
    who = f"{_export_tag(username)}_" if username else ""
    out_fn = f"for_upload_{who}{today.strftime('%Y-%m-%d')}.xlsx"
    posting_serial = (today - _XLSX_EPOCH).days

    src = zipfile.ZipFile(SAP_TEMPLATE)
    parts = {n: src.read(n) for n in src.namelist()}

    # Shared strings: keep existing entries (ids untouched), append new ones.
    sst_xml = parts["xl/sharedStrings.xml"].decode()
    existing = re.findall(r"<si>.*?</si>", sst_xml, re.S)
    plain = [re.sub(r"<[^>]+>", "", si) for si in existing]
    s_index = {}
    for i, t in enumerate(plain):
        s_index.setdefault(t, i)
    new_si: list = []

    def sref(text: str) -> str:
        text = str(text)
        esc = _xesc(text)
        if esc not in s_index:
            s_index[esc] = len(existing) + len(new_si)
            pre = ' xml:space="preserve"' if text.strip() != text else ""
            new_si.append(f"<si><t{pre}>{esc}</t></si>")
        return str(s_index[esc])

    def cell(col, row, value, kind, style=None):
        s = style or _SAP_COL_STYLE[col]
        if value is None or value == "":
            return f'<c r="{col}{row}" s="{s}"/>'
        if kind == "str":
            return f'<c r="{col}{row}" s="{s}" t="s"><v>{sref(value)}</v></c>'
        return f'<c r="{col}{row}" s="{s}"><v>{value}</v></c>'

    rows_xml = []
    rnum = 2
    for e in entries:
        inv_serial = ""
        if e.get("date"):
            try:
                inv_serial = (date.fromisoformat(e["date"]) - _XLSX_EPOCH).days
            except ValueError:
                pass
        masked, supp = _fmt_card(e.get("account", ""))
        last, first  = _name_parts(e.get("member", ""))
        amount = round(float(e.get("amount") or 0), 2)
        merchant = re.sub(r"\s+", " ", str(e.get("merchant") or "")).strip()
        try:
            gl = int(e.get("gl"))
        except (TypeError, ValueError):
            gl = ""
        ser_s = str(e.get("ser") or "").strip()
        ser_is_text = ser_s.startswith("0") or not ser_s.isdigit()
        purpose = e.get("purpose") or ""
        # Matched rows carry the note on the receipt snapshot; receipt-less
        # rows keep it on the pool transaction itself.
        companions = (e.get("receipt") or {}).get("companions") or e.get("companions")
        if companions:
            purpose = (purpose + " w/ " + companions).strip()

        c = [
            cell("A", rnum, e.get("receipt_type") or FIXED["receipt_type"], "str"),
            cell("B", rnum, FIXED["employee_id"], "num"),
            cell("C", rnum, FIXED["payee"], "str"),
            cell("D", rnum, "", "num"),
            cell("E", rnum, inv_serial, "num"),
            cell("F", rnum, FIXED["domestic"], "str"),
            cell("G", rnum, merchant, "str"),
            cell("H", rnum, posting_serial, "num"),
            cell("I", rnum, gl, "num"),
            (cell("J", rnum, ser_s, "str", _SAP_SER_TEXT_STYLE) if ser_is_text
             else cell("J", rnum, int(ser_s) if ser_s else "", "num")),
            cell("K", rnum, FIXED["currency"], "str"),
            cell("L", rnum, FIXED["tax_code"], "str"),
            cell("M", rnum, "", "num"),
            cell("N", rnum, amount, "num"),
            cell("O", rnum, FIXED["cost_center"], "str"),
            cell("P", rnum, "", "num"),
            cell("Q", rnum, "", "num"),
            cell("R", rnum, purpose, "str"),
            # 'Reason for Cash' — only meaningful on cash lines; free text in
            # the template. Blank means "fill in by hand before upload".
            cell("S", rnum, (e.get("cash_reason") or "") if e.get("cash") else "", "str"),
            cell("T", rnum, masked, "str"),
            cell("U", rnum, last, "str"),
            cell("V", rnum, first, "str"),
            cell("W", rnum, supp, "str"),
            cell("X", rnum, "", "num"),
            cell("Y", rnum, "", "num"),
            cell("Z", rnum, amount, "num"),
        ]
        rows_xml.append(f'<row r="{rnum}" spans="1:26">' + "".join(c) + "</row>")
        rnum += 1

    # Sheet: keep row 1 (header) verbatim, replace all data rows with ours.
    sheet = parts["xl/worksheets/sheet1.xml"].decode()
    header_row = re.search(r'<row r="1".*?</row>', sheet, re.S).group(0)
    sheet = re.sub(r"<sheetData>.*</sheetData>",
                   "<sheetData>" + header_row + "".join(rows_xml) + "</sheetData>",
                   sheet, flags=re.S)
    sheet = re.sub(r'<dimension ref="[^"]*"/>',
                   f'<dimension ref="A1:Z{max(rnum - 1, 1)}"/>', sheet)
    parts["xl/worksheets/sheet1.xml"] = sheet.encode()

    if new_si:
        total = len(existing) + len(new_si)
        sst_xml = re.sub(r'(<sst[^>]*count=")\d+(")', rf"\g<1>{total}\g<2>", sst_xml)
        sst_xml = re.sub(r'(uniqueCount=")\d+(")', rf"\g<1>{total}\g<2>", sst_xml)
        sst_xml = sst_xml.replace("</sst>", "".join(new_si) + "</sst>")
        parts["xl/sharedStrings.xml"] = sst_xml.encode()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as out:
        for name, data in parts.items():
            out.writestr(name, data)
    xlsx_bytes = buf.getvalue()
    (OUT_DIR / out_fn).write_bytes(xlsx_bytes)
    return xlsx_bytes, out_fn


# ── Drive OAuth helpers ───────────────────────────────────────────────────────

def _get_creds(username: str):
    """Load and auto-refresh OAuth credentials for user. Returns None if not connected."""
    _ensure_dirs()
    token_file = TOKENS_DIR / f"{username}.json"
    if not CREDS_FILE.exists() or not token_file.exists():
        return None
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        creds = Credentials.from_authorized_user_file(str(token_file), SCOPES)
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            token_file.write_text(creds.to_json())
        if creds and creds.valid:
            return creds
    except Exception:
        pass
    return None


def _is_drive_connected(username: str) -> bool:
    return _get_creds(username) is not None


def _get_drive_service(username: str):
    from googleapiclient.discovery import build
    creds = _get_creds(username)
    if not creds:
        return None
    return build('drive', 'v3', credentials=creds)


def _get_or_create_folder(service, name: str, parent_id: str = None) -> str:
    """Return Drive folder ID, creating it if it doesn't exist."""
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        q += f" and '{parent_id}' in parents"
    res = service.files().list(q=q, fields='files(id)').execute()
    files = res.get('files', [])
    if files:
        return files[0]['id']
    meta = {'name': name, 'mimeType': 'application/vnd.google-apps.folder'}
    if parent_id:
        meta['parents'] = [parent_id]
    return service.files().create(body=meta, fields='id').execute()['id']


def _get_receipts_folder_id(service, username: str) -> str:
    """Return the Wayfinder/Receipts/ folder id, creating it if needed.
    (The legacy Matched/ subfolder is no longer created or used.)"""
    wayfinder_id = _get_or_create_folder(service, 'Wayfinder')
    receipts_id  = _get_or_create_folder(service, 'Receipts', wayfinder_id)
    # Extra card profiles get their own subfolder (Wayfinder/Receipts/<name>);
    # the default profile keeps the legacy root so nothing moves.
    pid = _active_pid(username)
    if pid:
        profs, _ = _load_profiles(username)
        pname = next((p["name"] for p in profs if p["id"] == pid), pid)
        receipts_id = _get_or_create_folder(service, pname, receipts_id)
    # Cache receipts folder ID for UI link
    meta = _load_drive_meta(username)
    if meta.get('receipts_folder_id') != receipts_id:
        meta['receipts_folder_id'] = receipts_id
        _save_drive_meta(username, meta)
    return receipts_id


def _upload_file_to_drive(username: str, file_bytes: bytes, filename: str,
                           mime_type: str) -> tuple:
    """Upload to Drive under Wayfinder/Receipts/. Returns (file_id, drive_url)."""
    from googleapiclient.http import MediaIoBaseUpload
    service = _get_drive_service(username)
    if not service:
        return None, None
    receipts_id = _get_receipts_folder_id(service, username)
    media  = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type)
    result = service.files().create(
        body={'name': filename, 'parents': [receipts_id]},
        media_body=media,
        fields='id,webViewLink'
    ).execute()
    fid = result.get('id')
    url = result.get('webViewLink') or f'https://drive.google.com/file/d/{fid}/view'
    return fid, url


# ── OCR ───────────────────────────────────────────────────────────────────────

_OCR_PROMPT = (
    'This image may contain ONE OR MULTIPLE receipts (e.g. several small '
    'receipts scanned together on a single page). Identify EACH distinct receipt. '
    'STEP 1 — ANNOTATION SWEEP: before anything else, scan the WHOLE image and '
    'its margins (top edge, corners, over the printed text, beside the total) for '
    'ANY annotation that is NOT part of the original printed receipt: '
    '(a) handwriting — numbers, checkmarks, names, short notes; often faint, '
    'small, slanted, in pen over thermal print, or cut off at the edge; AND '
    '(b) TYPED text overlaid or captioned onto the photo/screenshot (e.g. a '
    'caption like "W/ SEA, D2C" added above the receipt). '
    'Transcribe every annotation you find, exactly as written, into the '
    'handwriting_notes field (null only if there is truly none). '
    'STEP 2 — FIELDS: for EACH receipt look CAREFULLY for handwritten numbers '
    '(e.g. tip amount, final total written by hand on top of the printed receipt). '
    'Inspect tip line, total line, and any margin notes for handwriting. '
    'TIP CHECKBOX RULE: when the receipt prints tip options (e.g. "[ ] 18% … '
    '[ ] 20% (Tip $75.69 Total $454.12)") and one option is marked with a '
    'checkmark, X or circle, the MARKED option\'s Total IS the final amount — '
    'return it as handwritten_amount, and keep printed_amount as the pre-tip '
    'base Amount. '
    'For each receipt extract: '
    '1) date (YYYY-MM-DD; if the receipt only shows a relative date such as "today", '
    'return the literal string "today" — do not guess an absolute date), '
    '1b) time: the transaction time printed on the receipt as "HH:MM" in 24-hour '
    'format (convert AM/PM, e.g. "1:45 PM" -> "13:45"); null if no time is printed, '
    '2) merchant name, '
    '3) printed_amount: the PRINTED/typed total (number only), '
    '4) handwritten_amount: any HAND-WRITTEN final amount including tip (number only, '
    'null if none visible). '
    'Handwritten amount, when present, is the REAL final amount and overrides printed. '
    '5) card_brand: the payment card brand used, normalized to one of "amex", '
    '"visa", "other", or null only if there is NO card information at all (e.g. '
    'cash payment). Determine it from BOTH of these signals: '
    '(a) brand text/logo on the receipt — "AMERICAN EXPRESS"/"AMEX"/"AMX" -> "amex", '
    '"VISA" -> "visa", any other brand (Mastercard, Discover, etc.) -> "other"; AND '
    '(b) the card account number, even when masked (e.g. "XXXX-XXXXXX-X1234", '
    '"************1234", "ending in 1234", "AETC 3759"): use the FIRST visible digit '
    '- a number starting with 3 (15-digit, 4-6-5 grouping) -> "amex"; starting with 4 '
    '-> "visa"; starting with 2 or 5 -> "other". If (a) and (b) disagree, prefer the '
    'explicit brand text. AMEX receipts often show "AMEX", "AETC", or a 15-digit / '
    '3-prefixed card number — treat any of these as "amex". '
    '6) currency: the ISO 4217 code of the amounts on the receipt. Infer from '
    'currency symbols, wording and locale: "$"/"USD" -> "USD"; "₩"/"원"/"KRW"/Korean '
    'receipt text -> "KRW"; "₹"/"Rs"/"INR"/Indian receipt (GST, rupees) -> "INR"; '
    '"HK$"/"HKD"/Hong Kong receipt (Chinese text, HK addresses) -> "HKD"; '
    'other clear signals -> that ISO code (e.g. "EUR", "JPY"). Use "USD" only when '
    'the receipt is clearly US-based or shows "$" with English/US formatting; a bare '
    '"$" on a Hong Kong receipt means HKD, not USD. '
    '7) handwriting_notes: the full transcription from STEP 1 — every annotation '
    '(handwritten or typed overlay) on this receipt as one string '
    '(e.g. "W/ SEA, D2C  ✓20%"), null if none. '
    '8) companions: from the annotations, any note naming who the meal/expense '
    'was shared with — usually "w/ NAME" or "with NAME", handwritten OR typed, '
    'any case ("W/ SEA, D2C" -> "SEA, D2C", "w/sds" -> "sds", "w/ John, Amy" -> '
    '"John, Amy"). The part after "w/" is often initials, a team code or a '
    'nickname (may contain digits like "D2C") — transcribe it literally, do not '
    '"correct" it. Return just the name part without the "w/" prefix, or null '
    'if no such note is visible. '
    'For each receipt, ALSO return its bounding box in the image as bbox: '
    '[ymin, xmin, ymax, xmax] using a 0-1000 normalized coordinate system '
    '(0=top/left, 1000=bottom/right). '
    'Return a JSON ARRAY ONLY, one object per receipt: '
    '[{"date":"YYYY-MM-DD","time":null,"merchant":"name","printed_amount":0.00,"handwritten_amount":null,'
    '"card_brand":"amex","currency":"USD","handwriting_notes":null,"companions":null,'
    '"bbox":[100,50,800,950]}]. '
    'If only one receipt is visible, return an array with a single element.'
)


def _normalize_ocr(result: dict) -> dict:
    """Derive the final `amount` (handwritten priority) and coerce numeric fields.

    Keeps backward compat with the old single-`amount` shape: if a model still returns
    only `amount`, it is treated as the printed total.
    """
    if not result:
        return result

    def _num(v):
        if v is None:
            return None
        try:
            return round(float(v), 2)
        except (ValueError, TypeError):
            return None

    printed = _num(result.get("printed_amount"))
    handwritten = _num(result.get("handwritten_amount"))
    legacy = _num(result.get("amount"))
    if printed is None and legacy is not None:
        printed = legacy
    result["printed_amount"] = printed
    result["handwritten_amount"] = handwritten
    # Handwritten (tip-inclusive) total wins; fall back to printed.
    result["amount"] = handwritten if handwritten is not None else printed
    # bbox: [ymin, xmin, ymax, xmax] in 0-1000 normalized coords (None if absent/invalid).
    result["bbox"] = _coerce_bbox(result.get("bbox"))
    result["time"] = _coerce_time(result.get("time"))
    result["card_brand"] = _coerce_card_brand(result.get("card_brand"))
    result["currency"] = _coerce_currency(result.get("currency"))
    hn = result.get("handwriting_notes")
    result["handwriting_notes"] = hn.strip()[:200] if isinstance(hn, str) and hn.strip() else None
    result["companions"] = _coerce_companions(result.get("companions"))
    # Deterministic fallback: the model often transcribes "w/ NAME" into the
    # handwriting sweep but misses the dedicated companions field.
    if not result["companions"] and result["handwriting_notes"]:
        m = re.search(r"\b(?:w/|with\s)\s*([A-Za-z가-힣][A-Za-z0-9가-힣 ,.&/-]*)",
                      result["handwriting_notes"], re.I)
        if m:
            result["companions"] = _coerce_companions(m.group(1))
    return result


def _coerce_companions(v):
    """Handwritten 'w/ NAME' note → the name part ('sds', 'John, Amy'), else None."""
    if not v or not isinstance(v, str):
        return None
    s = re.sub(r"^\s*(w/|with)\s*", "", v.strip(), flags=re.I).strip(" ,;")
    return s[:60] or None


def _coerce_time(v):
    """Normalize an OCR time to 'HH:MM' 24h, or None. Accepts HH:MM(:SS) and AM/PM."""
    if not isinstance(v, str):
        return None
    m = re.match(r"\s*(\d{1,2}):(\d{2})(?::\d{2})?\s*([AaPp][Mm])?\s*$", v)
    if not m:
        return None
    h, mnt = int(m.group(1)), int(m.group(2))
    ap = (m.group(3) or "").lower()
    if ap == "pm" and h < 12:
        h += 12
    if ap == "am" and h == 12:
        h = 0
    if not (0 <= h <= 23 and 0 <= mnt <= 59):
        return None
    return f"{h:02d}:{mnt:02d}"


def _coerce_currency(v):
    """Normalize a model's currency string to an ISO 4217 code (None if unknown)."""
    if not v:
        return None
    s = str(v).strip().upper()
    if s in ("NULL", "NONE", "UNKNOWN", ""):
        return None
    aliases = {"₩": "KRW", "원": "KRW", "WON": "KRW", "₹": "INR", "RS": "INR",
               "RUPEE": "INR", "RUPEES": "INR", "$": "USD", "US$": "USD", "HK$": "HKD"}
    s = aliases.get(s, s)
    return s if (len(s) == 3 and s.isalpha()) else None


# ── FX conversion (foreign-currency receipts → USD estimate) ─────────────────
# Card networks settle at their own rate (spread + AMEX FX fee ~2.5%), so the
# USD estimate is matched against statement amounts with a tolerance band.

FX_TOLERANCE = 0.05          # ±5% band around the ECB reference conversion
_FX_CACHE_FILE = DATA_DIR / "fx_cache.json"
_FX_FALLBACK = {"KRW": 1510.0, "INR": 94.0, "HKD": 7.8, "EUR": 0.86, "JPY": 146.0}  # offline approx (2026-07; HKD is USD-pegged)


def _fx_rate(currency: str, date_str: str = None):
    """Units of `currency` per 1 USD on `date_str` (YYYY-MM-DD, default latest).

    ECB reference rates via frankfurter.app (free, no key), cached on disk.
    Weekends/holidays resolve to the previous business day server-side.
    Returns None for unknown currencies with no fallback.
    """
    if not currency or currency == "USD":
        return 1.0
    key = f"{currency}:{date_str or 'latest'}"
    cache = {}
    try:
        cache = json.loads(_FX_CACHE_FILE.read_text())
        if key in cache:
            return cache[key]
    except Exception:
        cache = {}
    try:
        import urllib.request
        url = f"https://api.frankfurter.dev/v1/{date_str or 'latest'}?base=USD&symbols={currency}"
        # NB: frankfurter.dev returns 403 for the default Python-urllib UA.
        req = urllib.request.Request(url, headers={"User-Agent": "wayfinder-cardconv/1.0"})
        with urllib.request.urlopen(req, timeout=6) as resp:
            rate = json.load(resp)["rates"][currency]
        cache[key] = rate
        _ensure_dirs()
        _FX_CACHE_FILE.write_text(json.dumps(cache))
        return rate
    except Exception:
        return _FX_FALLBACK.get(currency)


def _fx_usd_estimate(amount, currency: str, date_str: str = None):
    """(usd_estimate, rate) for a foreign amount; (None, None) if not convertible."""
    if amount is None or not currency or currency == "USD":
        return None, None
    rate = _fx_rate(currency, date_str)
    if not rate:
        return None, None
    return round(float(amount) / rate, 2), rate


def _coerce_card_brand(v):
    """Normalize a model's card-brand string to 'amex'/'visa'/'other' or None."""
    if not v:
        return None
    s = str(v).strip().lower()
    if not s or s in ("null", "none", "unknown"):
        return None
    if "amex" in s or "american express" in s or s == "amx":
        return "amex"
    if "visa" in s:
        return "visa"
    return "other"


def _coerce_bbox(v):
    """Validate a model bbox into [ymin, xmin, ymax, xmax] ints, or None."""
    if not isinstance(v, (list, tuple)) or len(v) != 4:
        return None
    try:
        box = [int(round(float(x))) for x in v]
    except (ValueError, TypeError):
        return None
    # Clamp to the 0-1000 normalized range and require a positive area.
    box = [max(0, min(1000, c)) for c in box]
    ymin, xmin, ymax, xmax = box
    if ymax <= ymin or xmax <= xmin:
        return None
    return box


def _extract_ocr_list(text: str) -> list:
    """Parse a model OCR response into a list of normalized receipt dicts.

    Accepts either a JSON ARRAY (multi-receipt, current prompt) or a single
    JSON OBJECT (back-compat). Returns [] when nothing parseable is found.
    """
    if not text:
        return []
    text = text.strip()
    parsed = None
    # Prefer a top-level array (one image may hold multiple receipts).
    m = re.search(r'\[.*\]', text, re.DOTALL)
    if m:
        try:
            parsed = json.loads(m.group(0))
        except Exception:
            parsed = None
    if parsed is None:  # fall back to a single object
        m = re.search(r'\{.*\}', text, re.DOTALL)
        if m:
            try:
                parsed = json.loads(m.group(0))
            except Exception:
                parsed = None
    if parsed is None:
        return []
    if isinstance(parsed, dict):
        parsed = [parsed]
    return [_normalize_ocr(item) for item in parsed if isinstance(item, dict)]


def _ocr_receipt(file_bytes: bytes, mime_type: str) -> list:
    """OCR receipt(s) using Claude Vision. Returns a list of receipt dicts (may be empty)."""
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return []
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        b64 = base64.standard_b64encode(file_bytes).decode()
        if mime_type == 'application/pdf':
            block = {
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": b64}
            }
        else:
            block = {
                "type": "image",
                "source": {"type": "base64", "media_type": mime_type, "data": b64}
            }
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,  # room for multiple receipts in one image
            messages=[{"role": "user", "content": [
                block,
                {"type": "text", "text": _OCR_PROMPT}
            ]}]
        )
        results = _extract_ocr_list(resp.content[0].text)
        for r in results:
            r["_model"] = "Claude"
        return results
    except Exception:
        pass
    return []


# Configurable via GEMINI_OCR_MODEL (.env). gemini-2.5-flash: fast/cheap, strong on
# KR+EN receipts. Bump to gemini-2.5-pro if accuracy issues arise (~10x cost).
_DEFAULT_GEMINI_OCR_MODEL = "gemini-2.5-flash"


def _ocr_receipt_gemini(file_bytes: bytes, mime_type: str) -> list:
    """OCR receipt(s) using Gemini Vision. Returns a list of receipt dicts (may be empty)."""
    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        return []
    try:
        # New unified SDK (google-genai). Old google.generativeai is deprecated.
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=api_key)
        model_name = os.environ.get('GEMINI_OCR_MODEL', _DEFAULT_GEMINI_OCR_MODEL)
        # Inline bytes handle both images and PDFs (<20MB)
        resp = client.models.generate_content(
            model=model_name,
            contents=[
                types.Part.from_bytes(data=file_bytes, mime_type=mime_type),
                _OCR_PROMPT,
            ],
        )
        results = _extract_ocr_list(resp.text or "")
        for r in results:
            r["_model"] = "Gemini"
        return results
    except Exception:
        pass
    return []


def _ocr_receipt_auto(file_bytes: bytes, mime_type: str) -> list:
    """Primary OCR: Gemini first, fall back to Claude. Returns a list of receipt dicts.

    One image may contain multiple receipts, so callers must iterate the result.
    """
    results = _ocr_receipt_gemini(file_bytes, mime_type)
    if results and any(r.get("amount") is not None for r in results):
        return results
    claude = _ocr_receipt(file_bytes, mime_type)
    return claude or results


def _sub_entry_id(file_id: str, sub_index: int) -> str:
    """Stable ledger entry id for the Nth receipt found in one image."""
    base = (file_id or uuid.uuid4().hex)[:8]
    return f"rcpt_{base}_{sub_index}"


_ISO_DATE_RE    = re.compile(r'^\d{4}-\d{2}-\d{2}$')
_TODAY_TOKEN_RE = re.compile(r'today|오늘', re.IGNORECASE)


def _ocr_entry_fields(ocr: dict, upload_date: str = None) -> dict:
    """Map a normalized OCR dict to the ledger entry's ocr_* fields.

    Some receipts carry only a relative date (they literally print "today");
    `upload_date` (YYYY-MM-DD) fills those in — the first upload date is the
    closest absolute anchor we have. Any other non-ISO junk becomes None so
    the CSV-match backfill can supply the date instead."""
    ocr_date = ocr.get("date")
    if ocr_date == "YYYY-MM-DD":  # treat placeholder as missing
        ocr_date = None
    if isinstance(ocr_date, str) and not _ISO_DATE_RE.match(ocr_date):
        ocr_date = upload_date if _TODAY_TOKEN_RE.search(ocr_date) else None
    has_ocr = ocr.get("amount") is not None
    currency = ocr.get("currency")
    usd_est, fx_rate = _fx_usd_estimate(ocr.get("amount"), currency, ocr_date)
    return {
        "ocr_status":             "done" if has_ocr else "failed",
        "ocr_date":               ocr_date,
        "ocr_time":               ocr.get("time"),
        "ocr_amount":             ocr.get("amount"),
        "ocr_printed_amount":     ocr.get("printed_amount"),
        "ocr_handwritten_amount": ocr.get("handwritten_amount"),
        "ocr_merchant":           ocr.get("merchant"),
        "ocr_companions":         ocr.get("companions"),
        "ocr_handwriting":        ocr.get("handwriting_notes"),
        "ocr_model":              ocr.get("_model"),
        "ocr_bbox":               ocr.get("bbox"),
        "card_brand":             ocr.get("card_brand"),
        "ocr_currency":           currency,
        "usd_estimate":           usd_est,
        "fx_rate":                fx_rate,
    }


# ── Multipart file parser ─────────────────────────────────────────────────────

_MIME_MAP = {
    'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
    'png': 'image/png',  'pdf': 'application/pdf',
    'gif': 'image/gif',  'webp': 'image/webp',
}


def _parse_multipart_files(raw_handler) -> list:
    """Parse multipart body; returns list of (filename, bytes, mime_type)."""
    ct = raw_handler.headers.get("Content-Type", "")
    m  = re.search(r'boundary=([^\s;]+)', ct)
    if not m:
        return []
    boundary = ("--" + m.group(1)).encode()
    length   = int(raw_handler.headers.get("Content-Length", 0))
    data     = raw_handler.rfile.read(length)
    files    = []
    for part in data.split(boundary):
        if b'filename="' not in part:
            continue
        fn_m = re.search(rb'filename="([^"]+)"', part)
        if not fn_m:
            continue
        filename = fn_m.group(1).decode('utf-8', errors='replace')
        hdr_end  = part.find(b"\r\n\r\n")
        if hdr_end == -1:
            continue
        content = part[hdr_end + 4:]
        if content.endswith(b"\r\n"):
            content = content[:-2]
        ext  = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        mime = _MIME_MAP.get(ext, 'application/octet-stream')
        if content:
            files.append((filename, content, mime))
    return files


# ── Conversion logic ──────────────────────────────────────────────────────────

def _classify(merchant: str, rules: list):
    m = merchant.upper()
    for r in rules:
        if r["kw"].upper() in m:
            return r["gl"], r["ser"], r["purpose"]
    return None, None, None


def _fmt_card(acct: str):
    acct = re.sub(r'\D', '', acct)
    if len(acct) >= 8:
        masked = acct[:4] + "********" + acct[-4:]
        supp   = f"{acct[:4]}-{acct[4:10]}-{acct[10:]}" if len(acct) == 15 else acct
    else:
        masked = acct; supp = acct
    return masked, supp


def _name_parts(full: str):
    p = full.strip().split()
    return (p[-1], " ".join(p[:-1])) if len(p) >= 2 else (full, "")


def _parse_date(s: str):
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            pass
    return None


# ── Ledger ──────────────────────────────────────────────────────────────────

_SUPPORTED_MIME = {'image/jpeg', 'image/png', 'application/pdf', 'image/gif', 'image/webp'}


def _run_batch_ocr(username: str) -> dict:
    """Scan Drive, OCR new/pending files, update ledger. Returns stats dict."""
    service = _get_drive_service(username)
    if not service:
        return {"error": "drive not connected"}
    try:
        receipts_id = _get_receipts_folder_id(service, username)
        results = service.files().list(
            q=(f"'{receipts_id}' in parents "
               f"and trashed=false and mimeType!='application/vnd.google-apps.folder'"),
            fields="files(id,name,mimeType,webViewLink)"
        ).execute()
        drive_files = results.get('files', [])

        ledger  = _load_ledger(username)
        entries = ledger["entries"]

        # Skip files already confirmed in the ledger.
        done_fids = {e.get("file_id") for e in entries
                     if (e.get("multi_ocr") and e.get("ocr_amount") is not None)
                     or e.get("matched") or e.get("completed")}
        # Also skip files already waiting in the staging queue (accumulated from
        # previous runs the user hasn't reviewed yet).
        staging = _load_ocr_staging(username)
        staged_fids = {e.get("file_id") for e in staging.get("entries", [])}
        # And skip discarded tombstones — without this, every scheduled batch
        # run re-staged receipts the user had already discarded (2026-07-24).
        skip_fids = done_fids | staged_fids | set(_load_discarded_fids(username))

        new_staged = []
        now_disp = datetime.now().strftime("%Y-%m-%d %H:%M")
        now_iso  = datetime.now().isoformat()

        for f in drive_files:
            fid  = f.get('id')
            mime = f.get('mimeType', '')
            if mime not in _SUPPORTED_MIME:
                continue
            if fid in skip_fids:
                continue
            content  = service.files().get_media(fileId=fid).execute()
            ocr_list = _ocr_receipt_auto(content, mime) or [{}]
            url      = f.get('webViewLink') or f'https://drive.google.com/file/d/{fid}/view'
            for sub_index, ocr in enumerate(ocr_list):
                entry = {
                    "id":           _sub_entry_id(fid, sub_index),
                    "file_id":      fid,
                    "sub_index":    sub_index,
                    "filename":     f.get('name', ''),
                    "drive_url":    url,
                    "mime_type":    mime,
                    "match_status": "pending_match",
                    "multi_ocr":    True,
                    "uploaded_at":  now_disp,
                    "synced_at":    now_iso,
                }
                entry.update(_ocr_entry_fields(ocr, upload_date=now_iso[:10]))
                new_staged.append(entry)

        # Accumulate into staging queue — never replace, always append.
        if new_staged:
            staging.setdefault("entries", []).extend(new_staged)
            staging["staged_at"] = now_iso
            _save_ocr_staging(username, staging)

        ledger["last_batch_at"] = now_iso
        _save_ledger(username, ledger)

        if new_staged:
            total_pending = len(staging.get("entries", []))
            _send_push_notification(
                username,
                title="🧾 New receipts ready to review",
                body=f"{len(new_staged)} new receipt{'s' if len(new_staged) != 1 else ''} synced from Drive — {total_pending} total pending review.",
                url="/cardconv/ledger",
            )
        return {"staged": len(new_staged), "total_pending": len(staging.get("entries", []))}
    except Exception as e:
        return {"error": str(e)}


def _handle_batch_run(username: str, ctx):
    """POST /cardconv/batch/run — X-Batch-Secret authenticated batch OCR trigger."""
    secret = os.environ.get("CARDCONV_BATCH_SECRET", "")
    headers = (ctx or {}).get("headers")
    provided = headers.get("X-Batch-Secret", "") if headers else ""
    if not secret or provided != secret:
        return ("json", {"error": "unauthorized"}, 401)
    # sweep every card profile, not just the active one
    profs, _ = _load_profiles(username)
    results = []
    for p in profs:
        _PROFILE_CTX.force_pid = p["id"]
        try:
            r = _run_batch_ocr(username)
        finally:
            _PROFILE_CTX.force_pid = None
        r["profile"] = p["name"]
        results.append(r)
    if len(results) == 1:
        return ("json", results[0])
    return ("json", {"profiles": results})


def _apply_ledger_filters(entries: list, status: str, dfrom: str, dto: str,
                          card_brand: str = "", usage: str = "",
                          completed: str = "hide", merchant: str = "",
                          sort: str = "date", settle: str = "all") -> list:
    """Filter + sort ledger entries by status, OCR-date range, card brand, usage,
    completion state and merchant text.

    Shared by the JSON API, the PDF export and the xlsx export so all honor
    identical filters. Date filters keep entries without an OCR date always
    visible. `completed` is one of: "hide" (default — exclude completed),
    "only" (completed only), "all" (include both). `merchant` is a
    case-insensitive substring match. `sort` is "date" (desc, default) or
    "merchant" (A→Z).
    """
    filtered = entries
    if status == "unmatched_any":
        # Everything not yet matched — explicit unmatched plus pending_match.
        filtered = [e for e in filtered if e.get("match_status") in ("unmatched", "pending_match")]
    elif status and status != "all":
        filtered = [e for e in filtered if e.get("match_status") == status]
    if dfrom:
        filtered = [e for e in filtered if not e.get("ocr_date") or e["ocr_date"] >= dfrom]
    if dto:
        filtered = [e for e in filtered if not e.get("ocr_date") or e["ocr_date"] <= dto]
    if card_brand and card_brand != "all":
        if card_brand == "unknown":
            filtered = [e for e in filtered if not e.get("card_brand")]
        else:
            filtered = [e for e in filtered if e.get("card_brand") == card_brand]
    if usage and usage != "all":
        filtered = [e for e in filtered if (e.get("usage") or "Regular") == usage]
    if completed == "only":
        filtered = [e for e in filtered if e.get("completed")]
    elif completed != "all":  # "hide" (default)
        filtered = [e for e in filtered if not e.get("completed")]
    if merchant:
        m = merchant.strip().lower()
        filtered = [e for e in filtered if m in (e.get("ocr_merchant") or "").lower()]
    if settle and settle != "all":
        # "open" also covers receipts not linked to any transaction yet.
        if settle == "open":
            filtered = [e for e in filtered if (e.get("settle_status") or "open") == "open"]
        else:
            filtered = [e for e in filtered if e.get("settle_status") == settle]
    if sort == "merchant":
        return sorted(filtered, key=lambda e: (e.get("ocr_merchant") or "￿").lower())
    return sorted(
        filtered,
        key=lambda e: e.get("ocr_date") or e.get("uploaded_at") or "",
        reverse=True,
    )


def _times_close(a, b, tol_min: int = 5) -> bool:
    """True when either printed time is unknown or they differ ≤ tol_min minutes.

    Copies of the same receipt print the same transaction time; two separate
    same-amount purchases at one merchant rarely do. A small tolerance absorbs
    OCR misreads of a digit."""
    if not a or not b:
        return True
    try:
        ah, am = map(int, a.split(":"))
        bh, bm = map(int, b.split(":"))
    except ValueError:
        return True
    return abs((ah * 60 + am) - (bh * 60 + bm)) <= tol_min


def _mark_duplicates(entries: list):
    """Flag likely-duplicate receipts in-place.

    Two receipts are duplicates when their final amount and merchant match AND
    their dates are compatible — equal, or at least one side missing (OCR often
    fails to read a date on one copy of the same receipt). Date is deliberately
    NOT part of the bucket key so a date-less entry still groups with its dated
    twin; multi-receipt sub-entries (different file_id, same OCR values) group
    too since file_id is ignored. Within a group the keeper / head is chosen by
    match_status priority (matched > unmatched > pending_match), then a present
    date; the rest are flagged for easy bulk-deletion.
    Two refinements: entries the user marked as separate purchases
    (dup_exempt) never enter a group, and printed transaction times split a
    group — same-image copies print the same time, while two same-amount
    purchases at one merchant usually don't (see _times_close).
    Sets e['dup'] (bool), e['dup_keep'] (bool) and e['dup_group_id'] (str|None)
    on each entry. dup_group_id lets the UI collapse a group into one row.
    """
    for e in entries:
        e["dup"] = False
        e["dup_keep"] = False
        e["dup_group_id"] = None

    # Bucket by (amount, merchant); date/time handled per-pair below.
    buckets = {}
    for e in entries:
        amt = e.get("ocr_amount")
        if amt is None or e.get("dup_exempt"):
            continue
        merch = (e.get("ocr_merchant") or "").strip().lower()
        buckets.setdefault((round(amt, 2), merch), []).append(e)

    gi = 0
    for bucket in buckets.values():
        if len(bucket) < 2:
            continue
        # Partition into date-compatible groups (equal date, or a missing date
        # absorbed into the first dated group it meets).
        used = [False] * len(bucket)
        for i in range(len(bucket)):
            if used[i]:
                continue
            group, used[i] = [bucket[i]], True
            anchor = bucket[i].get("ocr_date")
            anchor_t = bucket[i].get("ocr_time")
            for j in range(i + 1, len(bucket)):
                if used[j]:
                    continue
                dj = bucket[j].get("ocr_date")
                if ((anchor is None or dj is None or anchor == dj)
                        and _times_close(anchor_t, bucket[j].get("ocr_time"))):
                    group.append(bucket[j])
                    used[j] = True
                    if anchor is None:
                        anchor = dj  # lock onto the first known date
                    if anchor_t is None:
                        anchor_t = bucket[j].get("ocr_time")
            if len(group) < 2:
                continue
            gid = f"dg_{gi}"
            gi += 1
            # Keeper / group head: status priority first
            # (matched > unmatched > pending_match), then a present date,
            # then the earliest uploaded_at / id for a stable tiebreak.
            status_rank = {"matched": 0, "unmatched": 1, "pending_match": 2}
            group.sort(key=lambda e: (
                status_rank.get(e.get("match_status"), 3),
                e.get("ocr_date") is None,
                str(e.get("uploaded_at") or ""),
                str(e.get("id") or ""),
            ))
            for k, e in enumerate(group):
                e["dup"] = True
                e["dup_keep"] = (k == 0)
                e["dup_group_id"] = gid


def _handle_ledger_delete(username: str, body: dict):
    """POST /cardconv/ledger/delete — remove entries by id from the ledger.

    Body: {"ids": [...], "also_drive": bool}. When also_drive is true, the
    corresponding Drive originals are moved to the trash (best-effort).
    """
    raw = body.get("ids", [])
    ids = {str(i) for i in (raw if isinstance(raw, list) else [raw]) if i}
    if not ids:
        return ("json", {"error": "no ids"}, 400)
    also_drive = bool(body.get("also_drive"))

    ledger = _load_ledger(username)
    removed_entries = [e for e in ledger["entries"] if e.get("id") in ids]
    ledger["entries"] = [e for e in ledger["entries"] if e.get("id") not in ids]
    _save_ledger(username, ledger)

    # Tombstone files that no longer have any ledger/staging reference —
    # otherwise Drive syncs resurrect deleted receipts (drive.file cannot
    # trash user-dropped files, so the file typically stays in the folder).
    live_fids = {e.get("file_id") for e in ledger["entries"] if e.get("file_id")}
    staged_fids = {e.get("file_id") for e in _load_ocr_staging(username).get("entries", [])}
    for e in removed_entries:
        fid = e.get("file_id")
        if fid and fid not in live_fids and fid not in staged_fids:
            _mark_discarded_fid(username, fid, e.get("filename"))

    # Deleting a matched receipt leaves its transaction pointing at a ghost —
    # rematch immediately so the tx unlinks and can pair with a surviving copy
    # (deleting the matched duplicate is the normal dup-cleanup flow).
    if any(e.get("matched") for e in removed_entries):
        _rematch_pool(username)

    trashed = 0
    if also_drive and removed_entries:
        service = _get_drive_service(username)
        if service:
            for e in removed_entries:
                fid = e.get("file_id")
                if not fid:
                    continue
                try:
                    service.files().update(fileId=fid, body={"trashed": True}).execute()
                    trashed += 1
                except Exception:
                    pass  # best-effort; ledger removal already succeeded
    return ("json", {"ok": True, "removed": len(removed_entries), "trashed": trashed})


def _revoke_drive_token(username: str) -> None:
    """Best-effort: ask Google to revoke the stored OAuth token before deletion."""
    token_file = TOKENS_DIR / f"{username}.json"
    if not token_file.exists():
        return
    try:
        import urllib.parse as _up, urllib.request as _ur
        data = json.loads(token_file.read_text())
        tok = data.get("refresh_token") or data.get("token")
        if not tok:
            return
        req = _ur.Request(
            "https://oauth2.googleapis.com/revoke",
            data=_up.urlencode({"token": tok}).encode(),
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )
        _ur.urlopen(req, timeout=5)
    except Exception:
        pass


def purge_user_data(username: str) -> None:
    """Delete ALL cardconv data for a user (Drive token + connection, ledger,
    receipts, uploads, settings, push subs). Called when an admin deletes the
    account so nothing — especially the Google Drive token — is left behind."""
    import shutil
    _revoke_drive_token(username)
    _PROFILE_CTX.force_pid = ""   # resolve helpers to the legacy default paths
    for f in [
        TOKENS_DIR / f"{username}.json",
        _drive_meta_file(username),
        _receipts_file(username),
        _review_file(username),
        _ocr_staging_file(username),
        _user_settings_file(username),
        _uploads_index_file(username),
        _push_subs_file(username),
    ]:
        try:
            Path(f).unlink()
        except FileNotFoundError:
            pass
        except Exception:
            pass
    shutil.rmtree(_uploads_dir(username), ignore_errors=True)
    _PROFILE_CTX.force_pid = None
    # profile-scoped files/dirs (…_{user}@{pid}.json) + the profile registry
    for pat in (f"*_{username}@*", f"profiles_{username}.json"):
        for f in DATA_DIR.glob(pat):
            try:
                shutil.rmtree(f, ignore_errors=True) if f.is_dir() else f.unlink()
            except Exception:
                pass


def _apply_receipt_completion(username: str, ids: set, completed: bool) -> dict:
    """Set the receipt-level completed flag on `ids`.

    Shared by every path that changes settlement state (Complete button, Ledger
    status action, Review mirror). Settlement state lives entirely in app data —
    the app never creates folders or moves files on Drive (policy 2026-07-22;
    the old Completed-folder archiving only ever worked for app-uploaded files
    anyway). `moved`/`attempted` stay in the response for the UI contract but
    are always zero now.
    """
    ledger = _load_ledger(username)
    now = datetime.now().isoformat()
    touched = []
    for e in ledger["entries"]:
        if e.get("id") in ids:
            e["completed"] = completed
            e["completed_at"] = now if completed else None
            touched.append(e)
    _save_ledger(username, ledger)
    return {"count": len(touched), "moved": 0, "attempted": 0}


def _set_linked_tx_status(username: str, receipt_ids: set, status: str) -> int:
    """Set the settlement status on transactions matched to `receipt_ids`.

    The other half of the receipt/transaction sync: unmatched receipts carry no
    transaction and are skipped. Returns how many transactions were touched."""
    pool = _load_tx_pool(username)
    now = datetime.now().isoformat()
    touched = 0
    for t in pool["entries"]:
        if (t.get("receipt") or {}).get("id") in receipt_ids:
            t["status"] = status
            t["completed_at"] = now if status == "completed" else None
            touched += 1
    if touched:
        _save_tx_pool(username, pool)
    return touched


def _reconcile_settle_status(username: str) -> int:
    """Heal receipt.completed ↔ tx.status disagreements left by pre-sync eras
    (the ✓ Complete button and Review statuses didn't mirror each other before
    2026-07-14). Runs lazily on Ledger/Review loads; a no-op when consistent.

    Per mismatched pair: in_progress wins (explicit workflow state → receipt
    un-completed); otherwise the completed side wins (archived receipt closes
    a stale open tx; a completed tx archives its active receipt)."""
    pool = _load_tx_pool(username)
    ledger = _load_ledger(username)
    comp = {e.get("id"): bool(e.get("completed")) for e in ledger["entries"]}
    now = datetime.now().isoformat()
    to_complete, to_uncomplete = set(), set()
    tx_fixed = 0
    for t in pool["entries"]:
        rid = (t.get("receipt") or {}).get("id")
        if rid not in comp:
            continue
        status = t.get("status") or "open"
        if (status == "completed") == comp[rid]:
            continue
        if status == "in_progress":
            to_uncomplete.add(rid)
        elif status == "open":
            t["status"] = "completed"
            t["completed_at"] = now
            tx_fixed += 1
        else:
            to_complete.add(rid)
    if tx_fixed:
        _save_tx_pool(username, pool)
    if to_complete:
        _apply_receipt_completion(username, to_complete, True)
    if to_uncomplete:
        _apply_receipt_completion(username, to_uncomplete, False)
    return tx_fixed + len(to_complete) + len(to_uncomplete)


def _handle_ledger_complete(username: str, body: dict):
    """POST /cardconv/ledger/complete — mark entries complete (or undo).

    Body: {"ids": [...], "undo": bool}. Completing flags entries, moves their
    Drive originals to the Completed folder (best-effort; the ledger flag is
    authoritative) and mirrors the settlement status onto linked transactions
    so Review agrees."""
    raw = body.get("ids", [])
    ids = {str(i) for i in (raw if isinstance(raw, list) else [raw]) if i}
    if not ids:
        return ("json", {"error": "no ids"}, 400)
    undo = bool(body.get("undo"))
    res = _apply_receipt_completion(username, ids, not undo)
    _set_linked_tx_status(username, ids, "open" if undo else "completed")
    return ("json", {"ok": True, **res})


def _parse_filter_params(query: dict) -> dict:
    """Extract the shared Ledger filter params from a query dict.

    Used by the JSON API, PDF export and xlsx export so all interpret the
    status/date/card-brand/usage/completed filters identically.
    """
    def _q(key, default):
        return (query.get(key, [default]) or [default])[0]
    return {
        "status":     _q("status", "all"),
        "dfrom":      _q("from", ""),
        "dto":        _q("to", ""),
        "card_brand": _q("card_brand", "all"),
        "usage":      _q("usage", "all"),
        "completed":  _q("completed", "hide"),
        "merchant":   _q("merchant", ""),
        "sort":       _q("sort", "date"),
        "settle":     _q("settle", "all"),
    }


def _annotate_settle_status(username: str, entries: list):
    """Attach the linked transaction's settlement status (open / in_progress /
    completed) to each matched receipt, live from the pool — Review changes
    reflect in the Ledger immediately without duplicated state."""
    pool = _load_tx_pool(username)
    by_receipt = {}
    for t in pool.get("entries", []):
        rid = (t.get("receipt") or {}).get("id")
        if rid:
            by_receipt[rid] = t.get("status") or "open"
    for e in entries:
        e["settle_status"] = by_receipt.get(e.get("id"))
    return entries


def _handle_ledger_api(username: str, query: dict):
    """GET /cardconv/ledger/api — filtered JSON data."""
    _reconcile_settle_status(username)   # self-heal pre-sync mismatches (no-op when consistent)
    _heal_orphan_matches(username)       # self-heal one-sided match links (no-op when consistent)
    ledger  = _load_ledger(username)
    entries = ledger["entries"]
    f = _parse_filter_params(query)
    try:
        page = max(1, int((query.get("page", ["1"]) or ["1"])[0]))
    except ValueError:
        page = 1
    # limit<=0 (the default) returns everything — the Ledger shows all rows at once.
    try:
        limit = int((query.get("limit", ["0"]) or ["0"])[0])
    except ValueError:
        limit = 0

    _annotate_settle_status(username, entries)
    filtered = _apply_ledger_filters(entries, f["status"], f["dfrom"], f["dto"],
                                     f["card_brand"], f["usage"], f["completed"],
                                     f["merchant"], f["sort"], f["settle"])
    _mark_duplicates(filtered)

    # Stat cards are view switchers — their numbers must not change with the
    # view dimensions they control (status/settle/completed). Compute stats on
    # a base that keeps only the user filters (date/card/usage/merchant).
    stats_base = _apply_ledger_filters(entries, "all", f["dfrom"], f["dto"],
                                       f["card_brand"], f["usage"], "hide",
                                       f["merchant"], f["sort"], "all")
    stats   = _ledger_stats(stats_base)
    total_f = len(filtered)
    if limit > 0:
        pages = max(1, (total_f + limit - 1) // limit)
        start = (page - 1) * limit
        page_entries = filtered[start:start + limit]
    else:
        pages, page, page_entries = 1, 1, filtered
    # Distinct usage tags across the whole ledger (for the filter dropdown), and
    # a completed count so the UI can surface how many are archived.
    usages = sorted({(e.get("usage") or "Regular") for e in entries})
    completed_n = sum(1 for e in entries if e.get("completed"))
    inprog_n = sum(1 for e in entries if e.get("settle_status") == "in_progress")
    return ("json", {
        "total":       stats["total"],
        "matched":     stats["matched"],
        "unmatched":   stats["unmatched"],
        "pending_match": stats["pending_match"],
        "completed":   completed_n,
        "in_progress": inprog_n,
        "usages":      usages,
        "page":        page,
        "pages":       pages,
        "last_synced": ledger.get("last_batch_at"),
        "entries":     page_entries,
    })


def _handle_status_change(username: str, entry_id: str, body: dict):
    """POST /cardconv/ledger/<id>/status — manual status override."""
    raw = body.get("status", "")
    status = (raw[0] if isinstance(raw, list) else str(raw)).strip()
    if status not in ("matched", "unmatched", "pending_match"):
        return ("json", {"error": "invalid status"}, 400)
    ledger = _load_ledger(username)
    for e in ledger["entries"]:
        if e.get("id") == entry_id:
            e["match_status"] = status
            e["matched"] = (status == "matched")
            if status == "matched":
                e["matched_at"] = datetime.now().isoformat()
                # Matched ⇒ AMEX unconditionally: the statement being matched
                # against IS the AMEX statement (cash can never match).
                e["card_brand"] = "amex"
            else:
                e["usd_settled"] = None  # settled figure came from the link
            _save_ledger(username, ledger)
            return ("json", {"ok": True})
    return ("json", {"error": "not found"}, 404)


def _handle_rematch(username: str, entry_id: str):
    """POST /cardconv/ledger/<id>/rematch — re-try CSV matching for a pending/unmatched entry."""
    ledger = _load_ledger(username)
    entries = ledger["entries"]
    entry = next((e for e in entries if e.get("id") == entry_id), None)
    if not entry:
        return ("json", {"error": "not found"}, 404)

    rdate = entry.get("ocr_date")
    if rdate == "YYYY-MM-DD":
        rdate = None
    # Handwritten amount takes priority over printed as the final match target.
    hw = entry.get("ocr_handwritten_amount")
    pr = entry.get("ocr_printed_amount") or entry.get("ocr_amount")
    raw_amount = hw if hw is not None else pr
    if raw_amount is None:
        return ("json", {"error": "no OCR amount to match against"}, 400)
    try:
        ramount = round(float(raw_amount), 2)
    except (ValueError, TypeError):
        return ("json", {"error": "invalid amount"}, 400)

    # Foreign-currency receipt: compare statement USD amounts against the USD
    # estimate with a ±FX_TOLERANCE band instead of exact cents.
    fx_cur = entry.get("ocr_currency")
    fx_usd = None
    if fx_cur and fx_cur != "USD":
        fx_usd = entry.get("usd_estimate")
        if fx_usd is None:
            fx_usd, _fxr = _fx_usd_estimate(ramount, fx_cur, rdate)
            if fx_usd is not None:
                entry["usd_estimate"], entry["fx_rate"] = fx_usd, _fxr

    uploads = _load_uploads(username)
    uploads_dir_path = _uploads_dir(username)
    target_names = set(_get_card_member_names(username))
    matched_tx = None

    for upload in uploads:
        csv_path = uploads_dir_path / upload.get("stored_name", "")
        if not csv_path.exists():
            continue
        try:
            csv_bytes = csv_path.read_bytes()
            reader = csv.DictReader(
                io.TextIOWrapper(io.BytesIO(csv_bytes), encoding='utf-8-sig', newline=''))
            rows = [r for r in reader
                    if r.get("Card Member Name", "").strip().upper() in target_names]
        except Exception:
            continue

        for row in rows:
            try:
                amount = round(float(row.get("Amount", 0)), 2)
            except ValueError:
                continue
            if fx_usd:
                if abs(amount - fx_usd) > fx_usd * FX_TOLERANCE:
                    continue
            elif abs(amount - ramount) > 0.01:
                continue

            inv_dt = _parse_date(row.get("Date", ""))
            inv_date_str = inv_dt.strftime("%Y-%m-%d") if inv_dt else None

            if rdate is None or inv_date_str is None:
                date_match = True
            elif rdate == inv_date_str:
                date_match = True
            else:
                try:
                    from datetime import timedelta as _td
                    rd  = date.fromisoformat(rdate)
                    id_ = date.fromisoformat(inv_date_str)
                    date_match = abs((rd - id_).days) <= 1
                except Exception:
                    date_match = False

            if date_match:
                merchant = row.get("Merchant Name", "").strip()
                dba      = row.get("Merchant Doing Business As", "").strip()
                matched_tx = {
                    "date":   inv_date_str,
                    "amount": amount,
                    "vendor": dba if (dba and dba != merchant) else merchant,
                }
                break
        if matched_tx:
            break

    now_iso = datetime.now().isoformat()
    if matched_tx:
        entry["matched"]             = True
        entry["match_status"]        = "matched"
        entry["matched_at"]          = now_iso
        entry["matched_transaction"] = matched_tx
        if entry.get("ocr_date") in (None, "", "unknown") and matched_tx.get("date"):
            entry["ocr_date_original"] = entry.get("ocr_date")
            entry["ocr_date"]          = matched_tx["date"]
    else:
        entry["match_status"] = "unmatched"
        entry["matched"]      = False

    _save_ledger(username, ledger)
    return ("json", {"ok": True, "matched": bool(matched_tx), "entry": entry})


def _handle_reocr(username: str, entry_id: str):
    """POST /cardconv/ledger/<id>/reocr — re-run OCR for all entries sharing the same file_id."""
    service = _get_drive_service(username)
    if not service:
        return ("json", {"error": "Drive not connected"}, 401)
    ledger = _load_ledger(username)
    entries = ledger["entries"]

    # Find the target entry to get its file_id
    target = next((e for e in entries if e.get("id") == entry_id), None)
    if not target:
        return ("json", {"error": "not found"}, 404)
    file_id = target.get("file_id")
    if not file_id:
        return ("json", {"error": "no file_id on entry"}, 400)

    try:
        meta    = service.files().get(fileId=file_id, fields="mimeType,name,webViewLink").execute()
        mime    = meta.get("mimeType", "image/jpeg")
        content = service.files().get_media(fileId=file_id).execute()
    except Exception as e:
        return ("json", {"error": f"Drive fetch failed: {e}"}, 500)

    ocr_list = _ocr_receipt_auto(content, mime) or [{}]

    # Preserve non-OCR fields from the existing sibling entries (matched at same sub_index)
    siblings = {e.get("sub_index", 0): e for e in entries if e.get("file_id") == file_id}

    # Remove all existing entries for this file_id
    entries[:] = [e for e in entries if e.get("file_id") != file_id]

    url     = meta.get("webViewLink") or f"https://drive.google.com/file/d/{file_id}/view"
    now_iso = datetime.now().isoformat()
    updated = []
    for sub_index, ocr in enumerate(ocr_list):
        sibling = siblings.get(sub_index, {})
        # Re-OCR keeps the FIRST upload date as the "today" anchor.
        fields  = _ocr_entry_fields(ocr, upload_date=str(sibling.get("uploaded_at") or now_iso)[:10])
        entry   = {
            "id":           _sub_entry_id(file_id, sub_index),
            "file_id":      file_id,
            "sub_index":    sub_index,
            "filename":     sibling.get("filename") or meta.get("name", ""),
            "drive_url":    url,
            "mime_type":    mime,
            "match_status": sibling.get("match_status", "pending_match"),
            "matched":      sibling.get("matched", False),
            "multi_ocr":    True,
            "uploaded_at":  sibling.get("uploaded_at", now_iso),
            "synced_at":    now_iso,
            "reocr_at":     now_iso,
        }
        if sibling.get("matched_at"):
            entry["matched_at"] = sibling["matched_at"]
        if sibling.get("matched_transaction"):
            entry["matched_transaction"] = sibling["matched_transaction"]
        entry.update(fields)
        entries.append(entry)
        updated.append(entry)

    _save_ledger(username, ledger)
    return ("json", {"ok": True, "updated": updated})


def _handle_image_proxy(username: str, file_id: str, bbox: list = None):
    """GET /cardconv/receipts/image/<file_id>[?bbox=ymin,xmin,ymax,xmax] — Drive proxy with optional crop.
    Always auto-rotates via EXIF so receipts display upright.
    """
    service = _get_drive_service(username)
    if not service:
        return ("html", "<p>Drive not connected</p>", 401)
    try:
        from PIL import Image as _Img, ImageOps as _IOP
        content = service.files().get_media(fileId=file_id).execute()
        if content[:5] == b"%PDF-":
            # PDF receipt: rasterize page 1 so <img> consumers can show it;
            # fall back to the raw PDF when pymupdf is unavailable.
            try:
                import fitz
                doc = fitz.open(stream=content, filetype="pdf")
                content = doc[0].get_pixmap(matrix=fitz.Matrix(2, 2)).tobytes("png")
            except Exception:
                return ("binary", content, "application/pdf", None)
        img = _Img.open(io.BytesIO(content))
        img = _IOP.exif_transpose(img)   # auto-rotate
        img = img.convert("RGB")

        if bbox and len(bbox) == 4:
            w, h = img.size
            ymin, xmin, ymax, xmax = [max(0, min(1000, v)) for v in bbox]
            left   = int(xmin / 1000 * w)
            upper  = int(ymin / 1000 * h)
            right  = int(xmax / 1000 * w)
            lower  = int(ymax / 1000 * h)
            pad = max(6, min(w, h) // 40)
            left, upper = max(0, left - pad), max(0, upper - pad)
            right, lower = min(w, right + pad), min(h, lower + pad)
            img = img.crop((left, upper, right, lower))

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=88, optimize=True)
        return ("binary", buf.getvalue(), "image/jpeg", None)
    except Exception as e:
        return ("html", f"<p>Image load error: {e}</p>", 404)


# ── PDF export ────────────────────────────────────────────────────────────────

_DEJAVU_DIR = "/usr/share/fonts/truetype/dejavu"
# Font candidates, preferred first. A CJK font (e.g. apt install fonts-nanum on
# the server) renders Korean merchant names; DejaVu covers Latin/symbols without
# crashing on non-ASCII. Each entry is (family, regular_path, bold_path).
_PDF_FONT_CANDIDATES = [
    ("Nanum", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
              "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
    ("NotoKR", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
               "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"),
    ("DejaVu", f"{_DEJAVU_DIR}/DejaVuSans.ttf",
               f"{_DEJAVU_DIR}/DejaVuSans-Bold.ttf"),
]
_PDF_STATUS_LABEL = {"matched": "Matched", "unmatched": "Unmatched",
                     "pending_match": "Pending Match"}
_PDF_STATUS_COLOR = {"matched": (22, 163, 74), "unmatched": (220, 38, 38),
                     "pending_match": (217, 119, 6)}


def _compress_receipt_image(raw: bytes, max_w: int = 700, quality: int = 72):
    """Resize + recompress a receipt image into a small JPEG.

    Returns (jpeg_bytes, (w, h)) or None on failure. Auto-rotates via EXIF,
    caps width at max_w px, and re-encodes as JPEG to keep the PDF small.
    """
    from PIL import Image, ImageOps
    try:
        img = Image.open(io.BytesIO(raw))
        img = ImageOps.exif_transpose(img)   # honour EXIF orientation tag
        img = img.convert("RGB")             # flatten alpha/CMYK/palette for JPEG
        if img.width > max_w:
            h = max(1, round(img.height * max_w / img.width))
            img = img.resize((max_w, h), Image.LANCZOS)
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=quality, optimize=True)
        return out.getvalue(), img.size
    except Exception:
        return None


def _fetch_drive_image(service, file_id: str):
    """Fetch raw image bytes from Drive, or None if unavailable."""
    if not service or not file_id:
        return None
    try:
        return service.files().get_media(fileId=file_id).execute()
    except Exception:
        return None


def _handle_ledger_pdf(username: str, query: dict):
    """GET /cardconv/ledger/download.pdf — landscape A4, 5 cols × 2 rows = 10 per page.
    Images only (no text), EXIF auto-rotated, maximally sized with minimal gaps.
    Multi-receipt source images span extra columns.
    """
    from fpdf import FPDF

    f = _parse_filter_params(query)
    status, dfrom, dto = f["status"], f["dfrom"], f["dto"]
    rids_q = (query.get("rids", [""]) or [""])[0]
    if rids_q:
        # Explicit receipt selection (Review "download selected") — bypasses
        # the ledger filters entirely.
        want = {r for r in rids_q.split(",") if r}
        entries = [e for e in _ledger_entries(username) if e.get("id") in want]
    else:
        entries = _apply_ledger_filters(_annotate_settle_status(username, _ledger_entries(username)), status, dfrom, dto,
                                        f["card_brand"], f["usage"], f["completed"],
                                        f["merchant"], f["sort"], f["settle"])
    stats   = _ledger_stats(entries)
    service = _get_drive_service(username)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False)
    MRG  = 5          # tight margins to maximise image area
    GAP  = 1.0        # gap between cells
    PAD  = 0.5        # inner padding within each cell
    pdf.set_margins(MRG, MRG, MRG)

    FONT, unicode_ok = "Helvetica", False
    for fam, reg, bold in _PDF_FONT_CANDIDATES:
        if os.path.exists(reg) and os.path.exists(bold):
            pdf.add_font(fam, "", reg)
            pdf.add_font(fam, "B", bold)
            FONT, unicode_ok = fam, True
            break

    def S(t):
        t = "" if t is None else str(t)
        return t if unicode_ok else t.encode("latin-1", "replace").decode("latin-1")

    # ── Group entries by file_id ──
    seen: dict  = {}
    groups: list = []
    for e in entries:
        fid = e.get("file_id") or ""
        if fid and fid in seen:
            groups[seen[fid]][1].append(e)
        else:
            seen[fid] = len(groups)
            groups.append((fid, [e]))

    # ── Grid: 5 cols × 2 rows, images only ──
    N_COLS = 5
    N_ROWS = 2
    HDR_H  = 9.0      # one-line header (first page only)
    PAGE_W = 297 - 2 * MRG   # 287mm
    PAGE_H = 210 - 2 * MRG   # 200mm
    CELL_W = (PAGE_W - GAP * (N_COLS - 1)) / N_COLS    # ≈56.6mm
    ROW_H  = (PAGE_H - HDR_H - GAP * (N_ROWS - 1)) / N_ROWS  # ≈94mm (first page)
    ROW_H_CONT = (PAGE_H - GAP * (N_ROWS - 1)) / N_ROWS      # ≈99.5mm (no header)

    def cell_inner_w(span: int) -> float:
        return CELL_W * span + GAP * (span - 1)

    def _best_orientation(jpeg: bytes, dim: tuple, cell_w: float, cell_h: float):
        """Return (jpeg, dim) rotated 90° if that gives significantly better cell coverage."""
        from PIL import Image as _PILImg
        def coverage(iw, ih, cw, ch):
            dw = cw
            dh = dw * ih / iw
            if dh > ch:
                dh, dw = ch, ch * iw / ih
            return dw * dh
        normal  = coverage(dim[0], dim[1], cell_w, cell_h)
        rotated = coverage(dim[1], dim[0], cell_w, cell_h)
        if rotated > normal * 1.15:          # rotate if ≥15% better fill
            img = _PILImg.open(io.BytesIO(jpeg)).rotate(90, expand=True)
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=82, optimize=True)
            return buf.getvalue(), img.size
        return jpeg, dim

    # ── Pre-fetch + compress images ──
    img_cache: dict = {}
    for fid, elist in groups:
        if not fid:
            img_cache[fid] = None
            continue
        span   = min(len(elist), N_COLS)
        cw     = cell_inner_w(span) - 2 * PAD
        ch     = ROW_H_CONT - 2 * PAD
        max_px = max(500, int(cell_inner_w(span) * 5))
        raw    = _fetch_drive_image(service, fid) or b""
        comp   = _compress_receipt_image(raw, max_w=max_px, quality=82)
        if comp:
            jpeg, dim = comp
            jpeg, dim = _best_orientation(jpeg, dim, cw, ch)
            img_cache[fid] = (jpeg, dim)
        else:
            img_cache[fid] = None

    # ── Pack groups into grid rows ──
    grid_rows: list = []
    cur_row:   list = []
    col = 0
    for fid, elist in groups:
        span = min(len(elist), N_COLS)
        if col + span > N_COLS:
            grid_rows.append(cur_row)
            cur_row, col = [], 0
        cur_row.append((fid, elist, span))
        col += span
    if cur_row:
        grid_rows.append(cur_row)

    if not grid_rows:
        pdf.add_page()
        pdf.set_font(FONT, "", 10)
        pdf.set_xy(MRG, MRG)
        pdf.cell(0, 10, S("No receipts for the selected filter."))
        out = pdf.output()
        return ("binary", bytes(out), "application/pdf",
                f"receipts_{_export_tag(username)}_{date.today().isoformat()}.pdf")

    # ── Draw ──
    def draw_header(first: bool):
        if not first:
            return
        pdf.set_font(FONT, "", 7)
        pdf.set_text_color(130, 130, 130)
        pdf.set_xy(MRG, MRG)
        flabel = {"all": "All"}.get(status) or _PDF_STATUS_LABEL.get(status, status)
        rng = f"{dfrom or '…'} ~ {dto or '…'}"
        pdf.cell(0, HDR_H - 1, S(
            f"Receipt Ledger  ·  {username}  ·  {flabel}  ·  {rng}  ·  "
            f"Total {stats['total']}  Matched {stats['matched']}  "
            f"Unmatched {stats['unmatched']}  Pending {stats['pending_match']}"
        ))
        pdf.set_text_color(0, 0, 0)

    first_page  = True
    row_in_page = 0   # which row on the current page (0 or 1)
    y           = 0.0
    n_rows      = len(grid_rows)

    for row_idx, row in enumerate(grid_rows):
        # Start a new page when both rows are filled
        if row_in_page == 0:
            pdf.add_page()
            draw_header(first_page)
            y = MRG + (HDR_H if first_page else 0)
            first_page = False

        # Determine row height: if this is the only row on this page (last item or
        # next item would start a new page), expand to full usable height.
        is_only_row_on_page = (row_in_page == 0 and (row_idx == n_rows - 1))
        if is_only_row_on_page:
            cur_row_h = PAGE_H - (HDR_H if pdf.page == 1 else 0) - 2 * MRG
        elif row_in_page == 0 and pdf.page == 1:
            cur_row_h = ROW_H
        else:
            cur_row_h = ROW_H_CONT

        x = MRG
        for fid, elist, span in row:
            iw = cell_inner_w(span)

            # Thin border
            pdf.set_draw_color(200, 200, 200)
            pdf.set_line_width(0.15)
            pdf.rect(x, y, iw, cur_row_h)

            # Image — fill the cell as fully as possible
            comp = img_cache.get(fid)
            if comp:
                jpeg, dim = comp
                avail_w = iw - 2 * PAD
                avail_h = cur_row_h - 2 * PAD
                dw = avail_w
                dh = dw * dim[1] / dim[0]
                if dh > avail_h:
                    dh, dw = avail_h, avail_h * dim[0] / dim[1]
                try:
                    pdf.image(io.BytesIO(jpeg),
                              x=x + (iw - dw) / 2,
                              y=y + (cur_row_h - dh) / 2,
                              w=dw, h=dh)
                except Exception:
                    pass
            else:
                pdf.set_xy(x, y)
                pdf.set_font(FONT, "", 6)
                pdf.set_text_color(180, 180, 180)
                pdf.cell(iw, cur_row_h, S("no image"), align="C")
                pdf.set_text_color(0, 0, 0)

            x += iw + GAP

        y += cur_row_h + GAP
        row_in_page = (row_in_page + 1) % N_ROWS

    out  = pdf.output()
    fname = f"receipts_{_export_tag(username)}_{date.today().isoformat()}.pdf"
    _add_hist({
        "type":     "pdf_download",
        "date":     datetime.now().strftime("%Y-%m-%d %H:%M"),
        "filename": fname,
        "filter":   flabel if 'flabel' in dir() else status,
        "count":    stats.get("total", 0),
        "user":     username,
    })
    return ("binary", bytes(out), "application/pdf", fname)


# ── HTTP handler ──────────────────────────────────────────────────────────────

def _get_client_info():
    import json as _json
    data = _json.loads(CREDS_FILE.read_text())
    return data.get("installed", data.get("web", {}))


def _handle_drive_auth(username: str, body):
    raw = body.get("code", "")
    code = (raw[0] if isinstance(raw, list) else str(raw)).strip()
    if not code:
        return ("redirect", "/cardconv/ledger")
    try:
        import json as _json, urllib.request as _req, urllib.parse as _up
        c = _get_client_info()
        token_data = _up.urlencode({
            "code":          code,
            "client_id":     c["client_id"],
            "client_secret": c["client_secret"],
            "redirect_uri":  "urn:ietf:wg:oauth:2.0:oob",
            "grant_type":    "authorization_code",
        }).encode()
        req = _req.Request("https://oauth2.googleapis.com/token", data=token_data,
                           headers={"Content-Type": "application/x-www-form-urlencoded"})
        resp = _req.urlopen(req)
        token_json = _json.loads(resp.read())
        if "error" in token_json:
            raise Exception(f"{token_json['error']}: {token_json.get('error_description','')}")
        c = _get_client_info()
        save_data = {
            "token":         token_json.get("access_token"),
            "refresh_token": token_json.get("refresh_token"),
            "token_uri":     "https://oauth2.googleapis.com/token",
            "client_id":     c["client_id"],
            "client_secret": c["client_secret"],
            "scopes":        SCOPES,
        }
        _ensure_dirs()
        (TOKENS_DIR / f"{username}.json").write_text(_json.dumps(save_data))
    except Exception as e:
        return ("html", f"<p style='padding:20px;color:var(--danger)'>Auth error: {e} "
                        f"<a href='/cardconv/ledger' style='color:var(--accent)'>Back</a></p>")
    # Eagerly create Wayfinder/Receipts so the user has a place to drop files,
    # and land them on a CTA page instead of straight back to the ledger.
    folder_url = ""
    try:
        service = _get_drive_service(username)
        if service:
            rid = _get_receipts_folder_id(service, username)
            folder_url = f"https://drive.google.com/drive/folders/{rid}"
    except Exception:
        pass
    from services._cardconv_render import _render_drive_connected
    return ("html", _render_drive_connected(folder_url))


import threading as _threading, uuid as _uuid

_sync_jobs: dict = {}   # job_id → {"status": "running"|"done"|"error", "staged": int, "error": str}

def _list_new_drive_files(username: str, service=None):
    """Drive files in Receipts not yet OCR-done in ledger nor staged (listing only,
    no download/OCR). None if Drive is not connected."""
    service = service or _get_drive_service(username)
    if not service:
        return None
    receipts_id = _get_receipts_folder_id(service, username)
    results = service.files().list(
        q=(f"'{receipts_id}' in parents "
           f"and trashed=false and mimeType!='application/vnd.google-apps.folder'"),
        fields="files(id,name,mimeType,webViewLink)"
    ).execute()
    existing = _load_receipts(username)
    done_fids = {r.get('file_id') for r in existing
                 if (r.get('multi_ocr') and r.get('ocr_amount') is not None)
                 or r.get('matched')}
    # Skip files already in the staging queue, plus discarded tombstones
    # (their Drive trash may have failed — see _mark_discarded_fid).
    staged_fids = {e.get("file_id") for e in _load_ocr_staging(username).get("entries", [])}
    skip_fids = done_fids | staged_fids | set(_load_discarded_fids(username))
    supported = {'image/jpeg', 'image/png', 'application/pdf', 'image/gif', 'image/webp'}
    return [f for f in results.get('files', [])
            if f.get('mimeType', '') in supported and f.get('id') not in skip_fids]


def _handle_drive_newcount(username: str):
    """GET /cardconv/drive/newcount — pending Drive file count for the Ledger banner."""
    try:
        files = _list_new_drive_files(username)
    except Exception:
        files = None
    if files is None:
        return ("json", {"connected": False, "new": 0})
    return ("json", {"connected": True, "new": len(files)})


def _do_drive_sync_work(username: str, job_id: str):
    """Background worker: scans Drive, OCRs new files, stages for review."""
    try:
        service = _get_drive_service(username)
        if not service:
            _sync_jobs[job_id] = {"status": "error", "staged": 0, "error": "Drive not connected"}
            return

        new_files = _list_new_drive_files(username, service)
        staging = _load_ocr_staging(username)

        new_staged = []
        now_disp = datetime.now().strftime("%Y-%m-%d %H:%M")
        now_iso  = datetime.now().isoformat()

        for f in new_files:
            fid = f.get('id')
            mime = f.get('mimeType', '')
            content  = service.files().get_media(fileId=fid).execute()
            ocr_list = _ocr_receipt_auto(content, mime) or [{}]
            url      = f.get('webViewLink') or f'https://drive.google.com/file/d/{fid}/view'
            for sub_index, ocr in enumerate(ocr_list):
                entry = {
                    "id":           _sub_entry_id(fid, sub_index),
                    "file_id":      fid,
                    "sub_index":    sub_index,
                    "filename":     f.get('name', ''),
                    "drive_url":    url,
                    "mime_type":    mime,
                    "match_status": "pending_match",
                    "multi_ocr":    True,
                    "uploaded_at":  now_disp,
                    "synced_at":    now_iso,
                }
                entry.update(_ocr_entry_fields(ocr, upload_date=now_iso[:10]))
                new_staged.append(entry)

        # Accumulate — append to existing staging queue, never replace.
        if new_staged:
            staging.setdefault("entries", []).extend(new_staged)
            staging["staged_at"] = now_iso
            _save_ocr_staging(username, staging)

        ledger = _load_ledger(username)
        ledger["last_batch_at"] = now_iso
        _save_ledger(username, ledger)

        _sync_jobs[job_id] = {"status": "done", "staged": len(new_staged)}

    except Exception as e:
        _sync_jobs[job_id] = {"status": "error", "staged": 0, "error": str(e)}


def _handle_drive_sync(username: str):
    """POST /cardconv/drive/sync — start background sync, return job_id immediately."""
    if not _get_drive_service(username):
        return ("json", {"error": "Drive not connected"}, 400)
    job_id = _uuid.uuid4().hex
    _sync_jobs[job_id] = {"status": "running", "staged": 0}
    t = _threading.Thread(target=_do_drive_sync_work, args=(username, job_id), daemon=True)
    t.start()
    return ("json", {"job_id": job_id})


def _handle_drive_sync_status(username: str, query: dict):
    """GET /cardconv/drive/sync/status?job=<id> — poll sync progress."""
    job_id = (query.get("job", [""])[0]).strip()
    job = _sync_jobs.get(job_id)
    if not job:
        return ("json", {"status": "not_found"})
    return ("json", job)


def _handle_receipt_upload(username: str, body):
    raw = body.get("__raw_handler__")
    if raw is None:
        return ("redirect", "/cardconv/ledger")
    if not _is_drive_connected(username):
        return ("html", "<p style='padding:20px;color:var(--danger)'>Connect Google Drive first. "
                        "<a href='/cardconv/ledger' style='color:var(--accent)'>Back</a></p>")

    files = _parse_multipart_files(raw)
    if not files:
        return ("redirect", "/cardconv/ledger")

    # DRM guard: a NASCA receipt keeps its original .jpg/.pdf extension, so it
    # would pass the mime check and upload ciphertext to Drive. Block the batch
    # with a friendly notice instead of silently storing an unreadable file.
    drm = next((fn for fn, content, _ in files if _is_nasca_drm(content)), None)
    if drm:
        return ("drm_blocked", drm, "receipt")

    receipts  = _load_receipts(username)
    supported = {'image/jpeg', 'image/png', 'application/pdf', 'image/gif', 'image/webp'}

    staged_entries = []
    now_disp = datetime.now().strftime("%Y-%m-%d %H:%M")
    now_iso  = datetime.now().isoformat()

    for filename, content, mime_type in files:
        if mime_type not in supported:
            continue
        file_id, drive_url = _upload_file_to_drive(username, content, filename, mime_type)
        ocr_list = _ocr_receipt_auto(content, mime_type) or [{}]
        for sub_index, ocr in enumerate(ocr_list):
            entry = {
                "id":           _sub_entry_id(file_id, sub_index),
                "file_id":      file_id,
                "sub_index":    sub_index,
                "filename":     filename,
                "drive_url":    drive_url,
                "mime_type":    mime_type,
                "uploaded_at":  now_disp,
                "synced_at":    now_iso,
                "match_status": "pending_match",
                "multi_ocr":    True,
            }
            entry.update(_ocr_entry_fields(ocr, upload_date=now_iso[:10]))
            staged_entries.append(entry)

    # Stage for visual review; only confirmed entries go to the ledger.
    existing = _load_ocr_staging(username)
    existing.setdefault("entries", []).extend(staged_entries)
    existing["staged_at"] = now_iso
    _save_ocr_staging(username, existing)
    return ("redirect", "/cardconv/receipts/review")


def _handle_manual_receipt_add(username: str, body: dict):
    """Add a manually-entered receipt sub-entry to an existing Drive image in the ledger."""
    file_id   = (body.get("file_id") or "").strip()
    filename  = (body.get("filename") or "").strip()
    drive_url = (body.get("drive_url") or "").strip()
    mime_type = (body.get("mime_type") or "image/jpeg").strip()

    def _num(v):
        try: return float(v) if v not in (None, "") else None
        except: return None

    ocr_date   = (body.get("ocr_date") or "").strip() or None
    merchant   = (body.get("ocr_merchant") or "").strip() or None
    printed    = _num(body.get("ocr_printed_amount"))
    handw      = _num(body.get("ocr_handwritten_amount"))
    final_amt  = handw if handw is not None else printed

    if not file_id:
        return ("json", {"error": "file_id required"}, 400)

    entries = _load_receipts(username)
    # Assign the next sub_index for this file_id.
    existing_subs = [e.get("sub_index", 0) for e in entries if e.get("file_id") == file_id]
    sub_index = (max(existing_subs) + 1) if existing_subs else 0

    now_disp = datetime.now().strftime("%Y-%m-%d %H:%M")
    now_iso  = datetime.now().isoformat()
    entry = {
        "id":                    _sub_entry_id(file_id, sub_index),
        "file_id":               file_id,
        "sub_index":             sub_index,
        "filename":              filename,
        "drive_url":             drive_url,
        "mime_type":             mime_type,
        "uploaded_at":           now_disp,
        "synced_at":             now_iso,
        "match_status":          "pending_match",
        "multi_ocr":             True,
        "ocr_status":            "manual",
        "ocr_model":             "Manual",
        "ocr_date":              ocr_date,
        "ocr_merchant":          merchant,
        "ocr_printed_amount":    printed,
        "ocr_handwritten_amount": handw,
        "ocr_amount":            final_amt,
        "ocr_bbox":              None,
    }
    entries.append(entry)
    _save_receipts(username, entries)
    return ("json", {"ok": True, "id": entry["id"]})


def _is_nasca_drm(data: bytes) -> bool:
    """A NASCA-encrypted file still carries a plaintext magic header even though
    the body is ciphertext. Catch it before parsing so the user gets a friendly
    'convert to normal document first' notice instead of a parser crash."""
    if not data:
        return False
    head = data.lstrip()[:64].upper()
    return head.startswith(b"<## NASCA") or b"NASCA DRM FILE" in head


def _handle_upload(body, user=None):
    raw = body.get("__raw_handler__")
    if raw is None:
        return ("html", "<p>Upload error: no raw handler</p>")
    ct = raw.headers.get("Content-Type", "")
    m  = re.search(r'boundary=([^\s;]+)', ct)
    if not m:
        return ("html", "<p>Upload error: no boundary</p>")
    boundary  = ("--" + m.group(1)).encode()
    length    = int(raw.headers.get("Content-Length", 0))
    data      = raw.rfile.read(length)

    csv_bytes = None
    csv_name  = "upload.csv"
    for part in data.split(boundary):
        if b'filename="' not in part:
            continue
        fn_m = re.search(rb'filename="([^"]+)"', part)
        if fn_m:
            csv_name = fn_m.group(1).decode()
        hdr_end = part.find(b"\r\n\r\n")
        if hdr_end == -1:
            continue
        content = part[hdr_end + 4:]
        if content.endswith(b"\r\n"):
            content = content[:-2]
        csv_bytes = content
        break

    if not csv_bytes:
        return ("redirect", "/cardconv/ledger")

    if not user:
        return ("redirect", "/cardconv/ledger")

    # (No early card-name guard: with zero registered names the ingest matches
    # nothing and the 0-match notice below lists the file's cardmembers with
    # one-click register buttons — far clearer than a silent bounce.)

    # DRM guard: a still-encrypted NASCA file would otherwise blow up the parser
    # and surface as a raw red error. Bounce it to a friendly notice instead.
    if _is_nasca_drm(csv_bytes):
        return ("drm_blocked", csv_name, "convert")
    try:
        # AMEX Master xlsx / Billing Support .xls uploads are adapted to CSV
        # shape, then flow through the same pipeline. Stored converted so
        # Re-run works unchanged.
        if (csv_name.lower().endswith((".xlsx", ".xls"))
                or csv_bytes[:4] in (b"PK\x03\x04", b"\xd0\xcf\x11\xe0")):
            csv_bytes = _master_xlsx_to_csv_bytes(csv_bytes)
        stats = _ingest_csv(user, csv_bytes, csv_name)
        if not stats["added"] and not stats["dup_skipped"]:
            # Nothing matched the registered card names — show who IS in the
            # file so the user knows which name to register (강프로 2026-07-24).
            counts = Counter()
            try:
                rdr = csv.DictReader(io.TextIOWrapper(
                    io.BytesIO(csv_bytes), encoding="utf-8-sig", newline=""))
                for r in rdr:
                    nm = (r.get("Card Member Name") or "").strip().upper()
                    if nm:
                        counts[nm] += 1
            except Exception:
                pass
            if counts:
                # keep the converted CSV so the register buttons can re-ingest
                # without a second upload
                entry = _save_uploaded_csv(user, csv_bytes, csv_name, 0, "")
                return ("html", _render_member_mismatch(user, csv_name, counts,
                                                        entry["id"]))
            import urllib.parse as _up
            return ("redirect", "/cardconv/convert?ingest_empty=" + _up.quote(csv_name))
        _save_uploaded_csv(user, csv_bytes, csv_name, stats["added"], "")
        _add_hist({
            "type":        "ingest",
            "source":      csv_name,
            "rows":        stats["added"],
            "dup_skipped": stats["dup_skipped"],
            "matched":     stats["matched"],
            "date":        datetime.now().strftime("%Y-%m-%d %H:%M"),
            "user":        user,
        })
        # New transactions join the pool; Review shows every open one.
        return ("redirect", "/cardconv/review")
    except Exception as e:
        return ("html", (
            '<!DOCTYPE html><html><head><meta charset="UTF-8">'
            '<meta name="viewport" content="width=device-width,initial-scale=1">'
            '<link rel="stylesheet" href="/static/style.css"></head><body>'
            '<div class="container" style="max-width:640px;padding-top:60px">'
            '<div class="notepad-card"><div class="notepad-body" style="padding:28px">'
            f'<h2 style="font-size:1.05rem;margin-bottom:10px">⚠️ Couldn\'t read {_esc(csv_name)}</h2>'
            '<p style="font-size:.86rem;color:var(--text-muted);line-height:1.7;margin-bottom:6px">'
            'This file doesn\'t look like an AMEX statement CSV (Posted_*.csv) or AMEX Master xlsx. '
            'Please check the file and try again.</p>'
            f'<p style="font-size:.74rem;color:var(--text-muted);margin-bottom:18px">Details: {_esc(str(e))}</p>'
            '<a href="/cardconv/convert" class="btn btn-primary">← Back to Convert</a>'
            '</div></div></div></body></html>'))


def _render_member_mismatch(username: str, csv_name: str, counts,
                            uid: str = "") -> str:
    """0-match notice: which Card Member Names the file actually contains vs
    what this profile has registered. Each found name is a one-click button
    that registers it and re-ingests the stored upload."""
    registered = _get_card_member_names(username)
    reg_html = (", ".join(f"<b>{_esc(n)}</b>" for n in registered)
                if registered else "<i>none registered yet</i>")
    chip_style = ('display:inline-flex;align-items:center;gap:6px;margin:3px;'
                  'padding:6px 14px;border:1px solid var(--accent);border-radius:99px;'
                  'font-size:.8rem;background:var(--surface-2);color:var(--text);'
                  'cursor:pointer;font-family:inherit')
    chips = "".join(
        f'<form method="POST" action="/cardconv/upload/register-name" style="display:inline">'
        f'<input type="hidden" name="uid" value="{_esc(uid)}">'
        f'<input type="hidden" name="name" value="{_esc(n)}">'
        f'<button type="submit" style="{chip_style}" '
        f'title="Register this cardmember and import their transactions">'
        f'＋ {_esc(n)} <span style="color:var(--text-muted);font-size:.7rem">{c} tx</span></button>'
        f'</form>'
        for n, c in counts.most_common(40))
    more = (f'<p style="font-size:.72rem;color:var(--text-muted)">…and '
            f'{len(counts) - 40} more</p>' if len(counts) > 40 else "")
    return (
        '<!DOCTYPE html><html><head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1">'
        '<link rel="stylesheet" href="/static/style.css"></head><body>'
        '<div class="container" style="max-width:720px;padding-top:60px">'
        '<div class="notepad-card"><div class="notepad-body" style="padding:28px">'
        f'<h2 style="font-size:1.05rem;margin-bottom:10px">🪪 No rows matched your card names — {_esc(csv_name)}</h2>'
        '<p style="font-size:.86rem;color:var(--text-muted);line-height:1.7;margin-bottom:14px">'
        'The file parsed fine, but none of its cardmembers match the names registered '
        'in this card profile. <b>Click the cardmember you manage below</b> — it will be '
        'registered to this profile and their transactions imported right away.</p>'
        f'<p style="font-size:.8rem;margin-bottom:8px">Registered in this profile: {reg_html}</p>'
        '<p style="font-size:.8rem;font-weight:700;margin:14px 0 6px">Cardmembers found in this file:</p>'
        f'<div style="line-height:2.4">{chips}</div>{more}'
        '<div style="margin-top:22px"><a href="/cardconv/convert" class="btn btn-secondary">← Back to Convert</a></div>'
        '</div></div></div></body></html>')


def _handle_register_name_and_rerun(username: str, body: dict):
    """POST /cardconv/upload/register-name — one-click from the 0-match page:
    add the cardmember name to this profile, then re-ingest the stored upload."""
    name = (body.get("name", [""])[0] or "").strip().upper()
    uid = (body.get("uid", [""])[0] or "").strip()
    if not name:
        return ("redirect", "/cardconv/convert")
    s = _load_user_settings(username)
    names = s.get("card_member_names") or []
    if name not in [n.strip().upper() for n in names]:
        names.append(name)
        s["card_member_names"] = names
        _save_user_settings(username, s)
    if uid:
        return _handle_upload_rerun(username, uid)
    return ("redirect", "/cardconv/convert")


def _handle_upload_rerun(username: str, uid: str):
    """Re-ingest a stored CSV — pulls in any transactions missing from the pool
    (e.g. after pool entries were removed); duplicates are skipped as usual."""
    items = _load_uploads(username)
    entry = next((i for i in items if i.get("id") == uid), None)
    if not entry:
        return ("redirect", "/cardconv/convert")
    stored = _uploads_dir(username) / entry.get("stored_name", "")
    if not stored.exists():
        return ("redirect", "/cardconv/convert")
    try:
        fn = entry.get("filename", "upload.csv")
        stats = _ingest_csv(username, stored.read_bytes(), fn)
        _add_hist({
            "type":        "ingest",
            "source":      fn,
            "rows":        stats["added"],
            "dup_skipped": stats["dup_skipped"],
            "matched":     stats["matched"],
            "date":        datetime.now().strftime("%Y-%m-%d %H:%M"),
            "user":        username,
        })
        return ("redirect", "/cardconv/review")
    except Exception as e:
        return ("html", f"<p style='color:red;padding:20px'>Error: {e}</p>")


def _handle_upload_delete(username: str, uid: str):
    """Delete a stored CSV and its index entry."""
    items = _load_uploads(username)
    entry = next((i for i in items if i.get("id") == uid), None)
    if entry:
        stored = _uploads_dir(username) / entry.get("stored_name", "")
        try:
            if stored.exists():
                stored.unlink()
        except Exception:
            pass
        _save_uploads(username, [i for i in items if i.get("id") != uid])
    return ("redirect", "/cardconv/convert")


# ── Render helpers ─────────────────────────────────────────────────────────────

def _handle_ledger_update(username: str, entry_id: str, body: dict):
    """POST /cardconv/ledger/<id>/update — manually edit OCR fields of an entry."""
    ledger = _load_ledger(username)
    updated = False
    for e in ledger["entries"]:
        if e.get("id") != entry_id:
            continue
        for field in ("ocr_date", "ocr_merchant"):
            raw = body.get(field)
            if raw is not None:
                val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
                e[field] = val or None
        # Companions ("w/ NAME" note): normalize and mirror into the matched
        # transaction's receipt snapshot so Review and the SAP purpose follow.
        raw = body.get("ocr_companions")
        if raw is not None:
            val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
            # '__clear__' sentinel: parse_qs drops blank values, so the client
            # sends an explicit marker to clear the note.
            e["ocr_companions"] = None if val == "__clear__" else _coerce_companions(val)
            if e.get("matched"):
                pool = _load_tx_pool(username)
                for t in pool["entries"]:
                    rc = t.get("receipt") or {}
                    if rc.get("id") == entry_id:
                        rc["companions"] = e["ocr_companions"]
                        _save_tx_pool(username, pool)
                        break
        # Card brand: normalize to amex/visa/other, or clear when blank.
        # Matched ⇒ AMEX by definition (the statement is AMEX), so a matched
        # receipt cannot be flipped to Cash — unmatch first.
        raw = body.get("card_brand")
        cash_blocked = False
        if raw is not None:
            val = (raw[0] if isinstance(raw, list) else str(raw)).strip().lower()
            if val == "other" and e.get("matched"):
                cash_blocked = True
            else:
                e["card_brand"] = val if val in ("amex", "visa", "other") else None
        # Reason for Cash: receipt is the source of truth; _sync_cash_pool
        # mirrors it onto the cash pool row for Review + the SAP column S.
        raw = body.get("cash_reason")
        if raw is not None:
            val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
            e["cash_reason"] = None if val == "__clear__" else (val[:120] or None)
        # Usage: free-text tag, defaults back to "Regular" when cleared.
        raw = body.get("usage")
        if raw is not None:
            val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
            e["usage"] = val or "Regular"
        # "Not a duplicate": pins the entry out of duplicate grouping ('1'/'0').
        raw = body.get("dup_exempt")
        if raw is not None:
            val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
            e["dup_exempt"] = val == "1"
        # Printed transaction time (HH:MM) — used by duplicate detection.
        raw = body.get("ocr_time")
        if raw is not None:
            val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
            e["ocr_time"] = _coerce_time(val)
        for field in ("ocr_printed_amount", "ocr_handwritten_amount"):
            raw = body.get(field)
            if raw is not None:
                val = (raw[0] if isinstance(raw, list) else str(raw)).strip()
                try:
                    e[field] = float(val) if val else None
                except ValueError:
                    pass
        # ocr_amount always reflects handwritten-priority final amount.
        hw = e.get("ocr_handwritten_amount")
        pr = e.get("ocr_printed_amount")
        e["ocr_amount"] = hw if hw is not None else pr
        updated = True
        break
    if not updated:
        return ("json", {"error": "not found"}, 404)
    _save_ledger(username, ledger)
    return ("json", {"ok": True, "cash_blocked": cash_blocked})


def _handle_ledger_bulk(username: str, body: dict):
    """POST /cardconv/ledger/bulk — apply one action to the selected receipts.

    JSON body: {"ids": [...], "action": "card"|"usage"|"companions"|"rematch",
    "value": ...}. companions mirrors into matched transactions' snapshots."""
    ids = set(body.get("ids") or [])
    action = body.get("action")
    value = body.get("value")
    if not ids or action not in ("card", "usage", "companions", "rematch", "settle"):
        return ("json", {"error": "bad request"}, 400)

    if action == "rematch":
        res = _rematch_pool(username, only_receipt_ids=ids)
        return ("json", {"ok": True, **res})

    if action == "settle":
        # One settlement status, both surfaces: set the linked transaction's
        # status (Review) AND the receipt completed flag (Ledger + Drive
        # archive) so the two views can never disagree.
        status = str(value or "").strip()
        if status not in ("open", "in_progress", "completed"):
            return ("json", {"error": "bad status"}, 400)
        tx_touched = _set_linked_tx_status(username, ids, status)
        res = _apply_receipt_completion(username, ids, status == "completed")
        return ("json", {"ok": True, "updated": res["count"], "tx_updated": tx_touched,
                         "status": status, "moved": res["moved"], "attempted": res["attempted"]})

    comp = _coerce_companions(str(value).strip()) if (action == "companions" and value) else None
    ledger = _load_ledger(username)
    updated = 0
    cash_blocked = 0
    for e in ledger["entries"]:
        if e.get("id") not in ids:
            continue
        if action == "card":
            v = str(value or "").strip().lower()
            if v == "other" and e.get("matched"):
                cash_blocked += 1  # matched ⇒ AMEX by definition
                continue
            e["card_brand"] = v if v in ("amex", "visa", "other") else None
        elif action == "usage":
            e["usage"] = str(value or "").strip() or "Regular"
        elif action == "companions":
            e["ocr_companions"] = comp
        updated += 1
    if updated:
        _save_ledger(username, ledger)
        if action == "companions":
            pool = _load_tx_pool(username)
            dirty = False
            for t in pool["entries"]:
                rc = t.get("receipt") or {}
                if rc.get("id") in ids:
                    rc["companions"] = comp
                    dirty = True
            if dirty:
                _save_tx_pool(username, pool)
    return ("json", {"ok": True, "updated": updated, "cash_blocked": cash_blocked})


def _handle_review_manual_match(username: str, body: dict):
    """POST /cardconv/review/match — manually link an open transaction to a receipt."""
    def val(k):
        v = body.get(k, "")
        return (v[0] if isinstance(v, list) else str(v)).strip()
    row_id   = val("row_id")
    rcpt_id  = val("receipt_id")
    if not row_id or not rcpt_id:
        return ("json", {"error": "missing params"}, 400)

    receipts = _load_receipts(username)
    receipt  = next((r for r in receipts if r.get("id") == rcpt_id), None)
    if not receipt:
        return ("json", {"error": "receipt not found"}, 404)

    pool = _load_tx_pool(username)
    entry = next((e for e in pool["entries"] if e.get("id") == row_id), None)
    if not entry:
        return ("json", {"error": "row not found"}, 404)

    _apply_receipt_match(entry, receipt, receipts)
    _save_receipts(username, receipts)
    _save_tx_pool(username, pool)
    return ("json", {"ok": True, "receipt": entry["receipt"]})


def _handle_review_reason(username: str, body: dict):
    """POST /cardconv/review/reason — save a loss reason for an unmatched transaction."""
    def _val(k):
        v = body.get(k, "")
        return (v[0] if isinstance(v, list) else str(v))
    rid = _val("id").strip()
    reason = _val("reason")
    if not rid:
        return ("json", {"error": "missing id"}, 400)
    pool = _load_tx_pool(username)
    for e in pool["entries"]:
        if e.get("id") == rid:
            e["loss_reason"] = reason
            _save_tx_pool(username, pool)
            return ("json", {"ok": True})
    return ("json", {"error": "not found"}, 404)


def _handle_review_no_receipt(username: str, body: dict):
    """POST /cardconv/review/no_receipt — file a transaction as receipt-less.

    Body: {id, on: '1'|'0'}. Marks an unmatched transaction as "no receipt
    exists" so it stops counting as an open matching problem. Cleared
    automatically if a receipt does match later."""
    def _val(k):
        v = body.get(k, "")
        return (v[0] if isinstance(v, list) else str(v)).strip()
    rid = _val("id")
    on = _val("on") == "1"
    if not rid:
        return ("json", {"error": "missing id"}, 400)
    pool = _load_tx_pool(username)
    entry = next((e for e in pool["entries"] if e.get("id") == rid), None)
    if entry is None:
        return ("json", {"error": "not found"}, 404)
    if on and entry.get("matched"):
        return ("json", {"error": "이미 영수증이 매칭된 거래입니다"}, 400)
    entry["no_receipt"] = on
    _save_tx_pool(username, pool)
    return ("json", {"ok": True})


def _handle_review_companions(username: str, body: dict):
    """POST /cardconv/review/companions — set the w/ note on a pool transaction.

    Lets Review annotate receipt-less transactions. Matched rows edit the
    ledger receipt instead (single source of truth), so this only writes to
    the pool entry; the note carries over if a receipt matches later."""
    def _val(k):
        v = body.get(k, "")
        return (v[0] if isinstance(v, list) else str(v)).strip()
    rid = _val("id")
    if not rid:
        return ("json", {"error": "missing id"}, 400)
    pool = _load_tx_pool(username)
    entry = next((e for e in pool["entries"] if e.get("id") == rid), None)
    if entry is None:
        return ("json", {"error": "not found"}, 404)
    entry["companions"] = _coerce_companions(_val("companions"))
    _save_tx_pool(username, pool)
    return ("json", {"ok": True})


def _handle_review_cash_reason(username: str, body: dict):
    """POST /cardconv/review/cash_reason — set the 'Reason for Cash' on a
    cash pool row. Flows into the SAP xlsx column S on export."""
    def _val(k):
        v = body.get(k, "")
        return (v[0] if isinstance(v, list) else str(v)).strip()
    rid = _val("id")
    if not rid:
        return ("json", {"error": "missing id"}, 400)
    pool = _load_tx_pool(username)
    entry = next((e for e in pool["entries"] if e.get("id") == rid), None)
    if entry is None:
        return ("json", {"error": "not found"}, 404)
    if not entry.get("cash"):
        return ("json", {"error": "not a cash row"}, 400)
    reason = _val("reason")[:120] or None
    entry["cash_reason"] = reason
    _save_tx_pool(username, pool)
    # Receipt is the source of truth (_sync_cash_pool mirrors receipt → pool
    # on every load) — write it there too or the edit reverts next render.
    rid = (entry.get("receipt") or {}).get("id")
    if rid:
        ledger = _load_ledger(username)
        for e in ledger["entries"]:
            if e.get("id") == rid:
                e["cash_reason"] = reason
                _save_ledger(username, ledger)
                break
    return ("json", {"ok": True})


def _handle_review_usage(username: str, body: dict):
    """POST /cardconv/review/usage — set the usage tag on a pool transaction.

    Lets Review tag receipt-less transactions. When the row is matched the
    ledger receipt stays the source of truth, so it is updated in the same
    call and both surfaces agree."""
    def _val(k):
        v = body.get(k, "")
        return (v[0] if isinstance(v, list) else str(v)).strip()
    rid   = _val("id")
    usage = _val("usage") or "Regular"
    if not rid:
        return ("json", {"error": "missing id"}, 400)
    pool = _load_tx_pool(username)
    entry = next((e for e in pool["entries"] if e.get("id") == rid), None)
    if entry is None:
        return ("json", {"error": "not found"}, 404)
    entry["usage"] = usage
    _save_tx_pool(username, pool)
    rcpt_id = (entry.get("receipt") or {}).get("id")
    if entry.get("matched") and rcpt_id:
        ledger = _load_ledger(username)
        for e in ledger["entries"]:
            if e.get("id") == rcpt_id:
                e["usage"] = usage
                _save_ledger(username, ledger)
                break
    return ("json", {"ok": True})


def _handle_review_set_status(username: str, body: dict):
    """POST /cardconv/review/status — set open / in_progress / completed.

    in_progress marks transactions submitted to SAP and awaiting approval,
    so the settlement state is trackable between upload and sign-off."""
    raw = body.get("ids", [])
    ids = {str(i) for i in (raw if isinstance(raw, list) else [raw]) if i}
    status = str(body.get("status", "")).strip()
    if not ids:
        return ("json", {"error": "no ids"}, 400)
    if status not in ("open", "in_progress", "completed"):
        return ("json", {"error": "bad status"}, 400)
    pool = _load_tx_pool(username)
    now = datetime.now().isoformat()
    touched = 0
    rids = set()
    for e in pool["entries"]:
        if e.get("id") in ids:
            e["status"] = status
            e["completed_at"] = now if status == "completed" else None
            touched += 1
            rid = (e.get("receipt") or {}).get("id")
            if rid:
                rids.add(rid)
    _save_tx_pool(username, pool)
    # Review → Ledger mirror: matched receipts follow the transaction.
    if rids:
        _apply_receipt_completion(username, rids, status == "completed")
    return ("json", {"ok": True, "touched": touched, "status": status})


def _handle_review_complete(username: str, body: dict):
    """POST /cardconv/review/complete — mark transactions completed (or undo).

    Body: {"ids": [...], "undo": bool}. Completed transactions leave the default
    Review view and the xlsx export; git-style history lives in the pool file."""
    raw = body.get("ids", [])
    ids = {str(i) for i in (raw if isinstance(raw, list) else [raw]) if i}
    if not ids:
        return ("json", {"error": "no ids"}, 400)
    undo = bool(body.get("undo"))
    pool = _load_tx_pool(username)
    now = datetime.now().isoformat()
    touched = 0
    rids = set()
    for e in pool["entries"]:
        if e.get("id") in ids:
            e["status"] = "open" if undo else "completed"
            e["completed_at"] = None if undo else now
            touched += 1
            rid = (e.get("receipt") or {}).get("id")
            if rid:
                rids.add(rid)
    _save_tx_pool(username, pool)
    # Review → Ledger mirror: matched receipts follow the transaction.
    if rids:
        _apply_receipt_completion(username, rids, not undo)
    return ("json", {"ok": True, "touched": touched, "undo": undo})


def _select_review_entries(username: str, query: dict) -> list:
    """Transactions targeted by a Review download: an explicit `ids` selection,
    or all OPEN transactions within the optional from/to date range (dateless
    rows always kept, matching the Ledger/Review convention)."""
    pool = _load_tx_pool(username)
    ids_q = (query.get("ids", [""]) or [""])[0]
    if ids_q:
        want = {i for i in ids_q.split(",") if i}
        return [e for e in pool["entries"] if e.get("id") in want]
    dfrom = (query.get("from", [""]) or [""])[0]
    dto   = (query.get("to", [""]) or [""])[0]
    entries = [e for e in pool["entries"] if e.get("status", "open") == "open"]
    if dfrom or dto:
        entries = [e for e in entries
                   if not e.get("date")
                   or ((not dfrom or e["date"] >= dfrom) and (not dto or e["date"] <= dto))]
    return entries


def _handle_review_download(username: str, query: dict):
    """GET /cardconv/review/download — xlsx of open (or explicitly selected)
    transactions, built on demand from the pool."""
    # Cash rows (synthetic pool mirrors of cash receipts) ARE exported: the
    # SAP template has a 'Reason for Cash' column for exactly these lines
    # (2026-07-21, 강프로 — reversal of the earlier exclusion).
    entries = _select_review_entries(username, query)
    if not entries:
        return ("html", "<h2 style='padding:40px'>No open transactions to download.</h2>", 404)
    # Receipt Type (col A): D = AMEX statement charge; everything else —
    # cash, personal Visa, any non-AMEX method — is reimbursed as cash → A.
    # Reason for Cash lives on the receipt (OCR confirm / Ledger edit) — read
    # it live so a fresh edit exports even before a Review render mirrors it.
    rcpts = {r.get("id"): r for r in _load_receipts(username)}
    for e in entries:
        rc = rcpts.get((e.get("receipt") or {}).get("id"))
        rb = (rc or {}).get("card_brand") or ""
        e["receipt_type"] = "A" if (e.get("cash") or rb in ("visa", "other")) else "D"
        if e.get("cash") and rc is not None:
            e["cash_reason"] = rc.get("cash_reason") or e.get("cash_reason")
    try:
        xlsx_bytes, out_fn = _build_xlsx_from_entries(entries, username)
    except FileNotFoundError as e:
        return ("html", f"<h2 style='padding:40px'>{e}</h2>", 404)
    return ("file_inline", xlsx_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", out_fn)


def _handle_review_pdf(username: str, query: dict):
    """GET /cardconv/review/download.pdf — receipt-image PDF for the matched
    receipts of the targeted transactions (selected ids, or all open)."""
    entries = _select_review_entries(username, query)
    rids = [(e.get("receipt") or {}).get("id") for e in entries if e.get("matched")]
    rids = [r for r in rids if r]
    if not rids:
        return ("html", "<h2 style='padding:40px'>No matched receipts among the "
                        "targeted transactions.</h2>", 404)
    return _handle_ledger_pdf(username, {"rids": [",".join(rids)]})


# GL account → human category for the expense report (fallback: Other).
GL_CATEGORY = {
    "53270377": "Transportation",
    "53410177": "Meals",
    "53470177": "Meals",
    "53410103": "Client",
    "53210177": "Office",
    "53290177": "Office",
    "53311577": "Subscription",
}


def _handle_expense_report(username: str, query: dict):
    """GET /cardconv/review/expense_report — human-readable report:
    summary table + one labeled receipt-image block per item (modeled on the
    reference trip report; single USD column, single Rcpt flag, no FX)."""
    entries = _select_review_entries(username, query)
    if not entries:
        return ("html", "<h2 style='padding:40px'>No transactions to export.</h2>", 404)
    xlsx = _build_expense_report(username, entries)
    # Same naming convention as the receipt PDF this report replaces.
    fn = f"receipts_{_export_tag(username)}_{date.today().isoformat()}.xlsx"
    return ("file_inline", xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fn)


def _build_expense_report(username: str, entries: list) -> bytes:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font
    from PIL import Image as PILImage
    from PIL import ImageOps

    entries = sorted(entries, key=lambda e: e.get("date") or "", reverse=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    for col, w in {"A": 26, "B": 12, "C": 16, "D": 36, "E": 30, "F": 7, "G": 12}.items():
        ws.column_dimensions[col].width = w

    bold = Font(bold=True)
    ws["A1"] = "EXPENSE REPORT"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:G1")
    dates = sorted(e["date"] for e in entries if e.get("date"))
    ws["A2"], ws["B2"] = "Dates", (f"{dates[0]} → {dates[-1]}" if dates else "")
    ws["A3"], ws["B3"] = "Company", "Cheil"
    ws["A2"].font = ws["A3"].font = bold

    def category(e) -> str:
        return GL_CATEGORY.get(str(e.get("gl", "")), "Other")

    def purpose(e) -> str:
        p = str(e.get("purpose", "") or "")
        comp = (e.get("receipt") or {}).get("companions") or e.get("companions") or ""
        return f"{p} w/ {comp}" if comp else p

    r = 5
    ws.cell(row=r, column=1, value=f"EXPENSES    {len(entries)} items").font = bold
    r += 1
    for ci, h in enumerate(["#", "Date", "Category", "Purpose", "Vendor", "Rcpt", "USD"], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = bold
    total = 0.0
    for n, e in enumerate(entries, 1):
        r += 1
        try:
            amt = float(e.get("amount", 0))
        except (TypeError, ValueError):
            amt = 0.0
        total += amt
        has_rcpt = bool(e.get("matched") and (e.get("receipt") or {}).get("file_id"))
        for ci, v in enumerate([n, e.get("date", ""), category(e), purpose(e),
                                e.get("merchant", ""), "✓" if has_rcpt else "—", amt], 1):
            ws.cell(row=r, column=ci, value=v)
        ws.cell(row=r, column=7).number_format = "#,##0.00"
    r += 1
    ws.cell(row=r, column=6, value="TOTAL").font = bold
    tc = ws.cell(row=r, column=7, value=round(total, 2))
    tc.font = bold
    tc.number_format = "#,##0.00"

    # ── Receipt blocks: 2-across grid (label line above each image) so the
    #    sheet doesn't scroll forever — was one full-width image per row ──
    r += 2
    ws.cell(row=r, column=1, value="RECEIPTS    2 per row").font = bold
    service = _get_drive_service(username)
    img_refs = []  # keep BytesIO alive until save

    # This report replaces the receipt PDF, so images must stay readable —
    # encode as large as the per-image share of a ~10 MB workbook allows,
    # stepping down the ladder only when a photo is too heavy.
    n_imgs = sum(1 for e in entries
                 if e.get("matched") and (e.get("receipt") or {}).get("file_id"))
    per_img_budget = 9_000_000 // max(1, n_imgs)
    LADDER = ((460, 520, 85), (460, 460, 72), (440, 390, 68), (360, 320, 62), (300, 260, 55))

    def encode(im):
        for w, h, q in LADDER:
            t = im.copy()
            t.thumbnail((w, h))
            buf = io.BytesIO()
            t.save(buf, format="JPEG", quality=q)
            if buf.tell() <= per_img_budget:
                break
        buf.seek(0)
        return t, buf

    # Anchor at B/E (not A) so the grid sits off the left edge, and the pair
    # reads as one block rather than two far-apart columns.
    ANCHORS = ((2, 4, "B"), (5, 7, "E"))  # (label span start, span end, image col)
    for i in range(0, len(entries), 2):
        label_row, img_row = r + 1, r + 2
        max_h = 60.0
        for slot, e in enumerate(entries[i:i + 2]):
            col0, col1, img_col = ANCHORS[slot]
            label = (f"#{i + slot + 1} · {e.get('date', '')} · {e.get('merchant', '')} · "
                     f"{category(e)} · {e.get('amount', '')} USD\n{purpose(e)}")
            ws.merge_cells(start_row=label_row, start_column=col0,
                           end_row=label_row, end_column=col1)
            c = ws.cell(row=label_row, column=col0, value=label)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = Font(bold=True, size=9)
            fid = (e.get("receipt") or {}).get("file_id") if e.get("matched") else None
            raw = _fetch_drive_image(service, fid) if fid else None
            if not raw:
                ws.cell(row=img_row, column=col0,
                        value="(no receipt image)" if not fid else "(image unavailable)")
                continue
            try:
                im = ImageOps.exif_transpose(PILImage.open(io.BytesIO(raw)))
                im, buf = encode(im.convert("RGB"))
                xi = XLImage(buf)
                xi.width, xi.height = im.width, im.height
                ws.add_image(xi, f"{img_col}{img_row}")
                img_refs.append(buf)
                max_h = max(max_h, im.height * 0.75 + 8)  # px→pt
            except Exception:
                ws.cell(row=img_row, column=col0, value="(image decode failed)")
        ws.row_dimensions[label_row].height = 28
        ws.row_dimensions[img_row].height = max_h
        r = img_row + 1  # one default-height spacer row between pairs

    r += 2
    ws.cell(row=r, column=1,
            value=f"Generated by Cheil AMEX Expense Assistant · {date.today().isoformat()}")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


_CARD_BRAND_LABEL = {"amex": "AMEX", "visa": "Visa", "other": "Other"}


def _handle_ledger_xlsx(username: str, query: dict):
    """GET /cardconv/ledger/download.xlsx — filtered ledger rows as a settlement xlsx.

    Maps the currently-filtered Ledger entries onto the same template `convert`
    uses, sourcing date/vendor/amount from each receipt (matched-transaction
    values preferred when present). Card Type and Usage are appended as columns
    28/29 for visibility. Honors all Ledger filters via the shared parser.
    """
    f = _parse_filter_params(query)
    entries = _apply_ledger_filters(_annotate_settle_status(username, _ledger_entries(username)),
                                    f["status"], f["dfrom"], f["dto"],
                                    f["card_brand"], f["usage"], f["completed"],
                                    f["merchant"], f["sort"], f["settle"])
    if TEMPLATE.exists():
        template_path = TEMPLATE
    elif TEMPLATE_FALLBACK.exists():
        template_path = TEMPLATE_FALLBACK
    else:
        return ("html", "<h2 style='padding:40px'>Template file not found.</h2>", 404)

    rules = _load_kw()
    today = date.today()
    posting_dt = datetime(today.year, today.month, today.day)
    wb = openpyxl.load_workbook(template_path)
    ws = wb["sheetMst"]
    # Extra columns for receipt-centric metadata.
    ws.cell(1, 28).value = "Card Type"
    ws.cell(1, 29).value = "Usage"

    start = 2
    while ws.cell(start, 1).value is not None:
        start += 1

    for e in entries:
        mt       = e.get("matched_transaction") or {}
        vendor   = mt.get("vendor") or e.get("ocr_merchant") or ""
        amount   = mt.get("amount")
        if amount is None:
            amount = e.get("ocr_amount") or 0.0
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            amount = 0.0
        date_str = mt.get("date") or e.get("ocr_date")
        inv_dt   = _parse_date(date_str) if date_str else None

        gl, ser, purpose = _classify(vendor, rules)
        if gl is None:
            gl, ser, purpose = 53410177, "160", "Coffee, Snack and meal"

        ws.cell(start,  1).value = FIXED["receipt_type"]
        ws.cell(start,  2).value = FIXED["employee_id"]
        ws.cell(start,  3).value = FIXED["payee"]
        ws.cell(start,  5).value = inv_dt
        ws.cell(start,  6).value = FIXED["domestic"]
        ws.cell(start,  7).value = vendor
        ws.cell(start,  8).value = posting_dt
        ws.cell(start,  9).value = gl
        ws.cell(start, 10).value = ser
        ws.cell(start, 11).value = FIXED["currency"]
        ws.cell(start, 12).value = FIXED["tax_code"]
        ws.cell(start, 14).value = amount
        ws.cell(start, 15).value = FIXED["cost_center"]
        ws.cell(start, 18).value = purpose
        ws.cell(start, 26).value = amount
        ws.cell(start, 28).value = _CARD_BRAND_LABEL.get(e.get("card_brand"), "")
        ws.cell(start, 29).value = e.get("usage") or "Regular"
        start += 1

    buf = io.BytesIO()
    wb.save(buf)
    out_fn = f"ledger_export_{_export_tag(username)}_{today.strftime('%Y-%m-%d')}.xlsx"
    return ("file_inline", _inline_to_shared_strings(buf.getvalue()),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", out_fn)


def _esc(s) -> str:
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;")) if s is not None else ""




# Auto re-export every module-level name (incl _underscore) for `import *`.
__all__ = [k for k in list(globals()) if not k.startswith('__')]
